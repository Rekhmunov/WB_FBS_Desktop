from __future__ import annotations

from collections.abc import Mapping
from datetime import UTC, datetime, timedelta
from html import escape
import json
import ipaddress
import io
import csv
import logging
from pathlib import Path
import random
import re
import secrets
import threading
import time
from urllib.parse import urlparse
import urllib.error
import urllib.request

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from .auth import create_session_token, hash_password, verify_password

_log = logging.getLogger(__name__)
from .config import AppConfig, load_app_config, sync_chats_enabled
from .repository import ReviewRepository
from .service import MarketplaceSyncError, ReviewAutomationService, _normalize_timestamp, _parse_ozon_message_text, _wb_image_url
from .models import ReviewInput
from .stock_service import StockScheduler, sync_stock_source
from . import wb_fbs as wb_fbs_mod

try:  # pragma: no cover - optional in sqlite-only environments
    import psycopg  # type: ignore
except Exception:  # pragma: no cover
    psycopg = None

def _apply_template_substitution(
    raw_tpl: str,
    author: str,
    vars_ctx: dict[str, object],
) -> str:
    """Apply author placeholders and custom variable substitution to a template string.

    Strips any remaining unreplaced %VAR% tokens at the end.
    """
    if not raw_tpl:
        return raw_tpl
    if not author:
        for placeholder in ("%USER%", "%AUTHOR%"):
            raw_tpl = raw_tpl.replace(f", {placeholder}", "").replace(f" {placeholder}", "")
    raw_tpl = raw_tpl.replace("%USER%", author).replace("%AUTHOR%", author)
    for key, value in vars_ctx.items():
        raw_tpl = raw_tpl.replace(str(key), str(value or ""))
    return re.sub(r'%[A-Z0-9_]{2,50}%', '', raw_tpl)


CATEGORIES = [
    "negative_delivery",
    "negative_product",
    "negative_other",
    "positive_quality",
    "positive_product",
    "neutral_other",
]

BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATES_DIR = BASE_DIR / "web_templates"
STATIC_DIR = BASE_DIR / "web_static"

TEMPLATE_GROUPS: list[dict[str, object]] = [
    {
        "id": "positive",
        "title": "Позитив",
        "subgroups": [
            "Общий",
            "Вкус",
            "Материал",
            "Общий позитив",
            "Позитив доставка",
            "Позитив запах",
            "Позитив конструкция",
            "Позитив упаковка",
            "Позитив цвет",
            "Эффект",
        ],
    },
    {
        "id": "product_dissatisfaction",
        "title": "Недовольство товаром",
        "subgroups": [
            "Общий",
            "Брак и Б/У",
            "Высокая цена",
            "Качество",
            "Негатив запах",
            "Негатив конструкция",
            "Негатив цвет",
            "Не подошел лично мне",
            "Не соответствует фото",
            "Не устраивает эффект",
            "Общий негатив",
            "Побочные эффекты",
            "Подделка",
            "Срок годности",
            "Текстура, консистенция, материал",
        ],
    },
    {
        "id": "delivery_problems",
        "title": "Проблемы при доставке",
        "subgroups": [
            "Общий",
            "Долгая доставка",
            "Испорченная упаковка",
            "Наклейка",
            "Недостающая упаковка / грязное / поврежденное и сломанное",
            "Некомплект",
            "Не тот товар",
            "Общие доставка",
        ],
    },
    {
        "id": "wrong_size",
        "title": "Неправильный размер",
        "subgroups": [
            "Общий",
            "Альтернативные измерения",
            "Большемерит/маломерит",
            "Не подошел размер",
        ],
    },
    {
        "id": "textless_ratings",
        "title": "Оценки без текста",
        "subgroups": [
            "1-3 звезды",
            "4-5 звезд",
        ],
    },
]

TEXTLESS_RATINGS_GROUP_ID = "textless_ratings"
TEXTLESS_LOCKED_SUBGROUPS: tuple[str, ...] = ("1-3 звезды", "4-5 звезд")
GENERAL_LOCKED_SUBGROUP = "Общий"
GENERAL_LOCKED_GROUP_IDS: tuple[str, ...] = (
    "positive",
    "product_dissatisfaction",
    "delivery_problems",
    "wrong_size",
)

DEFAULT_TEMPLATE_CONTENT: dict[str, list[str]] = {
    "Общий": ["Спасибо за ваш отзыв! Мы ценим обратную связь и уже работаем над улучшениями."],
    "Вкус": [
        "%USER%, добрый день. Мы рады, что вы довольны покупкой. Попробуйте еще %%RECO%% — вам точно понравится!",
        "Добрый день, %USER%! Благодарим за доверие и внимание к вкусу нашего продукта.",
        "Здравствуйте, %USER%! Спасибо за высокую оценку. Будем ждать вас снова!",
    ],
    "Материал": ["Спасибо за отзыв! Рады, что материал вам понравился."],
    "Общий позитив": [
        "Приветствуем, %USER%! Благодарим вас за высокую оценку нашей продукции. Ваше мнение очень важно для нас. С уважением, %BRAND%",
        "Здравствуйте, %USER%! Мы искренне благодарны за ваше время и положительный отзыв о нашем продукте. Спасибо, что выбираете нас. Желаем вам прекрасного дня!",
        "Здравствуйте! Спасибо за ваше доверие и положительный отзыв о нашем продукте. Хорошего вам дня! С уважением, команда бренда %BRAND%",
        "%USER%, добрый день! Благодарим за отзыв! Мы всегда рады помочь вам. Прекрасного вам дня!",
        "%USER%, добрый день! Мы рады, что вы остались довольны нашим брендом. Спасибо за вашу поддержку! С надеждой на ваши будущие покупки, %BRAND%",
        "%USER%, добрый день. Мы стараемся предлагать только качественные товары. Прекрасного дня!",
        "%USER%, добрый день! Большое спасибо за приятные слова о нашем бренде %BRAND%. Это важно для нас! С надеждой на ваши будущие покупки, %BRAND%",
        "%USER%, добрый день! Мы рады, что вы довольны нашей продукцией. Благодарим за высокую оценку! С надеждой на ваши будущие покупки, %BRAND%",
        "Благодарим вас за добрые слова и оценку нашего продукта.",
        "Добрый день, спасибо за обратную связь! Рады, что наш продукт оправдал ваши ожидания. Мы дорожим мнением каждого покупателя!",
        "Добрый день! Спасибо за вашу поддержку! Надеемся на долгосрочное сотрудничество. Хорошего настроения!",
        "Добрый день, спасибо за обратную связь! Спасибо за отличный отзыв и оценку нашего продукта.",
        "Добрый день! Благодарим вас за хороший отзыв о нашем продукте. Надеемся, что он приносит вам удовольствие! Мы дорожим мнением каждого покупателя!",
        "Добрый день! Благодарим за ваш отзыв. Мы рекомендуем попробовать %RECO% — это отличный вариант для новых открытий!",
        "Добрый день! Ваше мнение важно для нас и помогает нам развиваться. С уважением, ваш %BRAND%!",
        "Добрый день! Мы рады, что наш товар приносит вам удовольствие. Спасибо за отзыв. Прекрасного дня!",
        "Добрый день! Мы рады, что наша продукция соответствует вашим ожиданиям. Спасибо за отзыв!",
        "Добрый день! Мы ценим ваш отзыв и поддержку. Рекомендуем попробовать и %RECO% — это может принести вам новые впечатления.",
        "Добрый день! Мы ценим ваше доверие к нашему бренду. Попробуйте и %RECO% — вы останетесь довольны результатом.",
        "Добрый день! Мы ценим ваше мнение о нашей продукции. Спасибо за положительный отзыв.",
        "Добрый день! Мы ценим вашу поддержку. Рекомендуем вам попробовать %RECO% — это наш бестселлер!",
        "Добрый день. Мы рады, что вы оценили качество нашего товара. Спасибо за высокую оценку. С уважением, ваш %BRAND%!",
        "Добрый день. Спасибо за ваше доброе отношение к нашему бренду. Ваши слова — лучшая награда для нас. Хорошего дня!",
        "Здравствуйте, %USER%! Очень приятно получить от вас такой отзыв. Спасибо за высокую оценку! С надеждой на ваши будущие покупки, %BRAND%",
        "Здравствуйте, спасибо за обратную связь! Ваше доверие — наше главное признание. Прекрасного вам дня!",
        "Здравствуйте, спасибо за обратную связь! Ваш отзыв вдохновляет нас на новые достижения. С уважением, ваш %BRAND%!",
        "Здравствуйте, спасибо за обратную связь! Мы рады, что вы довольны качеством нашей продукции. Спасибо за отличный отзыв.",
        "Здравствуйте, спасибо за обратную связь! Мы рады, что вы оценили качество нашего товара. Спасибо за высокую оценку. Прекрасного дня!",
        "Здравствуйте, спасибо за обратную связь! Огромное спасибо за положительную оценку и поддержку нашего бренда %BRAND%!",
        "Здравствуйте! Благодарим за вашу поддержку и хороший отзыв о продукции бренда!",
        "Здравствуйте! Благодарим за положительный отзыв о нашем бренде %BRAND%! Мы дорожим мнением каждого покупателя!",
        "Здравствуйте! Большое спасибо за вашу поддержку и высокую оценку нашего бренда %BRAND%!",
        "Здравствуйте! Благодарим за отзыв!",
        "Здравствуйте! Мы рады, что вы остались довольны нашим товаром. Спасибо за отзыв! Мы дорожим мнением каждого покупателя!",
        "Здравствуйте! Мы ценим ваш отзыв и благодарим за высокую оценку нашего продукта. Хорошего дня!",
        "Здравствуйте! Мы ценим ваш отзыв. Попробуйте еще %RECO% — это один из самых популярных товаров бренда.",
        "Здравствуйте! Мы ценим ваше мнение о нашей продукции. Спасибо за отзыв!",
        "Здравствуйте! Рады, что вы остались довольны нашей продукцией. Спасибо за отзыв!",
        "Мы рады, что вы выбрали именно наш бренд. Спасибо за вашу поддержку!",
        "Мы всегда стараемся делать качественные товары. Спасибо за ваш отзыв!",
        "Приветствуем! Благодарим за вашу поддержку и добрые слова о нашем бренде!",
        "Приветствуем! Большое спасибо за высокую оценку нашего бренда!",
        "Приветствуем! Мы рады, что вам нравится наш продукт. Не упустите шанс попробовать и %RECO% — это один из лучших товаров бренда.",
        "Приветствуем! Мы рады, что вы оценили нашу продукцию. Если вы хотите разнообразить выбор, попробуйте %RECO%.",
        "Приветствуем! Мы ценим ваше мнение и благодарим за высокую оценку. Не забудьте попробовать %RECO% — это хит продаж!",
        "Здравствуйте! Очень приятно получить такой отзыв. Спасибо за добрые слова о нашей продукции!",
        "Здравствуйте! Спасибо за вашу поддержку! Мы всегда рады видеть вас в числе наших клиентов. Прекрасного вам дня!",
        "Здравствуйте! Спасибо за добрые слова и поддержку. Прекрасного дня!",
        "Добрый день! Рады, что вы оценили качество нашего товара. Спасибо за вашу поддержку.",
        "Добрый день! Спасибо за высокую оценку нашего продукта! Ваш отзыв — лучшая награда для нас.",
    ],
    "Позитив доставка": ["Спасибо! Очень рады, что доставка прошла отлично."],
    "Позитив запах": ["Спасибо за отзыв! Приятно, что аромат вам понравился."],
    "Позитив конструкция": ["Спасибо! Рады, что конструкция товара вам подошла."],
    "Позитив упаковка": ["Спасибо! Рады, что упаковка вам понравилась."],
    "Позитив цвет": ["Спасибо за высокую оценку! Рады, что цвет вам подошел."],
    "Эффект": ["Спасибо за отзыв! Рады, что вы заметили хороший эффект."],
    "Брак и Б/У": ["Нам очень жаль, что вы получили товар в таком состоянии. Уже разбираемся."],
    "Высокая цена": ["Спасибо за обратную связь. Учтем ваш комментарий по стоимости."],
    "Качество": ["Нам жаль, что качество не оправдало ожиданий. Передали информацию в отдел качества."],
    "Негатив запах": ["Сожалеем о ситуации. Проверим партию и вернемся с ответом."],
    "Негатив конструкция": ["Спасибо за сигнал. Мы уже передали информацию в отдел разработки."],
    "Негатив цвет": ["Сожалеем, что цвет не совпал с ожиданиями. Проверим карточку товара."],
    "Не подошел лично мне": ["Спасибо за отзыв. Нам жаль, что товар вам не подошел."],
    "Не соответствует фото": ["Сожалеем о несоответствии. Передали информацию для проверки карточки."],
    "Не устраивает эффект": ["Спасибо за отзыв. Передали ваше замечание технологам."],
    "Общий негатив": ["Нам очень жаль, что вы остались недовольны. Уже разбираемся с ситуацией."],
    "Побочные эффекты": ["Сожалеем о ситуации. Рекомендуем прекратить использование и написать нам в поддержку."],
    "Подделка": ["Спасибо за сигнал. Мы проведем дополнительную проверку партии."],
    "Срок годности": ["Спасибо за отзыв. Мы проверим товар и условия хранения."],
    "Текстура, консистенция, материал": ["Спасибо за обратную связь. Передали замечание в отдел качества."],
    "Долгая доставка": ["Сожалеем о задержке доставки. Проверим логистику по вашему заказу."],
    "Испорченная упаковка": ["Нам очень жаль. Передали информацию в логистику и отдел упаковки."],
    "Наклейка": ["Спасибо за сигнал. Проверим корректность маркировки."],
    "Недостающая упаковка / грязное / поврежденное и сломанное": [
        "Сожалеем о состоянии товара. Уже разбираемся и улучшим контроль отгрузки."
    ],
    "Некомплект": ["Сожалеем о неполной комплектации. Мы уже передали информацию на склад."],
    "Не тот товар": ["Нам жаль, что пришел не тот товар. Уже разбираемся с отгрузкой."],
    "Общие доставка": ["Спасибо за отзыв о доставке. Учтем замечание и исправим процесс."],
    "Альтернативные измерения": ["Спасибо за отзыв. Дополним информацию по размерам в карточке товара."],
    "Большемерит/маломерит": ["Сожалеем, что размер не подошел. Передадим замечание по размерной сетке."],
    "Не подошел размер": ["Спасибо за обратную связь. Учтем это при обновлении размерной таблицы."],
    "Общие теги": [
        "Спасибо за оценку и выбор тегов {теги}! Нам очень приятно, что вы отметили эти преимущества.",
        "Благодарим за отзыв с тегами {теги}. Ваши отметки помогают нам становиться лучше!",
    ],
    "1-3 звезды": ["Спасибо за оценку. Нам важно ваше мнение — мы улучшаем сервис каждый день."],
    "4-5 звезд": [
        "Спасибо за высокую оценку! Будем рады снова видеть вас среди покупателей.",
        "Спасибо за 5 звезд! Очень рады, что вам все понравилось.",
    ],
}


def _is_protected_default_subgroup(group_id: str, subgroup: str) -> bool:
    clean_group = str(group_id or "").strip()
    clean_subgroup = str(subgroup or "").strip()
    if clean_group == TEXTLESS_RATINGS_GROUP_ID and clean_subgroup in TEXTLESS_LOCKED_SUBGROUPS:
        return True
    if clean_group in GENERAL_LOCKED_GROUP_IDS and clean_subgroup == GENERAL_LOCKED_SUBGROUP:
        return True
    return False


class SyncRequest(BaseModel):
    account_id: int | None = Field(default=None, description="Specific marketplace account ID")
    all_accounts: bool = Field(default=True, description="Sync all active accounts")
    account_ids: list[int] | None = Field(default=None, description="Specific account IDs to sync (from preview checkboxes)")
    total_expected: int | None = Field(default=None, ge=0, description="Expected total items from preview")


class SalaryRatesRequest(BaseModel):
    rate_review: float = Field(ge=0, default=0.0, description="Ставка за один обработанный отзыв (вручную)")
    rate_question: float = Field(ge=0, default=0.0, description="Ставка за один обработанный вопрос")
    rate_chat: float = Field(ge=0, default=0.0, description="Ставка за один обработанный чат")


class SyncCapabilitiesRequest(BaseModel):
    account_id: int = Field(ge=1, description="Marketplace account ID for capabilities check")


class ChatQuickTemplateCreateRequest(BaseModel):
    template_name: str = Field(min_length=1, max_length=200)
    template_text: str = Field(min_length=1, max_length=2000)


class ChatQuickTemplateUpdateRequest(BaseModel):
    template_name: str = Field(min_length=1, max_length=200)
    template_text: str = Field(min_length=1, max_length=2000)


class ManualReplyRequest(BaseModel):
    operator_name: str = Field(min_length=2, max_length=120)
    response_text: str = Field(min_length=2, max_length=2000)


class AccountCreateRequest(BaseModel):
    marketplace: str = Field(description="wb|ozon|mock")
    account_name: str = Field(min_length=2, max_length=120)
    api_url: str | None = Field(default=None, max_length=2000)
    api_key: str | None = Field(default=None, max_length=2000)
    client_id: str | None = Field(default=None, max_length=200)
    integration: dict[str, object] | None = None


class AccountStatusRequest(BaseModel):
    is_active: bool


class ConversationStatusRequest(BaseModel):
    status: str = Field(description="open|waiting|closed")


class ConversationReplyRequest(BaseModel):
    response_text: str = Field(min_length=1, max_length=4000)
    idempotency_key: str | None = Field(default=None, max_length=120)


class TemplateUpsertRequest(BaseModel):
    category: str
    mode: str = Field(description="auto|manual|ignore")
    template_text: str = Field(max_length=4000)
    is_enabled: bool | None = None


class AISettingsRequest(BaseModel):
    provider: str = Field(description="rules|yandex")
    yandex_api_key: str | None = None
    yandex_folder_id: str | None = None
    yandex_model_uri: str | None = None
    group_processors: dict[str, str] | None = None
    default_sync_lookback_days: int = Field(default=7, ge=0, le=365)


class AIConnectionTestRequest(BaseModel):
    yandex_api_key: str | None = None
    yandex_folder_id: str | None = None


class AIReviewTestRequest(BaseModel):
    review_text: str = Field(min_length=1, max_length=8000)
    review_rating: int | None = Field(default=None, ge=1, le=5)
    yandex_api_key: str | None = None
    yandex_folder_id: str | None = None


class RoleUpdateRequest(BaseModel):
    role: str = Field(description="user|admin|feedback_manager")


class AdminUserCreateRequest(BaseModel):
    email: str = Field(min_length=5, max_length=255)
    password: str = Field(min_length=8, max_length=255)
    role: str = Field(default="user", description="user")
    plan_code: str = Field(default="starter", min_length=2, max_length=100)


class AdminUserPasswordUpdateRequest(BaseModel):
    password: str = Field(min_length=8, max_length=255)


class TeamMemberProfileUpdateRequest(BaseModel):
    full_name: str = ""


class TenantUserCreateRequest(BaseModel):
    email: str = Field(min_length=5, max_length=255)
    password: str = Field(min_length=8, max_length=255)
    role: str = Field(default="feedback_manager", description="feedback_manager")
    full_name: str | None = Field(default=None, max_length=200)
    permissions: list["ManagerPermissionItemRequest"] = Field(default_factory=list)


class TenantUserRoleUpdateRequest(BaseModel):
    role: str = Field(description="admin|feedback_manager")


class ManagerPermissionItemRequest(BaseModel):
    account_id: int = Field(ge=1)
    can_reviews: bool = False
    can_questions: bool = False
    can_chats: bool = False


class ManagerPermissionsUpdateRequest(BaseModel):
    permissions: list[ManagerPermissionItemRequest] = Field(default_factory=list)


class UserBlockUpdateRequest(BaseModel):
    blocked: bool
    reason: str | None = Field(default=None, max_length=500)


class UserDeleteRequest(BaseModel):
    confirm: bool = False


class SuperAdminSettingsRequest(BaseModel):
    payment_provider: str = Field(default="manual", max_length=80)
    payment_api_key: str | None = Field(default=None, max_length=2000)
    ai_provider: str = Field(description="rules|yandex")
    yandex_api_key: str | None = None
    yandex_folder_id: str | None = None
    yandex_model_uri: str | None = None
    group_processors: dict[str, str] | None = None
    use_sync_start_date: bool = False
    sync_start_date: str | None = None
    default_sync_lookback_days: int = Field(default=7, ge=0, le=365)


class TemplateVariableUpsertRequest(BaseModel):
    var_key: str = Field(min_length=3, max_length=120)
    title: str = Field(min_length=1, max_length=255)
    description: str | None = Field(default=None, max_length=2000)
    is_user_editable: bool = False
    source_type: str = Field(default="manual", max_length=40)
    source_path: str | None = Field(default=None, max_length=255)
    default_value: str | None = Field(default=None, max_length=4000)
    is_active: bool = True


class TemplateVariableDeleteRequest(BaseModel):
    var_key: str = Field(min_length=3, max_length=120)


class CreateSupplyProductionRequest(BaseModel):
    name: str
    head_name: str = ""
    address: str = ""
    load_contact: str = ""
    addr_index: str = ""
    addr_region_code: str = ""
    addr_district: str = ""
    addr_city: str = ""
    addr_settlement: str = ""
    addr_street: str = ""
    addr_house: str = ""
    addr_corpus: str = ""
    addr_flat: str = ""


class UpdateSupplyProductionRequest(BaseModel):
    name: str
    head_name: str = ""
    address: str = ""
    load_contact: str = ""
    addr_index: str = ""
    addr_region_code: str = ""
    addr_district: str = ""
    addr_city: str = ""
    addr_settlement: str = ""
    addr_street: str = ""
    addr_house: str = ""
    addr_corpus: str = ""
    addr_flat: str = ""


class CreatePoARecordRequest(BaseModel):
    legal_entity_id: int
    contractor_id: int
    driver_id: int = 0
    driver_manual_name: str = ""
    driver_manual_docs: str = ""


class UpdatePoARecordRequest(BaseModel):
    legal_entity_id: int
    contractor_id: int
    driver_id: int = 0
    driver_manual_name: str = ""
    driver_manual_docs: str = ""


class CreateSupplyContractorRequest(BaseModel):
    name: str
    requisites: str = ""


class UpdateSupplyContractorRequest(BaseModel):
    name: str
    requisites: str = ""


class CreateSupplySourceRequest(BaseModel):
    name: str
    api_key: str
    marketplace: str = "wb"
    client_id: str = ""


class UpsertSupplyEdoSettingsRequest(BaseModel):
    api_url: str = "https://logist-api.kontur.ru/"
    api_key: str | None = None  # None = keep previous
    diadoc_url: str = "https://diadoc-api.kontur.ru/"
    diadoc_client_id: str = ""
    diadoc_login: str = ""
    diadoc_password: str | None = None  # None = keep previous
    diadoc_from_box_id: str = ""
    diadoc_to_box_id: str = ""
    cert_thumbprint: str = ""
    is_enabled: bool = True


class UpsertSupplyChzSettingsRequest(BaseModel):
    """Minimal CHZ connection settings. True API has no static API key — only УКЭП.

    Optional fields (api_base, kpp, wb_analytics_api_key, …) keep previous values
    when omitted (None). Empty string for wb_analytics_api_key clears the token.
    """

    is_enabled: bool = False
    participant_inn: str = ""
    product_group: str = ""
    api_base: str | None = None
    kpp: str | None = None
    fias_id: str | None = None
    return_type: str | None = None
    cert_thumbprint: str | None = None
    wb_analytics_api_key: str | None = None


class ProductCategoryItemRequest(BaseModel):
    id: int | None = None
    name: str = ""
    boxes_per_pallet: int | None = None


class ProductCategoriesSaveRequest(BaseModel):
    items: list[ProductCategoryItemRequest] = Field(default_factory=list)


class WbKizCirculationSyncRequest(BaseModel):
    source_id: int
    date_from: str = ""
    date_to: str = ""


class WbKizCirculationSyncCancelRequest(BaseModel):
    source_id: int = 0
    run_id: int | None = None


class WbKizChzAuthRequest(BaseModel):
    uuid: str
    signature_base64: str
    inn: str = ""


class WbKizChzSubmitItem(BaseModel):
    doc_type: str
    product_group: str = ""
    event_keys: list[str] = Field(default_factory=list)
    product_document: dict[str, object] = Field(default_factory=dict)
    # Exact base64 of JSON bytes that were signed (preferred over re-serializing product_document).
    product_document_b64: str = ""
    signature_base64: str


class WbKizChzSubmitRequest(BaseModel):
    source_id: int
    token: str
    documents: list[WbKizChzSubmitItem] = Field(default_factory=list)
    run_id: int | None = None


class WbKizChzReconcileRequest(BaseModel):
    source_id: int
    token: str = ""


class WbKizChzCisStatusRequest(BaseModel):
    source_id: int
    token: str = ""
    event_keys: list[str] = Field(default_factory=list)


class WbKizChzPrepareRequest(BaseModel):
    source_id: int = 0
    event_keys: list[str] = Field(default_factory=list)


class OzonEdoSendRequest(BaseModel):
    doc_type: str  # zakaz | etrn
    signature_base64: str
    xml_base64: str | None = None  # optional; server rebuilds if omitted


class ToggleSupplySourceRequest(BaseModel):
    is_enabled: bool = True


class SyncSuppliesRequest(BaseModel):
    source_id: int | None = None


class SupplyManualFieldsRequest(BaseModel):
    pass_number: str | None = None
    pallets_count: str | None = None
    driver_name: str | None = None
    notes: str | None = None
    production: str | None = None
    drivers_json: str | None = None  # JSON array of {pass_number, driver_name, pallets_count}


class CreateSupplyDriverRequest(BaseModel):
    full_name: str = ""
    last_name: str = ""
    first_name: str = ""
    middle_name: str = ""
    phone: str = ""
    documents: str = ""
    in_person: str = ""
    vehicles: list = Field(default_factory=list)
    carrier: str = ""
    carrier_name: str = ""
    carrier_inn: str = ""
    carrier_kpp: str = ""
    carrier_phone: str = ""
    carrier_fns_id: str = ""
    carrier_addr_index: str = ""
    carrier_addr_region_code: str = ""
    carrier_addr_district: str = ""
    carrier_addr_city: str = ""
    carrier_addr_settlement: str = ""
    carrier_addr_street: str = ""
    carrier_addr_house: str = ""
    carrier_addr_corpus: str = ""
    carrier_addr_flat: str = ""
    carrier_addr_fias: str = ""
    doc_vu_series: str = ""
    doc_vu_number: str = ""
    doc_vu_issuer: str = ""
    doc_vu_date: str = ""
    doc_inn_fl: str = ""


class CreateSupplyWarehouseRequest(BaseModel):
    warehouse_name: str
    address: str = ""
    addr_index: str = ""
    addr_region_code: str = ""
    addr_district: str = ""
    addr_city: str = ""
    addr_settlement: str = ""
    addr_street: str = ""
    addr_house: str = ""
    addr_corpus: str = ""
    addr_flat: str = ""


class UpdateSupplyWarehouseRequest(BaseModel):
    warehouse_name: str
    address: str = ""
    addr_index: str = ""
    addr_region_code: str = ""
    addr_district: str = ""
    addr_city: str = ""
    addr_settlement: str = ""
    addr_street: str = ""
    addr_house: str = ""
    addr_corpus: str = ""
    addr_flat: str = ""


class CreateSupplyLegalEntityRequest(BaseModel):
    short_name: str
    full_name: str = ""
    requisites: str = ""
    signatories: str = ""
    in_person: str = ""
    basis: str = ""
    address: str = ""
    phone: str = ""
    addr_index: str = ""
    addr_region_code: str = ""
    addr_district: str = ""
    addr_city: str = ""
    addr_settlement: str = ""
    addr_street: str = ""
    addr_house: str = ""
    addr_corpus: str = ""
    addr_flat: str = ""
    addr_fias: str = ""
    signature_image: str | None = None


class UpdateSupplyLegalEntityRequest(BaseModel):
    short_name: str
    full_name: str = ""
    requisites: str = ""
    signatories: str = ""
    in_person: str = ""
    basis: str = ""
    address: str = ""
    phone: str = ""
    addr_index: str = ""
    addr_region_code: str = ""
    addr_district: str = ""
    addr_city: str = ""
    addr_settlement: str = ""
    addr_street: str = ""
    addr_house: str = ""
    addr_corpus: str = ""
    addr_flat: str = ""
    addr_fias: str = ""
    signature_image: str | None = None  # new base64 or None
    clear_signature: bool = False        # True = delete existing


class UpdateSupplyDriverRequest(BaseModel):
    full_name: str = ""
    last_name: str = ""
    first_name: str = ""
    middle_name: str = ""
    phone: str = ""
    documents: str = ""
    in_person: str = ""
    vehicles: list = Field(default_factory=list)
    carrier: str = ""
    carrier_name: str = ""
    carrier_inn: str = ""
    carrier_kpp: str = ""
    carrier_phone: str = ""
    carrier_fns_id: str = ""
    carrier_addr_index: str = ""
    carrier_addr_region_code: str = ""
    carrier_addr_district: str = ""
    carrier_addr_city: str = ""
    carrier_addr_settlement: str = ""
    carrier_addr_street: str = ""
    carrier_addr_house: str = ""
    carrier_addr_corpus: str = ""
    carrier_addr_flat: str = ""
    carrier_addr_fias: str = ""
    doc_vu_series: str = ""
    doc_vu_number: str = ""
    doc_vu_issuer: str = ""
    doc_vu_date: str = ""
    doc_inn_fl: str = ""


class ManagerSuppliesAccessRequest(BaseModel):
    can_supplies: bool = False
    can_supply_settings: bool = False
    can_supply_poa: bool = False
    can_supply_certs: bool = False
    can_supply_planning: bool = False
    can_supply_stock: bool = False
    stock_productions: list[str] = Field(default_factory=list)
    supply_sources: dict = {}  # {source_id: {"wb": bool, "wb_fbs": bool, "wb_fbs_tsd": bool, "ozon": bool}}


class FeedbackMaterialRequest(BaseModel):
    name: str = Field(default="", max_length=300)
    unit: str = Field(default="шт", max_length=32)


class SupplyBalanceSaveRequest(BaseModel):
    """Legacy editable-matrix save (kept for compat; UI uses ledger endpoints)."""
    production_id: int
    items: list[dict[str, object]] = Field(default_factory=list)


class SupplyBalanceVisibilityRequest(BaseModel):
    items: list[dict[str, object]] = Field(default_factory=list)


class SupplyStockReceiptRequest(BaseModel):
    date: str = Field(default="", max_length=20)
    comment: str = Field(default="", max_length=500)
    items: list[dict[str, object]] = Field(default_factory=list)


class SupplyStockAdjustmentRequest(BaseModel):
    """Opening balance or inventory adjustment.

    ``mode``: ``opening`` | ``adjustment``.
    ``quantity_mode``: ``absolute`` (target on-hand) or ``delta`` (signed change).
    """
    mode: str = Field(default="adjustment", max_length=32)
    quantity_mode: str = Field(default="absolute", max_length=32)
    date: str = Field(default="", max_length=20)
    comment: str = Field(default="", max_length=500)
    items: list[dict[str, object]] = Field(default_factory=list)


class ManagerSalaryAccessRequest(BaseModel):
    can_salary: bool = False
    can_salary_settings: bool = False
    can_salary_report: bool = False
    can_salary_zp_export: bool = False
    salary_productions: list[str] = Field(default_factory=list)


class SalaryWorkerCreateRequest(BaseModel):
    full_name: str = Field(min_length=1, max_length=200)
    position: str = Field(default="", max_length=200)
    birth_date: str = Field(default="", max_length=20)
    legal_entity: str = Field(default="", max_length=200)
    production: str = Field(default="", max_length=100)
    visible_for_accountant: bool = True


class SalaryProductCreateRequest(BaseModel):
    order_num: int = Field(default=0, ge=0)
    name: str = Field(min_length=1, max_length=300)
    roles: str = Field(default="", max_length=200)  # comma-separated: "Швея,Упаковщик,Закройщик"
    price_ivanovo: float = Field(default=0.0, ge=0)        # kept for compat
    price_kineshma: float = Field(default=0.0, ge=0)       # kept for compat
    price_nerl: float = Field(default=0.0, ge=0)           # kept for compat
    price_kineshma_poshiv: float = Field(default=0.0, ge=0)
    price_kineshma_raskroi: float = Field(default=0.0, ge=0)
    price_kineshma_upakovka: float = Field(default=0.0, ge=0)
    price_nerl_poshiv: float = Field(default=0.0, ge=0)
    price_nerl_raskroi: float = Field(default=0.0, ge=0)
    price_nerl_upakovka: float = Field(default=0.0, ge=0)


class SalaryEntryItem(BaseModel):
    product_id: int = Field(ge=1)
    quantity: float = Field(default=0.0, ge=0)
    price_snapshot: float = Field(default=0.0, ge=0)


class SalaryEntriesSaveRequest(BaseModel):
    worker_id: int = Field(ge=1)
    entry_date: str = Field(min_length=8, max_length=10)
    entries: list[SalaryEntryItem] = Field(default_factory=list)


class SalaryOkladSaveRequest(BaseModel):
    worker_id: int = Field(ge=1)
    entry_date: str = Field(min_length=8, max_length=10)
    amount: float = Field(default=0.0, ge=0)


class SalaryExtraItem(BaseModel):
    amount: float = Field(default=0.0, ge=0)
    note: str = Field(default="", max_length=500)


class SalaryExtrasSaveRequest(BaseModel):
    worker_id: int = Field(ge=1)
    entry_date: str = Field(min_length=8, max_length=10)
    extras: list[SalaryExtraItem] = Field(default_factory=list)


class SalaryWorkerLinkRequest(BaseModel):
    worker_id: int = Field(ge=1)
    linked_worker_id: int = Field(ge=1)


class SalaryExtraProdItem(BaseModel):
    prod_type: str = Field(..., pattern="^(poshiv|raskroi|upakovka)$")
    product_id: int = Field(ge=1)
    quantity: float = Field(default=0.0, ge=0)
    price_snapshot: float = Field(default=0.0, ge=0)


class SalaryExtraProdsRequest(BaseModel):
    worker_id: int = Field(ge=1)
    entry_date: str = Field(min_length=8, max_length=10)
    entries: list[SalaryExtraProdItem] = Field(default_factory=list)


class SalaryLinkedSnapshotItem(BaseModel):
    linked_worker_id: int = Field(ge=1)
    linked_worker_name: str = Field(default="", max_length=300)
    amount: float = Field(default=0.0, ge=0)


class SalaryLinkedSnapshotSaveRequest(BaseModel):
    worker_id: int = Field(ge=1)
    entry_date: str = Field(min_length=8, max_length=10)
    links: list[SalaryLinkedSnapshotItem] = Field(default_factory=list)


class UserTemplateVariableValuesSaveRequest(BaseModel):
    values: dict[str, str] = Field(default_factory=dict)


TEMPLATE_VARIABLE_KEY_RE = re.compile(r"^%[A-Z0-9_]{2,50}%$")


class UserSyncSettingsRequest(BaseModel):
    use_sync_start_date: bool = True
    sync_start_date: str | None = None


class TariffPlanUpsertRequest(BaseModel):
    code: str = Field(min_length=2, max_length=100)
    title: str = Field(min_length=2, max_length=200)
    monthly_price: float = Field(default=0)
    limits: dict[str, object] = Field(default_factory=dict)
    is_active: bool = True


class TariffPlanDeleteRequest(BaseModel):
    code: str = Field(min_length=2, max_length=100)


class TenantPlanUpdateRequest(BaseModel):
    owner_user_id: int
    plan_code: str = Field(min_length=2, max_length=100)
    limits_override: dict[str, object] = Field(default_factory=dict)


class UserPlanUpdateRequest(BaseModel):
    plan_code: str = Field(min_length=2, max_length=100)


class PaymentRecordCreateRequest(BaseModel):
    owner_user_id: int
    amount: float
    currency: str = Field(default="RUB", max_length=10)
    status: str = Field(default="pending", max_length=80)
    external_payment_id: str | None = Field(default=None, max_length=255)
    details: dict[str, object] = Field(default_factory=dict)
    paid_at: str | None = None
    months: int = Field(default=1, ge=1, le=36)
    grace_days: int = Field(default=3, ge=0, le=30)


class PaymentRecordDeleteRequest(BaseModel):
    id: int = Field(ge=1)


class ProfileUpdateRequest(BaseModel):
    full_name: str | None = Field(default=None, max_length=200)
    email: str | None = Field(default=None, max_length=255)
    current_password: str | None = Field(default=None, max_length=255)
    new_password: str | None = Field(default=None, max_length=255)
    new_password_repeat: str | None = Field(default=None, max_length=255)
    use_sync_start_date: bool | None = None
    sync_start_date: str | None = None


class ClearReviewsRequest(BaseModel):
    user_id: int | None = None


class ClearConversationsRequest(BaseModel):
    user_id: int | None = None
    kind: str | None = None
    source: str | None = None


class TemplateSubgroupSaveRequest(BaseModel):
    templates: list[str] = Field(default_factory=list)


class TemplateVariantCreateRequest(BaseModel):
    group_id: str = Field(min_length=2, max_length=100)
    subgroup: str = Field(min_length=1, max_length=255)
    template_text: str = Field(min_length=1, max_length=4000)


class DefaultTemplateSubgroupSaveRequest(BaseModel):
    templates: list[str] = Field(default_factory=list)


class DefaultTemplateVariantCreateRequest(BaseModel):
    group_id: str = Field(min_length=2, max_length=100)
    subgroup: str = Field(min_length=1, max_length=255)
    template_text: str = Field(min_length=1, max_length=4000)


class DefaultTemplateSubgroupManageRequest(BaseModel):
    group_id: str = Field(min_length=2, max_length=100)
    subgroup: str = Field(min_length=1, max_length=255)


class DefaultTemplateSubgroupRenameRequest(BaseModel):
    group_id: str = Field(min_length=2, max_length=100)
    subgroup: str = Field(min_length=1, max_length=255)
    new_subgroup: str = Field(min_length=1, max_length=255)


class DefaultTemplateBulkImportRequest(BaseModel):
    group_id: str = Field(min_length=2, max_length=100)
    subgroup: str = Field(min_length=1, max_length=255)
    templates: list[str] = Field(default_factory=list)


class ProcessingRuleItemRequest(BaseModel):
    group_id: str = Field(min_length=2, max_length=100)
    action_mode: str = Field(description="template|manual")
    auto_send: bool = False


class ProcessingRulesApplyRequest(BaseModel):
    rules: list[ProcessingRuleItemRequest] = Field(default_factory=list)


class RecommendationRowRequest(BaseModel):
    source_article: str = Field(default="", max_length=255)
    targets_csv: str = Field(default="", max_length=4000)


class RecommendationsSaveRequest(BaseModel):
    rows: list[RecommendationRowRequest] = Field(default_factory=list)


class StockSourceCreateRequest(BaseModel):
    marketplace: str = Field(min_length=1, max_length=20)
    account_name: str = Field(min_length=1, max_length=200)
    api_url: str = Field(default="", max_length=500)
    api_key: str = Field(default="", max_length=2000)
    client_id: str = Field(default="", max_length=200)
    interval_hours: int = Field(default=24, ge=1, le=24)
    retention_days: int = Field(default=30, ge=1, le=365)


class OzonCombinedDocsRequest(BaseModel):
    supply_ids: list[int]


class CertificateCreateRequest(BaseModel):
    legal_entity_short: str = ""
    category: str = ""
    number: str = ""
    expiry_date: str = ""
    verification_url: str = ""
    image_data: str | None = None
    doc_type: str = "Сертификат соответствия"


class StockSourceUpdateRequest(BaseModel):
    account_name: str | None = None
    api_key: str | None = None
    client_id: str | None = None
    interval_hours: int | None = Field(default=None, ge=1, le=24)
    retention_days: int | None = Field(default=None, ge=1, le=365)
    is_active: bool | None = None


class WbFbsAutoSyncSettingsRequest(BaseModel):
    enabled: bool = False
    # Preferred: minutes (10, 30, 60, …). Legacy clients may still send hours.
    interval_minutes: int | None = Field(default=None, ge=10, le=1440)
    interval_hours: int | None = Field(default=None, ge=1, le=24)
    # WB /api/v3/orders period window (days). Hard max 30.
    lookback_days: int = Field(default=3, ge=1, le=30)
    active_from: str = "12:00"
    active_to: str = "06:00"
    collect_mgt_enabled: bool = False
    collect_mgt_interval_minutes: int | None = Field(default=None, ge=10, le=1440)
    collect_mgt_interval_hours: int | None = Field(default=None, ge=1, le=24)
    collect_mgt_active_from: str = "12:00"
    collect_mgt_active_to: str = "06:00"

    @staticmethod
    def _validated_period_minutes(
        *,
        minutes_value: int | None,
        hours_value: int | None,
        label: str,
    ) -> int:
        allowed_minutes = (10, 30, 60, 120, 180, 360, 720, 1440)
        legacy_hours = (1, 2, 3, 6, 12, 24)
        if minutes_value is not None:
            minutes = int(minutes_value)
            if minutes not in allowed_minutes:
                raise ValueError(
                    f"{label} должен быть одним из: "
                    + ", ".join(str(v) for v in allowed_minutes)
                    + " минут"
                )
            return minutes
        if hours_value is not None:
            hours = int(hours_value)
            if hours not in legacy_hours:
                raise ValueError(
                    f"{label} должен быть одним из: "
                    + ", ".join(str(v) for v in legacy_hours)
                    + " часов"
                )
            return hours * 60
        return 60

    def validated_interval_minutes(self) -> int:
        return self._validated_period_minutes(
            minutes_value=self.interval_minutes,
            hours_value=self.interval_hours,
            label="Период синхронизации",
        )

    def validated_collect_interval_minutes(self) -> int:
        return self._validated_period_minutes(
            minutes_value=self.collect_mgt_interval_minutes,
            hours_value=self.collect_mgt_interval_hours,
            label="Период автосбора МГТ",
        )


ROLE_ADMIN = "admin"
ROLE_USER = "user"
ROLE_FEEDBACK_MANAGER = "feedback_manager"
ROLE_CAN_ACCESS_ANALYTICS = {ROLE_ADMIN, ROLE_USER}
ROLE_CAN_ACCESS_SETTINGS = {ROLE_ADMIN, ROLE_USER}
ROLE_ASSIGNABLE_BY_ADMIN = {ROLE_USER, ROLE_FEEDBACK_MANAGER}
TENANT_ROLE_OWNER = "admin"
TENANT_ROLE_MANAGER = "feedback_manager"
# All roles that are treated as "manager" (can have granular permissions configured)
TENANT_MANAGER_ROLES = {"feedback_manager", "production_manager", "manager"}
SESSION_TTL_SECONDS = 30 * 24 * 60 * 60
CSRF_COOKIE_NAME = "csrf_token"
CSRF_HEADER_NAME = "X-CSRF-Token"
RATE_LIMIT_API_READ_PER_MINUTE = 600
RATE_LIMIT_API_WRITE_PER_MINUTE = 180
RATE_LIMIT_SYNC_PER_MINUTE = 20
RATE_LIMIT_LOGIN_PER_10_MIN = 30
FAILED_LOGIN_LIMIT_PER_15_MIN = 10
AUTO_SYNC_INTERVAL_SECONDS = 60


def _normalize_role(raw_role: object) -> str:
    role = str(raw_role or "").strip().lower()
    if role == ROLE_ADMIN:
        return ROLE_ADMIN
    if role == ROLE_USER:
        return ROLE_USER
    if role == ROLE_FEEDBACK_MANAGER:
        return ROLE_FEEDBACK_MANAGER
    return ROLE_USER


def create_app(config: AppConfig | None = None) -> FastAPI:
    # Uvicorn sets the root logger level to WARNING, which silently drops our
    # INFO messages even if the child logger level is INFO.
    # Fix: attach a StreamHandler directly to our package logger and disable
    # propagation to root.  This guarantees INFO output to stderr → journald
    # regardless of uvicorn's root logger configuration.
    _rp_logger = logging.getLogger("review_processor")
    if not _rp_logger.handlers:
        _h = logging.StreamHandler()
        _h.setLevel(logging.DEBUG)
        _h.setFormatter(logging.Formatter(
            "%(asctime)s %(levelname)s %(name)s: %(message)s"
        ))
        _rp_logger.addHandler(_h)
    _rp_logger.setLevel(logging.INFO)
    _rp_logger.propagate = False  # bypass root logger whose level is WARNING
    app_config = config or load_app_config()
    repository = ReviewRepository(db_url=app_config.db_url)
    service = ReviewAutomationService(repository)
    self_registration_enabled = bool(app_config.self_registration_enabled)

    app = FastAPI(title="Marketplace Reviews Assistant", version="1.0.0")
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
    sync_stop_event = threading.Event()
    # Supply sync state — separate from main feedback sync
    supply_sync_lock = threading.Lock()
    supply_sync_state: dict[str, object] = {
        "in_progress": False,
        "page": 0,
        "synced": 0,
        "total": 0,
        "errors": [],
        "message": "",
        "started_at": None,
        "finished_at": None,
    }
    # TTN daily sequential counter — stored in DB, resets each new day
    @app.post("/api/ttn/next-number")
    def ttn_next_number(request: Request) -> dict[str, object]:
        _require_user(request)
        return {"number": repository.next_ttn_number()}

    sync_lock = threading.Lock()
    sync_state: dict[str, object] = {
        "in_progress": False,
        "cancel_requested": False,
        "last_started_at": None,
        "last_finished_at": None,
        "polling_enabled": False,
        "polling_user_id": None,
        "polling_account_ids": [],
        "polling_since_date": None,
        "polling_started_at": None,
        "last_poll_at": None,
        "last_poll_result": None,
        # Progress tracking (visible to all users via /api/sync/status)
        "progress_step": "",
        "progress_account": "",
        "progress_channel": "",
        "progress_loaded": 0,
        "progress_total_items": 0,
        "progress_total_accounts": 0,
        "progress_current_account": 0,
        # Sync result report (shown after completion)
        "last_sync_report": None,  # populated after manual sync finishes
        "sync_log": [],  # list of log lines accumulated during sync
    }
    auto_sync_stop_event = threading.Event()
    auto_sync_worker: dict[str, threading.Thread | None] = {"thread": None}
    rate_limit_lock = threading.Lock()
    rate_buckets: dict[str, list[float]] = {}
    failed_login_attempts: dict[str, list[float]] = {}
    stock_scheduler = StockScheduler(repository)
    wb_fbs_scheduler = wb_fbs_mod.WbFbsScheduler(repository)

    def _client_ip(request: Request) -> str:
        forwarded = str(request.headers.get("x-forwarded-for") or "").split(",")[0].strip()
        if forwarded:
            return forwarded
        if request.client and request.client.host:
            return str(request.client.host)
        return "unknown"

    def _allow_rate(scope: str, *, limit: int, window_seconds: int) -> bool:
        now = time.time()
        cutoff = now - float(window_seconds)
        with rate_limit_lock:
            bucket = [ts for ts in rate_buckets.get(scope, []) if ts >= cutoff]
            if len(bucket) >= limit:
                rate_buckets[scope] = bucket
                return False
            bucket.append(now)
            rate_buckets[scope] = bucket
            return True

    def _record_failed_login(login_key: str) -> None:
        now = time.time()
        cutoff = now - 15 * 60
        with rate_limit_lock:
            bucket = [ts for ts in failed_login_attempts.get(login_key, []) if ts >= cutoff]
            bucket.append(now)
            failed_login_attempts[login_key] = bucket

    def _clear_failed_login(login_key: str) -> None:
        with rate_limit_lock:
            failed_login_attempts.pop(login_key, None)

    def _is_login_blocked(login_key: str) -> bool:
        now = time.time()
        cutoff = now - 15 * 60
        with rate_limit_lock:
            bucket = [ts for ts in failed_login_attempts.get(login_key, []) if ts >= cutoff]
            failed_login_attempts[login_key] = bucket
            return len(bucket) >= FAILED_LOGIN_LIMIT_PER_15_MIN

    def _is_private_host(hostname: str) -> bool:
        host = hostname.strip().lower().rstrip(".")
        if not host:
            return True
        if host in {"localhost", "localhost.localdomain"} or host.endswith(".localhost") or host.endswith(".local"):
            return True
        try:
            ip = ipaddress.ip_address(host)
        except ValueError:
            return False
        return bool(
            ip.is_private
            or ip.is_loopback
            or ip.is_link_local
            or ip.is_multicast
            or ip.is_reserved
            or ip.is_unspecified
        )

    def _validate_account_api_url(marketplace: str, raw_url: str) -> str:
        normalized = raw_url.strip()
        parsed = urlparse(normalized)
        if parsed.scheme.lower() != "https":
            raise HTTPException(status_code=400, detail="Адрес интерфейса API должен начинаться с https://")
        host = (parsed.hostname or "").strip().lower()
        if not host:
            raise HTTPException(status_code=400, detail="Некорректный адрес интерфейса API")
        if _is_private_host(host):
            raise HTTPException(status_code=400, detail="Адрес интерфейса API указывает на недопустимый внутренний хост")
        if marketplace == "wb" and not (host == "feedbacks-api.wildberries.ru" or host.endswith(".wildberries.ru")):
            raise HTTPException(status_code=400, detail="Для WB разрешены только домены wildberries.ru")
        if marketplace == "ozon" and not (host == "api-seller.ozon.ru" or host.endswith(".ozon.ru")):
            raise HTTPException(status_code=400, detail="Для OZON разрешены только домены ozon.ru")
        if marketplace == "yandex" and not (host == "api.partner.market.yandex.ru" or host.endswith(".market.yandex.ru")):
            raise HTTPException(status_code=400, detail="Для Яндекс Маркета разрешены только домены market.yandex.ru")
        return normalized

    def _set_session_cookie(response: RedirectResponse, token: str) -> None:
        secure_cookie = bool(app_config.is_production)
        response.set_cookie(
            "session_token",
            token,
            httponly=True,
            samesite="lax",
            secure=secure_cookie,
            max_age=SESSION_TTL_SECONDS,
        )
        response.set_cookie(
            CSRF_COOKIE_NAME,
            secrets.token_urlsafe(32),
            httponly=False,
            samesite="lax",
            secure=secure_cookie,
            max_age=SESSION_TTL_SECONDS,
        )

    def _ensure_csrf_cookie(response: HTMLResponse | RedirectResponse, request: Request) -> None:
        if not request.cookies.get("session_token"):
            return
        if request.cookies.get(CSRF_COOKIE_NAME):
            return
        response.set_cookie(
            CSRF_COOKIE_NAME,
            secrets.token_urlsafe(32),
            httponly=False,
            samesite="lax",
            secure=bool(app_config.is_production),
            max_age=SESSION_TTL_SECONDS,
        )

    def _is_same_origin(request: Request, origin_value: str) -> bool:
        parsed = urlparse(origin_value)
        if not parsed.scheme or not parsed.netloc:
            return False
        expected_scheme = str(request.headers.get("x-forwarded-proto") or request.url.scheme).lower()
        expected_host = str(request.url.hostname or "").lower()
        expected_port = request.url.port or (443 if expected_scheme == "https" else 80)
        origin_host = str(parsed.hostname or "").lower()
        origin_port = parsed.port or (443 if parsed.scheme.lower() == "https" else 80)
        return parsed.scheme.lower() == expected_scheme and origin_host == expected_host and origin_port == expected_port

    def _check_csrf(request: Request) -> None:
        method = request.method.upper()
        if method not in {"POST", "PUT", "PATCH", "DELETE"}:
            return
        path = request.url.path
        if not path.startswith("/api/"):
            return
        # Only enforce CSRF for authenticated browser requests.
        if not request.cookies.get("session_token"):
            return
        cookie_token = str(request.cookies.get(CSRF_COOKIE_NAME) or "")
        header_token = str(request.headers.get(CSRF_HEADER_NAME) or "").strip()
        if not cookie_token or not header_token or not secrets.compare_digest(cookie_token, header_token):
            raise HTTPException(status_code=403, detail="CSRF токен отсутствует или неверен")
        origin = str(request.headers.get("origin") or "").strip()
        referer = str(request.headers.get("referer") or "").strip()
        if origin and not _is_same_origin(request, origin):
            raise HTTPException(status_code=403, detail="Недопустимый origin запроса")
        if not origin and referer and not _is_same_origin(request, referer):
            raise HTTPException(status_code=403, detail="Недопустимый referer запроса")

    def _check_rate_limit(request: Request) -> None:
        path = request.url.path
        method = request.method.upper()
        ip = _client_ip(request)
        if path == "/login" and method == "POST":
            login_scope = f"login:{ip}"
            if not _allow_rate(login_scope, limit=RATE_LIMIT_LOGIN_PER_10_MIN, window_seconds=10 * 60):
                raise HTTPException(status_code=429, detail="Слишком много попыток входа. Попробуйте позже.")
        if path.startswith("/api/"):
            if method in {"GET", "HEAD", "OPTIONS"}:
                limit = RATE_LIMIT_API_READ_PER_MINUTE
            elif path == "/api/sync":
                limit = RATE_LIMIT_SYNC_PER_MINUTE
            else:
                limit = RATE_LIMIT_API_WRITE_PER_MINUTE
            scope = f"api:{method}:{path}:{ip}"
            if not _allow_rate(scope, limit=limit, window_seconds=60):
                raise HTTPException(status_code=429, detail="Слишком много запросов. Попробуйте позже.")

    @app.middleware("http")
    async def hardening_middleware(request: Request, call_next):
        try:
            _check_rate_limit(request)
            _check_csrf(request)
        except HTTPException as exc:
            if request.url.path == "/login":
                return HTMLResponse(build_login_html(error=str(exc.detail)), status_code=exc.status_code)
            return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})
        response = await call_next(request)
        response.headers.setdefault("X-Frame-Options", "DENY")
        response.headers.setdefault("X-Content-Type-Options", "nosniff")
        response.headers.setdefault("Referrer-Policy", "same-origin")
        response.headers.setdefault("Permissions-Policy", "camera=(), microphone=(), geolocation=()")
        # CryptoPro ЭЦП Browser plug-in loads nmcades_plugin_api.js from the browser
        # extension and may use an NPAPI <object type="application/x-cades"> fallback.
        response.headers.setdefault(
            "Content-Security-Policy",
            "default-src 'self'; "
            "script-src 'self' 'unsafe-inline' 'unsafe-eval' chrome-extension: moz-extension: safari-extension:; "
            "style-src 'self' 'unsafe-inline'; "
            "img-src 'self' data:; "
            "connect-src 'self' chrome-extension: moz-extension: safari-extension:; "
            "object-src *; "
            "frame-ancestors 'none'; form-action 'self'; base-uri 'self'",
        )
        if app_config.is_production:
            response.headers.setdefault("Strict-Transport-Security", "max-age=63072000; includeSubDomains; preload")
        _ensure_csrf_cookie(response, request)
        return response

    def _now_iso() -> str:
        return datetime.now(UTC).isoformat()

    def _issue_session(user_id: int) -> str:
        token = create_session_token()
        expires = (datetime.now(UTC) + timedelta(days=30)).isoformat()
        repository.create_session(token=token, user_id=user_id, expires_at=expires)
        return token

    def _login_attempt_key(request: Request, email: str) -> str:
        return f"{_client_ip(request)}::{email.strip().lower()}"

    def _get_current_user(request: Request) -> dict[str, object] | None:
        token = request.cookies.get("session_token")
        if not token:
            return None
        repository.cleanup_expired_sessions(_now_iso())
        return repository.get_session_user(token)

    def _require_user(request: Request) -> dict[str, object]:
        user = _get_current_user(request)
        if user is None:
            raise HTTPException(status_code=401, detail="Требуется авторизация")
        return user

    def _require_admin(request: Request) -> dict[str, object]:
        user = _require_user(request)
        if user.get("role") != ROLE_ADMIN:
            raise HTTPException(status_code=403, detail="Доступ только для администратора")
        return user

    def _is_super_admin(user: dict[str, object]) -> bool:
        return bool(user.get("is_super_admin"))

    def _tenant_owner_id(user: dict[str, object]) -> int:
        owner_raw = user.get("owner_user_id")
        if owner_raw is None:
            return int(user["id"])
        try:
            return int(owner_raw)
        except (TypeError, ValueError):
            return int(user["id"])

    def _require_super_admin(request: Request) -> dict[str, object]:
        user = _require_admin(request)
        if not _is_super_admin(user):
            raise HTTPException(status_code=403, detail="Доступ только для супер-администратора")
        return user

    def _require_tenant_owner(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        if _is_super_admin(user):
            owner_scope_user_id = _tenant_owner_id(user)
            if owner_scope_user_id <= 0:
                owner_scope_user_id = int(user["id"])
            owner_scope_user = repository.get_user_by_id(owner_scope_user_id)
            return owner_scope_user or user
        if _tenant_owner_id(user) != int(user["id"]):
            raise HTTPException(status_code=403, detail="Недостаточно прав для управления командой")
        return user

    def _parse_sync_start_date_or_none(value: str | None, *, enabled: bool) -> str | None:
        if not enabled:
            return None
        raw = (value or "").strip()
        if not raw:
            raise HTTPException(status_code=400, detail="Укажите дату начала синхронизации")
        try:
            datetime.strptime(raw, "%Y-%m-%d")
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Дата должна быть в формате ГГГГ-ММ-ДД") from exc
        return raw

    def _target_user_for_admin_scope(*, actor: dict[str, object], target_user_id: int) -> dict[str, object]:
        target = repository.get_user_by_id(target_user_id)
        if target is None:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        if _is_super_admin(actor):
            return target
        actor_owner_id = _tenant_owner_id(actor)
        target_owner_id = _tenant_owner_id(target)
        if target_owner_id != actor_owner_id:
            raise HTTPException(status_code=403, detail="Пользователь не относится к вашему кабинету")
        if bool(target.get("is_super_admin")):
            raise HTTPException(status_code=403, detail="Недостаточно прав для управления этим пользователем")
        return target

    _TENANT_ALLOWED_ROLES = {TENANT_ROLE_OWNER} | TENANT_MANAGER_ROLES

    def _normalize_tenant_role_or_400(raw_role: str) -> str:
        role = str(raw_role or "").strip().lower()
        if role not in _TENANT_ALLOWED_ROLES:
            raise HTTPException(status_code=400, detail="Недопустимая роль")
        return role

    def _manager_permissions_context_for_user(user: dict[str, object]) -> list[dict[str, object]]:
        if str(user.get("role") or "").strip().lower() not in TENANT_MANAGER_ROLES:
            return []
        return repository.list_manager_permissions(manager_user_id=int(user["id"]))

    def _manager_allowed_review_account_ids(user: dict[str, object]) -> list[int] | None:
        if str(user.get("role") or "").strip().lower() not in TENANT_MANAGER_ROLES:
            return None
        rows = _manager_permissions_context_for_user(user)
        ids: list[int] = []
        seen: set[int] = set()
        for row in rows:
            if not bool(row.get("can_reviews")):
                continue
            try:
                account_id = int(row.get("account_id"))
            except (TypeError, ValueError):
                continue
            if account_id <= 0 or account_id in seen:
                continue
            seen.add(account_id)
            ids.append(account_id)
        return ids

    def _manager_allowed_conversation_accounts(user: dict[str, object]) -> dict[str, list[int]] | None:
        if str(user.get("role") or "").strip().lower() not in TENANT_MANAGER_ROLES:
            return None
        rows = _manager_permissions_context_for_user(user)
        scope: dict[str, list[int]] = {"question": [], "chat": []}
        seen: dict[str, set[int]] = {"question": set(), "chat": set()}
        for row in rows:
            try:
                account_id = int(row.get("account_id"))
            except (TypeError, ValueError):
                continue
            if account_id <= 0:
                continue
            if bool(row.get("can_questions")) and account_id not in seen["question"]:
                seen["question"].add(account_id)
                scope["question"].append(account_id)
            if bool(row.get("can_chats")) and account_id not in seen["chat"]:
                seen["chat"].add(account_id)
                scope["chat"].append(account_id)
        return scope

    def _manager_owner_account_ids(owner_user_id: int) -> set[int]:
        return {
            int(item.get("id"))
            for item in repository.list_marketplace_accounts(user_id=owner_user_id, include_secrets=False)
            if item.get("id") is not None
        }

    def _require_manager_scope_for_review(user: dict[str, object], review_uid: str) -> None:
        if str(user.get("role") or "").strip().lower() not in TENANT_MANAGER_ROLES:
            return
        allowed = set(_manager_allowed_review_account_ids(user) or [])
        if not allowed:
            raise HTTPException(status_code=403, detail="Менеджеру не назначены доступы к отзывам")
        review = repository.get_review(user_id=_tenant_owner_id(user), review_uid=review_uid)
        if review is None:
            raise HTTPException(status_code=404, detail="Отзыв не найден")
        try:
            account_id = int(review.get("account_id"))
        except (TypeError, ValueError):
            raise HTTPException(status_code=403, detail="Отзыв не привязан к разрешенному кабинету")
        if account_id not in allowed:
            raise HTTPException(status_code=403, detail="Нет доступа к этому кабинету отзывов")

    def _require_manager_scope_for_conversation(user: dict[str, object], conversation_uid: str) -> None:
        if str(user.get("role") or "").strip().lower() not in TENANT_MANAGER_ROLES:
            return
        scope = _manager_allowed_conversation_accounts(user) or {"question": [], "chat": []}
        conversation = repository.get_conversation(user_id=_tenant_owner_id(user), conversation_uid=conversation_uid)
        if conversation is None:
            raise HTTPException(status_code=404, detail="Диалог не найден")
        kind = str(conversation.get("kind") or "").strip().lower()
        if kind not in {"question", "chat"}:
            raise HTTPException(status_code=403, detail="Нет доступа к этому типу диалога")
        allowed = set(scope.get(kind, []))
        if not allowed:
            raise HTTPException(status_code=403, detail="Менеджеру не назначены доступы к этому типу диалогов")
        try:
            account_id = int(conversation.get("account_id"))
        except (TypeError, ValueError):
            raise HTTPException(status_code=403, detail="Диалог не привязан к разрешенному кабинету")
        if account_id not in allowed:
            raise HTTPException(status_code=403, detail="Нет доступа к этому кабинету диалогов")

    def _require_analytics_access(request: Request) -> dict[str, object]:
        user = _require_user(request)
        if str(user.get("role")) not in ROLE_CAN_ACCESS_ANALYTICS:
            raise HTTPException(status_code=403, detail="Недостаточно прав для просмотра аналитики")
        return user

    def _require_settings_access(request: Request) -> dict[str, object]:
        user = _require_user(request)
        if str(user.get("role")) not in ROLE_CAN_ACCESS_SETTINGS:
            raise HTTPException(status_code=403, detail="Недостаточно прав для раздела настроек")
        return user

    def _set_sync_in_progress(in_progress: bool) -> None:
        with sync_lock:
            sync_state["in_progress"] = in_progress
            if in_progress:
                sync_state["cancel_requested"] = False
                sync_state["last_started_at"] = _now_iso()
                sync_stop_event.clear()
            else:
                sync_state["last_finished_at"] = _now_iso()

    def _snapshot_active_account_ids_for_user(user_id: int) -> list[int]:
        ids: list[int] = []
        seen: set[int] = set()
        for account in repository.list_marketplace_accounts(user_id, include_secrets=False):
            try:
                account_id = int(account.get("id"))
            except (TypeError, ValueError):
                continue
            if account_id <= 0 or account_id in seen:
                continue
            if not bool(account.get("is_active")):
                continue
            seen.add(account_id)
            ids.append(account_id)
        return ids

    def _serialize_sync_error_details(raw_errors: object) -> list[dict[str, object]]:
        if not isinstance(raw_errors, list):
            return []
        result: list[dict[str, object]] = []
        for item in raw_errors:
            if not isinstance(item, dict):
                continue
            cleaned: dict[str, object] = {}
            for key, value in item.items():
                cleaned[str(key)] = value
            result.append(cleaned)
        return result

    def _channel_access_label(capabilities: Mapping[str, object]) -> str:
        reviews_ok = bool(capabilities.get("reviews"))
        questions_ok = bool(capabilities.get("questions"))
        chats_ok = bool(capabilities.get("chats"))
        if reviews_ok and questions_ok and chats_ok:
            return "По данному ключу доступны все каналы: отзывы, вопросы и чаты."
        if chats_ok and not reviews_ok and not questions_ok:
            return "По данному ключу вы можете работать только с чатами. К отзывам и вопросам нет доступа."
        if reviews_ok and questions_ok and not chats_ok:
            return "По данному ключу вы можете работать только с отзывами и вопросами, но у вас нет доступа к чатам."
        if reviews_ok and chats_ok and not questions_ok:
            return "По данному ключу вы можете работать только с отзывами и чатами, но у вас нет доступа к вопросам."
        if questions_ok and chats_ok and not reviews_ok:
            return "По данному ключу вы можете работать только с вопросами и чатами, но у вас нет доступа к отзывам."
        if reviews_ok and not questions_ok and not chats_ok:
            return "По данному ключу вы можете работать только с отзывами. К вопросам и чатам нет доступа."
        if questions_ok and not reviews_ok and not chats_ok:
            return "По данному ключу вы можете работать только с вопросами. К отзывам и чатам нет доступа."
        return "По данному ключу нет доступа к отзывам, вопросам и чатам."

    def _probe_account_capabilities(*, user_id: int, account_id: int, since_date: str | None) -> dict[str, object]:
        account = repository.get_marketplace_account(
            user_id=user_id,
            account_id=account_id,
            include_secrets=True,
        )
        if account is None:
            raise HTTPException(status_code=404, detail="Кабинет маркетплейса не найден")
        if not bool(account.get("is_active")):
            raise HTTPException(status_code=400, detail="Кабинет отключен")

        probe = service.probe_account_channels(account=account, since_date=since_date or None)
        channels_raw = probe.get("channels")
        channels: dict[str, dict[str, object]] = (
            channels_raw if isinstance(channels_raw, dict) else {}
        )
        capabilities: dict[str, bool] = {
            "reviews": bool((channels.get("reviews") or {}).get("available")),
            "questions": bool((channels.get("questions") or {}).get("available")),
            "chats": bool((channels.get("chats") or {}).get("available")),
        }
        channel_messages: dict[str, str] = {}
        all_errors: list[dict[str, object]] = []
        for channel in ("reviews", "questions", "chats"):
            channel_data = channels.get(channel)
            if not isinstance(channel_data, Mapping):
                continue
            if bool(channel_data.get("available")):
                continue
            message = str(channel_data.get("error") or "").strip()
            if message:
                channel_messages[channel] = message
            all_errors.append(
                {
                    "account_id": int(probe.get("account_id") or account_id),
                    "marketplace": str(probe.get("marketplace") or account.get("marketplace") or ""),
                    "channel": channel,
                    "scope": channel,
                    "error": message,
                    "access_denied": bool(channel_data.get("access_denied")),
                }
            )
        can_sync_any = any(capabilities.values())
        return {
            "account_id": int(probe.get("account_id") or account_id),
            "marketplace": str(probe.get("marketplace") or account.get("marketplace") or ""),
            "account_name": str(probe.get("account_name") or account.get("account_name") or ""),
            "is_active": bool(account.get("is_active")),
            "capabilities": capabilities,
            "can_sync_any": can_sync_any,
            "summary": _channel_access_label(capabilities),
            "channel_messages": channel_messages,
            "errors": all_errors,
            "all_channels_available": bool(probe.get("all_channels_available")),
        }

    def _update_sync_progress(
        *,
        step: str = "",
        account: str = "",
        channel: str = "",
        loaded: int = 0,
        total_accounts: int = 0,
        current_account: int = 0,
    ) -> None:
        with sync_lock:
            sync_state["progress_step"] = step
            sync_state["progress_account"] = account
            sync_state["progress_channel"] = channel
            if loaded:
                sync_state["progress_loaded"] = int(sync_state.get("progress_loaded") or 0) + loaded
            if total_accounts:
                sync_state["progress_total_accounts"] = total_accounts
            if current_account:
                sync_state["progress_current_account"] = current_account
            # Accumulate log line (keep last 200 lines)
            if step or account or channel:
                ts = _now_iso()[11:19]  # HH:MM:SS
                parts = [p for p in [account, channel, step] if p]
                line = f"{ts}  {' → '.join(parts)}"
                log_lines = list(sync_state.get("sync_log") or [])
                log_lines.append(line)
                sync_state["sync_log"] = log_lines[-200:]

    def _run_sync_for_user(
        *,
        user_id: int,
        since_date: str | None,
        account_ids: list[int] | None,
        run_started_at: str,
        apply_date_filter: bool = False,
    ) -> dict[str, object]:
        with sync_lock:
            if bool(sync_state.get("in_progress")):
                # If an auto-sync is running and this is a manual request,
                # cancel the auto-sync and let the manual one proceed after a short wait.
                if apply_date_filter and not bool(sync_state.get("is_manual")):
                    sync_stop_event.set()  # signal auto-sync to stop
                    # Will retry after releasing lock; auto-sync checks stop_requested
                else:
                    raise HTTPException(status_code=409, detail="Синхронизация уже выполняется")
        # Brief wait for auto-sync to see the stop signal before we acquire the slot
        if apply_date_filter and sync_stop_event.is_set():
            import time as _time
            _time.sleep(2)
        with sync_lock:
            if bool(sync_state.get("in_progress")):
                raise HTTPException(status_code=409, detail="Синхронизация уже выполняется. Попробуйте снова через несколько секунд.")
            sync_state["in_progress"] = True
            sync_state["is_manual"] = apply_date_filter  # True only for manual button clicks
            sync_state["cancel_requested"] = False
            sync_state["last_started_at"] = run_started_at
            sync_state["progress_step"] = "Подготовка..."
            sync_state["progress_account"] = ""
            sync_state["progress_channel"] = ""
            sync_state["progress_loaded"] = 0
            sync_state["progress_total_items"] = 0
            sync_state["progress_total_accounts"] = 0
            sync_state["progress_current_account"] = 0
            sync_state["sync_log"] = []  # reset log for new sync
        sync_stop_event.clear()
        try:
            result = service.sync_all_accounts(
                user_id=user_id,
                since_date=since_date or None,
                account_ids=account_ids,
                stop_requested=sync_stop_event.is_set,
                progress_callback=_update_sync_progress,
                apply_date_filter=apply_date_filter,
            )
            # Build detailed sync report for the completion modal
            if apply_date_filter and isinstance(result, dict):
                report_accounts = []
                for stat in (result.get("account_channel_stats") or []):
                    acct_id = stat.get("account_id")
                    acct_name = stat.get("account_name") or f"#{acct_id}"
                    channels = {}
                    for ch in ("reviews", "questions", "chats"):
                        ch_data = stat.get(ch) or {}
                        channels[ch] = {
                            "ok": bool(ch_data.get("ok")),
                            "loaded": int(ch_data.get("loaded") or 0),
                            "skipped": int(ch_data.get("skipped_old") or 0),
                            "error": str(ch_data.get("error") or ""),
                        }
                    report_accounts.append({
                        "account_id": acct_id,
                        "account_name": acct_name,
                        "channels": channels,
                    })
                with sync_lock:
                    sync_state["last_sync_report"] = {
                        "started_at": run_started_at,
                        "finished_at": _now_iso(),
                        "accounts": report_accounts,
                        "total_reviews": int(result.get("loaded_reviews") or result.get("loaded") or 0),
                        "total_questions": int(result.get("loaded_questions") or 0),
                        "total_chats": int(result.get("loaded_chats") or 0),
                        "cancelled": bool(result.get("cancelled")),
                        "errors": int(result.get("failed_accounts") or 0),
                        "log": list(sync_state.get("sync_log") or []),
                    }
            return result
        finally:
            with sync_lock:
                sync_state["in_progress"] = False
                sync_state["last_finished_at"] = _now_iso()
                sync_state["progress_step"] = "Завершено"
            sync_stop_event.clear()

    def _start_auto_sync_worker_if_needed() -> None:
        with sync_lock:
            existing = auto_sync_worker.get("thread")
            if isinstance(existing, threading.Thread) and existing.is_alive():
                return
            auto_sync_stop_event.clear()

            def _auto_sync_loop() -> None:
                _log.info("auto_sync_loop: started, first poll in %ds", AUTO_SYNC_INTERVAL_SECONDS)
                while not auto_sync_stop_event.is_set():
                    auto_sync_stop_event.wait(AUTO_SYNC_INTERVAL_SECONDS)
                    if auto_sync_stop_event.is_set():
                        break
                    _log.info("auto_sync_loop: poll iteration starting")
                    # Read sync target from DB on every iteration so the loop
                    # works even if in-memory sync_state was cleared (e.g. Stop
                    # button pressed, then the next auto-sync still fires).
                    # This makes polling resilient to manual Stop and restarts.
                    try:
                        owner_users_for_poll = repository.list_users(owner_only=True)
                    except Exception as exc:
                        _log.warning("auto_sync_loop: list_users failed: %s", exc)
                        continue
                    for poll_user in owner_users_for_poll:
                        try:
                            polling_user_id = int(poll_user.get("id") or 0)
                            if polling_user_id <= 0:
                                continue
                            poll_accounts = [
                                item for item in
                                repository.list_marketplace_accounts(polling_user_id, include_secrets=False)
                                if item.get("is_active")
                            ]
                            account_ids = [int(a["id"]) for a in poll_accounts if a.get("id")]
                            if not account_ids:
                                continue
                            poll_sync_settings = repository.get_user_sync_settings(user_id=polling_user_id)
                            polling_since_raw = (
                                str(poll_sync_settings.get("sync_start_date") or "").strip()
                                if bool(poll_sync_settings.get("use_sync_start_date"))
                                else None
                            )
                            # Update in-memory state so UI can see polling is active
                            with sync_lock:
                                sync_state["polling_enabled"] = True
                                sync_state["polling_user_id"] = polling_user_id
                                sync_state["polling_account_ids"] = account_ids
                                sync_state["polling_since_date"] = polling_since_raw
                        except Exception:
                            continue
                        run_started_at = _now_iso()
                        try:
                            result = _run_sync_for_user(
                                user_id=polling_user_id,
                                since_date=polling_since_raw or None,
                                account_ids=account_ids,
                                run_started_at=run_started_at,
                            )
                            with sync_lock:
                                sync_state["last_poll_at"] = _now_iso()
                                sync_state["last_poll_result"] = {
                                    "ok": True,
                                    "run_started_at": run_started_at,
                                    "accounts": int(result.get("accounts") or 0),
                                    "success_accounts": int(result.get("success_accounts") or 0),
                                    "failed_accounts": int(result.get("failed_accounts") or 0),
                                    "loaded": int(result.get("loaded") or 0),
                                    "loaded_conversations": int(result.get("loaded_conversations") or 0),
                                    "account_ids": list(account_ids),
                                    "errors": _serialize_sync_error_details(result.get("errors")),
                                    "cancelled": bool(result.get("cancelled")),
                                }
                        except HTTPException as exc:
                            with sync_lock:
                                sync_state["last_poll_at"] = _now_iso()
                                sync_state["last_poll_result"] = {
                                    "ok": False,
                                    "run_started_at": run_started_at,
                                    "error": str(exc.detail),
                                    "account_ids": list(account_ids),
                                }
                        except Exception as exc:
                            with sync_lock:
                                sync_state["last_poll_at"] = _now_iso()
                                sync_state["last_poll_result"] = {
                                    "ok": False,
                                    "run_started_at": run_started_at,
                                    "error": str(exc),
                                    "account_ids": list(account_ids),
                                }
                        # Continue to next owner user (no break — all tenants polled)

            worker = threading.Thread(
                target=_auto_sync_loop,
                name="feedpilot-auto-sync-worker",
                daemon=True,
            )
            auto_sync_worker["thread"] = worker
            worker.start()

    def _template_group_by_id(group_id: str) -> dict[str, object] | None:
        for item in TEMPLATE_GROUPS:
            if str(item.get("id")) == group_id:
                return item
        return None

    def _is_protected_subgroup(group_id: str, subgroup: str) -> bool:
        """Return True if this subgroup must not be deleted by anyone.

        The 'textless_ratings' group has fixed per-star subgroups that are
        required for the review processing pipeline and cannot be removed.
        """
        from .service import ReviewAutomationService as _RAS
        if str(group_id or "").strip() != _RAS.TEXTLESS_GROUP_ID:
            return False
        return str(subgroup or "").strip() in _RAS.TEXTLESS_SUBGROUPS

    def _base_subgroups_for_group(group_id: str) -> list[str]:
        group = _template_group_by_id(group_id)
        if group is None:
            return []
        subgroups_raw = group.get("subgroups")
        if not isinstance(subgroups_raw, list):
            return []
        result: list[str] = []
        for value in subgroups_raw:
            name = str(value).strip()
            if name and name not in result:
                result.append(name)
        return result

    def _all_subgroups_for_group(group_id: str) -> list[dict[str, object]]:
        custom_rows = repository.list_default_template_subgroups(group_id=group_id)
        result: list[dict[str, object]] = []
        seen: set[str] = set()
        for row in custom_rows:
            clean = str((row or {}).get("subgroup") or "").strip()
            if not clean or clean in seen:
                continue
            seen.add(clean)
            subgroup_id = str((row or {}).get("subgroup_id") or "").strip() or None
            result.append({"name": clean, "subgroup_id": subgroup_id})
        if group_id in GENERAL_LOCKED_GROUP_IDS:
            general = GENERAL_LOCKED_SUBGROUP
            reordered = [item for item in result if str(item.get("name") or "") != general]
            general_item = next((item for item in result if str(item.get("name") or "") == general), None)
            if general_item is None:
                general_item = {"name": general, "subgroup_id": None}
            result = [general_item, *reordered]
        if result:
            return result
        # Backward-compatible fallback for old datasets where subgroup registry
        # might still be empty.
        return [{"name": name, "subgroup_id": None} for name in _base_subgroups_for_group(group_id)]
        

    def _validate_subgroup(group_id: str, subgroup: str) -> bool:
        clean_group_id = str(group_id or "").strip()
        clean_subgroup = str(subgroup or "").strip()
        if not clean_group_id or not clean_subgroup:
            return False
        return any(
            str(item.get("name") or "") == clean_subgroup
            for item in _all_subgroups_for_group(clean_group_id)
        )

    def _default_template_seed_rows() -> tuple[list[dict[str, str]], list[dict[str, object]]]:
        rows: list[dict[str, str]] = []
        subgroup_rows: list[dict[str, object]] = []
        for group in TEMPLATE_GROUPS:
            group_id = str(group.get("id") or "")
            subgroups = group.get("subgroups")
            if not group_id or not isinstance(subgroups, list):
                continue
            for subgroup in subgroups:
                name = str(subgroup).strip()
                if not name:
                    continue
                subgroup_rows.append({"group_id": group_id, "subgroup": name, "is_system": True})
                defaults = DEFAULT_TEMPLATE_CONTENT.get(name) or [f"Спасибо за отзыв! Категория: {name}."]
                for text in defaults:
                    clean = str(text or "").strip()
                    if not clean:
                        continue
                    rows.append(
                        {
                            "group_id": group_id,
                            "subgroup": name,
                            "template_text": clean,
                        }
                    )
        return rows, subgroup_rows

    def _ensure_platform_default_templates() -> None:
        seed_rows, subgroup_rows = _default_template_seed_rows()
        if repository.count_default_template_subgroups() == 0:
            repository.ensure_default_template_subgroups(subgroup_rows)
        for group_id in GENERAL_LOCKED_GROUP_IDS:
            repository.ensure_default_template_subgroups(
                [{"group_id": group_id, "subgroup": GENERAL_LOCKED_SUBGROUP}]
            )
            existing = repository.list_default_template_variants(group_id=group_id, subgroup=GENERAL_LOCKED_SUBGROUP)
            if not existing:
                repository.replace_default_subgroup_templates(
                    group_id=group_id,
                    subgroup=GENERAL_LOCKED_SUBGROUP,
                    templates=DEFAULT_TEMPLATE_CONTENT.get(GENERAL_LOCKED_SUBGROUP)
                    or ["Спасибо за ваш отзыв! Мы ценим обратную связь и уже работаем над улучшениями."],
                )
        repository.sync_default_template_subgroups_from_variants()
        if repository.count_default_template_variants(include_inactive=True) > 0:
            return
        seeded = repository.seed_default_templates_from_user_templates()
        if seeded > 0:
            return
        repository.seed_default_template_variants(seed_rows)

    def _build_template_group_items(counts: dict[tuple[str, str], int]) -> list[dict[str, object]]:
        items: list[dict[str, object]] = []
        for group in TEMPLATE_GROUPS:
            group_id = str(group.get("id") or "")
            title = str(group.get("title") or group_id)
            base_subgroups = set(_base_subgroups_for_group(group_id))
            all_subgroups = _all_subgroups_for_group(group_id)
            subgroups: list[dict[str, object]] = []
            for subgroup_item in all_subgroups:
                subgroup_name = str(subgroup_item.get("name") or "").strip()
                if not subgroup_name:
                    continue
                subgroups.append(
                    {
                        "name": subgroup_name,
                        "count": counts.get((group_id, subgroup_name), 0),
                        "subgroup_id": str(subgroup_item.get("subgroup_id") or "").strip() or None,
                        "is_system": subgroup_name in base_subgroups,
                    }
                )
            items.append(
                {
                    "id": group_id,
                    "title": title,
                    "subgroups": subgroups,
                }
            )
        return items

    def _ensure_default_template_variants(user_id: int) -> None:
        _ensure_platform_default_templates()
        repository.copy_default_templates_to_user(user_id=user_id, only_if_empty=True)

    def _parse_recommendation_targets(raw_csv: str) -> list[str]:
        values = str(raw_csv or "").replace(";", ",").replace("\n", ",").split(",")
        result: list[str] = []
        seen: set[str] = set()
        for value in values:
            article = value.strip()
            if not article or article in seen:
                continue
            seen.add(article)
            result.append(article)
        return result

    def _registration_disabled_response() -> HTMLResponse:
        return HTMLResponse(
            build_login_html(error="Самостоятельная регистрация отключена. Пользователей добавляет администратор."),
            status_code=403,
        )

    @app.get("/", response_class=HTMLResponse)
    def landing(request: Request) -> HTMLResponse:
        user = _get_current_user(request)
        if user is not None:
            return RedirectResponse("/app", status_code=302)
        return HTMLResponse(build_landing_html())

    @app.get("/login", response_class=HTMLResponse)
    def login_page(request: Request) -> HTMLResponse:
        user = _get_current_user(request)
        if user is not None:
            return RedirectResponse("/app", status_code=302)
        return HTMLResponse(build_login_html())

    @app.post("/login")
    def login(request: Request, email: str = Form(...), password: str = Form(...)) -> HTMLResponse:
        login_key = _login_attempt_key(request, email)
        if _is_login_blocked(login_key):
            return HTMLResponse(build_login_html(error="Слишком много неудачных попыток входа. Повторите позже."), status_code=429)
        user = repository.get_user_by_email(email)
        if user is None or not verify_password(password, str(user["password_hash"])):
            _record_failed_login(login_key)
            return HTMLResponse(build_login_html(error="Неверная эл. почта или пароль"), status_code=401)

        _clear_failed_login(login_key)
        token = _issue_session(int(user["id"]))
        response = RedirectResponse("/app", status_code=302)
        _set_session_cookie(response, token)
        return response

    @app.get("/register", response_class=HTMLResponse)
    def register_page(request: Request) -> HTMLResponse:
        user = _get_current_user(request)
        if user is not None:
            return RedirectResponse("/app", status_code=302)
        if not self_registration_enabled:
            return _registration_disabled_response()
        return HTMLResponse(build_register_html())

    @app.post("/register")
    def register(email: str = Form(...), password: str = Form(...), password_repeat: str = Form(...)) -> HTMLResponse:
        if not self_registration_enabled:
            return _registration_disabled_response()
        email = email.strip().lower()
        if len(email) < 5 or "@" not in email:
            return HTMLResponse(build_register_html(error="Введите корректную эл. почту"), status_code=400)
        if len(password) < 8:
            return HTMLResponse(build_register_html(error="Пароль должен быть не короче 8 символов"), status_code=400)
        if password != password_repeat:
            return HTMLResponse(build_register_html(error="Пароли не совпадают"), status_code=400)
        if repository.get_user_by_email(email) is not None:
            return HTMLResponse(build_register_html(error="Пользователь уже существует"), status_code=409)

        role = ROLE_ADMIN if repository.count_users() == 0 else ROLE_USER
        user = repository.create_user(email=email, password_hash=hash_password(password), role=role)
        token = _issue_session(int(user["id"]))
        response = RedirectResponse("/app", status_code=302)
        _set_session_cookie(response, token)
        return response

    @app.get("/logout")
    def logout(request: Request) -> RedirectResponse:
        token = request.cookies.get("session_token")
        if token:
            repository.delete_session(token)
        response = RedirectResponse("/", status_code=302)
        secure_cookie = bool(app_config.is_production)
        response.delete_cookie("session_token", samesite="lax", secure=secure_cookie)
        response.delete_cookie(CSRF_COOKIE_NAME, samesite="lax", secure=secure_cookie)
        return response

    @app.get("/app", response_class=HTMLResponse)
    def app_dashboard(request: Request) -> HTMLResponse:
        user = _get_current_user(request)
        if user is None:
            return RedirectResponse("/login", status_code=302)
        return HTMLResponse(build_app_html(user, repository=repository))

    @app.get("/wb-fbs/tsd", response_class=HTMLResponse)
    @app.get("/wb-fbs/tsd/{path:path}", response_class=HTMLResponse)
    def wb_fbs_tsd_page(request: Request, path: str = "") -> HTMLResponse:
        """Lightweight TSD page for WB FBS assembly (warehouse handheld)."""
        del path  # hash routing client-side
        user = _get_current_user(request)
        if user is None:
            return RedirectResponse("/login", status_code=302)
        response = HTMLResponse(build_wb_fbs_tsd_html(user, repository=repository))
        _ensure_csrf_cookie(response, request)
        return response

    @app.get("/admin", response_class=HTMLResponse)
    def admin_page(request: Request) -> HTMLResponse:
        user = _get_current_user(request)
        if user is None:
            return RedirectResponse("/login", status_code=302)
        if user.get("role") != ROLE_ADMIN:
            return HTMLResponse("<h1>Доступ запрещен</h1><p>Нужны права администратора.</p>", status_code=403)
        return HTMLResponse(build_admin_html(user))

    @app.get("/api/me")
    def get_me(request: Request) -> dict[str, object]:
        user = _require_user(request)
        return {
            "id": user["id"],
            "email": user["email"],
            "full_name": user.get("full_name") or "",
            "role": user["role"],
        }

    @app.get("/api/profile")
    def get_profile(request: Request) -> dict[str, object]:
        user = _require_user(request)
        sync_settings = repository.get_user_sync_settings(user_id=int(user["id"]))
        template_variables = repository.list_user_template_variable_values(user_id=int(user["id"]))
        editable_template_variables = [
            item
            for item in template_variables
            if bool(item.get("is_user_editable")) and bool(item.get("is_active"))
        ]
        return {
            "full_name": user.get("full_name") or "",
            "email": user["email"],
            "use_sync_start_date": bool(sync_settings.get("use_sync_start_date")),
            "sync_start_date": str(sync_settings.get("sync_start_date") or "") or None,
            "default_sync_lookback_days": int(sync_settings.get("default_sync_lookback_days") or 7),
            "editable_template_variables": editable_template_variables,
        }

    @app.get("/api/user-sync-settings")
    def get_user_sync_settings(request: Request) -> dict[str, object]:
        user = _require_user(request)
        return repository.get_user_sync_settings(user_id=int(user["id"]))

    @app.put("/api/user-sync-settings")
    def update_user_sync_settings(request: Request, payload: UserSyncSettingsRequest) -> dict[str, object]:
        user = _require_user(request)
        # User-level settings always operate with explicit sync start date.
        # The checkbox toggle was removed from UI, so force enabled mode.
        enabled_sync_start_date = True
        sync_start_date = _parse_sync_start_date_or_none(
            payload.sync_start_date,
            enabled=enabled_sync_start_date,
        )
        updated = repository.save_user_sync_settings(
            user_id=int(user["id"]),
            use_sync_start_date=enabled_sync_start_date,
            sync_start_date=sync_start_date,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        settings = repository.get_user_sync_settings(user_id=int(user["id"]))
        return {"ok": True, "settings": settings}

    @app.get("/api/user/template-variables")
    def user_list_template_variables(request: Request) -> dict[str, object]:
        user = _require_user(request)
        rows = repository.list_user_template_variable_values(user_id=int(user["id"]))
        items = [
            item
            for item in rows
            if bool(item.get("is_active")) and bool(item.get("is_user_editable"))
        ]
        return {"items": items, "count": len(items)}

    @app.put("/api/user/template-variables")
    def user_save_template_variables(
        payload: UserTemplateVariableValuesSaveRequest,
        request: Request,
    ) -> dict[str, object]:
        user = _require_user(request)
        saved = repository.save_user_template_variable_values(
            user_id=int(user["id"]),
            values={str(k): str(v) for k, v in dict(payload.values).items()},
        )
        return {"ok": True, "saved": int(saved)}

    @app.put("/api/profile")
    def update_profile(request: Request, payload: ProfileUpdateRequest) -> dict[str, object]:
        user = _require_user(request)
        user_id = int(user["id"])
        stored_user = repository.get_user_by_id(user_id)
        if stored_user is None:
            raise HTTPException(status_code=404, detail="Пользователь не найден")

        new_email = (payload.email or str(stored_user.get("email") or "")).strip().lower()
        if not new_email or "@" not in new_email:
            raise HTTPException(status_code=400, detail="Введите корректную электронную почту")

        full_name = (payload.full_name or "").strip() or None

        wants_password_change = any(
            value is not None and value != ""
            for value in (payload.current_password, payload.new_password, payload.new_password_repeat)
        )
        password_hash: str | None = None
        if wants_password_change:
            if not payload.current_password:
                raise HTTPException(status_code=400, detail="Введите текущий пароль")
            if not verify_password(payload.current_password, str(stored_user.get("password_hash") or "")):
                raise HTTPException(status_code=400, detail="Текущий пароль неверный")
            if not payload.new_password or len(payload.new_password) < 8:
                raise HTTPException(status_code=400, detail="Новый пароль должен быть не короче 8 символов")
            if payload.new_password != (payload.new_password_repeat or ""):
                raise HTTPException(status_code=400, detail="Новый пароль и подтверждение не совпадают")
            password_hash = hash_password(payload.new_password)

        try:
            updated = repository.update_user_profile(
                user_id=user_id,
                email=new_email,
                full_name=full_name,
                password_hash=password_hash,
            )
        except Exception as exc:
            is_duplicate_error = False
            if psycopg is not None and isinstance(exc, getattr(psycopg, "IntegrityError", (Exception,))):
                is_duplicate_error = True
            if "integrityerror" in str(type(exc)).lower():
                is_duplicate_error = True
            if not is_duplicate_error:
                raise
            raise HTTPException(status_code=409, detail="Эта электронная почта уже используется другим аккаунтом") from exc
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.get("/api/reviews")
    def list_reviews(
        request: Request,
        source: str | None = None,
        priority: str | None = None,
        status: str | None = None,
        category: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        sort: str = "newest",
        page: int = 1,
        page_size: int = 30,
        bucket: str = "all",
        product_search: str = "",
        has_contradiction: int = 0,
    ) -> dict[str, object]:
        user = _require_user(request)
        allowed_page_sizes = {10, 30, 50, 100}
        normalized_page_size = page_size if page_size in allowed_page_sizes else 30
        normalized_bucket = bucket.strip().lower()
        if normalized_bucket not in {"all", "new", "processed"}:
            normalized_bucket = "all"
        normalized_sort = sort.strip().lower()
        if normalized_sort not in {"newest", "oldest", "rating_asc", "rating_desc", "category"}:
            normalized_sort = "newest"

        normalized_date_from = date_from.strip() if date_from else None
        normalized_date_to = date_to.strip() if date_to else None
        for date_value in [normalized_date_from, normalized_date_to]:
            if date_value:
                try:
                    datetime.strptime(date_value, "%Y-%m-%d")
                except ValueError as exc:
                    raise HTTPException(status_code=400, detail="Неверный формат даты. Ожидается YYYY-MM-DD") from exc
        if normalized_date_from and normalized_date_to and normalized_date_from > normalized_date_to:
            raise HTTPException(status_code=400, detail="Дата начала не может быть позже даты окончания")

        normalized_source = source.strip().lower() if source else None
        if normalized_source in {"all", "all_sources"}:
            normalized_source = None

        status_key = status.strip().lower() if status else ""
        status_map: dict[str, list[str] | None] = {
            "": None,
            "all": None,
            "waiting_send": ["waiting_send"],
            "processed_outside_spix": ["processed_outside_spix"],
            "rejected": ["rejected", "ignored"],
            "answered": ["answered_auto", "answered_manual", "answered"],
            "waiting_processing": ["queued_for_operator", "waiting_processing"],
            "generating_answer": ["generating_answer"],
        }
        status_values = status_map.get(status_key)
        if status_values is None and status_key not in {"", "all"}:
            status_values = [status_key]

        account_ids_filter = _manager_allowed_review_account_ids(user)
        # Use owner's user_id for data queries — managers share owner's data
        owner_user_id = _tenant_owner_id(user)
        page_data = service.list_reviews_paginated(
            user_id=owner_user_id,
            source=normalized_source,
            priority=priority,
            status=None,
            statuses=status_values,
            category=category,
            date_from=normalized_date_from,
            date_to=normalized_date_to,
            sort=normalized_sort,
            page=max(page, 1),
            page_size=normalized_page_size,
            bucket=normalized_bucket,
            account_ids=account_ids_filter,
            product_search=product_search.strip() or None,
            has_contradiction=bool(has_contradiction),
        )
        # Enrich each review with a suggested template text for the reply column.
        # Uses a single batch query instead of N separate queries (eliminates N+1).
        user_id_int = owner_user_id
        items = page_data.get("items") or []
        # First pass: collect pairs and expose classified_subgroup
        pairs: list[tuple[str, str | None]] = []
        for item in items:
            meta = item.get("metadata") or {}
            subgroup = str(meta.get("classified_subgroup") or "")
            item["classified_subgroup"] = subgroup
            group_id = str(item.get("category") or "")
            if group_id:
                pairs.append((group_id, subgroup or None))
        # Single DB call — load ALL templates for relevant groups
        # Each review picks independently so different reviews get different templates
        group_ids_needed = list({p[0] for p in pairs if p[0]})
        try:
            tmpl_pool = repository.get_template_pool_for_reviews(
                user_id=user_id_int, group_ids=group_ids_needed
            )
        except Exception:
            tmpl_pool = {}
        # Load contradiction rules once for dynamic check on existing reviews
        try:
            contradiction_map = repository.get_review_contradiction_map(user_id=user_id_int)
        except Exception:
            contradiction_map = {}
        # Second pass: each review gets its own random pick from the pool
        for item in items:
            group_id = str(item.get("category") or "")
            subgroup = str(item.get("classified_subgroup") or "")
            rating_val = item.get("rating")
            # Skip template if contradiction rule matches (flag in metadata OR dynamic check)
            meta = item.get("metadata") or {}
            has_contradiction_flag = bool(meta.get("rating_contradiction"))
            has_contradiction_rule = bool(
                group_id and rating_val is not None
                and int(rating_val) in contradiction_map.get(group_id, set())
            )
            if has_contradiction_flag or has_contradiction_rule:
                item["suggested_reply"] = ""
            elif group_id:
                texts = tmpl_pool.get((group_id, subgroup)) or tmpl_pool.get((group_id, "")) or []
                raw_tpl = random.choice(texts) if texts else ""
                if raw_tpl:
                    _author = str(item.get("author") or "").strip()
                    try:
                        _vars_ctx = repository.build_template_variables_context(
                            user_id=user_id_int,
                            review_author=_author,
                            review_rating=item.get("rating"),
                            review_category=group_id,
                            review_sentiment="",
                            review_tags=None,
                            review_metadata=item.get("metadata") if isinstance(item.get("metadata"), dict) else {},
                        )
                    except Exception:
                        _vars_ctx = {}
                    raw_tpl = _apply_template_substitution(raw_tpl, _author, _vars_ctx)
                item["suggested_reply"] = raw_tpl
            else:
                item["suggested_reply"] = ""

        source_options = service.list_review_sources(user_id=owner_user_id)
        # Enrich with product photo URLs
        try:
            _photo_map = repository.get_product_photo_map(user_id=owner_user_id)
            if _photo_map:
                for _item in page_data["items"]:
                    _meta = _item.get("metadata") or {}
                    _raw = _meta.get("raw") or {} if isinstance(_meta, dict) else {}
                    _pd = (_raw.get("productDetails") or {}) if isinstance(_raw, dict) else {}
                    _keys = [
                        str(_pd.get("supplierArticle") or "").strip(),
                        str(_pd.get("nmId") or "").strip(),
                        str(_raw.get("supplierArticle") or "").strip(),
                    ]
                    _item["product_photo_url"] = next(
                        (_photo_map[k] for k in _keys if k and k in _photo_map), None
                    )
        except Exception:
            pass
        return {
            "items": page_data["items"],
            "count": len(page_data["items"]),
            "total": page_data["total"],
            "page": page_data["page"],
            "page_size": page_data["page_size"],
            "pages": page_data["pages"],
            "new_count": page_data["new_count"],
            "processed_count": page_data["processed_count"],
            "bucket": normalized_bucket,
            "sort": normalized_sort,
            "date_from": normalized_date_from,
            "date_to": normalized_date_to,
            "source": normalized_source,
            "status": status_key or "all",
            "source_options": source_options,
        }

    @app.get("/api/reviews/random-template")
    def reviews_random_template(
        request: Request,
        group_id: str,
        subgroup: str = "",
        review_uid: str = "",
    ) -> dict[str, object]:
        """Return a random template for the given group/subgroup — used by the reply refresh button."""
        user = _require_user(request)
        owner_uid = _tenant_owner_id(user)
        tmpl = repository.get_random_template_variant(
            user_id=owner_uid,
            group_id=group_id.strip(),
            subgroup=subgroup.strip() or None,
        )
        raw_text = str(tmpl.get("template_text") or "") if tmpl else ""
        if raw_text and review_uid.strip():
            review_obj = repository.get_review(user_id=owner_uid, review_uid=review_uid.strip())
            if review_obj:
                _author = str(review_obj.get("author") or "").strip()
                try:
                    _vars_ctx = repository.build_template_variables_context(
                        user_id=owner_uid,
                        review_author=_author,
                        review_rating=review_obj.get("rating"),
                        review_category=group_id.strip(),
                        review_sentiment="",
                        review_tags=None,
                        review_metadata=review_obj.get("metadata") if isinstance(review_obj.get("metadata"), dict) else {},
                    )
                except Exception:
                    _vars_ctx = {}
                raw_text = _apply_template_substitution(raw_text, _author, _vars_ctx)
        else:
            raw_text = re.sub(r'%[A-Z0-9_]{2,50}%', '', raw_text)
        return {"template_text": raw_text}

    @app.post("/api/reviews/{review_uid}/reply")
    def reply_to_review(review_uid: str, request: Request, payload: ConversationReplyRequest) -> dict[str, object]:
        """Send a reply to a WB review directly from the review table."""
        user = _require_user(request)
        owner_uid = _tenant_owner_id(user)
        review_obj = repository.get_review(user_id=owner_uid, review_uid=review_uid)
        if review_obj is None:
            raise HTTPException(status_code=404, detail="Отзыв не найден")
        response_text = str(payload.response_text or "").strip()
        if not response_text:
            raise HTTPException(status_code=400, detail="Текст ответа не может быть пустым")
        account_id = review_obj.get("account_id")
        account = repository.get_marketplace_account(
            user_id=owner_uid,
            account_id=int(account_id),
            include_secrets=True,
        ) if account_id else None
        if account is None:
            raise HTTPException(status_code=404, detail="Кабинет не найден")
        client = service._build_client(account)
        review_input = ReviewInput(
            review_id=str(review_obj.get("external_review_id") or ""),
            text=str(review_obj.get("text") or ""),
            metadata=review_obj.get("metadata") or {},
        )
        try:
            sent = client.send_review_reply(review=review_input, response_text=response_text)
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"Не удалось отправить ответ: {exc}")
        if not sent:
            raise HTTPException(status_code=502, detail="Ответ не был отправлен")
        repository.update_review_manual_reply(
            user_id=owner_uid,
            review_uid=review_uid,
            operator_name=str(user.get("full_name") or user.get("email") or "Оператор"),
            reply_text=response_text,
        )
        return {"ok": True}

    @app.post("/api/reviews/{review_uid}/retry-send")
    def retry_review_send(review_uid: str, request: Request) -> dict[str, object]:
        """Retry a previously failed auto-reply using the saved auto_reply text."""
        user = _require_user(request)
        owner_uid = _tenant_owner_id(user)
        review_obj = repository.get_review(user_id=owner_uid, review_uid=review_uid)
        if review_obj is None:
            raise HTTPException(status_code=404, detail="Отзыв не найден")
        auto_reply_text = str(review_obj.get("auto_reply") or "").strip()
        if not auto_reply_text:
            raise HTTPException(status_code=400, detail="Нет сохранённого текста для повторной отправки")
        account_id = review_obj.get("account_id")
        account = repository.get_marketplace_account(
            user_id=owner_uid,
            account_id=int(account_id),
            include_secrets=True,
        ) if account_id else None
        if account is None:
            raise HTTPException(status_code=404, detail="Кабинет не найден")
        client = service._build_client(account)
        review_input = ReviewInput(
            review_id=str(review_obj.get("external_review_id") or ""),
            text=str(review_obj.get("text") or ""),
            metadata=review_obj.get("metadata") or {},
        )
        try:
            sent = client.send_review_reply(review=review_input, response_text=auto_reply_text)
        except Exception as exc:
            repository.mark_review_send_error(
                user_id=owner_uid,
                review_uid=review_uid,
                error_message=str(exc),
            )
            raise HTTPException(status_code=502, detail=f"Не удалось отправить ответ: {exc}")
        if not sent:
            raise HTTPException(status_code=502, detail="Ответ не был отправлен")
        repository.clear_review_send_error(user_id=owner_uid, review_uid=review_uid)
        repository.update_review_processing_result(
            user_id=owner_uid,
            review_uid=review_uid,
            status="answered_auto",
            auto_reply=auto_reply_text,
        )
        return {"ok": True}

    @app.get("/api/conversations")
    def list_conversations(
        request: Request,
        source: str | None = None,
        account_id: int | None = None,
        kind: str | None = None,
        status: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        sort: str = "newest",
        bucket: str = "new",
        page: int = 1,
        page_size: int = 30,
        search: str | None = None,
    ) -> dict[str, object]:
        user = _require_user(request)
        normalized_source = (source or "").strip().lower()
        if not normalized_source or normalized_source == "all":
            normalized_source = None
        kind_key = (kind or "").strip().lower()
        if not kind_key or kind_key == "all":
            normalized_kind = None
        elif kind_key in {"question", "chat"}:
            normalized_kind = kind_key
        else:
            raise HTTPException(status_code=400, detail="Тип должен быть: вопрос, чат или все")
        status_key = (status or "").strip().lower()
        if not status_key or status_key == "all":
            normalized_status = None
            status_key = "all"
        elif status_key in {"open", "waiting", "closed"}:
            normalized_status = status_key
        else:
            raise HTTPException(status_code=400, detail="Статус должен быть: открыт, ожидает, закрыт или все")
        normalized_sort = sort.strip().lower()
        if normalized_sort not in {"newest", "oldest"}:
            normalized_sort = "newest"
        normalized_date_from = date_from.strip() if date_from else None
        normalized_date_to = date_to.strip() if date_to else None
        for date_value in [normalized_date_from, normalized_date_to]:
            if date_value:
                try:
                    datetime.strptime(date_value, "%Y-%m-%d")
                except ValueError as exc:
                    raise HTTPException(status_code=400, detail="Неверный формат даты. Ожидается YYYY-MM-DD") from exc
        if normalized_date_from and normalized_date_to and normalized_date_from > normalized_date_to:
            raise HTTPException(status_code=400, detail="Дата начала не может быть позже даты окончания")
        normalized_bucket = bucket.strip().lower()
        if normalized_bucket not in {"all", "new", "processed"}:
            normalized_bucket = "new"
        normalized_page_size = page_size if page_size in {10, 30, 50, 100, 200, 500, 1000} else 30
        manager_conversation_scope = _manager_allowed_conversation_accounts(user)
        # Use owner's user_id for data queries — managers share owner's data
        conv_owner_user_id = _tenant_owner_id(user)
        page_data = repository.list_conversations_paginated(
            user_id=conv_owner_user_id,
            source=normalized_source,
            account_id=account_id,
            kind=normalized_kind,
            status=normalized_status,
            statuses=None,
            date_from=normalized_date_from,
            date_to=normalized_date_to,
            sort=normalized_sort,
            page=max(page, 1),
            page_size=normalized_page_size,
            bucket=normalized_bucket,
            search=str(search).strip() if search else None,
            account_permissions=manager_conversation_scope,
        )
        source_options = repository.list_conversation_sources(user_id=conv_owner_user_id)
        account_options = repository.list_conversation_accounts(
            user_id=conv_owner_user_id, kind=normalized_kind
        )
        # For answered questions: enrich with last sent text from conversation_messages
        # (text answered via our system) and portal reply from metadata.raw.answer.text
        items = page_data["items"]
        if normalized_kind == "question" and normalized_bucket == "processed":
            answered_uids = [
                item["conversation_uid"] for item in items
                if item.get("processed_by_operator") or item.get("last_sent_at")
            ]
            if answered_uids:
                try:
                    sent_texts = repository.get_last_sent_text_for_conversations(
                        user_id=conv_owner_user_id, conversation_uids=answered_uids
                    )
                    for item in items:
                        uid = item["conversation_uid"]
                        item["last_sent_text"] = sent_texts.get(uid, "")
                except Exception:
                    pass
            # For YM processed questions without reply text: try to fetch from YM API
            ym_items_no_reply = [
                item for item in items
                if str(item.get("source") or "").lower() == "yandex"
                and not str(item.get("last_sent_text") or "").strip()
            ]
            if ym_items_no_reply:
                try:
                    acct_id = int(ym_items_no_reply[0].get("account_id") or 0)
                    ym_acct = repository.get_marketplace_account(
                        user_id=conv_owner_user_id, account_id=acct_id, include_secrets=True
                    ) if acct_id else None
                    if ym_acct and str(ym_acct.get("marketplace") or "") == "yandex":
                        from .service import ReviewAutomationService as _RAS
                        _ym_client = _RAS._build_client(ym_acct)
                        for item in ym_items_no_reply:
                            ext_id = str(item.get("external_conversation_id") or "").strip()
                            if not ext_id:
                                continue
                            try:
                                answers = getattr(_ym_client, "fetch_question_answers", lambda _: [])(ext_id)
                                if answers:
                                    # Find the BUSINESS seller answer
                                    for ans in answers:
                                        if str((ans.get("author") or {}).get("type") or "").upper() in ("BUSINESS", "SELLER"):
                                            item["last_sent_text"] = str(ans.get("text") or "")
                                            break
                                    else:
                                        # First answer if no explicit business type
                                        item["last_sent_text"] = str(answers[0].get("text") or "")
                            except Exception:
                                pass
                except Exception:
                    pass
        # Enrich questions with product photo URLs
        if normalized_kind == "question":
            try:
                _photo_map = repository.get_product_photo_map(user_id=conv_owner_user_id)
                if _photo_map:
                    for _item in items:
                        _meta = _item.get("metadata") or {}
                        _raw = (_meta.get("raw") or {}) if isinstance(_meta, dict) else {}
                        _pd = (_raw.get("productDetails") or {}) if isinstance(_raw, dict) else {}
                        _qi = (_raw.get("questionIdentifiers") or {}) if isinstance(_raw, dict) else {}
                        _keys = [
                            str(_raw.get("sku") or "").strip(),                    # Ozon SKU
                            str(_pd.get("supplierArticle") or "").strip(),          # WB supplier article
                            str(_pd.get("nmId") or "").strip(),                     # WB nmId
                            str(_raw.get("supplierArticle") or "").strip(),         # WB supplier article alt
                            str(_qi.get("offerId") or "").strip(),                  # YM offerId
                            str(_raw.get("offerId") or "").strip(),                 # YM offerId alt
                        ]
                        _item["product_photo_url"] = next(
                            (_photo_map[k] for k in _keys if k and k in _photo_map), None
                        )
            except Exception:
                pass
        return {
            "items": items,
            "count": len(items),
            "total": page_data["total"],
            "page": page_data["page"],
            "page_size": page_data["page_size"],
            "pages": page_data["pages"],
            "new_count": page_data["new_count"],
            "processed_count": page_data["processed_count"],
            "bucket": normalized_bucket,
            "sort": normalized_sort,
            "date_from": normalized_date_from,
            "date_to": normalized_date_to,
            "source": normalized_source or "all",
            "status": status_key,
            "kind": normalized_kind or "all",
            "source_options": source_options,
            "account_options": account_options,
        }

    @app.post("/api/conversations/{conversation_uid}/status")
    def set_conversation_status(
        conversation_uid: str,
        payload: ConversationStatusRequest,
        request: Request,
    ) -> dict[str, object]:
        user = _require_user(request)
        _require_manager_scope_for_conversation(user, conversation_uid)
        owner_uid = _tenant_owner_id(user)
        status_value = payload.status.strip().lower()
        if status_value not in {"open", "waiting", "closed"}:
            raise HTTPException(status_code=400, detail="Статус должен быть: открыт, ожидает или закрыт")
        updated = repository.update_conversation_status(
            user_id=owner_uid,
            conversation_uid=conversation_uid,
            status=status_value,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Диалог не найден")
        repository.log_review_action(
            user_id=owner_uid,
            review_uid=conversation_uid,
            action_type="conversation_status",
            actor=str(user["email"]),
            details={"status": status_value},
        )
        return {"ok": True}

    @app.post("/api/conversations/{conversation_uid}/mark-answered")
    def mark_conversation_answered(
        conversation_uid: str,
        request: Request,
    ) -> dict[str, object]:
        """Move a chat to the 'answered' bucket by setting last_sent_at = now.

        Useful for ad/promo chats where the seller does not need to reply but
        wants to clear them from the 'needs reply' queue.
        """
        user = _require_user(request)
        _require_manager_scope_for_conversation(user, conversation_uid)
        owner_uid = _tenant_owner_id(user)
        conversation = repository.get_conversation(
            user_id=owner_uid, conversation_uid=conversation_uid
        )
        if conversation is None:
            raise HTTPException(status_code=404, detail="Диалог не найден")
        updated = repository.mark_conversation_answered(
            user_id=owner_uid,
            conversation_uid=conversation_uid,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Диалог не найден")
        repository.log_review_action(
            user_id=owner_uid,
            review_uid=conversation_uid,
            action_type="conversation_mark_answered",
            actor=str(user["email"]),
            details={"manual": True},
        )
        return {"ok": True}

    @app.post("/api/conversations/{conversation_uid}/move-to-new")
    def move_conversation_to_new(
        conversation_uid: str,
        request: Request,
    ) -> dict[str, object]:
        """Move a chat to the 'new' bucket by clearing last_sent_at.

        Used when the operator manually moves an answered chat back to New,
        e.g. if they want to re-process it.
        """
        user = _require_user(request)
        _require_manager_scope_for_conversation(user, conversation_uid)
        owner_uid = _tenant_owner_id(user)
        conversation = repository.get_conversation(
            user_id=owner_uid,
            conversation_uid=conversation_uid,
        )
        if conversation is None:
            raise HTTPException(status_code=404, detail="Диалог не найден")
        updated = repository.move_conversation_to_new(
            user_id=owner_uid,
            conversation_uid=conversation_uid,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Диалог не найден")
        repository.log_review_action(
            user_id=owner_uid,
            review_uid=conversation_uid,
            action_type="conversation_move_to_new",
            actor=str(user["email"]),
            details={"manual": True},
        )
        return {"ok": True}

    @app.post("/api/conversations/{conversation_uid}/reply")
    def reply_conversation(
        conversation_uid: str,
        payload: ConversationReplyRequest,
        request: Request,
    ) -> dict[str, object]:
        user = _require_user(request)
        _require_manager_scope_for_conversation(user, conversation_uid)
        owner_uid = _tenant_owner_id(user)
        operator_name = str(user.get("full_name") or user.get("email") or "").strip() or "Продавец"
        idempotency_key = (payload.idempotency_key or "").strip() or f"{conversation_uid}:{int(time.time() * 1000)}"
        result = service.send_conversation_reply(
            user_id=owner_uid,
            conversation_uid=conversation_uid,
            response_text=payload.response_text,
            operator_name=operator_name,
            idempotency_key=idempotency_key,
        )
        if not bool(result.get("ok")):
            raise HTTPException(status_code=502, detail=str(result.get("error") or "Не удалось отправить ответ в диалог"))
        return {
            "ok": True,
            "status": result.get("status"),
            "deduplicated": bool(result.get("deduplicated")),
            "idempotency_key": idempotency_key,
        }

    @app.get("/api/wb-image")
    def wb_image_proxy(request: Request, id: str, account_id: int) -> object:
        """Proxy WB chat images via /api/v1/seller/download/{id}.

        WB returns image URLs pointing to internal K8s addresses that are
        not publicly reachable. The downloadID field allows fetching via
        the public buyer-chat-api endpoint with Authorization header.
        """
        from fastapi.responses import Response as _Response
        user = _require_user(request)
        clean_id = str(id or "").strip()
        if not clean_id:
            raise HTTPException(status_code=400, detail="Missing image id")
        account = repository.get_marketplace_account(
            user_id=_tenant_owner_id(user),
            account_id=account_id,
            include_secrets=True,
        )
        if account is None or str(account.get("marketplace") or "") != "wb":
            raise HTTPException(status_code=404, detail="WB account not found")
        api_key = str(account.get("api_key") or "").strip()
        extra = account.get("extra") if isinstance(account.get("extra"), dict) else {}
        chats_api_url = str(extra.get("chats_api_url") or "https://buyer-chat-api.wildberries.ru").rstrip("/")
        if not api_key:
            raise HTTPException(status_code=400, detail="WB credentials missing")
        download_url = f"{chats_api_url}/api/v1/seller/download/{clean_id}"
        try:
            req = urllib.request.Request(
                download_url,
                method="GET",
                headers={"Authorization": api_key},
            )
            with urllib.request.urlopen(req, timeout=15) as resp:
                content = resp.read()
                content_type = resp.headers.get("Content-Type", "image/jpeg")
            return _Response(content=content, media_type=content_type)
        except urllib.error.HTTPError as exc:
            _log.warning("wb_image_proxy: HTTP %d for id=%s", exc.code, clean_id)
            raise HTTPException(status_code=502, detail=f"WB returned HTTP {exc.code}")
        except Exception as exc:
            _log.warning("wb_image_proxy: error %s for id=%s", exc, clean_id)
            raise HTTPException(status_code=502, detail="Failed to fetch WB image")

    @app.get("/api/ozon-image")
    def ozon_image_proxy(request: Request, url: str, account_id: int) -> object:
        """Proxy Ozon chat images that require Client-Id/Api-Key authentication.

        The browser cannot load Ozon image URLs directly (they return 401).
        This endpoint fetches the image server-side using the stored credentials
        and streams it back to the browser.
        """
        from fastapi.responses import Response as _Response
        user = _require_user(request)
        if not url.startswith("https://api-seller.ozon.ru/"):
            raise HTTPException(status_code=400, detail="Invalid Ozon image URL")
        account = repository.get_marketplace_account(
            user_id=_tenant_owner_id(user),
            account_id=account_id,
            include_secrets=True,
        )
        if account is None or str(account.get("marketplace") or "") != "ozon":
            raise HTTPException(status_code=404, detail="Ozon account not found")
        api_key = str(account.get("api_key") or "").strip()
        extra = account.get("extra") if isinstance(account.get("extra"), dict) else {}
        client_id = str(extra.get("client_id") or "").strip()
        _log.info(
            "ozon_image_proxy: account_id=%d client_id=%r api_key_len=%d",
            account_id, client_id, len(api_key),
        )
        if not api_key or not client_id:
            raise HTTPException(status_code=400, detail="Ozon credentials missing")
        try:
            req = urllib.request.Request(
                url,
                method="GET",
                headers={"Client-Id": client_id, "Api-Key": api_key},
            )
            with urllib.request.urlopen(req, timeout=15) as resp:
                content = resp.read()
                content_type = resp.headers.get("Content-Type", "image/jpeg")
            return _Response(content=content, media_type=content_type)
        except urllib.error.HTTPError as exc:
            _log.warning("ozon_image_proxy: HTTP %d for %s", exc.code, url[:80])
            raise HTTPException(status_code=502, detail=f"Ozon returned HTTP {exc.code}")
        except Exception as exc:
            _log.warning("ozon_image_proxy: error %s for %s", exc, url[:80])
            raise HTTPException(status_code=502, detail="Failed to fetch Ozon image")

    @app.get("/api/conversations/{conversation_uid}/messages")
    def conversation_messages(conversation_uid: str, request: Request, limit: int = 200, refresh: int = 0) -> dict[str, object]:
        user = _require_user(request)
        _require_manager_scope_for_conversation(user, conversation_uid)
        owner_uid = _tenant_owner_id(user)
        conversation = repository.get_conversation(user_id=owner_uid, conversation_uid=conversation_uid)
        if conversation is None:
            raise HTTPException(status_code=404, detail="Диалог не найден")
        messages = repository.list_conversation_messages(
            user_id=owner_uid,
            conversation_uid=conversation_uid,
            limit=limit,
        )
        # For WB chats: fetch events from WB API when:
        # - messages table is empty (first load), OR
        # - refresh=1 parameter passed (force reload of full history), OR
        # - conversation.last_message_at is newer than the newest message in DB
        #   (buyer sent a new message AFTER our reply), OR
        # - no inbound (buyer) messages in DB at all — buyer wrote BEFORE our reply
        #   and auto-sync never saved their message text, OR
        # - unread_count > 0 means buyer has messages we haven't shown yet
        _should_refresh = not messages or bool(refresh)
        if not _should_refresh and messages and str(conversation.get("source") or "") == "wb":
            conv_last_msg = str(conversation.get("last_message_at") or "").strip()
            db_newest = str(messages[-1].get("created_at") or "").strip() if messages else ""
            # If conversation updated more recently than newest DB message → refresh
            if conv_last_msg and db_newest and conv_last_msg > db_newest:
                _should_refresh = True
            # If no inbound (buyer) messages in DB → buyer wrote before our reply
            # and auto-sync never stored their message text → always fetch history
            if not _should_refresh:
                has_inbound = any(str(m.get("direction") or "") == "inbound" for m in messages)
                if not has_inbound:
                    _should_refresh = True
            # If WB still reports unread messages → buyer has messages not yet in DB
            if not _should_refresh:
                unread = int(conversation.get("unread_count") or 0)
                if unread > 0:
                    _should_refresh = True
        if _should_refresh and str(conversation.get("source") or "") == "wb":
            try:
                account_id = conversation.get("account_id")
                ext_id = str(conversation.get("external_conversation_id") or "")
                if account_id and ext_id:
                    account = repository.get_marketplace_account(
                        user_id=owner_uid,
                        account_id=int(account_id),
                        include_secrets=True,
                    )
                    if account:
                        client = service._build_client(account)
                        # Fetch events starting from since_date (or 30 days ago fallback).
                        # This gives only RECENT events — fast (1-2 requests instead of 83).
                        # Use sync_start_date from user settings as the start cursor.
                        user_sync_settings = repository.get_user_sync_settings(user_id=owner_uid)
                        event_since = (
                            str(user_sync_settings.get("sync_start_date") or "").strip()
                            if bool(user_sync_settings.get("use_sync_start_date"))
                            else None
                        )
                        if not event_since:
                            # Default: events from last 30 days
                            event_since = (datetime.now(UTC) - timedelta(days=30)).date().isoformat()
                        # Convert since date to ms timestamp for cursor
                        if hasattr(client, "_to_wb_unix_timestamp"):
                            since_ts = client._to_wb_unix_timestamp(event_since)  # type: ignore[attr-defined]
                            # WB events cursor is in milliseconds
                            resume_cursor = str(since_ts * 1000) if since_ts else None
                        else:
                            resume_cursor = None
                        # Limit to 5 pages max for on-demand fetch — prevents
                        # timeout on chats with thousands of historical events.
                        # The auto-sync cursor keeps messages up-to-date.
                        _page_limit = [0]
                        def _stop_after_5_pages(p: int, m: int) -> None:
                            _page_limit[0] = p
                        sender_map = client._fetch_last_sender_map(  # type: ignore[attr-defined]
                            resume_cursor=resume_cursor,
                            stop_requested=lambda: _page_limit[0] >= 5,
                            page_progress_callback=_stop_after_5_pages,
                        )
                        sender_map.pop("_final_cursor", None)
                        entry = sender_map.get(ext_id, {})
                        wb_events = entry.get("events") or []
                        # Status update is handled by the 60s auto-sync; skip here.
                        history: list[dict[str, object]] = []
                        for ev in wb_events:
                            if not isinstance(ev, dict):
                                continue
                            ev_id = str(ev.get("eventID") or "").strip()
                            ev_sender = str(ev.get("sender") or "").strip().lower()
                            msg = ev.get("message") or {}
                            ev_text = str(msg.get("text") or "").strip()
                            if not ev_text:
                                attachments = msg.get("attachments") or {}
                                images = attachments.get("images") or []
                                if images:
                                    img_parts = [f"[img:{_wb_image_url(img)}]" for img in images if img.get("url") or img.get("downloadID")]
                                    ev_text = " ".join(img_parts) if img_parts else f"[Фото: {len(images)} шт.]"
                                elif attachments.get("goodCard"):
                                    ev_text = f"[Товар: {attachments['goodCard'].get('name', '')}]"
                            ev_ts_raw = ev.get("addTimestamp")
                            ev_ts_ms = int(ev_ts_raw) if ev_ts_raw is not None else 0
                            # Fallback: images[0]['date'] if addTimestamp is 0
                            if not ev_ts_ms and images and images[0].get("date"):
                                ev_iso = _normalize_timestamp(str(images[0]["date"])) or datetime.now(UTC).isoformat()
                            else:
                                ev_iso = _normalize_timestamp(ev_ts_ms) or datetime.now(UTC).isoformat()
                            client_name = str(ev.get("clientName") or "").strip()
                            if not ev_id or not ev_text:
                                continue
                            history.append({
                                "direction": "inbound" if ev_sender == "client" else "outbound",
                                "message_text": ev_text,
                                "idempotency_key": f"wb-event-{ev_id}",
                                "created_at": ev_iso,
                                "operator_name": client_name if ev_sender == "client" else "Продавец",
                            })
                        # Migrate old internal K8s URLs to wb-download tokens
                        repository.fix_wb_internal_photo_urls(
                            user_id=owner_uid,
                            conversation_uid=conversation_uid,
                        )
                        if history:
                            repository.bulk_insert_chat_history_messages(
                                user_id=owner_uid,
                                conversation_uid=conversation_uid,
                                messages=history,
                            )
                            # Move chat to "New" bucket if buyer replied after our last reply
                            try:
                                repository.move_chat_to_new_if_buyer_replied(
                                    user_id=owner_uid,
                                    conversation_uid=conversation_uid,
                                )
                            except Exception:
                                pass
                            messages = repository.list_conversation_messages(
                                user_id=owner_uid,
                                conversation_uid=conversation_uid,
                                limit=limit,
                            )
            except Exception:
                pass
        # For Ozon chats: fetch history from /v3/chat/history when empty or refresh=1
        if _should_refresh and str(conversation.get("source") or "") == "ozon":
            try:
                account_id = conversation.get("account_id")
                ext_id = str(conversation.get("external_conversation_id") or "")
                if account_id and ext_id:
                    account = repository.get_marketplace_account(
                        user_id=owner_uid,
                        account_id=int(account_id),
                        include_secrets=True,
                    )
                    if account:
                            client = service._build_client(account)
                            if hasattr(client, "_request_json") and hasattr(client, "chats_history_path"):
                                hist_body = client._request_json(  # type: ignore[attr-defined]
                                    path=client.chats_history_path,  # type: ignore[attr-defined]
                                    payload={"chat_id": ext_id, "limit": 100, "direction": "Backward"},
                                )
                                ozon_msgs = hist_body.get("messages") or []
                                history_ozon: list[dict[str, object]] = []
                                buyer_uid_web: str = ""
                                order_num_web: str = ""
                                for msg in ozon_msgs:
                                    if not isinstance(msg, dict):
                                        continue
                                    user_info = msg.get("user") or {}
                                    user_type = str(user_info.get("type") or "").lower()
                                    msg_id = str(msg.get("message_id") or "").strip()
                                    msg_ts = str(msg.get("created_at") or "")
                                    msg_text = _parse_ozon_message_text(
                                        msg.get("data"), bool(msg.get("is_image"))
                                    )
                                    if user_type == "customer":
                                        uid = str(user_info.get("id") or "").strip()
                                        if uid and not buyer_uid_web:
                                            buyer_uid_web = uid
                                        ctx = msg.get("context") or {}
                                        on = str(ctx.get("order_number") or "").strip()
                                        if on and not order_num_web:
                                            order_num_web = on
                                    if not msg_id or not msg_text:
                                        continue
                                    direction = "inbound" if user_type == "customer" else "outbound"
                                    history_ozon.append({
                                        "direction": direction,
                                        "message_text": msg_text,
                                        "idempotency_key": f"ozon-msg-{msg_id}",
                                        "created_at": msg_ts,
                                        "operator_name": "" if direction == "inbound" else "Продавец",
                                    })
                                # Update customer_name if still missing
                                if not conversation.get("customer_name"):
                                    new_name = (
                                        f"Заказ {order_num_web}" if order_num_web
                                        else (f"Покупатель {buyer_uid_web}" if buyer_uid_web else None)
                                    )
                                    if new_name:
                                        repository.update_conversation_customer_name(
                                            user_id=owner_uid,
                                            conversation_uid=conversation_uid,
                                            customer_name=new_name,
                                        )
                                # Fix any previously saved messages with old Markdown format
                                repository.fix_ozon_photo_messages(
                                    user_id=owner_uid,
                                    conversation_uid=conversation_uid,
                                )
                                if history_ozon:
                                    repository.bulk_insert_chat_history_messages(
                                        user_id=owner_uid,
                                        conversation_uid=conversation_uid,
                                        messages=history_ozon,
                                    )
                                messages = repository.list_conversation_messages(
                                    user_id=owner_uid,
                                    conversation_uid=conversation_uid,
                                    limit=limit,
                                )
            except Exception:
                pass
        return {
            "conversation": conversation,
            "messages": messages,
            "count": len(messages),
        }

    @app.get("/api/chat-quick-templates")
    def list_chat_quick_templates(request: Request) -> dict[str, object]:
        user = _require_user(request)
        items = repository.list_chat_quick_templates(user_id=int(user["id"]))
        return {"items": items, "count": len(items)}

    @app.post("/api/chat-quick-templates")
    def create_chat_quick_template(request: Request, payload: ChatQuickTemplateCreateRequest) -> dict[str, object]:
        user = _require_user(request)
        name = str(payload.template_name or "").strip()
        text = str(payload.template_text or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Введите название шаблона")
        if not text:
            raise HTTPException(status_code=400, detail="Введите текст шаблона")
        item = repository.add_chat_quick_template(
            user_id=int(user["id"]), template_name=name, template_text=text
        )
        return {"ok": True, "item": item}

    @app.put("/api/chat-quick-templates/{template_id}")
    def update_chat_quick_template(
        template_id: int, request: Request, payload: ChatQuickTemplateUpdateRequest
    ) -> dict[str, object]:
        user = _require_user(request)
        name = str(payload.template_name or "").strip()
        text = str(payload.template_text or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Введите название шаблона")
        if not text:
            raise HTTPException(status_code=400, detail="Введите текст шаблона")
        item = repository.update_chat_quick_template(
            user_id=int(user["id"]),
            template_id=int(template_id),
            template_name=name,
            template_text=text,
        )
        if item is None:
            raise HTTPException(status_code=404, detail="Шаблон не найден")
        return {"ok": True, "item": item}

    @app.delete("/api/chat-quick-templates/{template_id}")
    def delete_chat_quick_template(template_id: int, request: Request) -> dict[str, object]:
        user = _require_user(request)
        deleted = repository.delete_chat_quick_template(user_id=int(user["id"]), template_id=int(template_id))
        if not deleted:
            raise HTTPException(status_code=404, detail="Шаблон не найден")
        return {"ok": True, "deleted": True}

    # ── Review contradiction rules endpoints ─────────────────────────────────

    @app.get("/api/contradiction-rules")
    def list_contradiction_rules(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        try:
            repository._ensure_contradiction_rules_table()
            items = repository.list_review_contradiction_rules(user_id=int(user["id"]))
        except Exception:
            items = []
        return {"items": items}

    @app.post("/api/contradiction-rules")
    def save_contradiction_rule(
        request: Request,
        group_id: str,
        ratings: str,
    ) -> dict[str, object]:
        user = _require_settings_access(request)
        try:
            import json as _json
            ratings_list = [int(r) for r in _json.loads(ratings) if 1 <= int(r) <= 5]
        except Exception:
            raise HTTPException(status_code=400, detail="ratings must be JSON array of ints 1-5")
        if not group_id.strip():
            raise HTTPException(status_code=400, detail="group_id is required")
        repository._ensure_contradiction_rules_table()
        repository.save_review_contradiction_rule(
            user_id=int(user["id"]),
            group_id=group_id.strip(),
            ratings=ratings_list,
        )
        return {"ok": True}

    @app.delete("/api/contradiction-rules")
    def delete_contradiction_rule(request: Request, group_id: str) -> dict[str, object]:
        user = _require_settings_access(request)
        deleted = repository.delete_review_contradiction_rule(
            user_id=int(user["id"]),
            group_id=group_id.strip(),
        )
        if not deleted:
            raise HTTPException(status_code=404, detail="Правило не найдено")
        return {"ok": True}

    # ── Question quick templates endpoints ───────────────────────────────────

    @app.get("/api/question-quick-templates")
    def list_question_quick_templates(request: Request) -> dict[str, object]:
        user = _require_user(request)
        items = repository.list_question_quick_templates(user_id=int(user["id"]))
        return {"items": items, "count": len(items)}

    @app.post("/api/question-quick-templates")
    def create_question_quick_template(request: Request, payload: ChatQuickTemplateCreateRequest) -> dict[str, object]:
        user = _require_user(request)
        name = str(payload.template_name or "").strip()
        text = str(payload.template_text or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Введите название шаблона")
        if not text:
            raise HTTPException(status_code=400, detail="Введите текст шаблона")
        item = repository.add_question_quick_template(
            user_id=int(user["id"]), template_name=name, template_text=text
        )
        return {"ok": True, "item": item}

    @app.put("/api/question-quick-templates/{template_id}")
    def update_question_quick_template(
        template_id: int, request: Request, payload: ChatQuickTemplateUpdateRequest
    ) -> dict[str, object]:
        user = _require_user(request)
        name = str(payload.template_name or "").strip()
        text = str(payload.template_text or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Введите название шаблона")
        if not text:
            raise HTTPException(status_code=400, detail="Введите текст шаблона")
        item = repository.update_question_quick_template(
            user_id=int(user["id"]),
            template_id=int(template_id),
            template_name=name,
            template_text=text,
        )
        if item is None:
            raise HTTPException(status_code=404, detail="Шаблон не найден")
        return {"ok": True, "item": item}

    @app.delete("/api/question-quick-templates/{template_id}")
    def delete_question_quick_template(template_id: int, request: Request) -> dict[str, object]:
        user = _require_user(request)
        deleted = repository.delete_question_quick_template(
            user_id=int(user["id"]), template_id=int(template_id)
        )
        if not deleted:
            raise HTTPException(status_code=404, detail="Шаблон не найден")
        return {"ok": True, "deleted": True}

    # ── Product photos catalog ────────────────────────────────────────────────

    import os as _os

    def _product_photos_dir() -> str:
        """Writable catalog for Settings → Products thumbnails.

        Prefer ``data/product_photos`` (deploy-owned), then legacy ``product_photos``,
        then ``FEEDPILOT_PRODUCT_PHOTOS_DIR`` override.
        """
        override = str(_os.environ.get("FEEDPILOT_PRODUCT_PHOTOS_DIR") or "").strip()
        if override:
            return _os.path.abspath(override)
        root = _os.path.abspath(_os.path.join(_os.path.dirname(__file__), ".."))
        preferred = _os.path.join(root, "data", "product_photos")
        legacy = _os.path.join(root, "product_photos")
        # Prefer the dir we can actually write when both exist.
        for path in (preferred, legacy):
            try:
                _os.makedirs(path, exist_ok=True)
                probe = _os.path.join(path, ".write_probe")
                with open(probe, "wb") as fh:
                    fh.write(b"ok")
                _os.remove(probe)
                return path
            except Exception:
                continue
        return preferred

    def _product_photo_candidate_dirs() -> list[str]:
        root = _os.path.abspath(_os.path.join(_os.path.dirname(__file__), ".."))
        dirs = [
            _product_photos_dir(),
            _os.path.join(root, "data", "product_photos"),
            _os.path.join(root, "product_photos"),
        ]
        out: list[str] = []
        for path in dirs:
            if path and path not in out:
                out.append(path)
        return out

    def _resolve_product_photo_file(photo_path: str) -> str | None:
        name = _os.path.basename(str(photo_path or "").strip())
        if not name or name in {".", ".."}:
            return None
        for folder in _product_photo_candidate_dirs():
            fpath = _os.path.join(folder, name)
            if _os.path.isfile(fpath):
                return fpath
        return None

    def _ensure_product_photos_table() -> None:
        try:
            with repository._connect() as conn:
                repository._migrate_product_photos(conn)
                repository._migrate_product_categories(conn)
        except Exception:
            pass

    def _parse_product_box_qty(raw: object) -> int | None:
        text = str(raw or "").strip()
        if not text:
            return None
        try:
            value = int(text)
        except (TypeError, ValueError) as exc:
            raise HTTPException(
                status_code=400, detail="Кратность в коробе должна быть целым числом"
            ) from exc
        if value < 0:
            raise HTTPException(
                status_code=400, detail="Кратность в коробе не может быть отрицательной"
            )
        return value

    def _parse_product_category(raw: object, *, owner_uid: int) -> str:
        value = str(raw or "").strip()
        if not value:
            return ""
        allowed = {
            str(c.get("name") or "").strip()
            for c in repository.list_product_categories(user_id=owner_uid, seed_defaults=True)
            if str(c.get("name") or "").strip()
        }
        if value not in allowed:
            raise HTTPException(status_code=400, detail="Неизвестная категория товара")
        return value

    async def _save_product_photo_upload(photo: UploadFile) -> str:
        """Resize upload to WebP and store under product photos dir. Raises HTTPException on failure."""
        import io as _io
        import uuid as _uuid

        filename = str(photo.filename or "").strip()
        if not filename:
            raise HTTPException(status_code=400, detail="Не выбран файл фото")
        try:
            content = await photo.read()
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Не удалось прочитать файл фото") from exc
        if not content:
            raise HTTPException(status_code=400, detail="Пустой файл фото")
        if len(content) > 12 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="Фото больше 12 МБ")
        photo_dir = _product_photos_dir()
        try:
            from PIL import Image as _PilImage

            resample = getattr(getattr(_PilImage, "Resampling", _PilImage), "LANCZOS", _PilImage.LANCZOS)
            img = _PilImage.open(_io.BytesIO(content))
            img = img.convert("RGB")
            img.thumbnail((200, 200), resample)
            _os.makedirs(photo_dir, exist_ok=True)
            fname = f"{_uuid.uuid4().hex}.webp"
            fpath = _os.path.join(photo_dir, fname)
            try:
                img.save(fpath, "WEBP", quality=85)
            except Exception:
                # Some hosts ship Pillow without WebP encoder.
                fname = f"{_uuid.uuid4().hex}.jpg"
                fpath = _os.path.join(photo_dir, fname)
                img.save(fpath, "JPEG", quality=85)
            if not _os.path.isfile(fpath):
                raise RuntimeError(f"file not written: {fpath}")
            return fname
        except HTTPException:
            raise
        except PermissionError as exc:
            _log.warning("product photo permission failed dir=%s: %s", photo_dir, exc)
            raise HTTPException(
                status_code=500,
                detail="Нет прав на запись фото на сервере. Обратитесь к администратору.",
            ) from exc
        except Exception as exc:
            _log.warning("product photo processing failed: %s", exc)
            raise HTTPException(
                status_code=400,
                detail=(
                    "Не удалось обработать фото. "
                    "Используйте JPG, PNG или WEBP (не HEIC/AVIF)."
                ),
            ) from exc

    @app.get("/api/product-categories")
    def list_product_categories_api(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        _ensure_product_photos_table()
        owner_uid = _tenant_owner_id(user)
        items = repository.list_product_categories(user_id=owner_uid, seed_defaults=True)
        return {"items": items}

    @app.put("/api/product-categories")
    def save_product_categories_api(
        request: Request, payload: ProductCategoriesSaveRequest
    ) -> dict[str, object]:
        user = _require_settings_access(request)
        _ensure_product_photos_table()
        owner_uid = _tenant_owner_id(user)
        try:
            items = repository.save_product_categories(
                user_id=owner_uid,
                items=[item.model_dump() for item in (payload.items or [])],
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return {"ok": True, "items": items}

    @app.get("/api/products")
    def _list_products_ensure(request: Request) -> dict[str, object]:
        _ensure_product_photos_table()
        user = _require_settings_access(request)
        items = repository.list_product_photos(user_id=_tenant_owner_id(user))
        for item in items:
            item["photo_url"] = f"/api/products/photo/{item['id']}" if item.get("photo_path") else None
        return {"items": items}

    @app.post("/api/products")
    async def add_product(
        request: Request,
        name: str = Form(""),
        supplier_article: str = Form(""),
        wb_nmid: str = Form(""),
        ozon_sku: str = Form(""),
        yandex_offer_id: str = Form(""),
        box_qty: str = Form(""),
        product_category: str = Form(""),
        skip_kiz_gtin_check: str = Form(""),
        photo: UploadFile | None = File(None),
    ) -> dict[str, object]:
        user = _require_settings_access(request)
        _ensure_product_photos_table()
        owner_uid = _tenant_owner_id(user)
        photo_path: str | None = None
        if photo is not None and str(photo.filename or "").strip():
            photo_path = await _save_product_photo_upload(photo)
        parsed_box_qty = _parse_product_box_qty(box_qty)
        parsed_category = _parse_product_category(product_category, owner_uid=owner_uid)
        skip_gtin = str(skip_kiz_gtin_check or "").strip().lower() in ("1", "true", "yes", "on")
        item = repository.add_product_photo(
            user_id=owner_uid, name=name.strip(), supplier_article=supplier_article.strip(),
            wb_nmid=wb_nmid.strip(), ozon_sku=ozon_sku.strip(),
            yandex_offer_id=yandex_offer_id.strip(), photo_path=photo_path,
            box_qty=parsed_box_qty, product_category=parsed_category,
            skip_kiz_gtin_check=skip_gtin,
        )
        if item:
            item["photo_url"] = f"/api/products/photo/{item['id']}" if item.get("photo_path") else None
        return {"ok": True, "item": item}

    @app.put("/api/products/{product_id}")
    async def update_product(
        product_id: int,
        request: Request,
        name: str = Form(""),
        supplier_article: str = Form(""),
        wb_nmid: str = Form(""),
        ozon_sku: str = Form(""),
        yandex_offer_id: str = Form(""),
        box_qty: str = Form(""),
        product_category: str = Form(""),
        skip_kiz_gtin_check: str = Form(""),
        photo: UploadFile | None = File(None),
    ) -> dict[str, object]:
        user = _require_settings_access(request)
        owner_uid = _tenant_owner_id(user)
        new_photo_path: str | None = None
        if photo is not None and str(photo.filename or "").strip():
            new_photo_path = await _save_product_photo_upload(photo)
        parsed_box_qty = _parse_product_box_qty(box_qty)
        parsed_category = _parse_product_category(product_category, owner_uid=owner_uid)
        skip_gtin = str(skip_kiz_gtin_check or "").strip().lower() in ("1", "true", "yes", "on")
        ok = repository.update_product_photo(
            user_id=owner_uid, product_id=product_id, name=name.strip(),
            supplier_article=supplier_article.strip(), wb_nmid=wb_nmid.strip(),
            ozon_sku=ozon_sku.strip(), yandex_offer_id=yandex_offer_id.strip(),
            photo_path=new_photo_path,
            box_qty=parsed_box_qty, product_category=parsed_category,
            skip_kiz_gtin_check=skip_gtin,
        )
        if not ok:
            raise HTTPException(status_code=404, detail="Товар не найден")
        item = next(
            (
                i
                for i in repository.list_product_photos(user_id=owner_uid)
                if int(i.get("id") or 0) == int(product_id)
            ),
            None,
        )
        if item:
            item["photo_url"] = f"/api/products/photo/{item['id']}" if item.get("photo_path") else None
        return {"ok": True, "item": item}

    @app.delete("/api/products/{product_id}")
    def delete_product(product_id: int, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        deleted = repository.delete_product_photo(user_id=_tenant_owner_id(user), product_id=product_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Товар не найден")
        # Delete physical file from any known photos dir.
        if deleted.get("photo_path"):
            fpath = _resolve_product_photo_file(str(deleted.get("photo_path") or ""))
            if fpath:
                try:
                    _os.remove(fpath)
                except Exception:
                    pass
        return {"ok": True}

    @app.get("/api/products/photo/{product_id}")
    def product_photo(product_id: int, request: Request) -> object:
        from fastapi.responses import FileResponse as _FileResp
        _require_user(request)
        items = repository.list_product_photos(user_id=_tenant_owner_id(_require_user(request)))
        item = next((i for i in items if i.get("id") == product_id), None)
        if not item or not item.get("photo_path"):
            raise HTTPException(status_code=404, detail="Фото не найдено")
        fpath = _resolve_product_photo_file(str(item.get("photo_path") or ""))
        if not fpath:
            raise HTTPException(status_code=404, detail="Файл не найден")
        ext = _os.path.splitext(fpath)[1].lower()
        media = {
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".webp": "image/webp",
            ".gif": "image/gif",
        }.get(ext, "image/webp")
        return _FileResp(fpath, media_type=media)

    # ── Feedback materials (Настройки → Материалы) ───────────────────────────

    @app.get("/api/materials")
    def list_materials(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        repository.ensure_supply_balances_tables()
        items = repository.list_feedback_materials(user_id=_tenant_owner_id(user))
        return {"items": items}

    @app.post("/api/materials")
    def create_material(request: Request, payload: FeedbackMaterialRequest) -> dict[str, object]:
        user = _require_settings_access(request)
        name = str(payload.name or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Введите наименование")
        repository.ensure_supply_balances_tables()
        item = repository.add_feedback_material(
            user_id=_tenant_owner_id(user),
            name=name,
            unit=str(payload.unit or "шт").strip() or "шт",
        )
        return {"ok": True, "item": item}

    @app.put("/api/materials/{material_id}")
    def update_material(
        material_id: int, request: Request, payload: FeedbackMaterialRequest
    ) -> dict[str, object]:
        user = _require_settings_access(request)
        name = str(payload.name or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Введите наименование")
        ok = repository.update_feedback_material(
            user_id=_tenant_owner_id(user),
            material_id=material_id,
            name=name,
            unit=str(payload.unit or "шт").strip() or "шт",
        )
        if not ok:
            raise HTTPException(status_code=404, detail="Материал не найден")
        return {"ok": True}

    @app.delete("/api/materials/{material_id}")
    def delete_material(material_id: int, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        ok = repository.delete_feedback_material(
            user_id=_tenant_owner_id(user), material_id=material_id
        )
        if not ok:
            raise HTTPException(status_code=404, detail="Материал не найден")
        return {"ok": True}

    # Enrich /api/reviews and /api/conversations with product_photo_url
    # (done inline in list_reviews and list_conversations endpoints)

    # ── Review quick templates ────────────────────────────────────────────────

    def _ensure_review_quick_templates_table() -> None:
        try:
            with repository._connect() as conn:
                repository._migrate_review_quick_templates(conn)
        except Exception:
            pass

    @app.get("/api/review-quick-templates")
    def list_review_quick_templates(request: Request) -> dict[str, object]:
        user = _require_user(request)
        _ensure_review_quick_templates_table()
        items = repository.list_review_quick_templates(user_id=int(user["id"]))
        return {"items": items, "count": len(items)}

    @app.post("/api/review-quick-templates")
    def create_review_quick_template(request: Request, payload: ChatQuickTemplateCreateRequest) -> dict[str, object]:
        user = _require_user(request)
        _ensure_review_quick_templates_table()
        name = str(payload.template_name or "").strip()
        text = str(payload.template_text or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Введите название шаблона")
        if not text:
            raise HTTPException(status_code=400, detail="Введите текст шаблона")
        item = repository.add_review_quick_template(
            user_id=int(user["id"]), template_name=name, template_text=text
        )
        return {"ok": True, "item": item}

    @app.put("/api/review-quick-templates/{template_id}")
    def update_review_quick_template(
        template_id: int, request: Request, payload: ChatQuickTemplateUpdateRequest
    ) -> dict[str, object]:
        user = _require_user(request)
        name = str(payload.template_name or "").strip()
        text = str(payload.template_text or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Введите название шаблона")
        if not text:
            raise HTTPException(status_code=400, detail="Введите текст шаблона")
        _ensure_review_quick_templates_table()
        with repository._connect() as conn:
            conn.execute(
                repository._sql("""
                UPDATE review_quick_templates
                SET template_name = ?, template_text = ?, updated_at = ?
                WHERE id = ? AND user_id = ?
                """),
                (name, text, _now_iso(), int(template_id), int(user["id"])),
            )
        items = repository.list_review_quick_templates(user_id=int(user["id"]))
        updated = next((i for i in items if i["id"] == int(template_id)), None)
        if updated is None:
            raise HTTPException(status_code=404, detail="Шаблон не найден")
        return {"ok": True, "item": updated}

    @app.delete("/api/review-quick-templates/{template_id}")
    def delete_review_quick_template(template_id: int, request: Request) -> dict[str, object]:
        user = _require_user(request)
        deleted = repository.delete_review_quick_template(
            user_id=int(user["id"]), template_id=int(template_id)
        )
        if not deleted:
            raise HTTPException(status_code=404, detail="Шаблон не найден")
        return {"ok": True, "deleted": True}

    @app.post("/api/admin/actions-purge-sync")
    def admin_purge_sync_actions(request: Request) -> dict[str, object]:
        """Delete all sync_review and sync_conversation entries — they are no longer logged."""
        _require_admin(request)
        deleted = repository.purge_sync_action_logs()
        _log.info("admin_purge_sync_actions: deleted %d rows", deleted)
        return {"ok": True, "deleted": deleted}

    @app.post("/api/admin/conversations-clear")
    def admin_clear_conversations(request: Request, payload: ClearConversationsRequest) -> dict[str, object]:
        actor = _require_admin(request)
        if payload.user_id is None:
            target_user_id = _tenant_owner_id(actor) if not _is_super_admin(actor) else int(actor["id"])
        else:
            target_user_id = int(payload.user_id)
            _target_user_for_admin_scope(actor=actor, target_user_id=target_user_id)
        deleted = repository.clear_conversations(user_id=target_user_id, kind=payload.kind, source=payload.source)
        return {"ok": True, "deleted": deleted, "user_id": target_user_id}

    @app.get("/api/analytics")
    def user_analytics(
        request: Request,
        source: str = "",
        date_from: str = "",
        date_to: str = "",
    ) -> dict[str, object]:
        user = _require_analytics_access(request)
        return repository.get_user_analytics(
            user_id=int(user["id"]),
            source=source.strip() or None,
            date_from=date_from.strip() or None,
            date_to=date_to.strip() or None,
        )

    # ── Contest analysis endpoints ────────────────────────────────────────────

    @app.get("/api/analytics/contest/prepare")
    def contest_prepare(
        request: Request,
        source: str = "",
        date_from: str = "",
        date_to: str = "",
    ) -> dict[str, object]:
        """Return review count + cached run info for the given filters."""
        user = _require_analytics_access(request)
        uid = int(user["id"])
        src = source.strip() or None
        df = date_from.strip() or None
        dt = date_to.strip() or None

        # Check Yandex AI config
        ai_settings = repository.get_ai_settings(include_secrets=True)
        has_gpt = bool(ai_settings.get("yandex_api_key") and ai_settings.get("yandex_folder_id"))

        reviews = repository.list_reviews_for_contest(user_id=uid, source=src, date_from=df, date_to=dt)
        count = len(reviews)

        # Check if already analyzed
        cached = repository.find_cached_contest_run(
            user_id=uid, source=src or "", date_from=df or "", date_to=dt or ""
        )
        return {
            "count": count,
            "has_gpt": has_gpt,
            "cached_run_id": int(cached["id"]) if cached else None,
            "cached_violations": int(cached["violations_found"]) if cached else 0,
        }

    @app.post("/api/analytics/contest/start")
    def contest_start(
        request: Request,
        source: str = "",
        date_from: str = "",
        date_to: str = "",
    ) -> dict[str, object]:
        """Start background contest analysis. Returns run_id immediately."""
        user = _require_analytics_access(request)
        uid = int(user["id"])
        src = source.strip() or None
        df = date_from.strip() or None
        dt = date_to.strip() or None

        ai_settings = repository.get_ai_settings(include_secrets=True)
        api_key = str(ai_settings.get("yandex_api_key") or "").strip()
        folder_id = str(ai_settings.get("yandex_folder_id") or "").strip()
        model_uri = str(ai_settings.get("yandex_model_uri") or "").strip()
        if not api_key or not folder_id:
            raise HTTPException(status_code=400, detail="Яндекс GPT не настроен: укажите API-ключ и ID каталога в настройках.")

        reviews = repository.list_reviews_for_contest(user_id=uid, source=src, date_from=df, date_to=dt)
        if not reviews:
            raise HTTPException(status_code=400, detail="Нет отзывов, подходящих под указанные фильтры.")

        run_id = repository.create_contest_run(
            user_id=uid, source=src or "", date_from=df or "", date_to=dt or "", total=len(reviews)
        )

        def _bg():
            service.analyze_reviews_for_contest(
                run_id=run_id, user_id=uid, reviews=reviews,
                api_key=api_key, folder_id=folder_id,
                model_uri=model_uri if model_uri else None,
            )

        import threading as _th
        _th.Thread(target=_bg, daemon=True).start()
        return {"ok": True, "run_id": run_id, "total": len(reviews)}

    @app.get("/api/analytics/contest/status/{run_id}")
    def contest_status(request: Request, run_id: int) -> dict[str, object]:
        user = _require_analytics_access(request)
        run = repository.get_contest_run(run_id=run_id, user_id=int(user["id"]))
        if not run:
            raise HTTPException(status_code=404, detail="Анализ не найден.")
        return {
            "run_id": run_id,
            "status": run.get("status"),
            "total": int(run.get("total") or 0),
            "processed": int(run.get("processed") or 0),
            "violations_found": int(run.get("violations_found") or 0),
            "error": run.get("error"),
        }

    @app.get("/api/analytics/contest/export/{run_id}")
    def contest_export(request: Request, run_id: int):
        import io as _io, urllib.parse as _uparse
        from openpyxl import Workbook as _WB
        from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align, Border as _Border, Side as _Side
        from openpyxl.utils import get_column_letter as _gcl
        from starlette.responses import StreamingResponse as _SR

        user = _require_analytics_access(request)
        uid = int(user["id"])
        run = repository.get_contest_run(run_id=run_id, user_id=uid)
        if not run:
            raise HTTPException(status_code=404, detail="Анализ не найден.")
        if run.get("status") != "completed":
            raise HTTPException(status_code=400, detail="Анализ ещё не завершён.")

        results = repository.get_contest_results_with_reviews(run_id=run_id, user_id=uid)
        src = str(run.get("source") or "")
        src_label = {"wb": "ВБ (Wildberries)", "yandex": "ЯМ (Яндекс Маркет)", "ozon": "ОЗОН"}.get(src, src.upper())
        date_from = str(run.get("date_from") or "")
        date_to = str(run.get("date_to") or "")

        def _ru_date(d: str) -> str:
            return f"{d[8:10]}.{d[5:7]}.{d[2:4]}" if d and len(d) >= 10 else d

        LABELS = {
            "profanity":     "Мат / оскорбления / угрозы",
            "offtopic":      "Не по теме товара (доставка МП)",
            "personal_data": "Персональные данные",
            "advertising":   "Реклама / ссылки",
            "spam":          "Спам",
            "false_facts":   "Ложные факты",
            "wrong_product": "Перепутан товар / SKU",
            "fake":          "Признаки фейка",
            "prohibited":    "Запрещённый контент",
        }
        APPEALS = {
            "profanity":
                "Отзыв содержит ненормативную лексику, оскорбления или угрозы, что нарушает правила площадки. Прошу удалить данный отзыв.",
            "offtopic":
                "Отзыв содержит претензии к работе службы доставки / ПВЗ / курьера маркетплейса, а не оценку товара продавца. Согласно правилам площадки, такой контент не должен влиять на рейтинг карточки товара. Прошу удалить отзыв из карточки товара.",
            "personal_data":
                "Отзыв содержит персональные данные третьих лиц (ФИО, адрес, телефон), что нарушает ФЗ №152 «О персональных данных» и правила площадки. Прошу удалить отзыв.",
            "advertising":
                "Отзыв содержит рекламные материалы и/или ссылки на сторонние ресурсы, что нарушает правила публикации отзывов. Прошу удалить отзыв.",
            "spam":
                "Отзыв является спамом и не содержит информации, связанной с товаром. Прошу удалить отзыв.",
            "false_facts":
                "Отзыв содержит заведомо ложные сведения, не соответствующие фактическим данным заказа. Прошу проверить и удалить отзыв.",
            "wrong_product":
                "Отзыв относится к другому товару — покупатель ошибочно привязал его к данной карточке (перепутан SKU). Прошу удалить отзыв из карточки данного товара.",
            "fake":
                "Отзыв имеет признаки фейка: содержание не соответствует реальному опыту покупки данного товара. Прошу проверить наличие подтверждённого заказа у автора отзыва.",
            "prohibited":
                "Отзыв содержит запрещённый контент (экстремизм, дискриминацию, материалы 18+ и т.д.), что нарушает законодательство РФ и правила площадки. Прошу незамедлительно удалить отзыв.",
        }

        # ── Helpers ──────────────────────────────────────────────────────────
        import json as _j
        wb = _WB()
        H_FILL  = _Fill(fill_type="solid", fgColor="1E40AF")
        H2_FILL = _Fill(fill_type="solid", fgColor="DBEAFE")
        H_FONT  = _Font(bold=True, color="FFFFFF", size=11)
        H2_FONT = _Font(bold=True, color="1E40AF", size=10)
        BOLD    = _Font(bold=True, size=10)
        NORM    = _Font(size=10)
        thin    = _Side(style="thin", color="E2E8F0")
        BORDER  = _Border(left=thin, right=thin, top=thin, bottom=thin)
        CENTER  = _Align(horizontal="center", vertical="center", wrap_text=True)
        VCENTER = _Align(vertical="center", wrap_text=True)

        def _hrow(ws, row, values, fill=H_FILL, font=H_FONT):
            for c, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.fill = fill; cell.font = font; cell.border = BORDER; cell.alignment = CENTER

        def _row(ws, row, values):
            for c, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = NORM; cell.border = BORDER; cell.alignment = VCENTER

        # ── Sheet 1: Сводка ───────────────────────────────────────────────────
        ws1 = wb.active; ws1.title = "Сводка"
        ws1.merge_cells("A1:B1")
        ws1["A1"] = "Отчёт для оспаривания отзывов"
        ws1["A1"].font = _Font(bold=True, size=13, color="1E3A8A")
        ws1["A1"].alignment = CENTER
        ws1.row_dimensions[1].height = 22

        meta = [
            ("Источник", src_label),
            ("Период", f"{_ru_date(date_from)} — {_ru_date(date_to)}"),
            ("Всего проверено", int(run.get("total") or 0)),
            ("Потенциально оспариваемых", int(run.get("violations_found") or 0)),
        ]
        for i, (k, v) in enumerate(meta, 3):
            ws1.cell(row=i, column=1, value=k).font = BOLD
            ws1.cell(row=i, column=1).border = BORDER
            ws1.cell(row=i, column=2, value=v).font = NORM
            ws1.cell(row=i, column=2).border = BORDER

        r = len(meta) + 4
        ws1.cell(row=r, column=1, value="По типам нарушений").font = BOLD
        r += 1
        _hrow(ws1, r, ["Тип нарушения", "Кол-во отзывов"], H2_FILL, H2_FONT)
        violation_counts: dict[str, int] = {}
        for res in results:
            for v in (res.get("violations") or []):
                violation_counts[v] = violation_counts.get(v, 0) + 1
        for code, cnt in sorted(violation_counts.items(), key=lambda x: -x[1]):
            r += 1
            _row(ws1, r, [LABELS.get(code, code), cnt])
        ws1.column_dimensions["A"].width = 30
        ws1.column_dimensions["B"].width = 22

        # ── Sheet 2: Отзывы ───────────────────────────────────────────────────
        ws2 = wb.create_sheet("Отзывы к оспариванию")
        _hrow(ws2, 1, ["Дата", "Источник", "Артикул", "Рейтинг", "Нарушения", "Текст отзыва", "Текст жалобы для МП"])
        ws2.row_dimensions[1].height = 18
        for i, res in enumerate(results, 2):
            raw_dt = str(res.get("created_at") or "")
            date_str = _ru_date(raw_dt[:10])
            violations = res.get("violations") or []
            violation_labels = "; ".join(LABELS.get(v, v) for v in violations)
            # Build appeal text from all violation codes
            appeal_parts = []
            for v in violations:
                appeal_text = APPEALS.get(v)
                if appeal_text and appeal_text not in appeal_parts:
                    appeal_parts.append(appeal_text)
            appeal_combined = "\n".join(appeal_parts)
            _row(ws2, i, [
                date_str, src_label,
                res.get("article") or "—",
                res.get("rating") or "—",
                violation_labels,
                res.get("text") or "",
                appeal_combined,
            ])
            ws2.row_dimensions[i].height = 60

        ws2.column_dimensions["A"].width = 12
        ws2.column_dimensions["B"].width = 14
        ws2.column_dimensions["C"].width = 16
        ws2.column_dimensions["D"].width = 9
        ws2.column_dimensions["E"].width = 28
        ws2.column_dimensions["F"].width = 50
        ws2.column_dimensions["G"].width = 55
        ws2.freeze_panes = "A2"

        buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"Оспаривание_{src_label.split('(')[0].strip()}_{_ru_date(date_from)}_{_ru_date(date_to)}.xlsx"
        encoded = _uparse.quote(fname, safe="")
        return _SR(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
        )

    @app.get("/api/analytics/contest/details/{run_id}")
    def contest_details(request: Request, run_id: int) -> list[dict[str, object]]:
        """Return all processed reviews with their GPT results for the detail modal."""
        user = _require_analytics_access(request)
        uid = int(user["id"])
        run = repository.get_contest_run(run_id=run_id, user_id=uid)
        if not run:
            raise HTTPException(status_code=404, detail="Анализ не найден.")
        rows = repository.get_contest_details(run_id=run_id, user_id=uid)
        return [
            {
                "review_uid": r.get("review_uid"),
                "created_at": str(r.get("created_at") or "")[:10],
                "rating": r.get("rating"),
                "article": r.get("article") or "",
                "text": (str(r.get("text") or ""))[:300],
                "violations": r.get("violations") or [],
                "can_contest": bool(r.get("can_contest")),
            }
            for r in rows
        ]

    @app.get("/api/analytics/trend")
    def analytics_trend(
        request: Request,
        source: str = "",
        granularity: str = "week",
    ) -> list[dict[str, object]]:
        user = _require_analytics_access(request)
        gran = granularity.strip().lower()
        if gran not in ("week", "month"):
            gran = "week"
        return repository.get_analytics_trend(
            user_id=int(user["id"]),
            source=source.strip() or None,
            granularity=gran,
        )

    @app.get("/api/analytics/export")
    def analytics_export(
        request: Request,
        source: str = "",
        date_from: str = "",
        date_to: str = "",
    ):
        import io as _io
        import urllib.parse as _uparse
        from openpyxl import Workbook as _WB
        from openpyxl.styles import Font as _Font, PatternFill as _Fill, Alignment as _Align, Border as _Border, Side as _Side
        from openpyxl.utils import get_column_letter as _gcl
        from starlette.responses import StreamingResponse as _SR

        user = _require_analytics_access(request)
        uid = int(user["id"])
        src = source.strip() or None
        df = date_from.strip() or None
        dt = date_to.strip() or None

        if not src or src not in ("wb", "yandex"):
            raise HTTPException(status_code=400, detail="Источник должен быть wb или yandex")

        # ── Data ──────────────────────────────────────────────────────────────
        summary = repository.get_user_analytics(user_id=uid, source=src, date_from=df, date_to=dt)
        reviews = repository.list_reviews_for_export(user_id=uid, source=src, date_from=df, date_to=dt)

        src_label = {"wb": "ВБ (Wildberries)", "yandex": "ЯМ (Яндекс Маркет)"}.get(src, src.upper())
        _CAT_RU = {
            "positive": "Позитив",
            "product_dissatisfaction": "Недовольство товаром",
            "delivery_problems": "Проблемы при доставке",
            "wrong_size": "Неправильный размер",
            "tagged_reviews": "Отзывы с тегами",
            "textless_ratings": "Оценки без текста",
            "negative_delivery": "Негатив: доставка",
            "negative_product": "Негатив: товар",
            "negative_other": "Негатив: прочее",
            "positive_quality": "Позитив: качество",
            "positive_product": "Позитив: товар",
            "neutral_other": "Нейтральный: прочее",
            "ai_unclassified": "Не классифицирован",
        }
        def _cat_ru(key: str) -> str:
            return _CAT_RU.get(str(key or ""), str(key or "—"))
        def _ru_date(d: str | None) -> str:
            if not d or len(d) < 10:
                return d or "—"
            return f"{d[8:10]}.{d[5:7]}.{d[2:4]}"
        period_label = f"{_ru_date(df)} — {_ru_date(dt)}"

        wb = _WB()

        # ── Helpers ───────────────────────────────────────────────────────────
        H_FILL  = _Fill(fill_type="solid", fgColor="1E40AF")
        H2_FILL = _Fill(fill_type="solid", fgColor="DBEAFE")
        H_FONT  = _Font(bold=True, color="FFFFFF", size=11)
        H2_FONT = _Font(bold=True, color="1E40AF", size=10)
        BOLD    = _Font(bold=True, size=10)
        NORM    = _Font(size=10)
        thin    = _Side(style="thin", color="E2E8F0")
        BORDER  = _Border(left=thin, right=thin, top=thin, bottom=thin)
        CENTER  = _Align(horizontal="center", vertical="center", wrap_text=True)
        VCENTER = _Align(vertical="center", wrap_text=True)

        def _hrow(ws, row, values, fill=H_FILL, font=H_FONT):
            for c, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.fill = fill; cell.font = font; cell.border = BORDER; cell.alignment = CENTER

        def _row(ws, row, values):
            for c, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = NORM; cell.border = BORDER; cell.alignment = VCENTER

        def _autofit(ws, min_w=10, max_w=50):
            for col in ws.columns:
                length = max((len(str(cell.value or "")) for cell in col), default=min_w)
                ws.column_dimensions[_gcl(col[0].column)].width = min(max(length + 2, min_w), max_w)

        # ── Sheet 1: Сводка ───────────────────────────────────────────────────
        ws1 = wb.active; ws1.title = "Сводка"
        ws1.merge_cells("A1:B1")
        ws1["A1"] = f"Аналитический отчёт: {src_label}"
        ws1["A1"].font = _Font(bold=True, size=13, color="1E3A8A")
        ws1["A1"].alignment = CENTER
        ws1.row_dimensions[1].height = 22

        meta = [("Период", period_label), ("Источник", src_label),
                ("Всего отзывов", summary.get("total_reviews", 0)),
                ("Обработано", f"{summary.get('processed_reviews',0)} ({summary.get('processed_percent',0)}%)"),
                ("Оценка 4–5 ★", summary.get("high_rating_count", 0)),
                ("Оценка 1–3 ★", summary.get("low_rating_count", 0)),
                ("Вопросов", summary.get("questions_count", 0)),
                ("Чатов", summary.get("chats_count", 0))]
        for i, (k, v) in enumerate(meta, 3):
            ws1.cell(row=i, column=1, value=k).font = BOLD
            ws1.cell(row=i, column=1).border = BORDER
            ws1.cell(row=i, column=2, value=v).font = NORM
            ws1.cell(row=i, column=2).border = BORDER

        r = len(meta) + 4
        ws1.cell(row=r, column=1, value="Распределение по оценкам").font = BOLD
        r += 1
        _hrow(ws1, r, ["Оценка", "Кол-во", "%"], H2_FILL, H2_FONT)
        by_rating = summary.get("by_rating", {})
        total_r = sum(by_rating.values()) or 1
        for star in [5, 4, 3, 2, 1]:
            r += 1
            cnt = by_rating.get(star, 0)
            _row(ws1, r, [f"{star} ★", cnt, f"{round(cnt/total_r*100)}%"])

        r += 2
        ws1.cell(row=r, column=1, value="Топ категорий отзывов").font = BOLD
        r += 1
        _hrow(ws1, r, ["Категория", "Кол-во", "%"], H2_FILL, H2_FONT)
        by_cat = summary.get("by_category", [])
        cat_total = sum(c["count"] for c in by_cat) or 1
        for cat in by_cat[:15]:
            r += 1
            _row(ws1, r, [_cat_ru(cat.get("category", "")), cat["count"], f"{round(cat['count']/cat_total*100)}%"])

        ws1.column_dimensions["A"].width = 28
        ws1.column_dimensions["B"].width = 20
        ws1.column_dimensions["C"].width = 10

        # ── Sheet 2: Отзывы ───────────────────────────────────────────────────
        ws2 = wb.create_sheet("Отзывы")
        _hrow(ws2, 1, ["Дата", "Источник", "Артикул", "Рейтинг", "Категория", "Текст отзыва", "Ответ магазина"])
        ws2.row_dimensions[1].height = 18
        for i, rev in enumerate(reviews, 2):
            raw_dt = str(rev.get("created_at", "") or "")
            date_str = raw_dt[8:10] + "." + raw_dt[5:7] + "." + raw_dt[2:4] if len(raw_dt) >= 10 else raw_dt[:10]
            _row(ws2, i, [
                date_str,
                src_label,
                rev.get("article") or "—",
                rev.get("rating") or "—",
                _cat_ru(rev.get("category") or ""),
                rev.get("text") or "",
                rev.get("reply_text") or "",
            ])
            ws2.row_dimensions[i].height = 40
        ws2.column_dimensions["A"].width = 12
        ws2.column_dimensions["B"].width = 14
        ws2.column_dimensions["C"].width = 18
        ws2.column_dimensions["D"].width = 9
        ws2.column_dimensions["E"].width = 24
        ws2.column_dimensions["F"].width = 50
        ws2.column_dimensions["G"].width = 40
        ws2.freeze_panes = "A2"

        # ── Sheet 3: По категориям ─────────────────────────────────────────────
        ws3 = wb.create_sheet("По категориям")
        _hrow(ws3, 1, ["Категория", "Кол-во отзывов", "% от общего"])
        for i, cat in enumerate(by_cat, 2):
            _row(ws3, i, [_cat_ru(cat.get("category", "")), cat["count"], f"{round(cat['count']/cat_total*100)}%"])
        ws3.column_dimensions["A"].width = 32; ws3.column_dimensions["B"].width = 18; ws3.column_dimensions["C"].width = 14

        buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"Аналитика_{src_label.split()[0]}_{(df or 'all').replace('-','')}_{(dt or 'all').replace('-','')}.xlsx"
        encoded = _uparse.quote(fname, safe="")
        return _SR(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
        )

    @app.post("/api/sync")
    def sync_reviews(request: Request, payload: SyncRequest) -> dict[str, object]:
        user = _require_user(request)
        user_id = int(user["id"])
        user_sync_settings = repository.get_user_sync_settings(user_id=user_id)
        since_date = (
            str(user_sync_settings.get("sync_start_date") or "").strip()
            if bool(user_sync_settings.get("use_sync_start_date"))
            else None
        )
        if payload.all_accounts or payload.account_ids:
            all_active = _snapshot_active_account_ids_for_user(user_id)
            if not all_active:
                raise HTTPException(status_code=400, detail="Нет активных кабинетов для синхронизации")
            # Filter to selected accounts if checkboxes were used
            if payload.account_ids:
                selected = set(int(x) for x in payload.account_ids if x)
                account_ids_snapshot = [aid for aid in all_active if aid in selected]
                if not account_ids_snapshot:
                    raise HTTPException(status_code=400, detail="Ни один из выбранных кабинетов не найден")
            else:
                account_ids_snapshot = all_active
            # Store expected total from preview for the progress bar
            if payload.total_expected is not None and payload.total_expected > 0:
                with sync_lock:
                    sync_state["progress_total_items"] = int(payload.total_expected)
            run_started_at = _now_iso()
            result = _run_sync_for_user(
                user_id=user_id,
                since_date=since_date or None,
                account_ids=account_ids_snapshot,
                run_started_at=run_started_at,
                apply_date_filter=True,  # manual sync applies date filter
            )
            with sync_lock:
                sync_state["polling_enabled"] = True
                sync_state["polling_user_id"] = user_id
                sync_state["polling_account_ids"] = list(account_ids_snapshot)
                sync_state["polling_since_date"] = since_date or None
                sync_state["polling_started_at"] = run_started_at
                sync_state["last_poll_at"] = run_started_at
                sync_state["last_poll_result"] = {
                    "ok": True,
                    "run_started_at": run_started_at,
                    "accounts": int(result.get("accounts") or 0),
                    "success_accounts": int(result.get("success_accounts") or 0),
                    "failed_accounts": int(result.get("failed_accounts") or 0),
                    "loaded": int(result.get("loaded") or 0),
                    "loaded_conversations": int(result.get("loaded_conversations") or 0),
                    "account_ids": list(result.get("account_ids") or account_ids_snapshot),
                    "errors": _serialize_sync_error_details(result.get("errors")),
                    "cancelled": bool(result.get("cancelled")),
                }
            _start_auto_sync_worker_if_needed()
            return result

        if payload.account_id is None:
            raise HTTPException(status_code=400, detail="Необходимо указать идентификатор кабинета")
        account = repository.get_marketplace_account(
            user_id=user_id,
            account_id=payload.account_id,
            include_secrets=True,
        )
        if account is None:
            raise HTTPException(status_code=404, detail="Кабинет маркетплейса не найден")
        marketplace = str(account["marketplace"])
        client = service._build_client(account)
        account_id_val = int(account["id"])
        loaded = 0
        loaded_conversations = 0
        errors: list[str] = []
        try:
            loaded = service.sync_reviews(
                user_id=user_id,
                source=marketplace,
                account_id=account_id_val,
                client=client,
                since_date=since_date or None,
                apply_date_filter=True,  # single-account manual sync: full YM catalog
            )
        except MarketplaceSyncError as exc:
            if not service._is_access_error(exc):
                raise HTTPException(status_code=502, detail=f"Ошибка синхронизации отзывов: {exc}") from exc
            errors.append(str(exc))
        try:
            loaded_conversations = service.sync_conversations(
                user_id=user_id,
                source=marketplace,
                account_id=account_id_val,
                client=client,
                since_date=since_date or None,
                apply_date_filter=True,
            )
        except MarketplaceSyncError as exc:
            if not service._is_access_error(exc):
                raise HTTPException(status_code=502, detail=f"Ошибка синхронизации диалогов: {exc}") from exc
            errors.append(str(exc))
        return {
            "accounts": 1,
            "loaded": loaded,
            "loaded_conversations": loaded_conversations,
            "errors": errors,
        }

    @app.post("/api/sync/capabilities")
    def sync_capabilities(request: Request, payload: SyncCapabilitiesRequest) -> dict[str, object]:
        user = _require_settings_access(request)
        user_id = int(user["id"])
        user_sync_settings = repository.get_user_sync_settings(user_id=user_id)
        since_date = (
            str(user_sync_settings.get("sync_start_date") or "").strip()
            if bool(user_sync_settings.get("use_sync_start_date"))
            else None
        )
        result = _probe_account_capabilities(
            user_id=user_id,
            account_id=int(payload.account_id),
            since_date=since_date or None,
        )
        return {"ok": True, "item": result}

    @app.get("/api/sync/capabilities")
    def sync_capabilities_all(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        user_id = int(user["id"])
        account_ids = _snapshot_active_account_ids_for_user(user_id)
        user_sync_settings = repository.get_user_sync_settings(user_id=user_id)
        since_date = (
            str(user_sync_settings.get("sync_start_date") or "").strip()
            if bool(user_sync_settings.get("use_sync_start_date"))
            else None
        )
        items: list[dict[str, object]] = []
        aggregate_errors: list[dict[str, object]] = []
        any_syncable = False
        for account_id in account_ids:
            item = _probe_account_capabilities(
                user_id=user_id,
                account_id=account_id,
                since_date=since_date or None,
            )
            items.append(item)
            any_syncable = any_syncable or bool(item.get("can_sync_any"))
            raw_item_errors = item.get("errors")
            if isinstance(raw_item_errors, list):
                aggregate_errors.extend(_serialize_sync_error_details(raw_item_errors))
        return {
            "ok": True,
            "items": items,
            "count": len(items),
            "any_syncable": any_syncable,
            "errors": aggregate_errors,
        }

    @app.get("/api/sync/preview")
    def sync_preview(request: Request) -> dict[str, object]:
        """Return estimated counts of items available to sync for all active accounts.

        Uses lightweight count endpoints (no full data download).  Results are
        used to populate the confirmation modal before starting a sync.
        """
        user = _require_user(request)
        user_id = int(user["id"])
        user_sync_settings = repository.get_user_sync_settings(user_id=user_id)
        since_date = (
            str(user_sync_settings.get("sync_start_date") or "").strip()
            if bool(user_sync_settings.get("use_sync_start_date"))
            else None
        )
        accounts = [
            item
            for item in repository.list_marketplace_accounts(user_id=user_id, include_secrets=True)
            if item["is_active"]
        ]
        items: list[dict[str, object]] = []
        total_reviews = 0
        total_questions = 0
        total_chats = 0
        for account in accounts:
            try:
                result = service.count_pending_for_account(
                    account=account,
                    since_date=since_date,
                )
            except Exception as _exc:
                _log.warning(
                    "sync_preview: count_pending_for_account failed account_id=%s: %s",
                    account.get("id"), _exc,
                )
                result = {
                    "account_id": int(account.get("id") or 0),
                    "account_name": str(account.get("account_name") or ""),
                    "marketplace": str(account.get("marketplace") or ""),
                    "reviews": 0,
                    "questions": 0,
                    "chats": 0,
                    "total": 0,
                }
            items.append(result)
            total_reviews += int(result.get("reviews") or 0)
            total_questions += int(result.get("questions") or 0)
            total_chats += int(result.get("chats") or 0)
        return {
            "ok": True,
            "since_date": since_date,
            "accounts": len(items),
            "items": items,
            "total_reviews": total_reviews,
            "total_questions": total_questions,
            "total_chats": total_chats,
            "total": total_reviews + total_questions + total_chats,
        }

    @app.get("/api/sync/status")
    def sync_status_public(request: Request) -> dict[str, object]:
        """Public sync progress endpoint accessible to all logged-in users."""
        _require_user(request)
        with sync_lock:
            return {
                "in_progress": bool(sync_state.get("in_progress")),
                "is_manual": bool(sync_state.get("is_manual")),
                "cancel_requested": bool(sync_state.get("cancel_requested")),
                "last_started_at": sync_state.get("last_started_at"),
                "last_finished_at": sync_state.get("last_finished_at"),
                "step": str(sync_state.get("progress_step") or ""),
                "account": str(sync_state.get("progress_account") or ""),
                "channel": str(sync_state.get("progress_channel") or ""),
                "loaded": int(sync_state.get("progress_loaded") or 0),
                "total_items": int(sync_state.get("progress_total_items") or 0),
                "total_accounts": int(sync_state.get("progress_total_accounts") or 0),
                "current_account": int(sync_state.get("progress_current_account") or 0),
                "last_sync_report": sync_state.get("last_sync_report"),
            }

    @app.get("/api/my/permissions")
    def get_my_permissions(request: Request) -> dict[str, object]:
        """Return the current permission flags for the logged-in user.

        Used by the client to detect permission changes without a full page
        reload.  Owners and admins always receive all-true flags so there is
        no polling overhead for them.
        """
        user = _require_user(request)
        role = str(user.get("role") or ROLE_USER)
        user_id = int(user.get("id") or 0)

        chats_on = sync_chats_enabled()
        if role in ROLE_CAN_ACCESS_SETTINGS:
            return {
                "can_view_feedback": True,
                "can_view_reviews": True,
                "can_view_questions": True,
                "can_view_chats": chats_on,
                "can_view_supplies": True,
                "can_view_any_supply": True,
                "can_view_salary": True,
                "sync_chats_enabled": chats_on,
            }

        _perms = repository.list_manager_permissions(manager_user_id=user_id)
        can_view_reviews = any(bool(p.get("can_reviews")) for p in _perms)
        can_view_questions = any(bool(p.get("can_questions")) for p in _perms)
        can_view_chats = any(bool(p.get("can_chats")) for p in _perms) and chats_on
        can_view_feedback = can_view_reviews or can_view_questions or can_view_chats

        can_view_supplies = bool(user.get("can_supplies"))
        can_view_any_supply = False
        if can_view_supplies:
            _supply_perms = repository.get_manager_supply_permissions(manager_user_id=user_id)
            _sp_sources = _supply_perms.get("sources") or {}
            can_view_any_supply = (
                any(v.get("wb") for v in _sp_sources.values())
                or any(v.get("wb_fbs") for v in _sp_sources.values())
                or any(v.get("wb_fbs_tsd") for v in _sp_sources.values())
                or any(v.get("ozon") for v in _sp_sources.values())
                or bool(_supply_perms.get("can_supply_poa"))
                or bool(_supply_perms.get("can_supply_settings"))
                or bool(_supply_perms.get("can_supply_certs"))
                or bool(user.get("can_supply_planning"))
                or bool(user.get("can_supply_stock"))
            )

        return {
            "can_view_feedback": can_view_feedback,
            "can_view_reviews": can_view_reviews,
            "can_view_questions": can_view_questions,
            "can_view_chats": can_view_chats,
            "can_view_supplies": can_view_supplies,
            "can_view_any_supply": can_view_any_supply,
            "can_supply_planning": bool(user.get("can_supply_planning")) or role in ROLE_CAN_ACCESS_SETTINGS,
            "can_supply_stock": bool(user.get("can_supply_stock")) or role in ROLE_CAN_ACCESS_SETTINGS,
            "stock_productions": (lambda: (
                __import__("json").loads(str(user.get("stock_productions") or "[]"))
            ) if not (role in ROLE_CAN_ACCESS_SETTINGS) else None)(),
            "can_view_salary": bool(user.get("can_salary")),
            "can_salary_report": bool(user.get("can_salary_report")),
            "can_salary_zp_export": bool(user.get("can_salary_zp_export")),
            "sync_chats_enabled": chats_on,
            "salary_productions": (lambda: (
                __import__("json").loads(str(user.get("salary_productions") or "[]"))
            ) if not (role in ROLE_CAN_ACCESS_SETTINGS) else None)(),
        }

    @app.get("/api/accounts")
    def list_accounts(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        items = repository.list_marketplace_accounts(user_id=int(user["id"]), include_secrets=True)
        return {"items": items, "count": len(items)}

    @app.post("/api/accounts")
    def create_account(request: Request, payload: AccountCreateRequest) -> dict[str, object]:
        user = _require_settings_access(request)
        marketplace = payload.marketplace.strip().lower()
        if marketplace not in {"wb", "ozon", "yandex", "mock"}:
            raise HTTPException(status_code=400, detail="Некорректный маркетплейс")
        integration = payload.integration if isinstance(payload.integration, dict) else {}
        default_api_urls = {
            "wb": "https://feedbacks-api.wildberries.ru/api/v1/feedbacks",
            "ozon": "https://api-seller.ozon.ru",
            "yandex": "https://api.partner.market.yandex.ru",
            "mock": "https://example.local/api/reviews",
        }
        api_url = (payload.api_url or "").strip() or str(integration.get("api_url") or default_api_urls[marketplace])
        api_url = _validate_account_api_url(marketplace, api_url)
        if marketplace in {"wb", "ozon", "yandex"} and not (payload.api_key or "").strip():
            raise HTTPException(status_code=400, detail="Для WB/OZON/ЯМ требуется ключ доступа")
        client_id_value = (payload.client_id or "").strip() or str(integration.get("client_id") or "").strip()
        if marketplace == "ozon" and not client_id_value:
            raise HTTPException(status_code=400, detail="Для OZON требуется идентификатор клиента")
        # Yandex Market: require business_id
        business_id_value = str(integration.get("business_id") or "").strip()
        if marketplace == "yandex" and not business_id_value:
            raise HTTPException(status_code=400, detail="Для Яндекс Маркета требуется Business ID")
        if client_id_value:
            integration["client_id"] = client_id_value
        if marketplace == "ozon":
            page_size = integration.get("page_size")
            if page_size is not None and (not isinstance(page_size, int) or page_size <= 0):
                raise HTTPException(status_code=400, detail="Размер страницы должен быть положительным целым числом")
        if marketplace == "wb":
            max_pages = integration.get("max_pages")
            if max_pages is not None and (not isinstance(max_pages, int) or max_pages <= 0):
                raise HTTPException(status_code=400, detail="Лимит страниц должен быть положительным целым числом")

        account = repository.create_marketplace_account(
            user_id=int(user["id"]),
            marketplace=marketplace,
            account_name=payload.account_name.strip(),
            api_url=api_url,
            api_key=(payload.api_key or "").strip() or None,
            extra=integration,
        )
        return {"ok": True, "item": account}

    @app.post("/api/accounts/{account_id}/status")
    def update_account_status(account_id: int, request: Request, payload: AccountStatusRequest) -> dict[str, object]:
        user = _require_settings_access(request)
        account = repository.get_marketplace_account(
            user_id=int(user["id"]),
            account_id=account_id,
            include_secrets=False,
        )
        if account is None:
            raise HTTPException(status_code=404, detail="Кабинет маркетплейса не найден")
        updated = repository.update_marketplace_account_status(
            user_id=int(user["id"]),
            account_id=account_id,
            is_active=payload.is_active,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Кабинет маркетплейса не найден")
        return {"ok": True}

    @app.delete("/api/accounts/{account_id}")
    def delete_account(account_id: int, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        account = repository.get_marketplace_account(
            user_id=int(user["id"]),
            account_id=account_id,
            include_secrets=False,
        )
        if account is None:
            raise HTTPException(status_code=404, detail="Кабинет маркетплейса не найден")
        repository.update_marketplace_account_status(
            user_id=int(user["id"]),
            account_id=account_id,
            is_active=False,
        )
        deleted = repository.delete_marketplace_account(
            user_id=int(user["id"]),
            account_id=account_id,
        )
        if not deleted:
            raise HTTPException(status_code=404, detail="Кабинет маркетплейса не найден")
        return {"ok": True}

    @app.get("/api/templates")
    def list_templates(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        items = repository.list_templates(user_id=int(user["id"]))
        return {"items": items, "count": len(items)}

    @app.get("/api/template-groups")
    def list_template_groups(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        user_id = int(user["id"])
        _ensure_default_template_variants(user_id)
        rows = repository.list_template_variants(user_id=user_id)
        counts: dict[tuple[str, str], int] = {}
        for row in rows:
            key = (str(row.get("group_id") or ""), str(row.get("subgroup") or ""))
            counts[key] = counts.get(key, 0) + 1
        items = _build_template_group_items(counts)
        return {"items": items, "count": len(items)}

    @app.get("/api/processing-rules")
    def list_processing_rules(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        user_id = int(user["id"])
        _ensure_default_template_variants(user_id)
        existing_rows = repository.list_processing_rules(user_id=user_id)
        existing_map = {str(row.get("group_id") or ""): row for row in existing_rows}
        items: list[dict[str, object]] = []
        for group in TEMPLATE_GROUPS:
            group_id = str(group.get("id") or "")
            title = str(group.get("title") or group_id)
            row = existing_map.get(group_id)
            mode = str((row or {}).get("action_mode") or "manual")
            if mode == "auto":
                mode = "template"
            if mode in {"ai", "ignore"}:
                mode = "manual"
            if mode not in {"template", "manual"}:
                mode = "manual"
            items.append(
                {
                    "group_id": group_id,
                    "title": title,
                    "action_mode": mode,
                    "auto_send": bool((row or {}).get("auto_send")),
                }
            )
        return {"items": items, "count": len(items)}

    @app.put("/api/processing-rules/apply")
    def apply_processing_rules(payload: ProcessingRulesApplyRequest, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        user_id = int(user["id"])
        normalized_rules: list[dict[str, object]] = []
        for item in payload.rules:
            group_id = item.group_id.strip()
            if _template_group_by_id(group_id) is None:
                raise HTTPException(status_code=400, detail=f"Неизвестная группа правил: {group_id}")
            mode = item.action_mode.strip().lower()
            if mode in {"ai", "ignore"}:
                mode = "manual"
            if mode not in {"template", "manual"}:
                raise HTTPException(status_code=400, detail=f"Некорректный режим правила: {mode}")
            normalized_rules.append(
                {
                    "group_id": group_id,
                    "action_mode": mode,
                    "auto_send": bool(item.auto_send),
                }
            )
        repository.replace_processing_rules(user_id=user_id, rules=normalized_rules)
        stats = service.apply_processing_rules_to_unprocessed(user_id=user_id)
        return {"ok": True, "applied": len(normalized_rules), "updated_reviews": stats}

    @app.get("/api/recommendations")
    def list_recommendations(request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        items = repository.list_recommendations(user_id=int(user["id"]))
        return {"items": items, "count": len(items)}

    @app.put("/api/recommendations")
    def save_recommendations(payload: RecommendationsSaveRequest, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        normalized_rows: list[dict[str, object]] = []
        unique_sources: set[str] = set()
        for row in payload.rows:
            source_article = row.source_article.strip()
            targets = _parse_recommendation_targets(row.targets_csv)
            if not source_article:
                continue
            if source_article in unique_sources:
                continue
            unique_sources.add(source_article)
            normalized_rows.append(
                {
                    "source_article": source_article,
                    "target_articles": targets,
                }
            )
        inserted_pairs = repository.replace_all_recommendations(
            user_id=int(user["id"]),
            rows=normalized_rows,
        )
        return {"ok": True, "sources": len(normalized_rows), "pairs": inserted_pairs}

    @app.post("/api/recommendations/import")
    async def import_recommendations(request: Request, file: UploadFile = File(...)) -> dict[str, object]:
        user = _require_settings_access(request)
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover - protected by dependency
            raise HTTPException(status_code=500, detail="Библиотека Excel не установлена") from exc

        filename = (file.filename or "").lower()
        if filename and not filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
            raise HTTPException(status_code=400, detail="Поддерживаются только файлы Excel формата .xlsx")
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Файл пустой")
        try:
            workbook = load_workbook(io.BytesIO(content), data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Не удалось прочитать Excel-файл") from exc
        sheet = workbook.active
        normalized_rows: list[dict[str, object]] = []
        unique_sources: set[str] = set()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            source_article = str(row[0] or "").strip() if len(row) > 0 else ""
            targets_csv = str(row[1] or "").strip() if len(row) > 1 else ""
            targets = _parse_recommendation_targets(targets_csv)
            if not source_article:
                continue
            if source_article in unique_sources:
                continue
            unique_sources.add(source_article)
            normalized_rows.append(
                {
                    "source_article": source_article,
                    "target_articles": targets,
                }
            )
        inserted_pairs = repository.replace_all_recommendations(
            user_id=int(user["id"]),
            rows=normalized_rows,
        )
        return {"ok": True, "sources": len(normalized_rows), "pairs": inserted_pairs}

    @app.get("/api/recommendations/export")
    def export_recommendations(request: Request) -> StreamingResponse:
        try:
            from openpyxl import Workbook
        except Exception as exc:  # pragma: no cover - protected by dependency
            raise HTTPException(status_code=500, detail="Библиотека Excel не установлена") from exc
        user = _require_settings_access(request)
        items = repository.list_recommendations(user_id=int(user["id"]))
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Рекомендации"
        sheet.append(["Артикул товара", "Рекомендуемые артикулы"])
        for item in items:
            sheet.append(
                [
                    str(item.get("source_article") or ""),
                    str(item.get("targets_csv") or ""),
                ]
            )
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        headers = {"Content-Disposition": 'attachment; filename="recommendations.xlsx"'}
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    @app.get("/api/template-subgroup")
    def get_template_subgroup(group_id: str, subgroup: str, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        user_id = int(user["id"])
        if not _validate_subgroup(group_id, subgroup):
            raise HTTPException(status_code=404, detail="Группа шаблонов или подгруппа не найдена")
        _ensure_default_template_variants(user_id)
        items = repository.list_template_variants(
            user_id=user_id,
            group_id=group_id,
            subgroup=subgroup,
        )
        return {"items": items, "count": len(items), "group_id": group_id, "subgroup": subgroup}

    @app.put("/api/template-subgroup")
    def save_template_subgroup(
        group_id: str,
        subgroup: str,
        payload: TemplateSubgroupSaveRequest,
        request: Request,
    ) -> dict[str, object]:
        user = _require_settings_access(request)
        if not _validate_subgroup(group_id, subgroup):
            raise HTTPException(status_code=404, detail="Группа шаблонов или подгруппа не найдена")
        repository.replace_subgroup_templates(
            user_id=int(user["id"]),
            group_id=group_id,
            subgroup=subgroup,
            templates=payload.templates,
        )
        return {"ok": True, "saved": len([x for x in payload.templates if x and x.strip()])}

    @app.post("/api/template-subgroup/item")
    def add_template_subgroup_item(payload: TemplateVariantCreateRequest, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        group_id = payload.group_id.strip()
        subgroup = payload.subgroup.strip()
        if not _validate_subgroup(group_id, subgroup):
            raise HTTPException(status_code=404, detail="Группа шаблонов или подгруппа не найдена")
        item = repository.add_template_variant(
            user_id=int(user["id"]),
            group_id=group_id,
            subgroup=subgroup,
            template_text=payload.template_text,
        )
        return {"ok": True, "item": item}

    @app.delete("/api/template-subgroup/item/{template_id}")
    def delete_template_subgroup_item(template_id: int, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        # Check if the template belongs to a protected subgroup before deleting
        existing = repository.get_template_variant_by_id(
            user_id=int(user["id"]),
            template_id=template_id,
        )
        if existing and _is_protected_subgroup(
            str(existing.get("group_id") or ""),
            str(existing.get("subgroup") or ""),
        ):
            raise HTTPException(status_code=403, detail="Шаблоны этой подгруппы защищены от удаления")
        deleted = repository.delete_template_variant(
            user_id=int(user["id"]),
            template_id=template_id,
        )
        if not deleted:
            raise HTTPException(status_code=404, detail="Шаблон не найден")
        return {"ok": True}

    @app.post("/api/templates/reset-to-defaults")
    def reset_templates_to_defaults(request: Request) -> dict[str, object]:
        """Reset all user templates to the current admin defaults.

        Only available to admin/owner — not managers.
        Deletes all existing user templates and copies from default_template_variants.
        """
        user = _require_admin(request)
        owner_id = _tenant_owner_id(user)
        deleted_and_replaced = repository.reset_templates_to_defaults(user_id=owner_id)
        return {"ok": True, "replaced": deleted_and_replaced}

    @app.put("/api/templates")
    def upsert_template(request: Request, payload: TemplateUpsertRequest) -> dict[str, object]:
        user = _require_settings_access(request)
        category = payload.category.strip().lower()
        mode = payload.mode.strip().lower()
        if category not in CATEGORIES:
            raise HTTPException(status_code=400, detail=f"Неизвестная категория: {category}")
        if mode not in {"auto", "manual", "ignore"}:
            raise HTTPException(status_code=400, detail="Режим должен быть: авто, вручную или игнор")
        repository.upsert_template(
            user_id=int(user["id"]),
            category=category,
            mode=mode,
            template_text=payload.template_text.strip(),
            is_enabled=payload.is_enabled,
        )
        return {"ok": True}

    @app.delete("/api/templates/{category}")
    def delete_template(category: str, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        normalized = category.strip().lower()
        if normalized not in CATEGORIES:
            raise HTTPException(status_code=400, detail=f"Неизвестная категория: {normalized}")
        deleted = repository.delete_template(user_id=int(user["id"]), category=normalized)
        if not deleted:
            raise HTTPException(status_code=404, detail="Правило не найдено")
        return {"ok": True}

    @app.post("/api/reviews/{review_id}/queue-manual")
    def queue_manual(review_id: str, request: Request) -> dict[str, object]:
        user = _require_user(request)
        _require_manager_scope_for_review(user, review_id)
        updated = service.queue_for_manual_processing_with_actor(
            actor_email=str(user.get("email") or ""),
            owner_user_id=_tenant_owner_id(user),
            review_uid=review_id,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Отзыв не найден")
        return {"ok": True}

    @app.post("/api/reviews/{review_id}/auto-reply")
    def auto_reply(review_id: str, request: Request) -> dict[str, object]:
        user = _require_user(request)
        _require_manager_scope_for_review(user, review_id)
        try:
            reply = service.generate_auto_reply(user_id=_tenant_owner_id(user), review_uid=review_id)
        except KeyError as exc:
            raise HTTPException(status_code=404, detail=str(exc) or "Отзыв не найден") from exc
        except MarketplaceSyncError as exc:
            raise HTTPException(status_code=502, detail=f"Не удалось отправить ответ в маркетплейс: {exc}") from exc
        return {"ok": True, "reply": reply}

    @app.post("/api/reviews/{review_id}/manual-reply")
    def manual_reply(review_id: str, payload: ManualReplyRequest, request: Request) -> dict[str, object]:
        user = _require_user(request)
        _require_manager_scope_for_review(user, review_id)
        updated = service.save_manual_reply_with_actor(
            actor_email=str(user.get("email") or ""),
            owner_user_id=_tenant_owner_id(user),
            review_uid=review_id,
            response_text=payload.response_text,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Отзыв не найден")
        return {"ok": True}

    @app.get("/api/admin/ai-settings")
    def get_ai_settings(request: Request) -> dict[str, object]:
        _require_super_admin(request)
        return repository.get_ai_settings()

    @app.put("/api/admin/ai-settings")
    def update_ai_settings(request: Request, payload: AISettingsRequest) -> dict[str, object]:
        _require_super_admin(request)
        provider = payload.provider.strip().lower()
        if provider not in {"rules", "yandex"}:
            raise HTTPException(status_code=400, detail="Провайдер должен быть: встроенные правила или Яндекс")
        lookback_days = int(payload.default_sync_lookback_days)
        repository.update_ai_settings(
            provider=provider,
            yandex_api_key=payload.yandex_api_key.strip() if payload.yandex_api_key is not None else None,
            yandex_folder_id=(payload.yandex_folder_id or "").strip() or None,
            yandex_model_uri=(payload.yandex_model_uri or "").strip() or None,
            group_processors=payload.group_processors,
            use_sync_start_date=False,
            sync_start_date=None,
        )
        repository.set_default_sync_lookback_days(days=lookback_days)
        return {"ok": True}

    @app.post("/api/admin/ai-settings/check")
    def check_ai_settings_connection(request: Request, payload: AIConnectionTestRequest) -> dict[str, object]:
        _require_super_admin(request)
        stored = repository.get_ai_settings(include_secrets=True)
        api_key = (payload.yandex_api_key or "").strip() or str(stored.get("yandex_api_key") or "").strip()
        folder_id = (payload.yandex_folder_id or "").strip() or str(stored.get("yandex_folder_id") or "").strip()
        if not api_key:
            raise HTTPException(status_code=400, detail="Укажите API-ключ Yandex Cloud.")
        if not folder_id:
            raise HTTPException(status_code=400, detail="Укажите ID каталога (folderId).")
        try:
            result = service.check_yandex_connection(api_key=api_key, folder_id=folder_id)
            return {
                "ok": True,
                "status": "ok",
                "message": str(result.get("message") or "Подключение успешно"),
                "model_uri": result.get("model_uri"),
                "response_preview": result.get("response_preview"),
            }
        except MarketplaceSyncError as exc:
            detail = str(exc).lower()
            error_code = "connection"
            if any(code in detail for code in ["401", "403", "unauthorized", "forbidden", "invalid api"]):
                error_code = "auth"
            elif any(code in detail for code in ["400", "404", "folder", "modeluri", "not found"]):
                error_code = "config"
            elif "429" in detail or "rate" in detail or "quota" in detail:
                error_code = "rate_limit"
            elif "timeout" in detail or "network" in detail:
                error_code = "network"
            return {"ok": False, "status": "error", "error_code": error_code, "error": str(exc)}

    @app.post("/api/admin/ai-settings/test-review")
    def test_ai_review_classification(request: Request, payload: AIReviewTestRequest) -> dict[str, object]:
        user = _require_super_admin(request)
        stored = repository.get_ai_settings(include_secrets=True)
        api_key = (payload.yandex_api_key or "").strip() or str(stored.get("yandex_api_key") or "").strip()
        folder_id = (payload.yandex_folder_id or "").strip() or str(stored.get("yandex_folder_id") or "").strip()
        review_text = str(payload.review_text or "").strip()
        if not api_key:
            raise HTTPException(status_code=400, detail="Укажите API-ключ Yandex Cloud.")
        if not folder_id:
            raise HTTPException(status_code=400, detail="Укажите ID каталога (folderId).")
        if not review_text:
            raise HTTPException(status_code=400, detail="Введите текст тестового отзыва.")
        try:
            result = service.classify_test_review_with_yandex(
                user_id=int(user["id"]),
                review_text=review_text,
                review_rating=payload.review_rating,
                settings={
                    "yandex_api_key": api_key,
                    "yandex_folder_id": folder_id,
                    "yandex_model_uri": str(stored.get("yandex_model_uri") or "") or None,
                },
            )
            return {
                "ok": True,
                "status": "ok",
                "group_id": result.get("group_id"),
                "group_title": result.get("group_title"),
                "subgroup_id": result.get("subgroup_id"),
                "subgroup": result.get("subgroup"),
                "model_uri": result.get("model_uri"),
                "raw_response": result.get("raw_response"),
            }
        except MarketplaceSyncError as exc:
            detail = str(exc).lower()
            error_code = "classification"
            if any(code in detail for code in ["401", "403", "unauthorized", "forbidden", "invalid api"]):
                error_code = "auth"
            elif any(code in detail for code in ["400", "404", "folder", "modeluri", "not found"]):
                error_code = "config"
            elif "429" in detail or "rate" in detail or "quota" in detail:
                error_code = "rate_limit"
            elif "timeout" in detail or "network" in detail:
                error_code = "network"
            return {
                "ok": False,
                "status": "error",
                "error_code": error_code,
                "error": str(exc),
                "debug": exc.details if isinstance(exc.details, dict) else {},
            }
        except Exception as exc:
            raise HTTPException(status_code=502, detail=f"Не удалось выполнить тестовый запрос к Yandex GPT: {exc}") from exc

    @app.get("/api/admin/ai-settings/active-ids")
    def list_ai_active_ids(request: Request) -> dict[str, object]:
        user = _require_super_admin(request)
        owner_user_id = _tenant_owner_id(user)
        options = service._list_group_subgroups_for_review_classification(
            repository=repository,
            user_id=owner_user_id,
        )
        items: list[dict[str, object]] = []
        for group in options:
            group_id = str(group.get("group_id") or "").strip()
            if not group_id or group_id == service.TEXTLESS_GROUP_ID:
                continue
            subgroup_items_raw = group.get("subgroup_items")
            subgroup_items: list[dict[str, str]] = []
            if isinstance(subgroup_items_raw, list):
                for subgroup_item in subgroup_items_raw:
                    if not isinstance(subgroup_item, dict):
                        continue
                    subgroup_id = str(subgroup_item.get("subgroup_id") or "").strip()
                    subgroup_title = str(subgroup_item.get("subgroup") or "").strip()
                    if not subgroup_id or not subgroup_title:
                        continue
                    subgroup_items.append(
                        {
                            "subgroup_id": subgroup_id,
                            "subgroup": subgroup_title,
                        }
                    )
            if not subgroup_items:
                continue
            items.append(
                {
                    "group_id": group_id,
                    "group_title": str(group.get("group_title") or group_id),
                    "subgroup_items": subgroup_items,
                }
            )
        return {"ok": True, "items": items, "count": len(items)}

    @app.get("/api/admin/ai-usage-stats")
    def get_ai_usage_stats(request: Request, days: int = 30) -> dict[str, object]:
        """Return daily Yandex GPT usage statistics for the last N days."""
        user = _require_admin(request)
        owner_id = _tenant_owner_id(user)
        rows = repository.get_ai_usage_stats(user_id=owner_id, days=min(max(days, 1), 90))
        # Estimate cost: Yandex YandexGPT Lite ≈ 0.20₽ per 1000 tokens input, 0.60₽ output
        # (approximate — actual pricing may differ)
        total_requests = sum(int(r.get("requests") or 0) for r in rows)
        total_input = sum(int(r.get("input_tokens") or 0) for r in rows)
        total_output = sum(int(r.get("output_tokens") or 0) for r in rows)
        return {
            "ok": True,
            "rows": rows,
            "totals": {
                "requests": total_requests,
                "input_tokens": total_input,
                "output_tokens": total_output,
                "total_tokens": total_input + total_output,
            },
        }

    @app.get("/api/admin/ai-request-log")
    def get_ai_request_log(request: Request, limit: int = 200) -> dict[str, object]:
        """Return recent Yandex GPT requests (last 1 day) for debugging."""
        user = _require_admin(request)
        owner_id = _tenant_owner_id(user)
        try:
            repository.purge_old_ai_request_logs(user_id=owner_id)
        except Exception:
            pass
        try:
            logs = repository.list_ai_request_logs(user_id=owner_id, limit=min(max(limit, 1), 500))
        except Exception as exc:
            _log.warning("ai-request-log: table not ready yet: %s", exc)
            logs = []
        return {"ok": True, "logs": logs, "count": len(logs)}

    @app.get("/api/admin/context")
    def get_admin_context(request: Request) -> dict[str, object]:
        user = _require_admin(request)
        user_id = int(user["id"])
        owner_user_id = _tenant_owner_id(user)
        manager_permissions: list[dict[str, object]] = []
        if str(user.get("role") or "").strip().lower() in TENANT_MANAGER_ROLES:
            manager_permissions = _manager_permissions_context_for_user(user)
        return {
            "user_id": user_id,
            "owner_user_id": owner_user_id,
            "is_super_admin": _is_super_admin(user),
            "is_tenant_owner": user_id == owner_user_id and not _is_super_admin(user),
            "manager_permissions": manager_permissions,
        }

    @app.get("/api/super-admin/settings")
    def super_admin_settings(request: Request) -> dict[str, object]:
        _require_super_admin(request)
        return repository.get_super_admin_settings()

    @app.put("/api/super-admin/settings")
    def super_admin_update_settings(payload: SuperAdminSettingsRequest, request: Request) -> dict[str, object]:
        _require_super_admin(request)
        ai_provider = payload.ai_provider.strip().lower()
        if ai_provider not in {"rules", "yandex"}:
            raise HTTPException(status_code=400, detail="Провайдер должен быть: встроенные правила или Яндекс")
        repository.save_super_admin_settings(
            payment_provider=(payload.payment_provider or "").strip() or "manual",
            payment_api_key=payload.payment_api_key.strip() if payload.payment_api_key is not None else None,
            ai_provider=ai_provider,
            yandex_api_key=payload.yandex_api_key.strip() if payload.yandex_api_key is not None else None,
            yandex_folder_id=(payload.yandex_folder_id or "").strip() or None,
            yandex_model_uri=(payload.yandex_model_uri or "").strip() or None,
            group_processors=payload.group_processors,
            use_sync_start_date=False,
            sync_start_date=None,
            default_sync_lookback_days=int(payload.default_sync_lookback_days),
        )
        return {"ok": True}

    @app.get("/api/super-admin/template-variables")
    def super_admin_list_template_variables(request: Request) -> dict[str, object]:
        _require_super_admin(request)
        items = repository.list_template_variables(only_active=False)
        return {"items": items, "count": len(items)}

    @app.put("/api/super-admin/template-variables")
    def super_admin_upsert_template_variable(
        payload: TemplateVariableUpsertRequest,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        normalized_key = payload.var_key.strip().upper()
        if not TEMPLATE_VARIABLE_KEY_RE.fullmatch(normalized_key):
            raise HTTPException(
                status_code=400,
                detail="Ключ переменной должен быть в формате %NAME% и содержать только A-Z, 0-9 и _ (2-50 символов).",
            )
        source_type = (payload.source_type or "").strip().lower() or "manual"
        if source_type not in {"manual", "review_field", "system"}:
            raise HTTPException(status_code=400, detail="source_type должен быть manual, review_field или system")
        item = repository.upsert_template_variable(
            var_key=normalized_key,
            title=payload.title.strip(),
            description=(payload.description or "").strip() or None,
            is_user_editable=bool(payload.is_user_editable),
            source_type=source_type,
            source_path=(payload.source_path or "").strip() or None,
            default_value=(payload.default_value or "").strip() or None,
            is_active=bool(payload.is_active),
        )
        return {"ok": True, "item": item}

    @app.delete("/api/super-admin/template-variables")
    def super_admin_delete_template_variable(
        payload: TemplateVariableDeleteRequest,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        deleted = repository.delete_template_variable(var_key=payload.var_key.strip().upper())
        if not deleted:
            raise HTTPException(status_code=404, detail="Переменная шаблона не найдена")
        return {"ok": True}

    @app.get("/api/super-admin/tariffs")
    def super_admin_list_tariffs(request: Request) -> dict[str, object]:
        _require_super_admin(request)
        items = repository.list_tariff_plans()
        return {"items": items, "count": len(items)}

    @app.put("/api/super-admin/tariffs")
    def super_admin_upsert_tariff(payload: TariffPlanUpsertRequest, request: Request) -> dict[str, object]:
        _require_super_admin(request)
        code = payload.code.strip().lower()
        if not code:
            raise HTTPException(status_code=400, detail="Код тарифа обязателен")
        repository.upsert_tariff_plan(
            code=code,
            title=payload.title.strip(),
            monthly_price=float(payload.monthly_price),
            limits=dict(payload.limits),
            is_active=bool(payload.is_active),
        )
        return {"ok": True}

    @app.delete("/api/super-admin/tariffs")
    def super_admin_delete_tariff(payload: TariffPlanDeleteRequest, request: Request) -> dict[str, object]:
        _require_super_admin(request)
        code = payload.code.strip().lower()
        if not code:
            raise HTTPException(status_code=400, detail="Код тарифа обязателен")
        deleted, in_use_count = repository.delete_tariff_plan(code=code)
        if not deleted and in_use_count > 0:
            raise HTTPException(
                status_code=409,
                detail=f"Тариф используется у {in_use_count} клиентов. Сначала смените им тариф.",
            )
        if not deleted:
            raise HTTPException(status_code=404, detail="Тариф не найден")
        return {"ok": True}

    @app.get("/api/super-admin/tenants")
    def super_admin_list_tenants(request: Request) -> dict[str, object]:
        _require_super_admin(request)
        items = repository.list_tenants_overview()
        return {"items": items, "count": len(items)}

    @app.get("/api/super-admin/default-template-groups")
    def super_admin_list_default_template_groups(request: Request) -> dict[str, object]:
        _require_super_admin(request)
        _ensure_platform_default_templates()
        rows = repository.list_default_template_variants()
        counts: dict[tuple[str, str], int] = {}
        for row in rows:
            key = (str(row.get("group_id") or ""), str(row.get("subgroup") or ""))
            counts[key] = counts.get(key, 0) + 1
        items = _build_template_group_items(counts)
        return {"items": items, "count": len(items)}

    @app.get("/api/super-admin/default-template-subgroup")
    def super_admin_get_default_template_subgroup(
        group_id: str,
        subgroup: str,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        if not _validate_subgroup(group_id, subgroup):
            raise HTTPException(status_code=404, detail="Группа шаблонов или подгруппа не найдена")
        _ensure_platform_default_templates()
        items = repository.list_default_template_variants(group_id=group_id, subgroup=subgroup)
        return {"items": items, "count": len(items), "group_id": group_id, "subgroup": subgroup}

    @app.put("/api/super-admin/default-template-subgroup")
    def super_admin_save_default_template_subgroup(
        group_id: str,
        subgroup: str,
        payload: DefaultTemplateSubgroupSaveRequest,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        if not _validate_subgroup(group_id, subgroup):
            raise HTTPException(status_code=404, detail="Группа шаблонов или подгруппа не найдена")
        repository.replace_default_subgroup_templates(
            group_id=group_id,
            subgroup=subgroup,
            templates=payload.templates,
        )
        return {"ok": True, "saved": len([x for x in payload.templates if x and x.strip()])}

    @app.post("/api/super-admin/default-template-subgroup")
    def super_admin_add_default_template_subgroup(
        payload: DefaultTemplateSubgroupManageRequest,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        group_id = payload.group_id.strip()
        subgroup = payload.subgroup.strip()
        if _template_group_by_id(group_id) is None:
            raise HTTPException(status_code=404, detail="Группа шаблонов не найдена")
        if not subgroup:
            raise HTTPException(status_code=400, detail="Название подгруппы обязательно")
        existing = {
            str(item.get("name") or "").strip()
            for item in _all_subgroups_for_group(group_id)
            if str(item.get("name") or "").strip()
        }
        if subgroup in existing:
            raise HTTPException(status_code=409, detail="Подгруппа с таким названием уже существует")
        repository.add_default_template_subgroup(group_id=group_id, subgroup=subgroup)
        return {"ok": True, "group_id": group_id, "subgroup": subgroup}

    @app.delete("/api/super-admin/default-template-subgroup")
    def super_admin_delete_default_template_subgroup(
        group_id: str,
        subgroup: str,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        clean_group_id = str(group_id or "").strip()
        clean_subgroup = str(subgroup or "").strip()
        if _template_group_by_id(clean_group_id) is None:
            raise HTTPException(status_code=404, detail="Группа шаблонов не найдена")
        if _is_protected_subgroup(clean_group_id, clean_subgroup):
            raise HTTPException(status_code=403, detail="Эта подгруппа защищена и не может быть удалена")
        if not clean_subgroup:
            raise HTTPException(status_code=400, detail="Название подгруппы обязательно")
        if _is_protected_default_subgroup(clean_group_id, clean_subgroup):
            raise HTTPException(
                status_code=400,
                detail="Подгруппы '1-3 звезды' и '4-5 звезд' в блоке 'Оценки без текста' удалять нельзя",
            )
        deleted = repository.delete_default_template_subgroup(group_id=clean_group_id, subgroup=clean_subgroup)
        if not deleted:
            raise HTTPException(status_code=404, detail="Подгруппа не найдена")
        return {"ok": True}

    @app.patch("/api/super-admin/default-template-subgroup")
    def super_admin_rename_default_template_subgroup(
        payload: DefaultTemplateSubgroupRenameRequest,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        group_id = payload.group_id.strip()
        subgroup = payload.subgroup.strip()
        new_subgroup = payload.new_subgroup.strip()
        if _template_group_by_id(group_id) is None:
            raise HTTPException(status_code=404, detail="Группа шаблонов не найдена")
        if not subgroup or not new_subgroup:
            raise HTTPException(status_code=400, detail="Название подгруппы обязательно")
        if subgroup == new_subgroup:
            current = repository.get_default_template_subgroup(group_id=group_id, subgroup=subgroup)
            return {
                "ok": True,
                "group_id": group_id,
                "subgroup": subgroup,
                "new_subgroup": new_subgroup,
                "subgroup_id": str((current or {}).get("subgroup_id") or "").strip() or None,
            }
        if not _validate_subgroup(group_id, subgroup):
            raise HTTPException(status_code=404, detail="Подгруппа не найдена")
        if _is_protected_default_subgroup(group_id, subgroup):
            raise HTTPException(status_code=400, detail="Эту системную подгруппу переименовывать нельзя")
        if _validate_subgroup(group_id, new_subgroup):
            raise HTTPException(status_code=409, detail="Подгруппа с таким названием уже существует")
        current = repository.get_default_template_subgroup(group_id=group_id, subgroup=subgroup)
        if current is None:
            raise HTTPException(status_code=404, detail="Подгруппа не найдена")
        preserved_subgroup_id = str(current.get("subgroup_id") or "").strip() or None
        renamed = repository.rename_default_template_subgroup(
            group_id=group_id,
            subgroup=subgroup,
            new_subgroup=new_subgroup,
        )
        if not renamed:
            raise HTTPException(status_code=409, detail="Не удалось переименовать подгруппу")
        return {
            "ok": True,
            "group_id": group_id,
            "subgroup": subgroup,
            "new_subgroup": new_subgroup,
            "subgroup_id": preserved_subgroup_id,
        }

    @app.post("/api/super-admin/default-template-subgroup/item")
    def super_admin_add_default_template_subgroup_item(
        payload: DefaultTemplateVariantCreateRequest,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        group_id = payload.group_id.strip()
        subgroup = payload.subgroup.strip()
        if not _validate_subgroup(group_id, subgroup):
            raise HTTPException(status_code=404, detail="Группа шаблонов или подгруппа не найдена")
        item = repository.add_default_template_variant(
            group_id=group_id,
            subgroup=subgroup,
            template_text=payload.template_text,
        )
        return {"ok": True, "item": item}

    @app.post("/api/super-admin/default-template-subgroup/bulk-import")
    def super_admin_bulk_import_default_template_subgroup_items(
        payload: DefaultTemplateBulkImportRequest,
        request: Request,
    ) -> dict[str, object]:
        _require_super_admin(request)
        group_id = payload.group_id.strip()
        subgroup = payload.subgroup.strip()
        if not _validate_subgroup(group_id, subgroup):
            raise HTTPException(status_code=404, detail="Группа шаблонов или подгруппа не найдена")
        templates = [str(item or "").strip() for item in payload.templates]
        clean_templates = [item for item in templates if item]
        if not clean_templates:
            raise HTTPException(status_code=400, detail="Передайте хотя бы один непустой шаблон")
        added = repository.add_default_template_variants_bulk(
            group_id=group_id,
            subgroup=subgroup,
            templates=clean_templates,
        )
        return {"ok": True, "added": int(added), "group_id": group_id, "subgroup": subgroup}

    @app.delete("/api/super-admin/default-template-subgroup/item/{template_id}")
    def super_admin_delete_default_template_subgroup_item(template_id: int, request: Request) -> dict[str, object]:
        _require_super_admin(request)
        deleted = repository.delete_default_template_variant(template_id=template_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Шаблон не найден")
        return {"ok": True}

    @app.post("/api/super-admin/tenant-plan")
    def super_admin_set_tenant_plan(payload: TenantPlanUpdateRequest, request: Request) -> dict[str, object]:
        _require_super_admin(request)
        tenant = repository.get_user_by_id(int(payload.owner_user_id))
        if tenant is None:
            raise HTTPException(status_code=404, detail="Пользователь кабинета не найден")
        if bool(tenant.get("is_super_admin")):
            raise HTTPException(status_code=400, detail="Нельзя назначать тариф супер-администратору")
        if _tenant_owner_id(tenant) != int(tenant["id"]):
            raise HTTPException(status_code=400, detail="Тариф можно назначать только владельцу кабинета")
        updated = repository.set_tenant_plan(
            owner_user_id=int(payload.owner_user_id),
            plan_code=payload.plan_code.strip().lower(),
            limits_override=dict(payload.limits_override),
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь кабинета не найден")
        return {"ok": True}

    @app.get("/api/super-admin/payments")
    def super_admin_list_payments(
        request: Request,
        owner_user_id: int | None = None,
        limit: int = 200,
    ) -> dict[str, object]:
        _require_super_admin(request)
        safe_limit = min(max(limit, 1), 1000)
        items = repository.list_billing_records(owner_user_id=owner_user_id, limit=safe_limit)
        return {"items": items, "count": len(items)}

    @app.post("/api/super-admin/payments")
    def super_admin_create_payment(payload: PaymentRecordCreateRequest, request: Request) -> dict[str, object]:
        _require_super_admin(request)
        tenant = repository.get_user_by_id(int(payload.owner_user_id))
        if tenant is None:
            raise HTTPException(status_code=404, detail="Пользователь кабинета не найден")
        if bool(tenant.get("is_super_admin")):
            raise HTTPException(status_code=400, detail="Нельзя привязывать оплату к супер-администратору")
        item, subscription = repository.save_payment_record_with_subscription_update(
            owner_user_id=int(payload.owner_user_id),
            amount=float(payload.amount),
            currency=payload.currency.strip().upper(),
            status=payload.status.strip().lower(),
            external_payment_id=(payload.external_payment_id or "").strip() or None,
            details=dict(payload.details),
            paid_at=(payload.paid_at or "").strip() or None,
            months=int(payload.months),
            grace_days=int(payload.grace_days),
        )
        return {"ok": True, "item": item, "subscription": subscription}

    @app.delete("/api/super-admin/payments")
    def super_admin_delete_payment(payload: PaymentRecordDeleteRequest, request: Request) -> dict[str, object]:
        _require_super_admin(request)
        deleted = repository.delete_payment_record(payment_id=int(payload.id))
        if not deleted:
            raise HTTPException(status_code=404, detail="Платеж не найден")
        return {"ok": True}

    @app.post("/api/super-admin/users/{target_user_id}/block")
    def super_admin_block_user(target_user_id: int, payload: UserBlockUpdateRequest, request: Request) -> dict[str, object]:
        actor = _require_super_admin(request)
        target = repository.get_user_by_id(target_user_id)
        if target is None:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        if int(target["id"]) == int(actor["id"]) and payload.blocked:
            raise HTTPException(status_code=400, detail="Нельзя заблокировать собственный аккаунт")
        updated = repository.set_user_blocked(
            user_id=target_user_id,
            blocked=bool(payload.blocked),
            reason=(payload.reason or "").strip() or None,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.post("/api/super-admin/users/{target_user_id}/delete")
    def super_admin_delete_user(target_user_id: int, payload: UserDeleteRequest, request: Request) -> dict[str, object]:
        actor = _require_super_admin(request)
        if not payload.confirm:
            raise HTTPException(status_code=400, detail="Требуется подтверждение удаления")
        if int(target_user_id) == int(actor["id"]):
            raise HTTPException(status_code=400, detail="Нельзя удалить собственный аккаунт")
        deleted = repository.soft_delete_user(user_id=target_user_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.get("/api/tenant/team")
    def tenant_list_team(request: Request) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        owner_id = _tenant_owner_id(owner)
        items = repository.list_tenant_users(owner_user_id=owner_id)
        for item in items:
            if str(item.get("role") or "").strip().lower() in TENANT_MANAGER_ROLES:
                item["manager_permissions"] = repository.list_manager_permissions(manager_user_id=int(item["id"]))
                item["can_supplies"] = bool(item.get("can_supplies"))
                item["can_salary"] = bool(item.get("can_salary"))
                item["can_salary_settings"] = bool(item.get("can_salary_settings"))
                try:
                    import json as _j_sp
                    item["salary_productions"] = _j_sp.loads(str(item.get("salary_productions") or "[]"))
                except Exception:
                    item["salary_productions"] = []
                _sp = repository.get_manager_supply_permissions(
                    manager_user_id=int(item["id"])
                )
                _sp["can_supply_planning"] = bool(item.get("can_supply_planning"))
                _sp["can_supply_stock"] = bool(item.get("can_supply_stock"))
                try:
                    import json as _j_stock
                    _sp["stock_productions"] = _j_stock.loads(
                        str(item.get("stock_productions") or "[]")
                    )
                except Exception:
                    _sp["stock_productions"] = []
                if not isinstance(_sp.get("stock_productions"), list):
                    _sp["stock_productions"] = []
                item["supply_permissions"] = _sp
                item["can_supply_stock"] = bool(item.get("can_supply_stock"))
                try:
                    import json as _j_stock2
                    item["stock_productions"] = _j_stock2.loads(
                        str(item.get("stock_productions") or "[]")
                    )
                except Exception:
                    item["stock_productions"] = []
        return {"items": items, "count": len(items)}

    @app.post("/api/tenant/team")
    def tenant_create_team_user(payload: TenantUserCreateRequest, request: Request) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        owner_id = _tenant_owner_id(owner)
        email = payload.email.strip().lower()
        if len(email) < 5 or "@" not in email:
            raise HTTPException(status_code=400, detail="Введите корректную эл. почту")
        if repository.get_user_by_email(email) is not None:
            raise HTTPException(status_code=409, detail="Пользователь с такой почтой уже существует")
        role = _normalize_tenant_role_or_400(payload.role) if payload.role else TENANT_ROLE_MANAGER
        created = repository.create_tenant_user(
            owner_user_id=owner_id,
            email=email,
            password_hash=hash_password(payload.password),
            role=role,
            full_name=(payload.full_name or "").strip() or None,
        )
        owner_account_ids = _manager_owner_account_ids(owner_id)
        normalized_permissions: list[dict[str, object]] = []
        for item in payload.permissions:
            account_id = int(item.account_id)
            if account_id not in owner_account_ids:
                raise HTTPException(
                    status_code=400,
                    detail=f"Кабинет {account_id} не относится к вашему профилю или недоступен",
                )
            normalized_permissions.append(
                {
                    "account_id": account_id,
                    "can_reviews": bool(item.can_reviews),
                    "can_questions": bool(item.can_questions),
                    "can_chats": bool(item.can_chats),
                }
            )
        saved_permissions = repository.replace_manager_permissions(
            manager_user_id=int(created["id"]),
            permissions=normalized_permissions,
        )
        return {
            "ok": True,
            "item": {
                "id": created.get("id"),
                "email": created.get("email"),
                "full_name": created.get("full_name"),
                "role": created.get("role"),
                "is_blocked": created.get("is_blocked"),
                "created_at": created.get("created_at"),
                "manager_permissions": repository.list_manager_permissions(manager_user_id=int(created["id"])),
                "permissions_saved": saved_permissions,
            },
        }

    @app.get("/api/tenant/team/{target_user_id}/permissions")
    def tenant_get_manager_permissions(target_user_id: int, request: Request) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        target = _target_user_for_admin_scope(actor=owner, target_user_id=target_user_id)
        if int(target["id"]) == int(owner["id"]):
            raise HTTPException(status_code=400, detail="Для владельца кабинета отдельные права менеджера не назначаются")
        if str(target.get("role") or "").strip().lower() not in TENANT_MANAGER_ROLES:
            raise HTTPException(status_code=400, detail="Права можно настраивать только для менеджера")
        permissions = repository.list_manager_permissions(manager_user_id=target_user_id)
        return {"items": permissions, "count": len(permissions)}

    @app.put("/api/tenant/team/{target_user_id}/permissions")
    def tenant_update_manager_permissions(
        target_user_id: int,
        payload: ManagerPermissionsUpdateRequest,
        request: Request,
    ) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        target = _target_user_for_admin_scope(actor=owner, target_user_id=target_user_id)
        if int(target["id"]) == int(owner["id"]):
            raise HTTPException(status_code=400, detail="Для владельца кабинета отдельные права менеджера не назначаются")
        if str(target.get("role") or "").strip().lower() not in TENANT_MANAGER_ROLES:
            raise HTTPException(status_code=400, detail="Права можно настраивать только для менеджера")
        owner_account_ids = _manager_owner_account_ids(_tenant_owner_id(owner))
        normalized_permissions: list[dict[str, object]] = []
        for item in payload.permissions:
            account_id = int(item.account_id)
            if account_id not in owner_account_ids:
                raise HTTPException(
                    status_code=400,
                    detail=f"Кабинет {account_id} не относится к вашему профилю или недоступен",
                )
            normalized_permissions.append(
                {
                    "account_id": account_id,
                    "can_reviews": bool(item.can_reviews),
                    "can_questions": bool(item.can_questions),
                    "can_chats": bool(item.can_chats),
                }
            )
        saved = repository.replace_manager_permissions(
            manager_user_id=target_user_id,
            permissions=normalized_permissions,
        )
        return {
            "ok": True,
            "saved": saved,
            "items": repository.list_manager_permissions(manager_user_id=target_user_id),
        }

    @app.get("/api/tenant/team/{target_user_id}/supply-permissions")
    def tenant_get_supply_permissions(target_user_id: int, request: Request) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        target = _target_user_for_admin_scope(actor=owner, target_user_id=target_user_id)
        repository._ensure_supply_tables()
        perms = repository.get_manager_supply_permissions(manager_user_id=target_user_id)
        perms["can_supply_planning"] = bool(target.get("can_supply_planning"))
        perms["can_supply_stock"] = bool(target.get("can_supply_stock"))
        try:
            import json as _j_stock_get
            stock_prods = _j_stock_get.loads(str(target.get("stock_productions") or "[]"))
        except Exception:
            stock_prods = []
        perms["stock_productions"] = stock_prods if isinstance(stock_prods, list) else []
        return {"ok": True, **perms}

    @app.put("/api/tenant/team/{target_user_id}/supplies-access")
    def tenant_set_manager_supplies_access(
        target_user_id: int,
        payload: ManagerSuppliesAccessRequest,
        request: Request,
    ) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        target = _target_user_for_admin_scope(actor=owner, target_user_id=target_user_id)
        if int(target["id"]) == int(owner["id"]):
            raise HTTPException(status_code=400, detail="Для владельца права не меняются")
        if str(target.get("role") or "").strip().lower() not in TENANT_MANAGER_ROLES:
            raise HTTPException(status_code=400, detail="Применимо только для менеджера")
        try:
            sources = {str(k): v for k, v in (payload.supply_sources or {}).items()}
            has_any_supply = (
                payload.can_supplies
                or payload.can_supply_settings
                or payload.can_supply_poa
                or payload.can_supply_certs
                or payload.can_supply_planning
                or payload.can_supply_stock
                or any(
                    (
                        v.get("wb")
                        or v.get("wb_fbs")
                        or v.get("wb_fbs_tsd")
                        or v.get("ozon")
                    )
                    for v in sources.values()
                    if isinstance(v, dict)
                )
            )
            repository.set_user_can_supplies(user_id=target_user_id, can_supplies=has_any_supply)
            repository.set_manager_supply_permissions(
                manager_user_id=target_user_id,
                can_supply_settings=payload.can_supply_settings,
                can_supply_poa=payload.can_supply_poa,
                can_supply_certs=payload.can_supply_certs,
                sources=sources,
            )
            # Save planning / Остатки flags on the user row
            with repository._connect() as _conn:
                _conn.execute(
                    repository._sql("UPDATE users SET can_supply_planning = ? WHERE id = ?"),
                    (repository._bool_db(payload.can_supply_planning), target_user_id),
                )
            repository._ensure_supply_tables()
            repository.set_user_can_supply_stock(
                user_id=target_user_id,
                can_supply_stock=bool(payload.can_supply_stock),
                stock_productions=list(payload.stock_productions or []),
            )
        except Exception as exc:
            _log.error("supplies-access save failed for user %s: %s", target_user_id, exc, exc_info=True)
            raise HTTPException(status_code=500, detail=f"Ошибка сохранения прав на поставки: {exc}") from exc
        return {"ok": True, "can_supplies": has_any_supply}

    @app.put("/api/tenant/team/{target_user_id}/salary-access")
    def tenant_set_manager_salary_access(
        target_user_id: int,
        payload: ManagerSalaryAccessRequest,
        request: Request,
    ) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        target = _target_user_for_admin_scope(actor=owner, target_user_id=target_user_id)
        if int(target["id"]) == int(owner["id"]):
            raise HTTPException(status_code=400, detail="Для владельца права не меняются")
        if str(target.get("role") or "").strip().lower() not in TENANT_MANAGER_ROLES:
            raise HTTPException(status_code=400, detail="Применимо только для менеджера")
        repository._ensure_supply_tables()
        repository.set_user_can_salary(
            user_id=target_user_id,
            can_salary=payload.can_salary,
            can_salary_settings=payload.can_salary_settings,
            can_salary_report=payload.can_salary_report,
            can_salary_zp_export=payload.can_salary_zp_export,
            salary_productions=list(payload.salary_productions),
        )
        return {"ok": True, "can_salary": payload.can_salary}

    @app.post("/api/tenant/team/{target_user_id}/role")
    def tenant_update_team_role(
        target_user_id: int,
        payload: TenantUserRoleUpdateRequest,
        request: Request,
    ) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        target = _target_user_for_admin_scope(actor=owner, target_user_id=target_user_id)
        role = _normalize_tenant_role_or_400(payload.role)
        if int(target["id"]) == int(owner["id"]) and role != TENANT_ROLE_OWNER:
            raise HTTPException(status_code=400, detail="Нельзя снять роль администратора у владельца кабинета")
        updated = repository.update_user_role(user_id=target_user_id, role=role)
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.patch("/api/tenant/team/{target_user_id}/profile")
    def tenant_update_team_profile(
        target_user_id: int,
        payload: TeamMemberProfileUpdateRequest,
        request: Request,
    ) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        target = _target_user_for_admin_scope(actor=owner, target_user_id=target_user_id)
        repository.update_user_profile(
            user_id=target_user_id,
            email=str(target.get("email") or ""),
            full_name=payload.full_name.strip() or None,
        )
        return {"ok": True}

    @app.post("/api/tenant/team/{target_user_id}/password")
    def tenant_update_team_password(
        target_user_id: int,
        payload: AdminUserPasswordUpdateRequest,
        request: Request,
    ) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        _target_user_for_admin_scope(actor=owner, target_user_id=target_user_id)
        updated = repository.update_user_password(
            user_id=target_user_id,
            password_hash=hash_password(payload.password),
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.post("/api/tenant/team/{target_user_id}/block")
    def tenant_set_user_block(
        target_user_id: int,
        payload: UserBlockUpdateRequest,
        request: Request,
    ) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        target = _target_user_for_admin_scope(actor=owner, target_user_id=target_user_id)
        if int(target["id"]) == int(owner["id"]) and payload.blocked:
            raise HTTPException(status_code=400, detail="Нельзя заблокировать собственный аккаунт")
        updated = repository.set_user_blocked(
            user_id=target_user_id,
            blocked=bool(payload.blocked),
            reason=(payload.reason or "").strip() or None,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.post("/api/tenant/team/{target_user_id}/delete")
    def tenant_delete_user(
        target_user_id: int,
        payload: UserDeleteRequest,
        request: Request,
    ) -> dict[str, object]:
        owner = _require_tenant_owner(request)
        if not payload.confirm:
            raise HTTPException(status_code=400, detail="Требуется подтверждение удаления")
        target = _target_user_for_admin_scope(actor=owner, target_user_id=target_user_id)
        if int(target["id"]) == int(owner["id"]):
            raise HTTPException(status_code=400, detail="Нельзя удалить владельца кабинета")
        deleted = repository.soft_delete_user(user_id=target_user_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.get("/api/tenant/me/plan")
    def tenant_me_plan(request: Request) -> dict[str, object]:
        user = _require_admin(request)
        owner_user_id = _tenant_owner_id(user)
        owner = repository.get_user_by_id(owner_user_id)
        if owner is None:
            raise HTTPException(status_code=404, detail="Владелец кабинета не найден")
        plans = repository.list_tariff_plans()
        current_plan_code = str(owner.get("plan_code") or "").strip().lower()
        current_plan = next((item for item in plans if str(item.get("code") or "").strip().lower() == current_plan_code), None)
        effective_limits: dict[str, object] = {}
        if current_plan and isinstance(current_plan.get("limits"), dict):
            effective_limits.update(current_plan["limits"])
        owner_override = owner.get("limits_override")
        if isinstance(owner_override, dict):
            effective_limits.update(owner_override)
        return {
            "owner_user_id": owner_user_id,
            "plan_code": current_plan_code,
            "plan": current_plan,
            "limits_override": owner_override if isinstance(owner_override, dict) else {},
            "effective_limits": effective_limits,
        }

    @app.get("/api/admin/users")
    def admin_list_users(request: Request) -> dict[str, object]:
        actor = _require_admin(request)
        if _is_super_admin(actor):
            items = repository.list_users(super_admin_only=False, owner_only=True)
        else:
            owner = _require_tenant_owner(request)
            items = repository.list_tenant_users(owner_user_id=int(owner["id"]))
        return {"items": items, "count": len(items)}

    @app.post("/api/admin/users")
    def admin_create_user(payload: AdminUserCreateRequest, request: Request) -> dict[str, object]:
        actor = _require_admin(request)
        email = payload.email.strip().lower()
        if len(email) < 5 or "@" not in email:
            raise HTTPException(status_code=400, detail="Введите корректную эл. почту")
        password = payload.password
        if len(password) < 8:
            raise HTTPException(status_code=400, detail="Пароль должен быть не короче 8 символов")
        if repository.get_user_by_email(email) is not None:
            raise HTTPException(status_code=409, detail="Пользователь с такой почтой уже существует")
        if _is_super_admin(actor):
            role = ROLE_USER
            plan_code = payload.plan_code.strip().lower()
            plans = repository.list_tariff_plans()
            all_codes = {str(item.get("code") or "").strip().lower() for item in plans}
            all_codes = {code for code in all_codes if code}
            if all_codes and plan_code not in all_codes:
                # Keep user creation resilient to stale UI state:
                # if selected plan was removed, fall back to any existing tariff.
                plan_code = sorted(all_codes)[0]
            if not plan_code:
                # Keep super-admin flow operational even when tariff catalog
                # is temporarily empty or not yet configured.
                plan_code = "starter"
            created = repository.create_user(
                email=email,
                password_hash=hash_password(password),
                role=role,
                plan_code=plan_code,
            )
        else:
            owner = _require_tenant_owner(request)
            created = repository.create_tenant_user(
                owner_user_id=int(owner["id"]),
                email=email,
                password_hash=hash_password(password),
                role=_normalize_tenant_role_or_400(payload.role),
                full_name=None,
            )
        return {
            "ok": True,
            "item": {
                "id": created.get("id"),
                "email": created.get("email"),
                "full_name": created.get("full_name"),
                "role": created.get("role"),
                "is_blocked": created.get("is_blocked"),
                "created_at": created.get("created_at"),
            },
        }

    @app.post("/api/admin/users/{target_user_id}/role")
    def admin_update_user_role(target_user_id: int, payload: RoleUpdateRequest, request: Request) -> dict[str, object]:
        current_user = _require_admin(request)
        target_user = _target_user_for_admin_scope(actor=current_user, target_user_id=target_user_id)
        if _is_super_admin(current_user):
            role = payload.role.strip().lower()
            if role not in ROLE_ASSIGNABLE_BY_ADMIN:
                raise HTTPException(
                    status_code=400,
                    detail="Роль должна быть: пользователь, менеджер обратной связи или администратор",
                )
            if role != ROLE_ADMIN:
                admin_rows = repository.raw_fetch(
                    """
                    SELECT id
                    FROM users
                    WHERE role = 'admin'
                      AND is_deleted = 0
                      AND is_super_admin = 0
                      AND owner_user_id = ?
                    """,
                    (_tenant_owner_id(target_user),),
                )
                if len(admin_rows) <= 1 and any(int(item["id"]) == target_user_id for item in admin_rows):
                    raise HTTPException(status_code=400, detail="Нельзя снять роль последнего администратора клиента")
        else:
            owner = _require_tenant_owner(request)
            role = _normalize_tenant_role_or_400(payload.role)
            if int(target_user["id"]) == int(owner["id"]) and role != TENANT_ROLE_OWNER:
                raise HTTPException(status_code=400, detail="Нельзя снять роль администратора у владельца кабинета")
        updated = repository.update_user_role(user_id=target_user_id, role=role)
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True, "by_admin": current_user["email"]}

    @app.post("/api/admin/users/{target_user_id}/password")
    def admin_update_user_password(
        target_user_id: int,
        payload: AdminUserPasswordUpdateRequest,
        request: Request,
    ) -> dict[str, object]:
        actor = _require_admin(request)
        if not _is_super_admin(actor):
            _require_tenant_owner(request)
        _target_user_for_admin_scope(actor=actor, target_user_id=target_user_id)
        updated = repository.update_user_password(
            user_id=target_user_id,
            password_hash=hash_password(payload.password),
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.post("/api/admin/users/{target_user_id}/plan")
    def admin_update_user_plan(
        target_user_id: int,
        payload: UserPlanUpdateRequest,
        request: Request,
    ) -> dict[str, object]:
        actor = _require_admin(request)
        if not _is_super_admin(actor):
            _require_tenant_owner(request)
        target_user = _target_user_for_admin_scope(actor=actor, target_user_id=target_user_id)
        if bool(target_user.get("is_super_admin")):
            raise HTTPException(status_code=400, detail="Нельзя менять тариф супер-администратора")
        plan_code = payload.plan_code.strip().lower()
        if not plan_code:
            raise HTTPException(status_code=400, detail="Код тарифа обязателен")
        plans = repository.list_tariff_plans()
        available_codes = {str(item.get("code") or "").strip().lower() for item in plans}
        if plan_code not in available_codes:
            raise HTTPException(status_code=404, detail="Тариф не найден")
        owner_user_id = _tenant_owner_id(target_user)
        updated = repository.set_tenant_plan(
            owner_user_id=owner_user_id,
            plan_code=plan_code,
            limits_override={},
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.post("/api/admin/users/{target_user_id}/block")
    def admin_block_user(
        target_user_id: int,
        payload: UserBlockUpdateRequest,
        request: Request,
    ) -> dict[str, object]:
        actor = _require_admin(request)
        if not _is_super_admin(actor):
            _require_tenant_owner(request)
        target_user = _target_user_for_admin_scope(actor=actor, target_user_id=target_user_id)
        actor_owner_id = _tenant_owner_id(actor)
        if payload.blocked and int(target_user["id"]) == actor_owner_id:
            raise HTTPException(status_code=400, detail="Нельзя заблокировать собственный аккаунт владельца")
        updated = repository.set_user_blocked(
            user_id=target_user_id,
            blocked=bool(payload.blocked),
            reason=(payload.reason or "").strip() or None,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.post("/api/admin/users/{target_user_id}/delete")
    def admin_delete_user(
        target_user_id: int,
        payload: UserDeleteRequest,
        request: Request,
    ) -> dict[str, object]:
        actor = _require_admin(request)
        if not _is_super_admin(actor):
            _require_tenant_owner(request)
        if not payload.confirm:
            raise HTTPException(status_code=400, detail="Требуется подтверждение удаления")
        target_user = _target_user_for_admin_scope(actor=actor, target_user_id=target_user_id)
        actor_owner_id = _tenant_owner_id(actor)
        if int(target_user["id"]) == actor_owner_id:
            raise HTTPException(status_code=400, detail="Нельзя удалить владельца кабинета")
        deleted = repository.soft_delete_user(user_id=target_user_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        return {"ok": True}

    @app.get("/api/admin/metrics")
    def admin_metrics(request: Request) -> dict[str, object]:
        actor = _require_admin(request)
        if _is_super_admin(actor):
            return repository.get_sla_metrics(user_id=None)
        return repository.get_sla_metrics(user_id=_tenant_owner_id(actor))

    @app.get("/api/admin/actions")
    def admin_actions(
        request: Request,
        page: int = 1,
        page_size: int = 50,
        action_type: str | None = None,
        actor: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        search: str | None = None,
    ) -> dict[str, object]:
        admin_user = _require_admin(request)
        safe_page = max(page, 1)
        safe_page_size = min(max(page_size, 1), 200)
        safe_offset = (safe_page - 1) * safe_page_size
        normalized_action_type = (action_type or "").strip() or None
        normalized_actor = (actor or "").strip() or None
        normalized_search = (search or "").strip() or None
        normalized_date_from = date_from.strip() if date_from else None
        normalized_date_to = date_to.strip() if date_to else None
        for date_value in [normalized_date_from, normalized_date_to]:
            if date_value:
                try:
                    datetime.strptime(date_value, "%Y-%m-%d")
                except ValueError as exc:
                    raise HTTPException(status_code=400, detail="Неверный формат даты. Ожидается YYYY-MM-DD") from exc
        if normalized_date_from and normalized_date_to and normalized_date_from > normalized_date_to:
            raise HTTPException(status_code=400, detail="Дата начала не может быть позже даты окончания")
        owner_scope_user_id = None if _is_super_admin(admin_user) else _tenant_owner_id(admin_user)
        if _is_super_admin(admin_user):
            rows, total = repository.list_recent_actions(
                user_id=None,
                limit=safe_page_size,
                offset=safe_offset,
                action_type=normalized_action_type,
                actor=normalized_actor,
                date_from=normalized_date_from,
                date_to=normalized_date_to,
                search=normalized_search,
            )
        else:
            rows, total = repository.list_recent_actions(
                user_id=owner_scope_user_id,
                limit=safe_page_size,
                offset=safe_offset,
                action_type=normalized_action_type,
                actor=normalized_actor,
                date_from=normalized_date_from,
                date_to=normalized_date_to,
                search=normalized_search,
            )
        filter_options = repository.list_action_filter_options(user_id=owner_scope_user_id)
        return {
            "items": rows,
            "count": len(rows),
            "total": int(total),
            "page": safe_page,
            "page_size": safe_page_size,
            "offset": safe_offset,
            "has_more": (safe_offset + len(rows)) < int(total),
            "filters": {
                "action_type": normalized_action_type or "all",
                "actor": normalized_actor or "all",
                "date_from": normalized_date_from,
                "date_to": normalized_date_to,
                "search": normalized_search or "",
            },
            "filter_options": filter_options,
        }

    @app.get("/api/admin/actions/export")
    def admin_actions_export(
        request: Request,
        format: str = "csv",
        action_type: str | None = None,
        actor: str | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        search: str | None = None,
    ) -> StreamingResponse:
        actor_user = _require_admin(request)
        export_format = format.strip().lower()
        if export_format not in {"csv", "xlsx"}:
            raise HTTPException(status_code=400, detail="Формат экспорта должен быть csv или xlsx")
        normalized_action_type = (action_type or "").strip() or None
        normalized_actor = (actor or "").strip() or None
        normalized_search = (search or "").strip() or None
        normalized_date_from = date_from.strip() if date_from else None
        normalized_date_to = date_to.strip() if date_to else None
        for date_value in [normalized_date_from, normalized_date_to]:
            if date_value:
                try:
                    datetime.strptime(date_value, "%Y-%m-%d")
                except ValueError as exc:
                    raise HTTPException(status_code=400, detail="Неверный формат даты. Ожидается YYYY-MM-DD") from exc
        if normalized_date_from and normalized_date_to and normalized_date_from > normalized_date_to:
            raise HTTPException(status_code=400, detail="Дата начала не может быть позже даты окончания")
        scope_user_id = None if _is_super_admin(actor_user) else _tenant_owner_id(actor_user)
        items, _total = repository.list_recent_actions(
            user_id=scope_user_id,
            limit=200000,
            offset=0,
            action_type=normalized_action_type,
            actor=normalized_actor,
            date_from=normalized_date_from,
            date_to=normalized_date_to,
            search=normalized_search,
        )
        normalized_rows: list[dict[str, str]] = []
        for item in items:
            details = item.get("details")
            details_text = ""
            if isinstance(details, dict):
                pairs: list[str] = []
                for key, value in details.items():
                    pairs.append(f"{key}={value}")
                details_text = "; ".join(pairs)
            row = {
                "created_at": str(item.get("created_at") or ""),
                "actor": str(item.get("actor") or ""),
                "review_uid": str(item.get("review_uid") or ""),
                "action_type": str(item.get("action_type") or ""),
                "details": details_text,
            }
            normalized_rows.append(row)
        if export_format == "csv":
            out = io.StringIO()
            writer = csv.DictWriter(out, fieldnames=["created_at", "actor", "review_uid", "action_type", "details"])
            writer.writeheader()
            for row in normalized_rows:
                writer.writerow(row)
            payload = io.BytesIO(out.getvalue().encode("utf-8-sig"))
            filename = f"admin-actions-{datetime.now(UTC).strftime('%Y%m%d-%H%M%S')}.csv"
            return StreamingResponse(
                payload,
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise HTTPException(
                status_code=500,
                detail="Для экспорта Excel нужен пакет openpyxl. Установите: pip install openpyxl",
            ) from exc
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Лента действий"
        sheet.append(["Время", "Пользователь", "Идентификатор", "Действие", "Детали"])
        for row in normalized_rows:
            sheet.append(
                [
                    row["created_at"],
                    row["actor"],
                    row["review_uid"],
                    row["action_type"],
                    row["details"],
                ]
            )
        out_xlsx = io.BytesIO()
        workbook.save(out_xlsx)
        out_xlsx.seek(0)
        xlsx_name = f"admin-actions-{datetime.now(UTC).strftime('%Y%m%d-%H%M%S')}.xlsx"
        return StreamingResponse(
            out_xlsx,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{xlsx_name}"'},
        )

    # ------------------------------------------------------------------
    # Salary endpoints
    # ------------------------------------------------------------------

    @app.get("/api/admin/salary/rates")
    def admin_salary_get_rates(request: Request) -> dict[str, object]:
        admin_user = _require_admin(request)
        owner_id = _tenant_owner_id(admin_user)
        rates = repository.get_salary_rates(owner_user_id=owner_id)
        return {"ok": True, "rates": rates}

    @app.put("/api/admin/salary/rates")
    def admin_salary_set_rates(request: Request, body: SalaryRatesRequest) -> dict[str, object]:
        admin_user = _require_admin(request)
        owner_id = _tenant_owner_id(admin_user)
        repository.set_salary_rates(
            owner_user_id=owner_id,
            rate_review=body.rate_review,
            rate_question=body.rate_question,
            rate_chat=body.rate_chat,
        )
        rates = repository.get_salary_rates(owner_user_id=owner_id)
        return {"ok": True, "rates": rates}

    @app.get("/api/admin/salary/report")
    def admin_salary_report(
        request: Request,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> dict[str, object]:
        admin_user = _require_admin(request)
        owner_id = _tenant_owner_id(admin_user)
        normalized_date_from = date_from.strip() if date_from else None
        normalized_date_to = date_to.strip() if date_to else None
        for date_value in [normalized_date_from, normalized_date_to]:
            if date_value:
                try:
                    datetime.strptime(date_value, "%Y-%m-%d")
                except ValueError as exc:
                    raise HTTPException(status_code=400, detail="Неверный формат даты. Ожидается YYYY-MM-DD") from exc
        if normalized_date_from and normalized_date_to and normalized_date_from > normalized_date_to:
            raise HTTPException(status_code=400, detail="Дата начала не может быть позже даты окончания")
        rows = repository.get_salary_report(
            owner_user_id=owner_id,
            date_from=normalized_date_from,
            date_to=normalized_date_to,
        )
        rates = repository.get_salary_rates(owner_user_id=owner_id)
        return {
            "ok": True,
            "rows": rows,
            "rates": rates,
            "date_from": normalized_date_from,
            "date_to": normalized_date_to,
        }

    @app.get("/api/admin/salary/report/export")
    def admin_salary_report_export(
        request: Request,
        format: str = "csv",
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> StreamingResponse:
        admin_user = _require_admin(request)
        owner_id = _tenant_owner_id(admin_user)
        export_format = format.strip().lower()
        if export_format not in {"csv", "xlsx"}:
            raise HTTPException(status_code=400, detail="Формат экспорта должен быть csv или xlsx")
        normalized_date_from = date_from.strip() if date_from else None
        normalized_date_to = date_to.strip() if date_to else None
        for date_value in [normalized_date_from, normalized_date_to]:
            if date_value:
                try:
                    datetime.strptime(date_value, "%Y-%m-%d")
                except ValueError as exc:
                    raise HTTPException(status_code=400, detail="Неверный формат даты. Ожидается YYYY-MM-DD") from exc
        if normalized_date_from and normalized_date_to and normalized_date_from > normalized_date_to:
            raise HTTPException(status_code=400, detail="Дата начала не может быть позже даты окончания")
        rows = repository.get_salary_report(
            owner_user_id=owner_id,
            date_from=normalized_date_from,
            date_to=normalized_date_to,
        )
        if export_format == "csv":
            out = io.StringIO()
            writer = csv.DictWriter(
                out,
                fieldnames=["actor", "review_count", "question_count", "chat_count", "total_amount"],
            )
            writer.writeheader()
            for row in rows:
                writer.writerow({
                    "actor": row.get("actor", ""),
                    "review_count": row.get("review_count", 0),
                    "question_count": row.get("question_count", 0),
                    "chat_count": row.get("chat_count", 0),
                    "total_amount": row.get("total_amount", 0),
                })
            payload = io.BytesIO(out.getvalue().encode("utf-8-sig"))
            filename = f"salary-report-{datetime.now(UTC).strftime('%Y%m%d-%H%M%S')}.csv"
            return StreamingResponse(
                payload,
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'},
            )
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise HTTPException(
                status_code=500,
                detail="Для экспорта Excel нужен пакет openpyxl. Установите: pip install openpyxl",
            ) from exc
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Зарплата операторов"
        sheet.append(["Оператор", "Отзывы", "Вопросы", "Чаты", "Сумма (₽)"])
        for row in rows:
            sheet.append([
                row.get("actor", ""),
                row.get("review_count", 0),
                row.get("question_count", 0),
                row.get("chat_count", 0),
                row.get("total_amount", 0),
            ])
        out_xlsx = io.BytesIO()
        workbook.save(out_xlsx)
        out_xlsx.seek(0)
        xlsx_name = f"salary-report-{datetime.now(UTC).strftime('%Y%m%d-%H%M%S')}.xlsx"
        return StreamingResponse(
            out_xlsx,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{xlsx_name}"'},
        )

    def _require_salary_access(request: Request) -> dict[str, object]:
        """Allow tenant owner OR manager with can_salary."""
        user = _require_user(request)
        role = str(user.get("role") or "").strip().lower()
        if role in ROLE_CAN_ACCESS_SETTINGS:
            return user
        if bool(user.get("can_salary")):
            return user
        raise HTTPException(status_code=403, detail="Нет доступа к начислению ЗП")

    def _salary_owner_id(user: dict[str, object]) -> int:
        return _tenant_owner_id(user)

    def _salary_allowed_productions(user: dict[str, object]) -> list[str] | None:
        """None = all productions (owner). List = restricted (manager)."""
        import json as _j
        role = str(user.get("role") or "").strip().lower()
        if role in ROLE_CAN_ACCESS_SETTINGS:
            return None
        try:
            prods = _j.loads(str(user.get("salary_productions") or "[]"))
            return list(prods) if isinstance(prods, list) else []
        except Exception:
            return []

    @app.post("/api/salary/products")
    def create_salary_product(
        payload: SalaryProductCreateRequest, request: Request
    ) -> dict[str, object]:
        user = _require_tenant_owner(request)
        owner_id = _tenant_owner_id(user)
        item = repository.create_salary_product(
            owner_user_id=owner_id,
            order_num=payload.order_num,
            name=payload.name,
            roles=payload.roles,
            price_kineshma_poshiv=round(float(payload.price_kineshma_poshiv), 2),
            price_kineshma_raskroi=round(float(payload.price_kineshma_raskroi), 2),
            price_kineshma_upakovka=round(float(payload.price_kineshma_upakovka), 2),
            price_nerl_poshiv=round(float(payload.price_nerl_poshiv), 2),
            price_nerl_raskroi=round(float(payload.price_nerl_raskroi), 2),
            price_nerl_upakovka=round(float(payload.price_nerl_upakovka), 2),
        )
        return {"ok": True, "item": item}

    @app.put("/api/salary/products/{product_id}")
    def update_salary_product(
        product_id: int,
        payload: SalaryProductCreateRequest,
        request: Request,
    ) -> dict[str, object]:
        user = _require_tenant_owner(request)
        owner_id = _tenant_owner_id(user)
        updated = repository.update_salary_product(
            owner_user_id=owner_id,
            product_id=product_id,
            order_num=payload.order_num,
            name=payload.name,
            roles=payload.roles,
            price_kineshma_poshiv=round(float(payload.price_kineshma_poshiv), 2),
            price_kineshma_raskroi=round(float(payload.price_kineshma_raskroi), 2),
            price_kineshma_upakovka=round(float(payload.price_kineshma_upakovka), 2),
            price_nerl_poshiv=round(float(payload.price_nerl_poshiv), 2),
            price_nerl_raskroi=round(float(payload.price_nerl_raskroi), 2),
            price_nerl_upakovka=round(float(payload.price_nerl_upakovka), 2),
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Товар не найден")
        return {"ok": True}

    class ProductReorderItem(BaseModel):
        id: int = Field(ge=1)
        order_num: int = Field(ge=0)

    class ProductReorderRequest(BaseModel):
        order: list[ProductReorderItem] = Field(default_factory=list)

    @app.put("/api/salary/products/reorder")
    def reorder_salary_products(
        payload: ProductReorderRequest, request: Request
    ) -> dict[str, object]:
        user = _require_tenant_owner(request)
        owner_id = _tenant_owner_id(user)
        repository.reorder_salary_products(
            owner_user_id=owner_id,
            order=[{"id": item.id, "order_num": item.order_num} for item in payload.order],
        )
        return {"ok": True}

    @app.delete("/api/salary/products/{product_id}")
    def delete_salary_product(product_id: int, request: Request) -> dict[str, object]:
        user = _require_tenant_owner(request)
        owner_id = _tenant_owner_id(user)
        deleted = repository.delete_salary_product(owner_user_id=owner_id, product_id=product_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Товар не найден")
        return {"ok": True}

    @app.get("/api/salary/workers")
    def list_salary_workers(request: Request) -> dict[str, object]:
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        workers = repository.list_salary_workers(owner_user_id=owner_id)
        allowed = _salary_allowed_productions(user)
        if allowed is not None:
            workers = [w for w in workers if str(w.get("production") or "") in allowed]
        return {"items": workers, "count": len(workers)}

    @app.post("/api/salary/workers")
    def create_salary_worker(
        payload: SalaryWorkerCreateRequest, request: Request
    ) -> dict[str, object]:
        user = _require_tenant_owner(request)
        owner_id = _tenant_owner_id(user)
        worker = repository.create_salary_worker(
            owner_user_id=owner_id,
            full_name=payload.full_name,
            position=payload.position,
            birth_date=payload.birth_date,
            legal_entity=payload.legal_entity,
            production=payload.production,
            visible_for_accountant=payload.visible_for_accountant,
        )
        return {"ok": True, "item": worker}

    @app.put("/api/salary/workers/{worker_id}")
    def update_salary_worker(
        worker_id: int,
        payload: SalaryWorkerCreateRequest,
        request: Request,
    ) -> dict[str, object]:
        user = _require_tenant_owner(request)
        owner_id = _tenant_owner_id(user)
        updated = repository.update_salary_worker(
            owner_user_id=owner_id,
            worker_id=worker_id,
            full_name=payload.full_name,
            position=payload.position,
            birth_date=payload.birth_date,
            legal_entity=payload.legal_entity,
            production=payload.production,
            visible_for_accountant=payload.visible_for_accountant,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Работник не найден")
        return {"ok": True}

    @app.delete("/api/salary/workers/{worker_id}")
    def delete_salary_worker(worker_id: int, request: Request) -> dict[str, object]:
        user = _require_tenant_owner(request)
        owner_id = _tenant_owner_id(user)
        deleted = repository.delete_salary_worker(owner_user_id=owner_id, worker_id=worker_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Работник не найден")
        return {"ok": True}

    def _fmt_birth_date(raw: str) -> str:
        """Convert YYYY-MM-DD → DD.MM.YYYY; pass through anything else."""
        s = str(raw or "").strip()
        if len(s) == 10 and s[4] == "-":
            y, m, d = s.split("-")
            return f"{d}.{m}.{y}"
        return s

    def _parse_birth_date_import(raw: str) -> tuple:
        """Returns (iso_str, warning_or_None).
        Accepts: DD.MM.YYYY, DD.MM.YY, YYYY-MM-DD, YYYY-MM-DD HH:MM:SS (datetime).
        """
        s = str(raw or "").strip()
        if not s:
            return "", None
        # Strip time component: "1991-08-31 00:00:00" → "1991-08-31"
        if " " in s:
            s = s.split(" ")[0].strip()
        # DD.MM.YYYY or DD.MM.YY
        parts = s.split(".")
        if len(parts) == 3:
            dd, mm, yy = parts
            try:
                year = int(yy) + 2000 if len(yy) == 2 else int(yy)
                return f"{year}-{mm.zfill(2)}-{dd.zfill(2)}", None
            except ValueError:
                pass
        # YYYY-MM-DD
        if len(s) == 10 and s[4] == "-":
            return s, None
        return "", f"нераспознанная дата рождения «{str(raw).strip()}»"

    @app.get("/api/salary/workers/export")
    def export_salary_workers(request: Request):
        import io
        from datetime import date as _date
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        user = _require_tenant_owner(request)
        owner_id = _tenant_owner_id(user)
        workers = repository.list_salary_workers(owner_user_id=owner_id)

        today_str = _date.today().strftime("%d.%m.%Y")
        fname = f"Работники {today_str}"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Работники"

        hdr_font  = Font(bold=True, name="Calibri", size=11)
        hdr_fill  = PatternFill("solid", fgColor="D6E4FF")
        center    = Alignment(horizontal="center", vertical="center")
        left_al   = Alignment(horizontal="left", vertical="center")
        thin      = Side(style="thin", color="BBCCE8")
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ["ФИО", "Должность", "Дата рождения", "Юр. принадлежность", "Производство", "Видимость для бухгалтера"]
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = center; cell.border = border

        for ri, w in enumerate(workers, start=2):
            vals = [
                w.get("full_name") or "",
                w.get("position") or "",
                _fmt_birth_date(str(w.get("birth_date") or "")),  # DD.MM.YYYY
                w.get("legal_entity") or "",
                w.get("production") or "",
                "Да" if w.get("visible_for_accountant") is not False else "Нет",
            ]
            for ci, v in enumerate(vals, start=1):
                cell = ws.cell(row=ri, column=ci, value=str(v))
                cell.alignment = left_al; cell.border = border
                if ci == 3:
                    cell.number_format = "@"
                    cell.quotePrefix = True

        # Auto-fit columns
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len * 1.15 + 2, 10), 50)

        ws.freeze_panes = "A2"
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        from urllib.parse import quote as _q
        cd = "attachment; filename=\"workers.xlsx\"; filename*=UTF-8''" + _q(fname + ".xlsx")
        return StreamingResponse(
            iter([buf.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": cd},
        )

    @app.get("/api/salary/payroll/template")
    def download_payroll_template(
        request: Request,
        date_from: str = "",
        date_to: str = "",
    ):
        """Download a blank payroll XLSX template pre-filled with workers and date columns."""
        import io
        from datetime import date as _date, timedelta
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)

        start = _date(2026, 1, 7)
        end_limit = _date.today() + timedelta(days=14)
        dates: list[str] = []
        d = start
        while d <= end_limit:
            iso = d.isoformat()
            if (not date_from or iso >= date_from) and (not date_to or iso <= date_to):
                dates.append(iso)
            d += timedelta(days=7)

        workers = repository.list_salary_workers(owner_user_id=owner_id)
        allowed = _salary_allowed_productions(user)
        if allowed is not None:
            workers = [w for w in workers if str(w.get("production") or "") in allowed]

        def date_ru(iso: str) -> str:
            y, m, dd = iso.split("-")
            return f"{dd}.{m}.{y}"  # DD.MM.YYYY — 4-digit year

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Шаблон начисления ЗП"

        hdr_font  = Font(bold=True, name="Calibri", size=11)
        fixed_fill = PatternFill("solid", fgColor="D6E4FF")
        date_fill  = PatternFill("solid", fgColor="F0F7FF")
        center = Alignment(horizontal="center", vertical="center")
        left_al = Alignment(horizontal="left", vertical="center")
        thin = Side(style="thin", color="BBCCE8")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        fixed_headers = ["ФИО","Должность","Дата рождения","Юр. принадлежность","Производство"]
        all_headers = fixed_headers + [date_ru(d) for d in dates]

        for ci, h in enumerate(all_headers, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = hdr_font
            cell.fill = fixed_fill if ci <= len(fixed_headers) else date_fill
            cell.alignment = center; cell.border = border
            ws.column_dimensions[cell.column_letter].width = max(len(h)*1.15+2, 8)
            if ci > len(fixed_headers):
                cell.number_format = "@"; cell.quotePrefix = True

        for ri, w in enumerate(workers, start=2):
            fixed_vals = [
                w.get("full_name") or "",
                w.get("position") or "",
                w.get("birth_date") or "",
                w.get("legal_entity") or "",
                w.get("production") or "",
            ]
            for ci, v in enumerate(fixed_vals, start=1):
                cell = ws.cell(row=ri, column=ci, value=str(v))
                cell.alignment = left_al; cell.border = border
                if ci == 3:
                    cell.number_format = "@"; cell.quotePrefix = True
            for ci in range(len(fixed_vals)+1, len(all_headers)+1):
                cell = ws.cell(row=ri, column=ci, value=None)
                cell.border = border; cell.alignment = center

        ws.freeze_panes = "F2"
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return StreamingResponse(
            iter([buf.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=\"payroll_template.xlsx\""},
        )

    def _parse_upload_to_rows(raw: bytes, filename: str) -> list[list[str]]:
        """Parse CSV / XLSX / XLS upload into a list of string rows."""
        import io, csv
        ext = (filename or "").lower().rsplit(".", 1)[-1]

        if ext in ("xlsx",):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            rows = [
                [str(cell.value if cell.value is not None else "").strip() for cell in row]
                for row in ws.iter_rows()
            ]
            wb.close()
            return rows

        if ext in ("xls",):
            try:
                import xlrd
                wb = xlrd.open_workbook(file_contents=raw)
                ws = wb.sheet_by_index(0)
                rows = [
                    [str(ws.cell_value(ri, ci) if ws.cell_value(ri, ci) is not None else "").strip()
                     for ci in range(ws.ncols)]
                    for ri in range(ws.nrows)
                ]
                return rows
            except ImportError:
                raise HTTPException(
                    status_code=400,
                    detail="Формат .xls не поддерживается. Сохраните файл в формате .xlsx и загрузите снова."
                )

        # Default: treat as CSV
        text: str | None = None
        for enc in ("utf-8-sig", "utf-8", "cp1251", "windows-1251", "latin-1"):
            try:
                text = raw.decode(enc); break
            except (UnicodeDecodeError, LookupError):
                continue
        if text is None:
            raise HTTPException(status_code=400, detail="Не удалось определить кодировку файла")
        text = text.replace("\r\n", "\n").replace("\r", "\n")
        delimiter = ";" if text.count(";") >= text.count(",") else ","
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        return list(reader)

    @app.post("/api/salary/workers/import")
    async def import_salary_workers(
        request: Request,
        file: UploadFile = File(...),
        preview: bool = False,
    ) -> dict[str, object]:
        """Import workers from XLSX (same format as export).
        preview=true  → parse + validate, return summary WITHOUT saving.
        preview=false → parse + validate + upsert, return results.
        """
        try:
            user = _require_tenant_owner(request)
            owner_id = _tenant_owner_id(user)
            raw = await file.read()
            rows = _parse_upload_to_rows(raw, file.filename or "")
            if not rows:
                raise HTTPException(status_code=400, detail="Файл пустой")

            VALID_POSITIONS = {
                "Нач. производства", "Менеджер", "Бухгалтер", "Закройщик",
                "Упаковщик", "Швея", "Технический директор", "Генеральный директор",
                "Грузчик", "Комплектовщик", "Разнорабочий",
            }
            VALID_PRODUCTIONS = {"Иваново", "Кинешма", "Нерль"}
            VALID_LEGAL = {"ООО Варфабрик", "ИП Авдеева М.Ю.", "ИП Рехмунов Д.О."}

            # Build name→worker map for upsert logic
            existing = repository.list_salary_workers(owner_user_id=owner_id)
            by_name = {str(w.get("full_name") or "").strip().lower(): w for w in existing}

            # Determine header row
            start_row = 1
            if rows and rows[0]:
                first = str(rows[0][0] or "").strip().lower().replace(" ", "")
                if first not in ("фио", "fullname", "name"):
                    start_row = 0

            warnings = []
            pending = []  # list of dicts ready to upsert

            for i, row in enumerate(rows[start_row:], start=start_row + 1):
                if not row or not any(str(c).strip() for c in row):
                    continue
                full_name = str(row[0]).strip() if len(row) > 0 else ""
                if not full_name or full_name.lower().replace(" ", "") in ("фио", "fullname"):
                    continue

                row_warns = []
                position = str(row[1]).strip() if len(row) > 1 else ""
                if position and position not in VALID_POSITIONS:
                    row_warns.append(f"должность «{position}» не из справочника — будет сохранена как есть")

                raw_bd = str(row[2]).strip() if len(row) > 2 else ""
                birth_date, bd_warn = _parse_birth_date_import(raw_bd)
                if bd_warn:
                    row_warns.append(bd_warn + " — дата рождения будет пустой")
                    birth_date = ""

                legal_entity = str(row[3]).strip() if len(row) > 3 else ""
                if legal_entity and legal_entity not in VALID_LEGAL:
                    row_warns.append(f"юр. принадлежность «{legal_entity}» не из справочника — будет сохранена как есть")

                production = str(row[4]).strip() if len(row) > 4 else ""
                if production and production not in VALID_PRODUCTIONS:
                    row_warns.append(f"производство «{production}» не из справочника — будет пустым")
                    production = ""

                vis_raw = str(row[5]).strip().lower() if len(row) > 5 else ""
                visible = vis_raw not in ("нет", "no", "false", "0")

                if row_warns:
                    warnings.append(f"Строка {i} ({full_name}): {'; '.join(row_warns)}")

                existing_worker = by_name.get(full_name.lower())
                pending.append({
                    "full_name": full_name,
                    "position": position,
                    "birth_date": birth_date,
                    "legal_entity": legal_entity,
                    "production": production,
                    "visible_for_accountant": visible,
                    "existing_id": int(existing_worker["id"]) if existing_worker else None,
                    "action": "update" if existing_worker else "create",
                })

            if preview:
                created_count = sum(1 for p in pending if p["action"] == "create")
                updated_count = sum(1 for p in pending if p["action"] == "update")
                return {
                    "preview": True,
                    "total_rows": len(pending),
                    "to_create": created_count,
                    "to_update": updated_count,
                    "warnings": warnings,
                }

            created = updated = 0
            errors = []
            for p in pending:
                try:
                    if p["action"] == "update":
                        repository.update_salary_worker(
                            owner_user_id=owner_id,
                            worker_id=p["existing_id"],
                            full_name=p["full_name"],
                            position=p["position"],
                            birth_date=p["birth_date"],
                            legal_entity=p["legal_entity"],
                            production=p["production"],
                            visible_for_accountant=p["visible_for_accountant"],
                        )
                        updated += 1
                    else:
                        repository.create_salary_worker(
                            owner_user_id=owner_id,
                            full_name=p["full_name"],
                            position=p["position"],
                            birth_date=p["birth_date"],
                            legal_entity=p["legal_entity"],
                            production=p["production"],
                            visible_for_accountant=p["visible_for_accountant"],
                        )
                        created += 1
                except Exception as exc:
                    errors.append(f"{p['full_name']}: {exc}")

            return {"ok": True, "created": created, "updated": updated, "warnings": warnings, "errors": errors}
        except HTTPException:
            raise
        except Exception as exc:
            _log.error("salary workers import error: %s", exc, exc_info=True)
            raise HTTPException(status_code=500, detail=f"Ошибка импорта: {exc}") from exc

    @app.get("/api/salary/entries")
    def get_salary_entries(
        request: Request,
        worker_id: int = 0,
        entry_date: str = "",
    ) -> dict[str, object]:
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        if not worker_id or not entry_date:
            raise HTTPException(status_code=400, detail="worker_id и entry_date обязательны")
        # Verify worker belongs to allowed production
        allowed = _salary_allowed_productions(user)
        if allowed is not None:
            workers = repository.list_salary_workers(owner_user_id=owner_id)
            target = next((w for w in workers if w.get("id") == worker_id), None)
            if not target or str(target.get("production") or "") not in allowed:
                raise HTTPException(status_code=403, detail="Нет доступа к данному работнику")
        entries = repository.get_salary_entries(
            owner_user_id=owner_id, worker_id=worker_id, entry_date=entry_date
        )
        return {"items": entries, "count": len(entries)}

    @app.get("/api/salary/totals")
    def get_salary_totals(request: Request) -> dict[str, object]:
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        totals = repository.get_salary_totals(owner_user_id=owner_id)
        allowed = _salary_allowed_productions(user)
        if allowed is not None:
            # Filter totals to only workers in allowed productions
            workers = repository.list_salary_workers(owner_user_id=owner_id)
            allowed_ids = {w["id"] for w in workers if str(w.get("production") or "") in allowed}
            totals = [t for t in totals if int(t.get("worker_id") or 0) in allowed_ids]
        return {"items": totals, "count": len(totals)}

    def _payroll_export_cd(name: str) -> str:
        from urllib.parse import quote
        safe = quote(name + ".xlsx")
        return "attachment; filename=\"payroll.xlsx\"; filename*=UTF-8''" + safe

    # Signatories per legal entity (hardcoded per business requirement)
    _LEGAL_SIGNATORIES: dict[str, str] = {
        "ООО Варфабрик":    "Рехмунова Екатерина Анатольевна",
        "ИП Авдеева М.Ю.":  "Авдеева Марина Юрьевна",
        "ИП Рехмунов Д.О.": "Рехмунов Дмитрий Олегович",
    }

    @app.get("/api/salary/payroll/report")
    def export_payroll_report(
        request: Request,
        legal_entity: str = "",
        entry_date: str = "",   # the Wednesday payroll date (YYYY-MM-DD)
        date_from: str = "",    # prev Friday
        date_to: str = "",      # next Thursday
    ):
        """Export a 'Расчёт начислений' XLSX for a specific legal entity and payroll date."""
        import io
        from datetime import date as _date
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        # Check report permission: managers need explicit can_salary_report flag
        _is_manager = str(user.get("role") or "").strip().lower() in TENANT_MANAGER_ROLES
        if _is_manager and not bool(user.get("can_salary_report")):
            raise HTTPException(status_code=403, detail="Нет доступа к экспорту расчёта начислений")

        if not legal_entity or not entry_date:
            raise HTTPException(status_code=400, detail="legal_entity и entry_date обязательны")

        signatory = _LEGAL_SIGNATORIES.get(legal_entity, "")

        # Format dates for header display
        def _fmt_date(iso: str) -> str:
            if not iso: return ""
            parts = iso.split("-")
            if len(parts) != 3: return iso
            return f"{parts[2]}.{parts[1]}.{parts[0]}"

        date_from_display = _fmt_date(date_from) if date_from else _fmt_date(entry_date)
        date_to_display   = _fmt_date(date_to)   if date_to   else _fmt_date(entry_date)

        # Get workers for this legal entity, filtered by manager's allowed productions
        all_workers = repository.list_salary_workers(owner_user_id=owner_id)
        allowed_prods = _salary_allowed_productions(user)  # None = all, [] = none
        workers = [
            w for w in all_workers
            if str(w.get("legal_entity") or "") == legal_entity
            and (allowed_prods is None or str(w.get("production") or "") in allowed_prods)
        ]

        # Get totals for entry_date
        totals = repository.get_salary_totals(owner_user_id=owner_id)
        totals_map: dict[str, float] = {
            f"{t['worker_id']}_{t['entry_date']}": float(t.get("total") or 0)
            for t in totals
        }

        # Build rows: only workers with salary > 0 on this date, sorted by full_name
        rows = []
        for w in sorted(workers, key=lambda x: str(x.get("full_name") or "")):
            amount = totals_map.get(f"{w['id']}_{entry_date}", 0.0)
            rows.append({"name": str(w.get("full_name") or ""), "amount": amount})

        # Build XLSX (portrait, A4)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Расчёт начислений"
        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = 9  # A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToPage = True

        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_al = Alignment(horizontal="left", vertical="center")

        # Column widths filling A4 portrait (≈17 cm usable, units ≈ char width)
        # A = ФИО wide, B = Всего compact
        COL_A_WIDTH = 50   # ~ФИО
        COL_B_WIDTH = 18   # ~Всего, руб.
        ws.column_dimensions["A"].width = COL_A_WIDTH
        ws.column_dimensions["B"].width = COL_B_WIDTH

        # Row 1: УТВЕРЖДАЮ ___ + signatory on same row, two-line via wrap
        ws.merge_cells("A1:B1")
        c = ws.cell(row=1, column=1,
                    value=f"УТВЕРЖДАЮ _____________________________  {signatory}")
        c.font = Font(size=11)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30  # tall enough for possible wrap

        # Row 2: title
        ws.merge_cells("A2:B2")
        c = ws.cell(row=2, column=1,
                    value=f"Расчет начислений с {date_from_display} по {date_to_display}")
        c.font = Font(bold=True, size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22

        # Rows 3-4: blank spacers
        ws.row_dimensions[3].height = 6
        ws.row_dimensions[4].height = 6

        # Row 5: header
        hdr_row = 5
        ws.row_dimensions[hdr_row].height = 28
        for ci, val in enumerate(["ФИО", "Всего, руб."], start=1):
            c = ws.cell(row=hdr_row, column=ci, value=val)
            c.font = Font(bold=True, size=11)
            c.alignment = center
            c.border = border
            c.fill = PatternFill("solid", fgColor="D6E4FF")

        # Data rows
        for i, row in enumerate(rows, start=hdr_row + 1):
            ws.row_dimensions[i].height = 18
            c1 = ws.cell(row=i, column=1, value=row["name"])
            c1.border = border
            c1.alignment = Alignment(horizontal="left", vertical="center")
            c1.font = Font(size=11)

            c2 = ws.cell(row=i, column=2,
                         value=round(row["amount"], 2) if row["amount"] else None)
            c2.border = border
            c2.alignment = Alignment(horizontal="center", vertical="center")
            c2.font = Font(size=11)
            if row["amount"]:
                c2.number_format = '#,##0.00'

        # Print settings: fit columns to one page width
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToPage  = True
        ws.print_options.horizontalCentered = False

        # Margins (inches): left 1.5cm, right 1.5cm, top/bottom 2cm
        ws.page_margins.left   = 0.59
        ws.page_margins.right  = 0.59
        ws.page_margins.top    = 0.79
        ws.page_margins.bottom = 0.79

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)

        from urllib.parse import quote as _q
        # Filename: "Расчёт начислений <юр.лицо> <дата из таблицы дд.мм.гггг>"
        entry_date_display = _fmt_date(entry_date)  # дд.мм.гггг
        fname = f"Расчёт начислений {legal_entity} {entry_date_display}"
        cd = "attachment; filename=\"report.xlsx\"; filename*=UTF-8''" + _q(fname + ".xlsx")
        return StreamingResponse(
            iter([buf.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": cd},
        )

    @app.get("/api/salary/payroll/export")
    def export_payroll_table(
        request: Request,
        date_from: str = "",
        date_to: str = "",
        legal_entity: str = "",
    ):
        """Export the payroll table as XLSX. Dates filtered by date_from/date_to."""
        import io
        from datetime import date as _date, timedelta
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        # Managers need explicit can_salary_zp_export permission
        _is_mgr = str(user.get("role") or "").strip().lower() in TENANT_MANAGER_ROLES
        if _is_mgr and not bool(user.get("can_salary_zp_export")):
            raise HTTPException(status_code=403, detail="Нет доступа к экспорту ЗП")

        # Build 7-day series from Jan 7 2026 to today + 14 days, newest first
        start = _date(2026, 1, 7)
        end_limit = _date.today() + timedelta(days=14)
        dates: list[str] = []
        d = start
        while d <= end_limit:
            iso = d.isoformat()
            if (not date_from or iso >= date_from) and (not date_to or iso <= date_to):
                dates.append(iso)
            d += timedelta(days=7)
        # ascending: oldest left → newest right (freeze pane lets scrolling land on recent data)

        workers = repository.list_salary_workers(owner_user_id=owner_id)
        allowed = _salary_allowed_productions(user)
        if allowed is not None:
            workers = [w for w in workers if str(w.get("production") or "") in allowed]
        # Filter by legal entity if specified
        if legal_entity:
            workers = [w for w in workers if str(w.get("legal_entity") or "") == legal_entity]

        totals = repository.get_salary_totals(owner_user_id=owner_id)
        totals_map: dict[str, float] = {
            f"{t['worker_id']}_{t['entry_date']}": float(t.get("total") or 0)
            for t in totals
        }

        def date_ru(iso: str) -> str:
            y, m, dd = iso.split("-")
            return f"{dd}.{m}.{y[2:]}"

        today_str = _date.today().strftime("%d.%m.%Y")
        # Include legal entity in filename/title if filtered
        report_name = f"Отчет ЗП {legal_entity} {today_str}" if legal_entity else f"Отчет ЗП {today_str}"

        wb = openpyxl.Workbook()
        ws = wb.active
        # Excel sheet name limit is 31 characters
        ws.title = report_name[:31]

        # Styles
        header_font = Font(bold=True, name="Calibri", size=11)
        header_fill = PatternFill("solid", fgColor="D6E4FF")
        date_fill   = PatternFill("solid", fgColor="F0F7FF")
        center      = Alignment(horizontal="center", vertical="center", wrap_text=False)
        left        = Alignment(horizontal="left", vertical="center")
        thin        = Side(style="thin", color="BBCCE8")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header row — 4 fixed columns (no Производство)
        fixed_headers = ["ФИО", "Должность", "Дата рождения", "Юр. принадлежность"]
        date_headers  = [date_ru(d) for d in dates]
        all_headers   = fixed_headers + date_headers

        for ci, h in enumerate(all_headers, start=1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font    = header_font
            cell.fill    = header_fill if ci <= len(fixed_headers) else date_fill
            cell.alignment = center
            cell.border  = border
            if ci > len(fixed_headers):
                cell.number_format = "@"
                cell.quotePrefix = True

        # Data rows
        for wi, w in enumerate(workers, start=2):
            fixed_vals = [
                str(w.get("full_name") or ""),
                str(w.get("position") or ""),
                str(w.get("birth_date") or ""),
                str(w.get("legal_entity") or ""),
            ]
            date_vals = [totals_map.get(f"{w['id']}_{d}", None) for d in dates]
            BIRTH_DATE_COL = 3
            for ci, v in enumerate(fixed_vals, start=1):
                cell = ws.cell(row=wi, column=ci, value=v)
                cell.alignment = left
                cell.border = border
                if ci == BIRTH_DATE_COL:
                    cell.number_format = "@"
                    cell.quotePrefix = True
            for ci, v in enumerate(date_vals, start=len(fixed_vals)+1):
                display_val = round(v, 2) if v else 0.0
                cell = ws.cell(row=wi, column=ci, value=display_val)
                cell.number_format = '#,##0.00'
                cell.alignment = center
                cell.border = border

        # Auto-fit column widths: scan every cell to find the widest content
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                try:
                    val = cell.value
                    if val is None:
                        continue
                    # Format numbers the same way they appear in Excel
                    if isinstance(val, (int, float)):
                        cell_str = f"{val:,.2f}"
                    else:
                        cell_str = str(val)
                    # Account for multi-line content (take widest line)
                    line_len = max(len(line) for line in cell_str.split("\n"))
                    if line_len > max_len:
                        max_len = line_len
                except Exception:
                    pass
            # +2 padding; Calibri 11pt ≈ 1.15× char units; minimum 8, maximum 60
            ws.column_dimensions[col_letter].width = min(max(max_len * 1.15 + 2, 8), 60)

        # Freeze first row + first 4 data columns (ФИО, Должность, Дата рождения, Юр.)
        ws.freeze_panes = "E2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            iter([buf.read()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": _payroll_export_cd(report_name)},
        )

    # ── Payroll data clear ──────────────────────────────────────────────────
    @app.post("/api/salary/clear")
    def clear_salary_data(
        request: Request,
        scope: str = "all",        # all | date | production | legal
        entry_date: str = "",
        production: str = "",
        legal_entity: str = "",
    ) -> dict[str, object]:
        """Clear payroll data by scope. Requires tenant owner."""
        user = _require_tenant_owner(request)
        owner_id = _tenant_owner_id(user)

        worker_ids: list[int] | None = None

        if scope == "date":
            if not entry_date:
                raise HTTPException(status_code=400, detail="entry_date обязателен")
            deleted = repository.clear_salary_data(
                owner_user_id=owner_id, entry_date=entry_date
            )
        elif scope == "production":
            if not production:
                raise HTTPException(status_code=400, detail="production обязателен")
            workers = repository.list_salary_workers(owner_user_id=owner_id)
            worker_ids = [int(w["id"]) for w in workers if str(w.get("production") or "") == production]
            if not worker_ids:
                return {"ok": True, "deleted": 0, "message": "Работники не найдены"}
            deleted = repository.clear_salary_data(
                owner_user_id=owner_id,
                worker_ids=worker_ids,
                entry_date=entry_date or None,
            )
        elif scope == "legal":
            if not legal_entity:
                raise HTTPException(status_code=400, detail="legal_entity обязателен")
            workers = repository.list_salary_workers(owner_user_id=owner_id)
            worker_ids = [int(w["id"]) for w in workers if str(w.get("legal_entity") or "") == legal_entity]
            if not worker_ids:
                return {"ok": True, "deleted": 0, "message": "Работники не найдены"}
            deleted = repository.clear_salary_data(
                owner_user_id=owner_id,
                worker_ids=worker_ids,
                entry_date=entry_date or None,
            )
        else:  # all
            deleted = repository.clear_salary_data(owner_user_id=owner_id)

        return {"ok": True, "deleted": deleted}

    # ── Distribution import (Стежка / Мулетон format) ──────────────────────

    @app.post("/api/salary/payroll/import-distribution")
    async def import_payroll_distribution(
        request: Request,
        file: UploadFile = File(...),
        entry_date: str = "",
        preview: bool = False,
    ) -> dict[str, object]:
        """Parse a multi-sheet distribution XLSX (Стежка/Мулетон) and upsert salary_entries."""
        try:
            import io, openpyxl

            user = _require_salary_access(request)
            owner_id = _salary_owner_id(user)

            if not entry_date:
                raise HTTPException(status_code=400, detail="entry_date обязателен")

            raw = await file.read()
            # keep_vba=True allows opening .xlsm (macro-enabled) files
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, keep_vba=True)

            workers = repository.list_salary_workers(owner_user_id=owner_id)
            by_name = {str(w.get("full_name") or "").strip().lower(): w for w in workers}

            products = repository.list_salary_products(owner_user_id=owner_id)
            prod_by_name = {str(p.get("name") or "").strip().lower(): p for p in products}

            def get_price_snapshot(worker: dict, product: dict) -> float:
                pos = str(worker.get("position") or "")
                prod_loc = str(worker.get("production") or "").lower()
                prefix = "kineshma" if "кинешма" in prod_loc else "nerl" if "нерль" in prod_loc else "ivanovo"
                suffix = "poshiv" if pos == "Швея" else "raskroi" if pos == "Закройщик" else "upakovka" if pos == "Упаковщик" else None
                if not suffix:
                    return 0.0
                key = f"price_{prefix}_{suffix}"
                return float(product.get(key) or 0)

            sheet_names_lower = {s.lower() for s in wb.sheetnames}

            def parse_sheet(ws) -> tuple:
                """Returns (preview_rows, product_headers, worker_product_map, warnings_list)"""
                # Build product column ranges from row 1 (handle merged cells)
                merged_map: dict[int, str] = {}  # col_idx(0-based) → product name
                for mr in ws.merged_cells.ranges:
                    if mr.min_row == 1:
                        val = ws.cell(row=1, column=mr.min_col).value
                        name = str(val or "").strip()
                        for c in range(mr.min_col, mr.max_col + 1):
                            merged_map[c - 1] = name  # 0-based

                header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

                # Build full header list resolving merged cells
                full_headers: list[str] = []
                for i, h in enumerate(header_row):
                    if i in merged_map:
                        full_headers.append(merged_map[i])
                    else:
                        full_headers.append(str(h or "").strip())

                # Find worker column index
                worker_col = 0
                for i, h in enumerate(full_headers):
                    if h and "работник" in h.lower():
                        worker_col = i
                        break

                # Build product → [col_indices]
                # Skip worker col, empty headers, and any "итог*" / "total" column
                product_cols: dict[str, list[int]] = {}
                for i, h in enumerate(full_headers):
                    if i == worker_col:
                        continue
                    hl = h.lower().strip()
                    if not hl or hl.startswith("итог") or hl in ("total", "итого", "итог"):
                        continue
                    product_cols.setdefault(h, []).append(i)

                product_names = list(product_cols.keys())

                # Parse data rows
                warnings: list[str] = []
                worker_data: dict[str, dict[str, int]] = {}  # worker_name → {prod → qty}
                preview_rows: list[list] = []

                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    name_val = row[worker_col] if worker_col < len(row) else None
                    if name_val is None:
                        continue
                    worker_name = str(name_val).strip()
                    wl = worker_name.lower()
                    if not worker_name or wl.startswith("итог") or wl in sheet_names_lower:
                        continue

                    prods: dict[str, int] = {}
                    preview_row: list = [worker_name]
                    for pname in product_names:
                        qty = 0
                        for ci in product_cols[pname]:
                            if ci < len(row):
                                v = row[ci]
                                if isinstance(v, (int, float)) and v > 0:
                                    qty += int(v)
                        prods[pname] = qty
                        preview_row.append(qty if qty else "")

                    worker_data[worker_name] = prods
                    preview_rows.append(preview_row)

                return preview_rows, product_names, worker_data, warnings

            sheets_preview: list[dict] = []
            all_warnings: list[str] = []
            pending: list[dict] = []  # {worker_id, product_id, quantity, price_snapshot}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                prev_rows, prod_names, worker_data, sh_warns = parse_sheet(ws)
                all_warnings.extend(sh_warns)

                sheets_preview.append({
                    "name": sheet_name,
                    "headers": ["Работник"] + prod_names,
                    "rows": prev_rows[:50],  # limit preview to 50 rows
                })

                for worker_name, prods in worker_data.items():
                    w = by_name.get(worker_name.lower())
                    if not w:
                        all_warnings.append(f"[{sheet_name}] Работник «{worker_name}» не найден")
                        continue
                    for prod_name, qty in prods.items():
                        if qty <= 0:
                            continue
                        p = prod_by_name.get(prod_name.lower())
                        if not p:
                            all_warnings.append(f"[{sheet_name}] Товар «{prod_name}» не найден в каталоге")
                            continue
                        price = get_price_snapshot(w, p)
                        pending.append({
                            "worker_id": int(w["id"]),
                            "product_id": int(p["id"]),
                            "quantity": qty,
                            "price_snapshot": price,
                        })

            if preview:
                return {
                    "preview": True,
                    "sheets": sheets_preview,
                    "to_save": len(pending),
                    "warnings": all_warnings,
                }

            # Save — group by worker, upsert salary_entries per worker
            from collections import defaultdict
            by_worker: dict[int, list[dict]] = defaultdict(list)
            for item in pending:
                by_worker[item["worker_id"]].append(item)

            saved = 0
            for worker_id, entries in by_worker.items():
                repository.upsert_salary_entries(
                    owner_user_id=owner_id,
                    worker_id=worker_id,
                    entry_date=entry_date,
                    entries=entries,
                )
                saved += len(entries)

            return {"ok": True, "saved": saved, "warnings": all_warnings}

        except HTTPException:
            raise
        except Exception as exc:
            _log.error("distribution import error: %s", exc, exc_info=True)
            raise HTTPException(status_code=500, detail=f"Ошибка импорта: {exc}") from exc

    @app.post("/api/salary/payroll/import")
    async def import_payroll_table(
        request: Request,
        file: UploadFile = File(...),
        preview: bool = False,
    ) -> dict[str, object]:
        """Import payroll XLSX — upserts total overrides per worker/date.
        preview=true  → parse + validate, return summary WITHOUT saving.
        preview=false → parse + validate + save, return results.
        """
        try:
            user = _require_tenant_owner(request)
            owner_id = _tenant_owner_id(user)

            raw = await file.read()
            rows = _parse_upload_to_rows(raw, file.filename or "")
            if not rows:
                raise HTTPException(status_code=400, detail="Файл пустой")

            header = rows[0]
            from datetime import date as _date, timedelta
            start = _date(2026, 1, 7)
            end_limit = _date.today() + timedelta(days=180)  # allow future dates
            series_dates_set = set()
            d = start
            while d <= end_limit:
                series_dates_set.add(d.isoformat())
                d += timedelta(days=7)

            def parse_date_header(s: str) -> str | None:
                s = str(s).strip()
                parts = s.split(".")
                if len(parts) != 3:
                    return None
                dd, mm, yy = parts
                try:
                    year = int(yy) + 2000 if len(yy) == 2 else int(yy)
                    iso = f"{year}-{mm.zfill(2)}-{dd.zfill(2)}"
                except ValueError:
                    return None
                return iso if iso in series_dates_set else None

            date_cols = []
            unrecognised_dates = []
            for ci, h in enumerate(header[5:], start=5):
                iso = parse_date_header(h)
                if iso:
                    date_cols.append((ci, iso))
                elif str(h).strip():
                    unrecognised_dates.append(str(h).strip())

            workers = repository.list_salary_workers(owner_user_id=owner_id)
            workers_by_name = {
                str(w.get("full_name") or "").strip().lower(): w for w in workers
            }

            VALID_POSITIONS = {
                "Нач. производства", "Менеджер", "Бухгалтер", "Закройщик",
                "Упаковщик", "Швея", "Технический директор", "Генеральный директор",
                "Грузчик", "Комплектовщик", "Разнорабочий",
            }

            warnings = []
            row_results = []   # [{name, matched, cells_found, cells_zero, issues}]
            pending_saves = []  # [{worker_id, entry_date, total_amount}]

            if unrecognised_dates:
                warnings.append(
                    f"Нераспознанные заголовки дат (пропущены): {', '.join(unrecognised_dates)}"
                )

            for ri, row in enumerate(rows[1:], start=2):
                if not row:
                    continue
                full_name = str(row[0]).strip() if row else ""
                if not full_name:
                    continue

                row_issues = []
                w = workers_by_name.get(full_name.lower())
                if not w:
                    warnings.append(f"Строка {ri}: работник «{full_name}» не найден — строка пропущена")
                    row_results.append({"row": ri, "name": full_name, "matched": False, "issues": ["Работник не найден"]})
                    continue

                # Validate optional metadata columns (warn, don't block)
                if len(row) > 1:
                    file_pos = str(row[1]).strip()
                    if file_pos and file_pos not in VALID_POSITIONS:
                        row_issues.append(f"должность «{file_pos}» не из справочника")

                cells_found = 0
                cells_zero = 0
                for ci, iso in date_cols:
                    raw_cell = str(row[ci]).strip() if ci < len(row) else ""
                    if not raw_cell:
                        continue
                    try:
                        amount = float(raw_cell.replace(",", ".").replace("\xa0", "").replace(" ", ""))
                    except ValueError:
                        row_issues.append(f"не удалось разобрать сумму «{raw_cell}» для {iso}")
                        continue
                    if amount == 0.0:
                        cells_zero += 1
                        continue  # skip zero-value cells — don't create override
                    cells_found += 1
                    pending_saves.append({
                        "worker_id": int(w["id"]),
                        "entry_date": iso,
                        "total_amount": amount,
                    })

                if row_issues:
                    warnings.append(f"Строка {ri} ({full_name}): {'; '.join(row_issues)}")
                row_results.append({
                    "row": ri,
                    "name": full_name,
                    "matched": True,
                    "cells_found": cells_found,
                    "cells_zero": cells_zero,
                    "issues": row_issues,
                })

            if preview:
                return {
                    "preview": True,
                    "total_rows": len(row_results),
                    "matched_workers": sum(1 for r in row_results if r.get("matched")),
                    "cells_to_save": len(pending_saves),
                    "warnings": warnings,
                    "rows": row_results,
                }

            # Actually save — store imported amount as "Дополнительные затраты"
            # and clear any previous salary_totals_override to avoid double-counting
            saved = 0
            for item in pending_saves:
                # Remove hard override (if any) so extras are not masked
                repository.set_salary_total_override(
                    owner_user_id=owner_id,
                    worker_id=item["worker_id"],
                    entry_date=item["entry_date"],
                    total_amount=0,  # 0 → deletes the override
                )
                # Replace extras with one "Импортировано" entry
                repository.replace_salary_entry_extras(
                    owner_user_id=owner_id,
                    worker_id=item["worker_id"],
                    entry_date=item["entry_date"],
                    extras=[{"amount": item["total_amount"], "note": "Импортировано"}],
                )
                saved += 1

            return {"ok": True, "saved": saved, "warnings": warnings, "rows": row_results}

        except HTTPException:
            raise
        except Exception as exc:
            _log.error("payroll import error: %s", exc, exc_info=True)
            raise HTTPException(status_code=500, detail=f"Ошибка импорта: {exc}") from exc

    @app.get("/api/salary/products")
    def _list_salary_products_for_payroll(request: Request) -> dict[str, object]:
        """Public to salary users (not just owners)."""
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        items = repository.list_salary_products(owner_user_id=owner_id)
        return {"items": items, "count": len(items)}

    @app.post("/api/salary/entries")
    def save_salary_entries(
        payload: SalaryEntriesSaveRequest, request: Request
    ) -> dict[str, object]:
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        # Verify worker belongs to allowed production
        allowed = _salary_allowed_productions(user)
        if allowed is not None:
            workers = repository.list_salary_workers(owner_user_id=owner_id)
            target = next((w for w in workers if w.get("id") == payload.worker_id), None)
            if not target or str(target.get("production") or "") not in allowed:
                raise HTTPException(status_code=403, detail="Нет доступа к данному работнику")
        repository.upsert_salary_entries(
            owner_user_id=owner_id,
            worker_id=payload.worker_id,
            entry_date=payload.entry_date,
            entries=[e.model_dump() for e in payload.entries],
        )
        return {"ok": True}

    # ── Vacation ───────────────────────────────────────────────────────────
    @app.get("/api/salary/vacations")
    def list_salary_vacations(request: Request) -> dict[str, object]:
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        items = repository.list_salary_vacations(owner_user_id=owner_id)
        allowed = _salary_allowed_productions(user)
        if allowed is not None:
            workers = repository.list_salary_workers(owner_user_id=owner_id)
            allowed_ids = {w["id"] for w in workers if str(w.get("production") or "") in allowed}
            items = [v for v in items if int(v.get("worker_id") or 0) in allowed_ids]
        return {"items": items}

    @app.post("/api/salary/vacation")
    def set_salary_vacation(
        request: Request,
        worker_id: int = 0,
        entry_date: str = "",
        on: bool = True,
    ) -> dict[str, object]:
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        if not worker_id or not entry_date:
            raise HTTPException(status_code=400, detail="worker_id и entry_date обязательны")
        repository.set_salary_vacation(
            owner_user_id=owner_id, worker_id=worker_id, entry_date=entry_date, on=on
        )
        return {"ok": True}

    # ── Oklad ──────────────────────────────────────────────────────────────
    @app.get("/api/salary/oklad")
    def get_salary_oklad(
        request: Request,
        worker_id: int = 0,
        entry_date: str = "",
    ) -> dict[str, object]:
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        if not worker_id or not entry_date:
            raise HTTPException(status_code=400, detail="worker_id и entry_date обязательны")
        amount = repository.get_salary_oklad(
            owner_user_id=owner_id, worker_id=worker_id, entry_date=entry_date
        )
        return {"amount": amount}

    @app.post("/api/salary/oklad")
    def save_salary_oklad(
        payload: SalaryOkladSaveRequest, request: Request
    ) -> dict[str, object]:
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        repository.upsert_salary_oklad(
            owner_user_id=owner_id,
            worker_id=payload.worker_id,
            entry_date=payload.entry_date,
            amount=payload.amount,
        )
        return {"ok": True}

    # ── Extras ─────────────────────────────────────────────────────────────
    @app.get("/api/salary/extras")
    def get_salary_extras(
        request: Request,
        worker_id: int = 0,
        entry_date: str = "",
    ) -> dict[str, object]:
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        if not worker_id or not entry_date:
            raise HTTPException(status_code=400, detail="worker_id и entry_date обязательны")
        items = repository.list_salary_entry_extras(
            owner_user_id=owner_id, worker_id=worker_id, entry_date=entry_date
        )
        return {"items": items, "count": len(items)}

    @app.post("/api/salary/extras")
    def save_salary_extras(
        payload: SalaryExtrasSaveRequest, request: Request
    ) -> dict[str, object]:
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        repository.replace_salary_entry_extras(
            owner_user_id=owner_id,
            worker_id=payload.worker_id,
            entry_date=payload.entry_date,
            extras=[e.model_dump() for e in payload.extras],
        )
        return {"ok": True}

    # ── Worker links ────────────────────────────────────────────────────────
    @app.get("/api/salary/links")
    def get_salary_links(
        request: Request,
        worker_id: int = 0,
    ) -> dict[str, object]:
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        if not worker_id:
            raise HTTPException(status_code=400, detail="worker_id обязателен")
        items = repository.list_salary_worker_links(
            owner_user_id=owner_id, worker_id=worker_id
        )
        return {"items": items, "count": len(items)}

    # ── Extra-production entries ────────────────────────────────────────────
    @app.get("/api/salary/extra-prods")
    def get_salary_extra_prods(
        request: Request,
        worker_id: int = 0,
        entry_date: str = "",
    ) -> dict[str, object]:
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        if not worker_id or not entry_date:
            raise HTTPException(status_code=400, detail="worker_id и entry_date обязательны")
        items = repository.get_salary_extra_prods(
            owner_user_id=owner_id, worker_id=worker_id, entry_date=entry_date
        )
        return {"items": items, "count": len(items)}

    @app.post("/api/salary/extra-prods")
    def save_salary_extra_prods(
        payload: SalaryExtraProdsRequest, request: Request
    ) -> dict[str, object]:
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        repository.upsert_salary_extra_prods(
            owner_user_id=owner_id,
            worker_id=payload.worker_id,
            entry_date=payload.entry_date,
            entries=[e.model_dump() for e in payload.entries],
        )
        return {"ok": True}

    # ── Linked snapshot (historical) ────────────────────────────────────────
    @app.get("/api/salary/linked-snapshot")
    def get_linked_snapshot(
        request: Request,
        worker_id: int = 0,
        entry_date: str = "",
    ) -> dict[str, object]:
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        if not worker_id or not entry_date:
            raise HTTPException(status_code=400, detail="worker_id и entry_date обязательны")
        items = repository.get_salary_linked_snapshot(
            owner_user_id=owner_id, worker_id=worker_id, entry_date=entry_date
        )
        return {"items": items, "count": len(items)}

    @app.post("/api/salary/linked-snapshot")
    def save_linked_snapshot(
        payload: SalaryLinkedSnapshotSaveRequest, request: Request
    ) -> dict[str, object]:
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        repository.save_salary_linked_snapshot(
            owner_user_id=owner_id,
            worker_id=payload.worker_id,
            entry_date=payload.entry_date,
            links=[lnk.model_dump() for lnk in payload.links],
        )
        return {"ok": True}

    @app.get("/api/salary/links/used")
    def get_salary_links_used(request: Request) -> dict[str, object]:
        """Return all linked_worker_ids currently used in any link for this tenant."""
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        used_ids = repository.list_all_salary_linked_ids(owner_user_id=owner_id)
        return {"ids": sorted(used_ids)}

    @app.post("/api/salary/links")
    def add_salary_link(
        payload: SalaryWorkerLinkRequest, request: Request
    ) -> dict[str, object]:
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        if payload.worker_id == payload.linked_worker_id:
            raise HTTPException(status_code=400, detail="Нельзя привязать работника к самому себе")
        repository.add_salary_worker_link(
            owner_user_id=owner_id,
            worker_id=payload.worker_id,
            linked_worker_id=payload.linked_worker_id,
        )
        return {"ok": True}

    @app.delete("/api/salary/links/{link_id}")
    def delete_salary_link(link_id: int, request: Request) -> dict[str, object]:
        user = _require_salary_access(request)
        owner_id = _salary_owner_id(user)
        deleted = repository.delete_salary_worker_link(
            owner_user_id=owner_id, link_id=link_id
        )
        if not deleted:
            raise HTTPException(status_code=404, detail="Связь не найдена")
        return {"ok": True}

    @app.get("/api/salary/my")
    def my_salary(
        request: Request,
        date_from: str | None = None,
        date_to: str | None = None,
    ) -> dict[str, object]:
        user = _require_user(request)
        normalized_date_from = date_from.strip() if date_from else None
        normalized_date_to = date_to.strip() if date_to else None
        for date_value in [normalized_date_from, normalized_date_to]:
            if date_value:
                try:
                    datetime.strptime(date_value, "%Y-%m-%d")
                except ValueError as exc:
                    raise HTTPException(status_code=400, detail="Неверный формат даты. Ожидается YYYY-MM-DD") from exc
        if normalized_date_from and normalized_date_to and normalized_date_from > normalized_date_to:
            raise HTTPException(status_code=400, detail="Дата начала не может быть позже даты окончания")
        owner_id = _tenant_owner_id(user)
        actor = str(user.get("email") or "")
        stats = repository.get_salary_stats_for_actor(
            owner_user_id=owner_id,
            actor=actor,
            date_from=normalized_date_from,
            date_to=normalized_date_to,
        )
        return {
            "ok": True,
            "stats": stats,
            "date_from": normalized_date_from,
            "date_to": normalized_date_to,
        }

    @app.get("/api/admin/sync-status")
    def admin_sync_status(request: Request) -> dict[str, object]:
        _require_super_admin(request)
        with sync_lock:
            return {
                "in_progress": bool(sync_state.get("in_progress")),
                "cancel_requested": bool(sync_state.get("cancel_requested")),
                "last_started_at": sync_state.get("last_started_at"),
                "last_finished_at": sync_state.get("last_finished_at"),
                "polling_enabled": bool(sync_state.get("polling_enabled")),
                "polling_user_id": sync_state.get("polling_user_id"),
                "polling_account_ids": list(sync_state.get("polling_account_ids") or []),
                "polling_since_date": sync_state.get("polling_since_date"),
                "polling_started_at": sync_state.get("polling_started_at"),
                "last_poll_at": sync_state.get("last_poll_at"),
                "last_poll_result": sync_state.get("last_poll_result"),
            }

    @app.post("/api/admin/sync-stop")
    def admin_stop_sync(request: Request) -> dict[str, object]:
        _require_admin(request)  # Any admin (not just super-admin) can stop sync
        was_running = False
        was_polling = False
        with sync_lock:
            was_running = bool(sync_state.get("in_progress"))
            was_polling = bool(sync_state.get("polling_enabled"))
        sync_stop_event.set()
        auto_sync_stop_event.set()
        with sync_lock:
            sync_state["cancel_requested"] = True
            sync_state["polling_enabled"] = False
            sync_state["polling_user_id"] = None
            sync_state["polling_account_ids"] = []
            sync_state["polling_since_date"] = None
            sync_state["polling_started_at"] = None
            sync_state["last_poll_result"] = {
                "ok": True,
                "cancelled": True,
                "message": "Синхронизация остановлена администратором",
                "run_started_at": _now_iso(),
            }
        return {
            "ok": True,
            "was_running": bool(was_running or was_polling),
            "already_stopped": not bool(was_running or was_polling),
        }

    @app.on_event("startup")
    def restore_auto_sync_on_startup() -> None:
        """Resume background polling for all users who have active accounts.

        When the server restarts the in-memory sync_state is lost.  This hook
        reads the database to find tenant-owner users with at least one active
        marketplace account, initialises sync_state, and starts the auto-sync
        worker.  The first poll fires after AUTO_SYNC_INTERVAL_SECONDS (60 s)
        to avoid hammering the marketplace APIs right at startup.
        """
        _log.info("restore_auto_sync_on_startup: starting")
        try:
            owner_users = repository.list_users(owner_only=True)
            _log.info("restore_auto_sync_on_startup: found %d owner users", len(owner_users))
            for user in owner_users:
                uid = int(user.get("id") or 0)
                if uid <= 0:
                    continue
                try:
                    accounts = [
                        item
                        for item in repository.list_marketplace_accounts(uid, include_secrets=False)
                        if item.get("is_active")
                    ]
                    _log.info(
                        "restore_auto_sync_on_startup: user %d has %d active accounts",
                        uid, len(accounts),
                    )
                    if not accounts:
                        continue
                    account_ids = [int(a["id"]) for a in accounts if a.get("id")]
                    sync_settings = repository.get_user_sync_settings(user_id=uid)
                    since_date = (
                        str(sync_settings.get("sync_start_date") or "").strip()
                        if bool(sync_settings.get("use_sync_start_date"))
                        else None
                    )
                    # Set the first user with active accounts as the polling target.
                    # If more tenants exist they will activate their own polling when
                    # they manually trigger a sync.
                    with sync_lock:
                        if not bool(sync_state.get("polling_enabled")):
                            sync_state["polling_enabled"] = True
                            sync_state["polling_user_id"] = uid
                            sync_state["polling_account_ids"] = account_ids
                            sync_state["polling_since_date"] = since_date
                            sync_state["polling_started_at"] = _now_iso()
                    _log.info(
                        "restore_auto_sync_on_startup: starting auto-sync worker for user %d "
                        "accounts=%s since=%s",
                        uid, account_ids, since_date,
                    )
                    _start_auto_sync_worker_if_needed()
                    # Repair any chats whose answered status got lost
                    try:
                        service.repair_all_chat_statuses(user_id=uid)
                    except Exception:
                        pass
                    # Purge review_actions older than 30 days to prevent
                    # unbounded table growth with 200k+ reviews per sync
                    try:
                        purged = repository.purge_old_review_actions(keep_days=30)
                        if purged:
                            _log.info("startup: purged %d old review_actions (>30 days)", purged)
                    except Exception:
                        pass
                    # Purge AI request debug logs older than 1 day
                    try:
                        repository.purge_old_ai_request_logs()
                    except Exception:
                        pass
                    # Purge AI usage stats older than 30 days
                    try:
                        repository.purge_old_ai_usage_logs(keep_days=30)
                    except Exception:
                        pass
                    # Add manually_closed_at column if missing
                    try:
                        with repository._connect() as _conn:
                            repository._migrate_manually_closed_at(_conn)
                    except Exception:
                        pass
                    break
                except Exception as _inner_exc:
                    _log.warning("restore_auto_sync_on_startup: inner error: %s", _inner_exc)
                    continue
        except Exception as _outer_exc:
            _log.error("restore_auto_sync_on_startup: fatal error: %s", _outer_exc)
        # StockScheduler intentionally not started: «Остатки» nav has been hidden
        # since 2026-05-10 (c0e13f2). Keep stock API/manual sync endpoints intact.
        _log.info(
            "restore_auto_sync_on_startup: stock_scheduler left stopped "
            "(Остатки UI hidden; use POST /api/stock/sync for manual runs)"
        )
        try:
            wb_fbs_scheduler.start()
        except Exception as _exc:
            _log.warning("restore_auto_sync_on_startup: wb_fbs_scheduler start failed: %s", _exc)

    # ── Stock module endpoints ────────────────────────────────────────────────

    @app.get("/api/stock/sources")
    def list_stock_sources(request: Request) -> dict[str, object]:
        user = _require_user(request)
        sources = repository.list_stock_sources(user_id=int(user["id"]))
        return {"items": sources, "count": len(sources)}

    @app.post("/api/stock/sources")
    def create_stock_source(request: Request, payload: StockSourceCreateRequest) -> dict[str, object]:
        user = _require_settings_access(request)
        extra = {}
        if payload.client_id:
            extra["client_id"] = payload.client_id
        source = repository.create_stock_source(
            user_id=int(user["id"]),
            marketplace=payload.marketplace.lower().strip(),
            account_name=payload.account_name.strip(),
            api_url=payload.api_url.strip(),
            api_key=payload.api_key.strip(),
            extra=extra,
            interval_hours=payload.interval_hours,
            retention_days=payload.retention_days,
        )
        return {"ok": True, "item": source}

    @app.put("/api/stock/sources/{source_id}")
    def update_stock_source(source_id: int, request: Request, payload: StockSourceUpdateRequest) -> dict[str, object]:
        user = _require_settings_access(request)
        uid = int(user["id"])
        extra_update = None
        if payload.client_id is not None:
            src = repository.get_stock_source(user_id=uid, source_id=source_id, include_secrets=False)
            if src:
                ex = dict(src.get("extra") or {})
                ex["client_id"] = payload.client_id
                extra_update = ex
        updated = repository.update_stock_source(
            user_id=uid,
            source_id=source_id,
            account_name=payload.account_name,
            api_key=payload.api_key,
            interval_hours=payload.interval_hours,
            retention_days=payload.retention_days,
            is_active=payload.is_active,
            extra=extra_update,
        )
        if not updated:
            raise HTTPException(status_code=404, detail="Источник не найден")
        return {"ok": True}

    @app.delete("/api/stock/sources/{source_id}")
    def delete_stock_source(source_id: int, request: Request) -> dict[str, object]:
        user = _require_settings_access(request)
        deleted = repository.delete_stock_source(user_id=int(user["id"]), source_id=source_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Источник не найден")
        return {"ok": True}

    @app.post("/api/stock/sync")
    def sync_stock_sources(request: Request) -> dict[str, object]:
        """Manually trigger stock sync for all active sources."""
        user = _require_settings_access(request)
        uid = int(user["id"])
        sources = repository.list_stock_sources(user_id=uid, include_secrets=True)
        results = []
        for source in sources:
            if not source.get("is_active"):
                continue
            result = sync_stock_source(source, repository)
            results.append(result)
        return {
            "ok": True,
            "synced": len(results),
            "results": results,
        }

    @app.get("/api/stock/reports")
    def list_stock_reports(request: Request, source_id: int | None = None) -> dict[str, object]:
        user = _require_user(request)
        reports = repository.list_stock_reports(user_id=int(user["id"]), source_id=source_id, limit=100)
        return {"items": reports, "count": len(reports)}

    @app.delete("/api/stock/reports")
    def delete_stock_reports(request: Request, source_id: int | None = None) -> dict[str, object]:
        user = _require_settings_access(request)
        deleted = repository.delete_all_stock_reports(user_id=int(user["id"]), source_id=source_id)
        return {"ok": True, "deleted": deleted}

    @app.get("/api/stock/reports/{report_id}/download")
    def download_stock_report(report_id: int, request: Request) -> object:
        from fastapi.responses import FileResponse as _FileResp
        user = _require_user(request)
        reports = repository.list_stock_reports(user_id=int(user["id"]), limit=1000)
        report = next((r for r in reports if int(r.get("id") or 0) == report_id), None)
        if not report:
            raise HTTPException(status_code=404, detail="Отчёт не найден")
        file_path = str(report.get("file_path") or "")
        if not file_path or not Path(file_path).exists():
            raise HTTPException(status_code=404, detail="Файл отчёта не найден на сервере")
        return _FileResp(file_path, filename=Path(file_path).name, media_type="application/json")

    @app.get("/api/stock/data")
    def get_stock_data(request: Request, source_id: int) -> dict[str, object]:
        """Return pivot stock data enriched with product catalog names and zero-fill."""
        user = _require_user(request)
        user_id = int(user["id"])
        owner_id = _tenant_owner_id(user)
        dates = repository.get_stock_data_dates(user_id=user_id, source_id=source_id)
        rows = repository.get_stock_data_pivot(user_id=user_id, source_id=source_id)

        # Determine source marketplace to choose correct catalog lookup key
        source = repository.get_stock_source(user_id=user_id, source_id=source_id)
        is_ozon = str((source or {}).get("marketplace") or "").lower() == "ozon"

        if is_ozon:
            catalog = repository.get_product_catalog_map_ozon(user_id=owner_id)
            # For Ozon: match by wb_article field which stores ozon seller article
            art_key = "wb_article"
        else:
            catalog = repository.get_product_catalog_map(user_id=owner_id)
            art_key = "wb_article"

        if catalog:
            # Build existing (warehouse, article) pairs from report
            existing: set[tuple[str, str]] = {
                (r["warehouse_name"], r.get(art_key, "")) for r in rows
            }
            # Substitute product names
            for r in rows:
                art = r.get(art_key, "")
                if art in catalog:
                    r["seller_article"] = catalog[art]["product_name"] or art
            # Zero-fill: for each warehouse that has data, add missing catalog articles
            warehouses: list[str] = []
            seen_wh: set[str] = set()
            for r in rows:
                wh = r["warehouse_name"]
                if wh not in seen_wh:
                    warehouses.append(wh)
                    seen_wh.add(wh)
            for wh in warehouses:
                for art, cat_item in catalog.items():
                    if (wh, art) not in existing:
                        rows.append({
                            "warehouse_name": wh,
                            "wb_article": art,
                            "seller_article": cat_item["product_name"] or art,
                            "dates": {d: 0 for d in dates},
                        })
        return {"dates": dates, "rows": rows, "count": len(rows)}

    # ── Product catalog endpoints ─────────────────────────────────────────────

    class ProductCatalogItemRequest(BaseModel):
        product_name: str = ""
        wb_article: str = ""
        ozon_article: str = ""

    @app.get("/api/stock/products")
    def list_products(request: Request) -> dict[str, object]:
        user = _require_user(request)
        owner_id = _tenant_owner_id(user)
        items = repository.list_product_catalog(user_id=owner_id)
        return {"ok": True, "items": items, "count": len(items)}

    @app.post("/api/stock/products")
    def upsert_product(request: Request, payload: ProductCatalogItemRequest) -> dict[str, object]:
        user = _require_user(request)
        owner_id = _tenant_owner_id(user)
        if not str(payload.wb_article or "").strip():
            raise HTTPException(status_code=400, detail="wb_article is required")
        item = repository.upsert_product_catalog_item(
            user_id=owner_id,
            wb_article=payload.wb_article,
            product_name=payload.product_name,
            ozon_article=payload.ozon_article,
        )
        return {"ok": True, "item": item}

    @app.delete("/api/stock/products/{item_id}")
    def delete_product(request: Request, item_id: int) -> dict[str, object]:
        user = _require_user(request)
        owner_id = _tenant_owner_id(user)
        deleted = repository.delete_product_catalog_item(user_id=owner_id, item_id=item_id)
        return {"ok": deleted}

    @app.post("/api/stock/products/import")
    async def import_products_excel(
        request: Request,
        file: UploadFile = File(...),
    ) -> dict[str, object]:
        """Import product catalog from Excel. Columns: Наименование товара, Артикул ВБ, Артикул ОЗОН."""
        user = _require_user(request)
        owner_id = _tenant_owner_id(user)
        try:
            import openpyxl  # type: ignore
            data = await file.read()
            import io
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            # Detect header row
            header = None
            col_name = col_wb = col_ozon = None
            for row in rows_iter:
                cells = [str(c or "").strip().lower() for c in row]
                for i, c in enumerate(cells):
                    if any(x in c for x in ("наименование", "название", "name", "товар")):
                        col_name = i
                    elif any(x in c for x in ("артикул вб", "wb_article", "wb article", "артикул wb", "артикул ВБ".lower())):
                        col_wb = i
                    elif any(x in c for x in ("артикул ozon", "артикул озон", "ozon_article", "ozon article")):
                        col_ozon = i
                if col_name is not None or col_wb is not None:
                    break
            # Fallback: assume columns 0=name, 1=wb, 2=ozon
            if col_name is None: col_name = 0
            if col_wb is None: col_wb = 1
            if col_ozon is None: col_ozon = 2
            imported = 0
            for row in rows_iter:
                if not row or all(c is None for c in row):
                    continue
                def _cell(idx: int) -> str:
                    if idx >= len(row): return ""
                    return str(row[idx] or "").strip()
                wb_art = _cell(col_wb)
                if not wb_art:
                    continue
                repository.upsert_product_catalog_item(
                    user_id=owner_id,
                    wb_article=wb_art,
                    product_name=_cell(col_name),
                    ozon_article=_cell(col_ozon),
                )
                imported += 1
            return {"ok": True, "imported": imported}
        except Exception as exc:
            _log.warning("import_products_excel error: %s", exc)
            raise HTTPException(status_code=400, detail=f"Ошибка парсинга файла: {exc}")

    # ── End stock endpoints ───────────────────────────────────────────────────

    # ── Supply module endpoints ───────────────────────────────────────────────

    def _supply_owner_id(user: dict[str, object]) -> int:
        """Return the owner's user_id for supply queries (same tenant logic)."""
        return _tenant_owner_id(user)

    def _can_view_supplies(user: dict[str, object]) -> bool:
        role = str(user.get("role") or ROLE_USER)
        if role in ROLE_CAN_ACCESS_SETTINGS:
            return True
        return bool(user.get("can_supplies"))

    def _can_view_supply_stock(user: dict[str, object]) -> bool:
        role = str(user.get("role") or ROLE_USER)
        if role in ROLE_CAN_ACCESS_SETTINGS:
            return True
        return bool(user.get("can_supply_stock"))

    def _allowed_stock_production_ids(user: dict[str, object]) -> list[int] | None:
        """None = all productions (owner). Otherwise list of allowed production ids."""
        role = str(user.get("role") or ROLE_USER)
        if role in ROLE_CAN_ACCESS_SETTINGS:
            return None
        try:
            import json as _j

            raw = _j.loads(str(user.get("stock_productions") or "[]"))
        except Exception:
            raw = []
        out: list[int] = []
        for x in raw if isinstance(raw, list) else []:
            try:
                pid = int(x)
            except (TypeError, ValueError):
                continue
            if pid > 0:
                out.append(pid)
        return out

    def _moscow_today() -> str:
        try:
            from zoneinfo import ZoneInfo

            return datetime.now(ZoneInfo("Europe/Moscow")).date().isoformat()
        except Exception:
            return datetime.now(UTC).date().isoformat()

    def _can_view_wb_fbs(user: dict[str, object]) -> bool:
        """True if user may open Поставки → ВБ ФБС (owner or explicit wb_fbs grant)."""
        role = str(user.get("role") or ROLE_USER)
        if role in ROLE_CAN_ACCESS_SETTINGS:
            return True
        if not bool(user.get("can_supplies")):
            return False
        perms = repository.get_manager_supply_permissions(manager_user_id=int(user["id"]))
        sources = perms.get("sources") or {}
        return any(
            bool(v.get("wb_fbs"))
            for v in sources.values()
            if isinstance(v, dict)
        )

    def _can_view_wb_fbs_tsd(user: dict[str, object]) -> bool:
        """True if user may open ТСД page for WB FBS assembly (owner or wb_fbs_tsd)."""
        role = str(user.get("role") or ROLE_USER)
        if role in ROLE_CAN_ACCESS_SETTINGS:
            return True
        if not bool(user.get("can_supplies")):
            return False
        perms = repository.get_manager_supply_permissions(manager_user_id=int(user["id"]))
        sources = perms.get("sources") or {}
        return any(
            bool(v.get("wb_fbs_tsd"))
            for v in sources.values()
            if isinstance(v, dict)
        )

    def _require_wb_fbs_tsd(user: dict[str, object]) -> None:
        if not _can_view_wb_fbs_tsd(user):
            raise HTTPException(status_code=403, detail="Нет доступа к ТСД")

    def _wb_fbs_tsd_allowed_source_ids(user: dict[str, object]) -> set[str] | None:
        """None = all sources (owner/settings). Else set of allowed source id strings."""
        role = str(user.get("role") or ROLE_USER)
        if role in ROLE_CAN_ACCESS_SETTINGS:
            return None
        perms = repository.get_manager_supply_permissions(manager_user_id=int(user["id"]))
        return {
            str(sid)
            for sid, sv in (perms.get("sources") or {}).items()
            if isinstance(sv, dict) and sv.get("wb_fbs_tsd")
        }

    def _is_wb_fbs_tenant_owner(user: dict[str, object]) -> bool:
        """Главный пользователь кабинета (или супер-админ)."""
        if _is_super_admin(user):
            return True
        role = str(user.get("role") or ROLE_USER)
        if role not in ROLE_CAN_ACCESS_SETTINGS:
            return False
        try:
            return _tenant_owner_id(user) == int(user["id"])
        except (TypeError, ValueError):
            return False

    def _require_wb_fbs_kiz_owner(user: dict[str, object]) -> None:
        """Вывод КИЗ — только главному пользователю (как шестерёнка auto-sync)."""
        if not _is_wb_fbs_tenant_owner(user):
            raise HTTPException(
                status_code=403,
                detail="Вывод КИЗ доступен только главному пользователю",
            )

    def _require_wb_fbs_owner_tab(user: dict[str, object], tab: str | None) -> None:
        # Tabs finished/cancelled/archive are disabled for everyone (incl. owner).
        if wb_fbs_mod.is_hidden_tab(tab):
            raise HTTPException(
                status_code=403,
                detail="Вкладки «Завершённые», «Отменённые» и «Архив» временно отключены",
            )

    def _sanitize_wb_fbs_owner_counts(
        user: dict[str, object], payload: dict[str, object]
    ) -> dict[str, object]:
        counts = payload.get("counts")
        if not isinstance(counts, dict):
            return payload
        sanitized = dict(counts)
        for tab in wb_fbs_mod.HIDDEN_TABS:
            sanitized[tab] = 0
        out = dict(payload)
        out["counts"] = sanitized
        return out

    @app.get("/api/supply-sources")
    def list_supply_sources(request: Request) -> list[dict[str, object]]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        return repository.list_supply_sources(user_id=owner_id)

    @app.post("/api/supply-sources")
    def create_supply_source(request: Request, payload: CreateSupplySourceRequest) -> dict[str, object]:
        user = _require_user(request)
        if str(user.get("role") or "") not in ROLE_CAN_ACCESS_SETTINGS:
            raise HTTPException(status_code=403, detail="Только владелец может добавлять источники")
        if not payload.name.strip():
            raise HTTPException(status_code=400, detail="Название не может быть пустым")
        if not payload.api_key.strip():
            raise HTTPException(status_code=400, detail="API-ключ не может быть пустым")
        try:
            repository._ensure_supply_tables()
            return repository.create_supply_source(
                user_id=int(user["id"]),
                name=payload.name.strip(),
                api_key=payload.api_key.strip(),
                marketplace=payload.marketplace,
                client_id=payload.client_id,
            )
        except Exception as exc:
            _log.error("create_supply_source error: %s", exc, exc_info=True)
            raise HTTPException(status_code=500, detail=f"Ошибка сервера: {exc}")

    @app.patch("/api/supply-sources/{source_id}/toggle")
    def toggle_supply_source(request: Request, source_id: int, payload: ToggleSupplySourceRequest) -> dict[str, object]:
        user = _require_user(request)
        if str(user.get("role") or "") not in ROLE_CAN_ACCESS_SETTINGS:
            raise HTTPException(status_code=403, detail="Нет доступа")
        ok = repository.toggle_supply_source(
            user_id=int(user["id"]), source_id=source_id, is_enabled=payload.is_enabled
        )
        if not ok:
            raise HTTPException(status_code=404, detail="Источник не найден")
        return {"ok": True, "source_id": source_id, "is_enabled": payload.is_enabled}

    @app.delete("/api/supply-sources/{source_id}")
    def delete_supply_source(request: Request, source_id: int) -> dict[str, object]:
        user = _require_user(request)
        if str(user.get("role") or "") not in ROLE_CAN_ACCESS_SETTINGS:
            raise HTTPException(status_code=403, detail="Нет доступа")
        ok = repository.delete_supply_source(user_id=int(user["id"]), source_id=source_id)
        if not ok:
            raise HTTPException(status_code=404, detail="Источник не найден")
        return {"ok": True}

    @app.get("/api/supplies")
    def list_supplies(
        request: Request,
        source_id: int | None = None,
        status_id: int | None = None,
        date_from: str | None = None,
        date_to: str | None = None,
        production: str | None = None,
        warehouse: str | None = None,
        search: str | None = None,
        page: int = 1,
        page_size: int = 50,
    ) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        return repository.list_supply_items(
            user_id=owner_id,
            source_id=source_id,
            status_id=status_id,
            date_from=date_from,
            production=production or None,
            warehouse=warehouse or None,
            search=search or None,
            date_to=date_to,
            page=page,
            page_size=page_size,
        )

    def _fetch_supply_goods_cached(owner_id: int, supply_id: int) -> list:
        """Return goods from DB; if empty, fetch from WB API and cache (same logic as get_supply_goods endpoint)."""
        cached = repository.get_supply_goods(user_id=owner_id, supply_id=supply_id)
        if cached:
            return cached
        try:
            import urllib.request as _ul2, json as _jm2, ssl as _sl2
            row = repository.get_supply_item_row(user_id=owner_id, supply_id=supply_id)
            if not row:
                return []
            src = repository.get_supply_source_with_key(user_id=owner_id, source_id=int(row["source_id"]))
            if not src or not src.get("api_key"):
                return []
            api_key = str(src["api_key"])
            ctx2 = _sl2.create_default_context()
            req2 = _ul2.Request(
                f"https://supplies-api.wildberries.ru/api/v1/supplies/{supply_id}/goods",
                headers={"Authorization": api_key, "Content-Type": "application/json", "User-Agent": "FeedPilot/1.0"},
                method="GET",
            )
            with _ul2.urlopen(req2, timeout=15, context=ctx2) as r2:
                goods = _jm2.loads(r2.read() or b"[]")
            if isinstance(goods, list) and goods:
                item_row = repository.get_supply_item_row(user_id=owner_id, supply_id=supply_id)
                if item_row:
                    repository.upsert_supply_goods(supply_item_id=int(item_row["id"]), goods=goods)
                return repository.get_supply_goods(user_id=owner_id, supply_id=supply_id)
        except Exception as _ex:
            _log.warning("_fetch_supply_goods_cached supply_id=%d: %s", supply_id, _ex)
        return []

    @app.get("/api/supplies/{supply_id}/goods")
    def get_supply_goods(request: Request, supply_id: int) -> list[dict[str, object]]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        name_map = repository.get_product_name_by_article(user_id=owner_id)

        def _enrich_goods(goods: list[dict]) -> list[dict]:
            for g in goods:
                vc = str(g.get("vendor_code") or "").strip()
                g["product_name"] = name_map.get(vc) or vc or ""
            return goods

        # Check if we have goods cached; if not, fetch from WB and cache
        cached = repository.get_supply_goods(user_id=owner_id, supply_id=supply_id)
        if cached:
            return _enrich_goods(cached)
        # Lazy-fetch from WB API
        try:
            import urllib.request as _ul, json as _jm, ssl as _sl
            def _wb_get(url: str, key: str):
                req = _ul.Request(url, headers={
                    "Authorization": key,
                    "Content-Type": "application/json",
                    "User-Agent": "FeedPilot/1.0",
                }, method="GET")
                ctx = _sl.create_default_context()
                for attempt in range(3):
                    try:
                        with _ul.urlopen(req, timeout=15, context=ctx) as r:
                            return r.status, _jm.loads(r.read() or b"{}")
                    except Exception as e:
                        code = getattr(e, "code", None)
                        if code in (429, 503):
                            import time as _t; _t.sleep((attempt + 1) * 2)
                            continue
                        return (int(code) if code else 0), {}
                return 0, {}

            row = repository.get_supply_item_row(user_id=owner_id, supply_id=supply_id)
            if not row:
                return []
            src = repository.get_supply_source_with_key(user_id=owner_id, source_id=int(row["source_id"]))
            if not src or not src.get("api_key"):
                return []
            api_key = str(src["api_key"])
            # Fetch details (warehouse, quantity)
            det_status, det_data = _wb_get(
                f"https://supplies-api.wildberries.ru/api/v1/supplies/{supply_id}", api_key
            )
            if det_status == 200 and isinstance(det_data, dict):
                det_data["supplyID"] = supply_id
                repository.upsert_supply_item(source_id=int(row["source_id"]), data=det_data)
            # Fetch goods
            g_status, goods = _wb_get(
                f"https://supplies-api.wildberries.ru/api/v1/supplies/{supply_id}/goods", api_key
            )
            if g_status == 200 and isinstance(goods, list):
                item_row = repository.get_supply_item_row(user_id=owner_id, supply_id=supply_id)
                if item_row:
                    repository.upsert_supply_goods(supply_item_id=int(item_row["id"]), goods=goods)
                    return _enrich_goods(repository.get_supply_goods(user_id=owner_id, supply_id=supply_id))
        except Exception as exc:
            _log.warning("lazy supply goods fetch error supply_id=%d: %s", supply_id, exc)
        return []

    @app.get("/api/supplies/{supply_id}/packing-list.pdf")
    def get_packing_list_pdf(request: Request, supply_id: int, slot_index: int = 0):
        """Generate packing list HTML → PDF via LibreOffice."""
        import subprocess as _sp, tempfile as _tf, pathlib as _pl, os as _os
        import html as _hm, json as _jsl4
        from fastapi.responses import Response
        from datetime import datetime as _dtt

        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)

        item_row = repository.get_supply_item_row(user_id=owner_id, supply_id=supply_id)
        if not item_row:
            raise HTTPException(status_code=404, detail="Поставка не найдена")
        item = dict(item_row)

        # Legal entity full name
        entities = repository.list_supply_legal_entities(user_id=owner_id)
        supplier_short = str(item.get("supplier_name") or "")
        le = next((e for e in entities if e.get("short_name") == supplier_short), None) or {}
        full_legal_name = le.get("full_name") or supplier_short

        # Warehouse address (composed from structured fields when present)
        warehouses = repository.list_supply_warehouses(user_id=owner_id)
        wh_map = {
            w["warehouse_name"]: repository.warehouse_address_line(w)
            for w in warehouses if w.get("warehouse_name")
        }
        dest_wh = str(item.get("warehouse_name") or "").strip()
        transit_wh = str(item.get("transit_warehouse_name") or "").strip()
        if transit_wh:
            wh_for_pickup = wh_map.get(transit_wh, transit_wh)
            wh_for_dest   = wh_map.get(dest_wh, dest_wh)
        else:
            wh_for_pickup = wh_map.get(dest_wh, dest_wh)
            wh_for_dest   = wh_for_pickup

        box_labels = {0:"Не указан",1:"Короба",2:"Короба",5:"Монопаллеты / СГТ",6:"Паллеты"}
        box_label = box_labels.get(int(item.get("box_type_id") or 0), "")

        # Pick pass_number from slot (slot_index), pallets = that slot's pallets
        _dj4 = item.get("drivers_json")
        _slots4 = []
        if _dj4:
            try: _slots4 = _jsl4.loads(_dj4)
            except Exception: pass
        if _slots4 and slot_index < len(_slots4):
            _s4 = _slots4[slot_index]
            pass_number   = str(_s4.get("pass_number") or item.get("pass_number") or "")
            pallets_count = str(_s4.get("pallets_count") or item.get("pallets_count") or "")
        elif _slots4:
            total_pallets = sum(int(s.get("pallets_count") or 0) for s in _slots4)
            pass_number   = str(item.get("pass_number") or "")
            pallets_count = str(total_pallets) if total_pallets else str(item.get("pallets_count") or "")
        else:
            pass_number   = str(item.get("pass_number") or "")
            pallets_count = str(item.get("pallets_count") or "")
        supply_id_str = str(supply_id)

        raw_sd = str(item.get("supply_date") or "")
        try:
            sd = _dtt.fromisoformat(raw_sd.replace("Z","").split("T")[0]) if raw_sd else _dtt.now()
            date_display = sd.strftime("%d.%m.%Y")
        except Exception:
            date_display = ""

        e = _hm.escape

        # Use explicit HTML width attributes for LibreOffice (CSS width is ignored)
        td = 'width="60%" style="border:1px solid black;padding:6pt 8pt;vertical-align:middle;font-size:11pt"'
        td_lbl = 'width="40%" style="border:1px solid black;padding:6pt 8pt;vertical-align:middle;font-size:11pt"'
        td_tall = 'width="60%" height="240" style="border:1px solid black;padding:6pt 8pt;vertical-align:top;font-size:11pt"'

        html_content = f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8">
<style>
  @page {{ size: 210mm 297mm; margin: 20mm 15mm 20mm 25mm; }}
  body {{ font-family: "Times New Roman", serif; font-size: 12pt; }}
  h1 {{ text-align: center; font-size: 13pt; font-weight: bold; margin: 0 0 4pt; }}
  h2 {{ text-align: center; font-size: 22pt; font-weight: bold; margin: 12pt 0 8pt; text-transform: uppercase; }}
</style>
</head>
<body>
<h1>Упаковочный лист {e(supplier_short)}</h1>
<h1>(поставка №{e(supply_id_str)}, {e(pass_number)})</h1>
<h2>{e(box_label)}</h2>
<table border="1" cellspacing="0" cellpadding="0" width="100%" style="border-collapse:collapse;margin-top:8pt;table-layout:fixed">
  <colgroup><col width="40%"><col width="60%"></colgroup>
  <tr><td {td_lbl}>Порядковый номер паллеты</td><td {td}>&nbsp;</td></tr>
  <tr><td {td_lbl}>Количество паллет</td><td {td}>{e(pallets_count)}</td></tr>
  <tr><td {td_lbl}>Количество коробок на паллете</td><td {td}>&nbsp;</td></tr>
  <tr><td {td_lbl}>Склад</td><td {td}>{e(wh_for_pickup)}</td></tr>
  <tr><td {td_lbl}>Склад назначения</td><td {td}>{e(wh_for_dest)}</td></tr>
  <tr><td {td_lbl}>Тип поставки</td><td {td}><b>{e(box_label)}</b></td></tr>
  <tr><td {td_lbl}>Наименование юридического лица</td><td {td}>{e(full_legal_name)}</td></tr>
  <tr><td {td_lbl}>Дата поставки</td><td {td}>{e(date_display)}</td></tr>
  <tr height="240"><td {td_lbl}>Штрих-код поставки</td><td {td_tall}><div style="height:240px;min-height:240px">&nbsp;</div></td></tr>
</table>
</body></html>"""

        tmp_dir  = _tf.mkdtemp()
        html_path = _pl.Path(tmp_dir) / f"pl_{supply_id}.html"
        pdf_path  = _pl.Path(tmp_dir) / f"pl_{supply_id}.pdf"
        html_path.write_text(html_content, encoding="utf-8")

        lo_env = dict(_os.environ)
        lo_env["HOME"] = tmp_dir
        lo_env["XDG_CACHE_HOME"]  = tmp_dir
        lo_env["XDG_CONFIG_HOME"] = tmp_dir
        lo_env["XDG_RUNTIME_DIR"] = tmp_dir
        lo_env["DCONF_PROFILE"]   = "/dev/null"

        lo_ok = False
        for binary in ("/usr/bin/soffice", "/usr/lib/libreoffice/program/soffice", "soffice", "libreoffice"):
            try:
                result = _sp.run(
                    [binary, "--headless", "--norestore",
                     f"-env:UserInstallation=file://{tmp_dir}/lo_profile",
                     "--convert-to", "pdf",
                     "--outdir", tmp_dir, str(html_path)],
                    capture_output=True, timeout=60, env=lo_env
                )
                if result.returncode == 0 and pdf_path.exists():
                    lo_ok = True
                    break
            except FileNotFoundError:
                continue
            except _sp.TimeoutExpired:
                raise HTTPException(status_code=504, detail="Таймаут конвертации")

        if not lo_ok:
            raise HTTPException(status_code=500, detail="Ошибка конвертации упаковочного листа в PDF")

        return Response(
            content=pdf_path.read_bytes(),
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="PackingList_{supply_id}.pdf"'},
        )

    @app.get("/api/supplies/{supply_id}/poa.pdf")
    def get_poa_pdf(request: Request, supply_id: int, slot_index: int = 0):
        """Generate Power of Attorney HTML → PDF via LibreOffice."""
        import subprocess as _sp, tempfile as _tf, pathlib as _pl, os as _os
        import html as _hm, json as _jsl
        from fastapi.responses import Response
        from datetime import datetime as _dtt

        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)

        item_row = repository.get_supply_item_row(user_id=owner_id, supply_id=supply_id)
        if not item_row:
            raise HTTPException(status_code=404, detail="Поставка не найдена")
        item = dict(item_row)

        entities = repository.list_supply_legal_entities(user_id=owner_id)
        supplier_short = str(item.get("supplier_name") or "").strip()
        # Exact match first, then case-insensitive strip match, then full_name match — never fall back to entities[0]
        le = (
            next((e for e in entities if str(e.get("short_name") or "").strip() == supplier_short), None)
            or next((e for e in entities if str(e.get("short_name") or "").strip().lower() == supplier_short.lower()), None)
            or next((e for e in entities if str(e.get("full_name") or "").strip() == supplier_short), None)
            or {}
        )
        org_full = le.get("full_name") or supplier_short
        org_req  = le.get("requisites") or ""
        org_line = ", ".join(filter(None, [org_full, org_req]))
        signatories = le.get("signatories") or supplier_short

        drivers = repository.list_supply_drivers(user_id=owner_id)
        # Multi-driver: pick slot by slot_index
        _dj = item.get("drivers_json")
        _slots = []
        if _dj:
            try: _slots = _jsl.loads(_dj)
            except Exception: pass
        if _slots and slot_index < len(_slots):
            _slot = _slots[slot_index]
            # manual_driver_name takes priority over dropdown driver_name
            driver_name = str(_slot.get("manual_driver_name") or _slot.get("driver_name") or "")
            pallets_slot = str(_slot.get("pallets_count") or item.get("pallets_count") or "")
            _manual_docs = str(_slot.get("manual_driver_docs") or "")
        else:
            driver_name = str(item.get("driver_name") or "")
            pallets_slot = str(item.get("pallets_count") or "")
            _manual_docs = ""
        # Override item pallets for this slot
        item = dict(item)
        item["pallets_count"] = pallets_slot
        driver_obj  = next((d for d in drivers if d.get("full_name") == driver_name), {})
        # Use manual docs if provided, otherwise composed line from registry (VU fields).
        driver_docs = _manual_docs if _manual_docs else repository.driver_documents_line(driver_obj)

        now = _dtt.now()
        date_display = now.strftime("%d.%m.%Y")
        supply_id_str = str(supply_id)

        goods_list = _fetch_supply_goods_cached(owner_id, supply_id)
        name_map = repository.get_product_name_by_article(user_id=owner_id)
        for g in goods_list:
            vc = str(g.get("vendor_code") or "")
            g["product_name"] = name_map.get(vc) or vc or ""
        if not goods_list:
            pallets_raw = int(item.get("pallets_count") or 0)
            goods_list = [{"product_name": "Текстильные товары", "quantity": pallets_raw}]

        e = _hm.escape
        _data_rows = "".join(
            f'<tr>'
            f'<td style="border:1px solid black;padding:0 2pt;text-align:center;white-space:nowrap;line-height:1.1">{i+1}</td>'
            f'<td class="mat-name" style="border:1px solid black;padding:0 2pt;text-align:left;white-space:normal;line-height:1.1">{e(g.get("product_name") or "Товар")}</td>'
            f'<td style="border:1px solid black;padding:0 2pt;text-align:center;white-space:nowrap;line-height:1.1">шт.</td>'
            f'<td style="border:1px solid black;padding:0 2pt;text-align:center;white-space:nowrap;line-height:1.1">{g.get("quantity") or "—"}</td>'
            f'</tr>'
            for i, g in enumerate(goods_list)
        )
        # Single unified table so all columns stay aligned
        goods_rows = _data_rows

        # Use underscores for signature lines — LibreOffice renders these reliably
        UL = "_" * 15  # underline substitute
        sig_name = e(signatories) if signatories and signatories != "—" else ""

        html_content = f"""<!DOCTYPE html>
<html xmlns:o="urn:schemas-microsoft-com:office:office"
      xmlns:w="urn:schemas-microsoft-com:office:word"
      xmlns="http://www.w3.org/TR/REC-html40">
<head><meta charset="utf-8">
<style>
  @page {{ size: 210mm 297mm; margin: 15mm 10mm 15mm 25mm; }}
  @page Section1 {{
    size: 210.0mm 297.0mm;
    margin: 15.0mm 10.0mm 15.0mm 25.0mm;
    mso-header-margin: 0mm;
    mso-footer-margin: 0mm;
    mso-paper-source: 0;
  }}
  div.Section1 {{ page: Section1; }}
  body {{ font-family: "Times New Roman", serif; font-size: 9pt; line-height: 1.05; }}
  .small {{ font-size: 7pt; text-align: center; }}
  .underline {{ text-decoration: underline; }}
  .bold {{ font-weight: bold; }}
  table.outer {{ width: 100%; border-collapse: collapse; margin-bottom: 4pt; }}
  table.codes {{ border-collapse: collapse; margin-left: auto; font-size: 8pt; border: 1px solid #000; }}
  table.codes td {{ border: 1px solid #000; padding: 1pt 4pt; }}
  table.mat {{ width: 100%; border-collapse: collapse; margin-top: 2pt; font-size: 9pt; border: 1px solid #000; }}
  table.mat td, table.mat th {{ border: 1px solid #000; padding: 0 2pt; line-height: 1.1; }}
  table.mat td.mat-name, table.mat th.mat-name {{ text-align: left; white-space: normal; }}
  table.mat td, table.mat th {{ text-align: center; white-space: nowrap; }}
  .dotline {{ display: inline-block; border-bottom: 1px solid #000; min-width: 120pt; }}
  p {{ margin: 0; padding: 0; }}
  p {{ margin: 2pt 0; }}
</style>
</head>
<body><div class="Section1">

<table class="outer">
  <tr>
    <td style="width:55%;vertical-align:top">{e(org_full)}</td>
    <td style="width:45%;vertical-align:top;text-align:right;font-size:8pt">
      Типовая межотраслевая форма № М-2<br>
      Утверждена постановлением Госстата России от 30.10.97 № 71а<br>
      <table class="codes" border="1" cellspacing="0">
        <tr><td colspan="2" align="center"><b>Коды</b></td></tr>
        <tr><td>Форма по ОКУД</td><td>0315001</td></tr>
        <tr><td>по ОКПО</td><td>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</td></tr>
      </table>
    </td>
  </tr>
</table>

<p align="center" style="font-size:12pt;font-weight:bold;text-align:center"><b>Доверенность № {e(supply_id_str)}</b></p>
<p>Дата выдачи <b><u>{e(date_display)}</u></b></p>
<p>Доверенность действительна 14 дней с даты подписания.</p>
<p><span class="underline">{e(org_line)}</span></p>
<p class="small">(наименование потребителя и его адрес)</p>
<p><span class="underline">{e(org_line)}</span></p>
<p class="small">(наименование плательщика и его адрес)</p>
<p>Доверенность выдана &nbsp; <u>водителю</u> &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; <u>{e(driver_name)}</u></p>
<p class="small">(должность) &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; (фамилия, имя, отчество)</p>
{f"<p>{e(driver_docs)}</p>" if driver_docs else ""}
<p>На отправку груза от &nbsp;&nbsp; <u>&nbsp;{e(supplier_short)}&nbsp;</u></p>
<p class="small">(наименование поставщика)</p>
<p>материальных ценностей по транспортной накладной &nbsp; <b><u>{e(supply_id_str)}</u></b> &nbsp; от &nbsp; <b><u>{e(date_display)}</u></b></p>
<p class="small">(наименование, номер и дата документа)</p>
<p>Перечень материальных ценностей, подлежащих доставке</p>
<table class="mat" border="1" cellspacing="0" width="100%" style="width:100%;table-layout:fixed">
  <colgroup><col style="width:3%"><col style="width:77%"><col style="width:10%"><col style="width:10%"></colgroup>
  <tr>
    <th style="border:1px solid black;padding:0 2pt;white-space:nowrap;line-height:1.1">№ по порядку</th>
    <th class="mat-name" style="border:1px solid black;padding:0 2pt;text-align:left;white-space:normal;line-height:1.1">Материальные ценности</th>
    <th style="border:1px solid black;padding:0 2pt;white-space:nowrap;line-height:1.1">Ед. изм.</th>
    <th style="border:1px solid black;padding:0 2pt;white-space:nowrap;line-height:1.1">Кол-во</th>
  </tr>
  {goods_rows}
</table>

<p style="margin-top:4pt">Подпись лица, получившего доверенность удостоверяем. &nbsp;&nbsp;&nbsp;&nbsp; {UL} &nbsp;&nbsp; ({e(driver_name)})</p>
<table width="100%" cellspacing="0" cellpadding="2" style="margin-top:4pt">
  <tr>
    <td width="25%" valign="bottom">Руководитель<br><small>М.П.</small></td>
    <td width="30%" valign="bottom" align="center">{UL}<br><small>подпись</small></td>
    <td width="45%" valign="bottom" align="center">{sig_name}<br><small>расшифровка подписи</small></td>
  </tr>
</table>
<table width="100%" cellspacing="0" cellpadding="2" style="margin-top:3pt">
  <tr>
    <td width="25%" valign="bottom">Главный бухгалтер</td>
    <td width="30%" valign="bottom" align="center">{UL}<br><small>подпись</small></td>
    <td width="45%" valign="bottom" align="center">{sig_name}<br><small>расшифровка подписи</small></td>
  </tr>
</table>
</div></body></html>"""

        tmp_dir   = _tf.mkdtemp()
        html_path = _pl.Path(tmp_dir) / f"poa_{supply_id}.html"
        pdf_path  = _pl.Path(tmp_dir) / f"poa_{supply_id}.pdf"
        html_path.write_text(html_content, encoding="utf-8")

        lo_env = dict(_os.environ)
        for k,v in [("HOME",tmp_dir),("XDG_CACHE_HOME",tmp_dir),("XDG_CONFIG_HOME",tmp_dir),
                    ("XDG_RUNTIME_DIR",tmp_dir),("DCONF_PROFILE","/dev/null")]:
            lo_env[k] = v

        lo_ok = False
        for binary in ("/usr/bin/soffice","/usr/lib/libreoffice/program/soffice","soffice","libreoffice"):
            try:
                r = _sp.run([binary,"--headless","--norestore",
                             f"-env:UserInstallation=file://{tmp_dir}/lo_profile",
                             "--convert-to","pdf","--outdir",tmp_dir,str(html_path)],
                            capture_output=True,timeout=60,env=lo_env)
                if r.returncode == 0 and pdf_path.exists():
                    lo_ok = True; break
            except FileNotFoundError:
                continue
            except _sp.TimeoutExpired:
                raise HTTPException(status_code=504, detail="Таймаут конвертации")

        if not lo_ok:
            raise HTTPException(status_code=500, detail="Ошибка конвертации доверенности в PDF")

        return Response(
            content=pdf_path.read_bytes(),
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="PoA_{supply_id}.pdf"'},
        )

    @app.get("/api/supplies/{supply_id}/nm-prices")
    def get_supply_nm_prices(request: Request, supply_id: int) -> dict[str, object]:
        """Return {nmID: discountedPrice} for all goods using the source api_key (same token, prices scope)."""
        import urllib.request as _ul, json as _jm, ssl as _sl
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        row = repository.get_supply_item_row(user_id=owner_id, supply_id=supply_id)
        if not row:
            return {"prices": {}}
        src = repository.get_supply_source_with_key(user_id=owner_id, source_id=int(row["source_id"]))
        if not src or not src.get("api_key"):
            return {"prices": {}}
        api_key = str(src["api_key"])
        ctx = _sl.create_default_context()
        prices: dict[str, float] = {}
        offset = 0
        try:
            while True:
                url = f"https://discounts-prices-api.wildberries.ru/api/v2/list/goods/filter?limit=1000&offset={offset}"
                req = _ul.Request(url, method="GET", headers={
                    "Authorization": api_key, "User-Agent": "Mozilla/5.0"
                })
                with _ul.urlopen(req, context=ctx, timeout=15) as r:
                    data = _jm.loads(r.read())
                page = data.get("data", {}).get("listGoods", [])
                for g in page:
                    nm = g.get("nmID")
                    sizes = g.get("sizes") or []
                    dp = float(sizes[0].get("discountedPrice", 0)) if sizes else 0.0
                    if nm and dp > 0:
                        prices[str(nm)] = dp
                offset += len(page)
                if len(page) < 1000:
                    break
        except Exception as exc:
            _log.warning("nm-prices fetch error supply_id=%d: %s", supply_id, exc)
        return {"prices": prices}

    @app.get("/api/supplies/{supply_id}/ttn.pdf")
    def get_ttn_pdf(request: Request, supply_id: int, slot_index: int = 0):
        """Generate TTN DOCX from same template as download button, convert to PDF via LibreOffice."""
        import subprocess as _sp, tempfile as _tf, zipfile as _zf, io as _io
        import pathlib as _pl
        import urllib.request as _ul, json as _jm, ssl as _sl
        import html as _html_mod
        from fastapi.responses import Response

        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)

        # ── Fetch supply data ──────────────────────────────────────────────
        item_row = repository.get_supply_item_row(user_id=owner_id, supply_id=supply_id)
        if not item_row:
            raise HTTPException(status_code=404, detail="Поставка не найдена")

        item = dict(item_row)
        supply_id_str = str(supply_id)

        # ── Legal entity ───────────────────────────────────────────────────
        entities = repository.list_supply_legal_entities(user_id=owner_id)
        supplier_short = str(item.get("supplier_name") or "").strip()
        le = (
            next((e for e in entities if str(e.get("short_name") or "").strip() == supplier_short), None)
            or next((e for e in entities if str(e.get("short_name") or "").strip().lower() == supplier_short.lower()), None)
            or next((e for e in entities if str(e.get("full_name") or "").strip() == supplier_short), None)
            or {}
        )
        org_full = le.get("full_name") or supplier_short
        org_req  = le.get("requisites") or ""
        org_line = ", ".join(filter(None, [org_full, org_req]))

        # ── Dates ──────────────────────────────────────────────────────────
        from datetime import datetime as _dtt
        now = _dtt.now()
        date_disp = now.strftime("%d.%m.%Y")
        raw_sd = str(item.get("supply_date") or "")
        try:
            sd = _dtt.fromisoformat(raw_sd.replace("Z","").split("T")[0]) if raw_sd else now
            supply_date_disp = sd.strftime("%d.%m.%Y")
        except Exception:
            supply_date_disp = date_disp

        # Multi-driver: pick slot by slot_index
        import json as _jsl2
        _dj2 = item.get("drivers_json")
        _slots2 = []
        if _dj2:
            try: _slots2 = _jsl2.loads(_dj2)
            except Exception: pass
        if _slots2 and slot_index < len(_slots2):
            _slot2 = _slots2[slot_index]
            driver_name = str(_slot2.get("manual_driver_name") or _slot2.get("driver_name") or "")
            pallets     = int(_slot2.get("pallets_count") or 0)
        else:
            driver_name = str(item.get("driver_name") or "")
            pallets     = int(item.get("pallets_count") or 0)
        VAT_RATE        = 0.22
        wh              = str(item.get("warehouse_name") or "").strip()
        transit_wh      = str(item.get("transit_warehouse_name") or "").strip()
        # Recipient address = initial (transit) warehouse; fallback to destination
        pickup_wh       = transit_wh or wh

        # ── Recipient (consignee) line from warehouse settings ─────────────
        warehouses = repository.list_supply_warehouses(user_id=owner_id)
        wh_addr = next(
            (repository.warehouse_address_line(w) for w in warehouses
             if str(w.get("warehouse_name") or "").strip() == pickup_wh),
            "",
        )
        recipient_line = "ООО «РВБ»" + (f", {wh_addr}" if wh_addr else "")

        # ── Goods list ─────────────────────────────────────────────────────
        goods_list = _fetch_supply_goods_cached(owner_id, supply_id)
        name_map = repository.get_product_name_by_article(user_id=owner_id)
        for g in goods_list:
            vc = str(g.get("vendor_code") or "")
            g["product_name"] = name_map.get(vc) or vc or ""

        # ── Prices from WB ─────────────────────────────────────────────────
        nm_prices: dict[int, float] = {}
        try:
            src = repository.get_supply_source_with_key(user_id=owner_id, source_id=int(item_row["source_id"]))
            if src and src.get("api_key"):
                api_key = str(src["api_key"])
                ctx = _sl.create_default_context()
                offset = 0
                while True:
                    url = f"https://discounts-prices-api.wildberries.ru/api/v2/list/goods/filter?limit=1000&offset={offset}"
                    req = _ul.Request(url, method="GET", headers={"Authorization": api_key, "User-Agent": "Mozilla/5.0"})
                    with _ul.urlopen(req, context=ctx, timeout=15) as r:
                        data = _jm.loads(r.read())
                    page = data.get("data", {}).get("listGoods", [])
                    for g in page:
                        nm = int(g.get("nmID") or 0)
                        sizes = g.get("sizes") or []
                        dp = float(sizes[0].get("discountedPrice", 0)) if sizes else 0.0
                        if nm and dp > 0:
                            nm_prices[nm] = dp
                    offset += len(page)
                    if len(page) < 1000:
                        break
        except Exception as ex:
            _log.warning("ttn-pdf prices fetch: %s", ex)

        # ── Number to Russian words ────────────────────────────────────────
        def _rubles_in_words(n: int) -> str:
            ones_m = ["","один","два","три","четыре","пять","шесть","семь","восемь","девять"]
            ones_f = ["","одна","две","три","четыре","пять","шесть","семь","восемь","девять"]
            teens  = ["десять","одиннадцать","двенадцать","тринадцать","четырнадцать",
                      "пятнадцать","шестнадцать","семнадцать","восемнадцать","девятнадцать"]
            tens   = ["","","двадцать","тридцать","сорок","пятьдесят","шестьдесят","семьдесят","восемьдесят","девяносто"]
            hunds  = ["","сто","двести","триста","четыреста","пятьсот","шестьсот","семьсот","восемьсот","девятьсот"]
            def chunk(x, fem):
                r,w = x%100,[]
                h = x//100
                if h: w.append(hunds[h])
                if r>=10 and r<=19: w.append(teens[r-10])
                else:
                    if r//10: w.append(tens[r//10])
                    d = r%10
                    if d: w.append((ones_f if fem else ones_m)[d])
                return w
            if n==0: return "ноль рублей 00 копеек"
            w=[]
            bn = n//1000000000
            mn = (n//1000000)%1000
            th = (n//1000)%1000
            ru = n%1000
            if bn:
                bw=chunk(bn,False)
                w.extend(bw)
                w.append(["миллиардов","миллиард","миллиарда","миллиардов"][1 if bn%10==1 and bn%100!=11 else 2 if bn%10 in(2,3,4) and bn%100 not in range(12,15) else 0 if bn==0 else 3])
            if mn:
                mw=chunk(mn,False)
                w.extend(mw)
                w.append(["миллионов","миллион","миллиона","миллионов"][1 if mn%10==1 and mn%100!=11 else 2 if mn%10 in(2,3,4) and mn%100 not in range(12,15) else 0 if mn==0 else 3])
            if th:
                tw2=chunk(th,True)
                w.extend(tw2)
                w.append(["тысяч","тысяча","тысячи","тысяч"][1 if th%10==1 and th%100!=11 else 2 if th%10 in(2,3,4) and th%100 not in range(12,15) else 0 if th==0 else 3])
            if ru:
                w.extend(chunk(ru,False))
            rub_w = ["рублей","рубль","рубля","рублей"][1 if ru%10==1 and ru%100!=11 else 2 if ru%10 in(2,3,4) and ru%100 not in range(12,15) else 0 if ru==0 else 3]
            w.append(rub_w)
            w.append("00 копеек")
            return " ".join(w)

        def fmt2(x: float) -> str:
            return f"{x:,.2f}".replace(",", " ").replace(".", ",")

        # ── Build per-row data ─────────────────────────────────────────────
        total_excl = total_vat = total_incl = 0.0
        qty_total = sum(int(g.get("quantity") or 0) for g in goods_list) or pallets

        rows_data = []
        for i, g in enumerate(goods_list):
            qty = int(g.get("quantity") or 0)
            nm  = int(g.get("nm_id") or 0)
            price_incl = nm_prices.get(nm)
            price_excl = price_incl / (1 + VAT_RATE) if price_incl else None
            amt_excl   = price_excl * qty if price_excl is not None else None
            vat_amt    = amt_excl * VAT_RATE if amt_excl is not None else None
            amt_incl   = amt_excl + vat_amt if amt_excl is not None else None
            if amt_excl is not None:
                total_excl += amt_excl; total_vat += vat_amt; total_incl += amt_incl
            rows_data.append({
                "num": i + 1,
                "name": g.get("product_name") or g.get("vendor_code") or "Товар",
                "qty": qty,
                "price_excl": fmt2(price_excl) if price_excl is not None else "—",
                "amt_excl":   fmt2(amt_excl)   if amt_excl   is not None else "—",
                "vat_amt":    fmt2(vat_amt)     if vat_amt    is not None else "—",
                "amt_incl":   fmt2(amt_incl)    if amt_incl   is not None else "—",
            })

        t_excl = fmt2(total_excl) if total_excl else "—"
        t_vat  = fmt2(total_vat)  if total_vat  else "—"
        t_incl = fmt2(total_incl) if total_incl else "—"
        amt_words = _rubles_in_words(round(total_incl)) if total_incl else "—"

        # Generate DOCX from same template as "Скачать ТТН" button
        import zipfile as _zf, io as _io, re as _re, html as _html_esc
        tpl_path = STATIC_DIR / "torg12_tpl.docx"
        with open(tpl_path, "rb") as f:
            tpl_bytes = f.read()
        with _zf.ZipFile(_io.BytesIO(tpl_bytes)) as zin:
            all_files = {name: zin.read(name) for name in zin.namelist()}
        doc_xml = all_files["word/document.xml"].decode("utf-8")

        # Row duplication
        row_rx = _re.compile(r'(<w:tr[\s>](?:(?!</w:tr>).)*?\{\{GOODS_NAME\}\}.*?</w:tr>)', _re.DOTALL)
        m = row_rx.search(doc_xml)
        if m and rows_data:
            row_tpl = m.group(1)
            multi = ""
            for rd in rows_data:
                r = row_tpl
                r = r.replace("{{ROW_NUM}}",         str(rd["num"]))
                r = r.replace("{{GOODS_NAME}}",       _html_esc.escape(rd["name"]))
                r = r.replace("{{PRICE}}",            _html_esc.escape(rd["price_excl"]))
                r = r.replace("{{ROW_AMOUNT_EXCL}}",  _html_esc.escape(rd["amt_excl"]))
                r = r.replace("{{ROW_VAT_SUM}}",      _html_esc.escape(rd["vat_amt"]))
                r = r.replace("{{ROW_AMOUNT_INCL}}",  _html_esc.escape(rd["amt_incl"]))
                r = r.replace("{{ROW_QTY}}",          str(rd["qty"]))
                multi += r
            doc_xml = doc_xml.replace(row_tpl, multi, 1)

        for ph, val in [
            ("{{TTN_NUMBER}}",      supply_id_str),
            ("{{ORG_FULL}}",        org_line),
            ("{{SUPPLIER}}",        org_line),
            ("{{PAYER}}",           org_line),
            ("{{RECIPIENT}}",       recipient_line),
            ("{{ORDER_DATE}}",      supply_id_str),
            ("{{DOC_NUM_VAL}}",     supply_id_str),
            ("{{DOC_DATE_VAL}}",    supply_date_disp),
            ("{{GOODS_NAME}}",      rows_data[0]["name"] if rows_data else "Товар"),
            ("{{ROW_NUM}}",         "1"),
            ("{{PRICE}}",           rows_data[0]["price_excl"] if rows_data else "—"),
            ("{{ROW_AMOUNT_EXCL}}", rows_data[0]["amt_excl"] if rows_data else "—"),
            ("{{ROW_VAT_SUM}}",     rows_data[0]["vat_amt"] if rows_data else "—"),
            ("{{ROW_AMOUNT_INCL}}", rows_data[0]["amt_incl"] if rows_data else "—"),
            ("{{QTY}}",             str(qty_total)),
            ("{{QTY_SHT}}",         f"{qty_total} шт"),
            ("{{TOTAL_EXCL}}",      t_excl),
            ("{{TOTAL_VAT}}",       t_vat),
            ("{{TOTAL_INCL}}",      t_incl),
            ("{{AMOUNT}}",          t_excl),
            ("{{VAT_SUM}}",         t_vat),
            ("{{AMOUNT_WITH_VAT}}", t_incl),
            ("{{AMOUNT_WORDS}}",    amt_words),
            ("{{PAGES_COUNT}}",     "1"),
            ("{{ITEMS_COUNT}}",     str(len(rows_data) or 1)),
            ("{{TOTAL_RUB}}",       str(int(total_incl)) if total_incl else "0"),
            ("{{TOTAL_KOP}}",       str(round((total_incl % 1) * 100)).zfill(2) if total_incl else "00"),
            ("{{SUPPLY_ID}}",       supply_id_str),
            ("{{DOC_DATE_FULL}}",   f"«{now.strftime('%d')}» {['января','февраля','марта','апреля','мая','июня','июля','августа','сентября','октября','ноября','декабря'][now.month-1]} {now.year}"),
            ("{{ISSUED_BY}}",       supplier_short or "—"),
            ("{{SIGNATORIES}}",      le.get("signatories") or supplier_short or "—"),
            ("{{PROD_HEAD}}",        next((p.get("head_name") or p.get("name") for p in repository.list_supply_productions(user_id=owner_id) if p.get("name") == str(item.get("production") or "")), "—")),
            ("{{SIGN_SUPPLIER}}",   supplier_short),
            ("{{SIGN_DRIVER}}",     driver_name),
        ]:
            doc_xml = doc_xml.replace(ph, val)
        doc_xml = doc_xml.replace("{{ROW_QTY}}", str(qty_total))

        all_files["word/document.xml"] = doc_xml.encode("utf-8")

        # Write DOCX to temp file and convert to PDF via LibreOffice
        import subprocess as _sp, tempfile as _tf, pathlib as _pl
        tmp_dir   = _tf.mkdtemp()
        docx_path = _pl.Path(tmp_dir) / f"ttn_{supply_id}.docx"
        pdf_path  = _pl.Path(tmp_dir) / f"ttn_{supply_id}.pdf"

        buf = _io.BytesIO()
        with _zf.ZipFile(buf, "w", _zf.ZIP_DEFLATED) as zout:
            for name, data in all_files.items():
                zout.writestr(name, data)
        docx_path.write_bytes(buf.getvalue())

        # Try soffice / libreoffice
        import os as _os
        # LibreOffice requires writable XDG dirs — force all to tmp_dir
        lo_env = dict(_os.environ)
        lo_env["HOME"]            = tmp_dir
        lo_env["TMPDIR"]          = tmp_dir
        lo_env["XDG_CACHE_HOME"]  = tmp_dir
        lo_env["XDG_CONFIG_HOME"] = tmp_dir
        lo_env["XDG_RUNTIME_DIR"] = tmp_dir
        lo_env["DCONF_PROFILE"]   = "/dev/null"
        lo_env["UserInstallation"] = f"file://{tmp_dir}/lo_profile"

        _lo_binaries = (
            "/usr/bin/soffice",
            "/usr/lib/libreoffice/program/soffice",
            "/usr/local/bin/soffice",
            "soffice",
            "/usr/bin/libreoffice",
            "libreoffice",
        )
        lo_ok = False
        for binary in _lo_binaries:
            try:
                result = _sp.run(
                    [binary, "--headless", "--norestore",
                     f"-env:UserInstallation=file://{tmp_dir}/lo_profile",
                     "--convert-to", "pdf",
                     "--outdir", tmp_dir, str(docx_path)],
                    capture_output=True, timeout=60, env=lo_env
                )
                _log.info("soffice %s exit=%d stdout=%s stderr=%s",
                          binary, result.returncode,
                          result.stdout.decode()[:200], result.stderr.decode()[:200])
                if result.returncode == 0 and pdf_path.exists():
                    lo_ok = True
                    break
            except FileNotFoundError:
                continue
            except _sp.TimeoutExpired:
                raise HTTPException(status_code=504, detail="Таймаут конвертации PDF")

        if not lo_ok:
            raise HTTPException(status_code=500,
                detail=f"LibreOffice не смог конвертировать DOCX в PDF. Убедитесь что libreoffice-writer установлен")

        if not pdf_path.exists():
            raise HTTPException(status_code=500, detail="PDF не создан")

        pdf_bytes = pdf_path.read_bytes()
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'inline; filename="TTN_{supply_id}.pdf"'},
        )

    # ── WB FBS Orders (marketplace-api; isolated from FBW supplies) ──────────

    @app.get("/api/wb-fbs/sources")
    def list_wb_fbs_sources(request: Request) -> list[dict[str, object]]:
        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        sources = [
            s
            for s in repository.list_supply_sources(user_id=owner_id)
            if (s.get("marketplace") or "wb").lower() == "wb"
            and s.get("is_enabled")
            and wb_fbs_mod.is_fbs_source_name(s.get("name"))
        ]
        role = str(user.get("role") or ROLE_USER)
        if role not in ROLE_CAN_ACCESS_SETTINGS:
            perms = repository.get_manager_supply_permissions(manager_user_id=int(user["id"]))
            allowed = {
                str(sid)
                for sid, sv in (perms.get("sources") or {}).items()
                if isinstance(sv, dict) and sv.get("wb_fbs")
            }
            sources = [s for s in sources if str(s.get("id")) in allowed]
        return sources

    @app.get("/api/wb-fbs/tsd/sources")
    def list_wb_fbs_tsd_sources(request: Request) -> list[dict[str, object]]:
        """WB FBS sources allowed for ТСД (warehouse page)."""
        user = _require_user(request)
        _require_wb_fbs_tsd(user)
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        sources = [
            s
            for s in repository.list_supply_sources(user_id=owner_id)
            if (s.get("marketplace") or "wb").lower() == "wb"
            and s.get("is_enabled")
            and wb_fbs_mod.is_fbs_source_name(s.get("name"))
        ]
        allowed = _wb_fbs_tsd_allowed_source_ids(user)
        if allowed is not None:
            sources = [s for s in sources if str(s.get("id")) in allowed]
        return sources

    @app.get("/api/wb-fbs/tsd/supplies")
    def list_wb_fbs_tsd_supplies(
        request: Request,
        source_id: int | None = None,
        search: str | None = None,
        page: int = 1,
        page_size: int = 100,
    ) -> dict[str, object]:
        """Assembly supplies for ТСД list."""
        user = _require_user(request)
        _require_wb_fbs_tsd(user)
        if not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id")
        allowed = _wb_fbs_tsd_allowed_source_ids(user)
        if allowed is not None and str(int(source_id)) not in allowed:
            raise HTTPException(status_code=403, detail="Нет доступа к источнику")
        owner_id = _supply_owner_id(user)
        return wb_fbs_mod.list_assembly_supplies(
            repository,
            user_id=owner_id,
            source_id=int(source_id),
            search=search or None,
            page=page,
            page_size=page_size,
        )

    @app.get("/api/wb-fbs/tsd/supplies/{supply_id}/summary")
    def wb_fbs_tsd_supply_summary(
        request: Request,
        supply_id: str,
        source_id: int,
    ) -> dict[str, object]:
        """Hub card: name/QR/orders/warehouse + KIZ/pick progress (local DB)."""
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        _require_wb_fbs_tsd(user)
        allowed = _wb_fbs_tsd_allowed_source_ids(user)
        if allowed is not None and str(int(source_id)) not in allowed:
            raise HTTPException(status_code=403, detail="Нет доступа к источнику")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid:
            raise HTTPException(status_code=400, detail="Укажите supply_id")
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        # Prefer exact supply_id match (search is ILIKE and can paginate away).
        assembly = wb_fbs_mod.list_assembly_supplies(
            repository,
            user_id=owner_id,
            source_id=int(source_id),
            search=sid,
            page=1,
            page_size=200,
        )
        supply_row = next(
            (
                x
                for x in (assembly.get("items") or [])
                if str(x.get("supply_id") or "") == sid
            ),
            None,
        )
        if supply_row is None:
            # Fallback: scan first page of assembly without search filter.
            assembly_all = wb_fbs_mod.list_assembly_supplies(
                repository,
                user_id=owner_id,
                source_id=int(source_id),
                search=None,
                page=1,
                page_size=200,
            )
            supply_row = next(
                (
                    x
                    for x in (assembly_all.get("items") or [])
                    if str(x.get("supply_id") or "") == sid
                ),
                None,
            )
        # Hub counters must match KIZ/pick scan tiles. Use the same order set +
        # live orders/meta classification as the builders, but skip stickers /
        # Content (those load only when opening a scan mode).
        progress = wb_detail.build_tsd_hub_progress(
            repository,
            user_id=owner_id,
            source_id=int(source_id),
            api_key=api_key,
            supply_id=sid,
        )
        kiz = progress.get("kiz") if isinstance(progress, dict) else None
        pick = progress.get("pick") if isinstance(progress, dict) else None
        if not isinstance(kiz, dict):
            kiz = {"total": 0, "done": 0}
        if not isinstance(pick, dict):
            pick = {"total": 0, "done": 0}
        base = dict(supply_row or {})
        base.setdefault("supply_id", sid)
        base.setdefault("source_id", int(source_id))
        # Align «N заказов» with tile totals (same partition as scan modes).
        base["order_count"] = int(progress.get("order_count") or 0)
        base["kiz"] = {
            "total": int(kiz.get("total") or 0),
            "done": int(kiz.get("done") or 0),
        }
        base["pick"] = {
            "total": int(pick.get("total") or 0),
            "done": int(pick.get("done") or 0),
        }
        return base

    @app.get("/api/wb-fbs/tsd/supplies/{supply_id}/kiz/status")
    def wb_fbs_tsd_kiz_status(
        request: Request,
        supply_id: str,
        source_id: int,
    ) -> dict[str, object]:
        """Live КИЗ check for ТСД hub refresh (same rules as desktop Маркировка)."""
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        _require_wb_fbs_tsd(user)
        allowed = _wb_fbs_tsd_allowed_source_ids(user)
        if allowed is not None and str(int(source_id)) not in allowed:
            raise HTTPException(status_code=403, detail="Нет доступа к источнику")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid or not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и supply_id")
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            return wb_detail.check_supply_kiz_status(
                repository,
                user_id=owner_id,
                source_id=int(source_id),
                api_key=api_key,
                supply_id=sid,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.get("/api/wb-fbs/tsd/supplies/{supply_id}/kiz")
    def wb_fbs_tsd_kiz_list(
        request: Request,
        supply_id: str,
        source_id: int,
    ) -> dict[str, object]:
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        _require_wb_fbs_tsd(user)
        allowed = _wb_fbs_tsd_allowed_source_ids(user)
        if allowed is not None and str(int(source_id)) not in allowed:
            raise HTTPException(status_code=403, detail="Нет доступа к источнику")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid:
            raise HTTPException(status_code=400, detail="Укажите supply_id")
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            return wb_detail.build_kiz_marking_payload(
                repository,
                user_id=owner_id,
                source_id=int(source_id),
                api_key=api_key,
                supply_id=sid,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.put("/api/wb-fbs/tsd/supplies/{supply_id}/kiz")
    async def wb_fbs_tsd_kiz_save(
        request: Request,
        supply_id: str,
        source_id: int,
    ) -> dict[str, object]:
        """KIZ save for ТСД: scan autosave is local_only; «Сохранить» pushes to WB."""
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        _require_wb_fbs_tsd(user)
        allowed_sources = _wb_fbs_tsd_allowed_source_ids(user)
        if allowed_sources is not None and str(int(source_id)) not in allowed_sources:
            raise HTTPException(status_code=403, detail="Нет доступа к источнику")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid:
            raise HTTPException(status_code=400, detail="Укажите supply_id")
        try:
            body = await request.json()
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Некорректный JSON") from exc
        items = body.get("items") if isinstance(body, dict) else None
        if not isinstance(items, list):
            raise HTTPException(status_code=400, detail="Укажите items[]")
        # Preserve client local_only for scan autosave. Explicit Save omits it → WB.
        normalized_items: list[dict[str, object]] = []
        for it in items:
            if not isinstance(it, dict):
                continue
            row = dict(it)
            row["local_only"] = bool(row.get("local_only"))
            normalized_items.append(row)
        items = normalized_items
        if not items:
            raise HTTPException(status_code=400, detail="Укажите items[]")
        only_local = bool(items) and all(
            isinstance(it, dict) and bool(it.get("local_only")) for it in items
        )
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            # Scope to this supply's local orders (same as TSD autosave).
            # Avoid get_supply_detail here — stickers/Content would slow «Сохранить».
            allowed = set(
                wb_detail._local_order_ids_for_supply(
                    repository,
                    user_id=owner_id,
                    source_id=int(source_id),
                    supply_id=sid,
                )
            )
            result = wb_detail.save_kiz_marking(
                api_key=api_key,
                items=items,
                allowed_order_ids=allowed,
                repo=repository,
                user_id=owner_id,
                source_id=int(source_id),
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        if not only_local:
            wb_detail.invalidate_supply_detail_cache(
                user_id=owner_id, source_id=int(source_id), supply_id=sid
            )
        return result

    @app.get("/api/wb-fbs/tsd/supplies/{supply_id}/pick-verify")
    def wb_fbs_tsd_pick_verify_list(
        request: Request,
        supply_id: str,
        source_id: int,
    ) -> dict[str, object]:
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        _require_wb_fbs_tsd(user)
        allowed = _wb_fbs_tsd_allowed_source_ids(user)
        if allowed is not None and str(int(source_id)) not in allowed:
            raise HTTPException(status_code=403, detail="Нет доступа к источнику")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid:
            raise HTTPException(status_code=400, detail="Укажите supply_id")
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            return wb_detail.build_pick_verify_payload(
                repository,
                user_id=owner_id,
                source_id=int(source_id),
                api_key=api_key,
                supply_id=sid,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.put("/api/wb-fbs/tsd/supplies/{supply_id}/pick-verify")
    async def wb_fbs_tsd_pick_verify_save(
        request: Request,
        supply_id: str,
        source_id: int,
    ) -> dict[str, object]:
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        _require_wb_fbs_tsd(user)
        allowed_sources = _wb_fbs_tsd_allowed_source_ids(user)
        if allowed_sources is not None and str(int(source_id)) not in allowed_sources:
            raise HTTPException(status_code=403, detail="Нет доступа к источнику")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid:
            raise HTTPException(status_code=400, detail="Укажите supply_id")
        try:
            body = await request.json()
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Некорректный JSON") from exc
        items = body.get("items") if isinstance(body, dict) else None
        if not isinstance(items, list):
            raise HTTPException(status_code=400, detail="Укажите items[]")
        try:
            allowed = set(
                wb_detail._local_order_ids_for_supply(
                    repository,
                    user_id=owner_id,
                    source_id=int(source_id),
                    supply_id=sid,
                )
            )
            return wb_detail.save_pick_verify(
                repo=repository,
                user_id=owner_id,
                source_id=int(source_id),
                items=items,
                allowed_order_ids=allowed,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.get("/api/wb-fbs/orders")
    def list_wb_fbs_orders(
        request: Request,
        source_id: int | None = None,
        tab: str | None = None,
        search: str | None = None,
        page: int = 1,
        page_size: int = 50,
    ) -> dict[str, object]:
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        _require_wb_fbs_owner_tab(user, tab)
        owner_id = _supply_owner_id(user)
        payload = _sanitize_wb_fbs_owner_counts(
            user,
            wb_fbs_mod.list_orders(
                repository,
                user_id=owner_id,
                source_id=source_id,
                tab=tab or None,
                search=search or None,
                page=page,
                page_size=page_size,
            ),
        )
        # Enrich current page with QR sticker parts (same as supply detail modal).
        items = payload.get("items") if isinstance(payload, dict) else None
        if source_id and isinstance(items, list) and items:
            try:
                api_key = _wb_fbs_source_key(owner_id, int(source_id))
                client = wb_fbs_mod.WbFbsClient(api_key)
                wb_detail.attach_sticker_parts_to_orders(
                    client, items, api_key=api_key
                )
            except Exception as exc:
                _log.warning("wb-fbs orders sticker enrich: %s", exc)
        return payload

    @app.get("/api/wb-fbs/orders/ids")
    def list_wb_fbs_order_ids(
        request: Request,
        source_id: int | None = None,
        tab: str | None = None,
        search: str | None = None,
    ) -> dict[str, object]:
        """IDs for current filters — used by «select all matching» in the UI."""
        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        _require_wb_fbs_owner_tab(user, tab)
        owner_id = _supply_owner_id(user)
        return wb_fbs_mod.list_order_ids(
            repository,
            user_id=owner_id,
            source_id=source_id,
            tab=tab or None,
            search=search or None,
        )

    @app.get("/api/wb-fbs/orders/lookup")
    def lookup_wb_fbs_order(
        request: Request,
        source_id: int,
        order_id: int | None = None,
        search: str | None = None,
    ) -> dict[str, object]:
        """Find one order by number across all tabs; if missing locally, query WB API.

        Used by the toolbar search when the order is not in Новые / На сборке /
        В доставке (e.g. finished or cancelled — sync intentionally skips those).
        Local DB lookup must work even when the marketplace API key is missing.
        """
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        if not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id")
        oid = order_id
        if oid is None:
            oid = wb_fbs_mod.parse_order_id_query(search)
        if oid is None:
            raise HTTPException(
                status_code=400,
                detail="Укажите номер заказа (не менее 6 цифр)",
            )
        src_full = repository.get_supply_source_with_key(
            user_id=owner_id, source_id=int(source_id)
        )
        if not src_full:
            raise HTTPException(status_code=400, detail="Источник не найден")
        if not wb_fbs_mod.is_fbs_source_name(src_full.get("name")):
            raise HTTPException(status_code=400, detail="Источник не является ФБС")
        api_key = str(src_full.get("api_key") or "").strip()
        payload = wb_fbs_mod.lookup_order_by_id(
            repository,
            user_id=owner_id,
            source_id=int(source_id),
            order_id=int(oid),
            api_key=api_key or None,
            allow_remote=bool(api_key),
        )
        payload = _sanitize_wb_fbs_owner_counts(user, payload)
        item = payload.get("item") if isinstance(payload, dict) else None
        # Stickers need marketplace API — skip quietly when key is absent.
        if api_key and isinstance(item, dict):
            try:
                client = wb_fbs_mod.WbFbsClient(api_key)
                wb_detail.attach_sticker_parts_to_orders(
                    client, [item], api_key=api_key
                )
            except Exception as exc:
                _log.warning("wb-fbs order lookup sticker enrich: %s", exc)
        return payload

    @app.get("/api/wb-fbs/supplies")
    def list_wb_fbs_supplies(
        request: Request,
        source_id: int | None = None,
        only_open: bool = False,
        tab: str | None = None,
        search: str | None = None,
        page: int = 1,
        page_size: int = 50,
    ) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        _require_wb_fbs_owner_tab(user, tab)
        owner_id = _supply_owner_id(user)
        # «На сборке» / «В доставке» — rows are supplies (поставки), not orders.
        # Serve from DB only (no live WB enrich) so tab switches stay fast;
        # boxes/status are filled during sync.
        tab_key = (tab or "").strip().lower()
        if tab_key == "delivery":
            return _sanitize_wb_fbs_owner_counts(
                user,
                wb_fbs_mod.list_delivery_supplies(
                    repository,
                    user_id=owner_id,
                    source_id=source_id,
                    search=search or None,
                    page=page,
                    page_size=page_size,
                ),
            )
        if tab_key == "assembly":
            return _sanitize_wb_fbs_owner_counts(
                user,
                wb_fbs_mod.list_assembly_supplies(
                    repository,
                    user_id=owner_id,
                    source_id=source_id,
                    search=search or None,
                    page=page,
                    page_size=page_size,
                ),
            )
        items = wb_fbs_mod.list_supplies(
            repository,
            user_id=owner_id,
            source_id=source_id,
            only_open=only_open,
        )
        return {"items": items, "total": len(items)}

    def _wb_fbs_source_key(owner_id: int, source_id: int) -> str:
        src_full = repository.get_supply_source_with_key(user_id=owner_id, source_id=source_id)
        if not src_full or not src_full.get("api_key"):
            raise HTTPException(status_code=400, detail="Источник не найден")
        if not wb_fbs_mod.is_fbs_source_name(src_full.get("name")):
            raise HTTPException(status_code=400, detail="Источник не является ФБС")
        return str(src_full["api_key"])

    @app.get("/api/wb-fbs/supplies/{supply_id}/detail")
    def wb_fbs_supply_detail(
        request: Request,
        supply_id: str,
        source_id: int,
    ) -> dict[str, object]:
        """Portal-like supply card for «На сборке» modal."""
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid or not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и supply_id")
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            return wb_detail.get_supply_detail(
                repository,
                user_id=owner_id,
                source_id=int(source_id),
                api_key=api_key,
                supply_id=sid,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.get("/api/wb-fbs/supplies/{supply_id}/trbx")
    def wb_fbs_list_supply_trbx(
        request: Request,
        supply_id: str,
        source_id: int,
    ) -> dict[str, object]:
        """List cargo places from WB (always live) and refresh local cache."""
        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid or not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и supply_id")
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            return wb_fbs_mod.list_supply_trbx(
                repository,
                user_id=owner_id,
                source_id=int(source_id),
                api_key=api_key,
                supply_id=sid,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.get("/api/wb-fbs/supplies/{supply_id}/trbx/stickers-print")
    def wb_fbs_trbx_stickers_print(
        request: Request,
        supply_id: str,
        source_id: int,
        box_ids: str = "",
    ) -> Response:
        """HTML print page for cargo-place QR stickers (all or selected)."""
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid or not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и supply_id")
        ids = [x.strip() for x in str(box_ids or "").split(",") if x.strip()]
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            stickers = wb_fbs_mod.fetch_trbx_stickers(
                api_key=api_key,
                supply_id=sid,
                box_ids=ids or None,
                sticker_type="png",
            )
            html_doc = wb_detail.render_trbx_stickers_html(
                supply_id=sid, stickers=stickers
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return Response(content=html_doc, media_type="text/html; charset=utf-8")

    @app.post("/api/wb-fbs/supplies/{supply_id}/trbx")
    async def wb_fbs_create_supply_trbx(
        request: Request,
        supply_id: str,
    ) -> dict[str, object]:
        """Create cargo places (короба) on an open FBS supply."""
        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        payload = await request.json()
        try:
            source_id = int(payload.get("source_id") or 0)
        except (TypeError, ValueError):
            source_id = 0
        try:
            amount = int(payload.get("amount") or 0)
        except (TypeError, ValueError):
            amount = 0
        if not sid or not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и supply_id")
        if amount < 1:
            raise HTTPException(status_code=400, detail="Укажите количество коробов")
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            result = wb_fbs_mod.create_supply_trbx(
                repository,
                user_id=owner_id,
                source_id=int(source_id),
                api_key=api_key,
                supply_id=sid,
                amount=amount,
                fetch_stickers=False,
                sticker_type="png",
            )
            # Never return sticker binaries to the browser (print is a separate page).
            if isinstance(result, dict):
                result = dict(result)
                result["stickers"] = []
                result["stickers_error"] = ""
            return result
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.delete("/api/wb-fbs/supplies/{supply_id}/trbx")
    async def wb_fbs_delete_supply_trbx(
        request: Request,
        supply_id: str,
    ) -> dict[str, object]:
        """Delete cargo places from an open FBS supply and return refreshed list."""
        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        payload = await request.json()
        try:
            source_id = int(payload.get("source_id") or 0)
        except (TypeError, ValueError):
            source_id = 0
        box_ids = [str(x).strip() for x in (payload.get("box_ids") or []) if str(x).strip()]
        if not sid or not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и supply_id")
        if not box_ids:
            raise HTTPException(status_code=400, detail="Укажите ID грузомест")
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            return wb_fbs_mod.delete_supply_trbx(
                repository,
                user_id=owner_id,
                source_id=int(source_id),
                api_key=api_key,
                supply_id=sid,
                box_ids=box_ids,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.get("/api/wb-fbs/supplies/{supply_id}/picking-list.pdf")
    def wb_fbs_supply_picking_list_pdf(
        request: Request,
        supply_id: str,
        source_id: int,
        variant: str = "summary",
    ) -> Response:
        """A4 picking list PDF for direct print (WB has no public picking-list API).

        ``variant``: ``summary`` (default) or ``extended``.
        """
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid or not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и supply_id")
        pick_variant = str(variant or "summary").strip().lower()
        if pick_variant not in {"summary", "extended"}:
            pick_variant = "summary"
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            payload = wb_detail.build_article_groups_for_print(
                repository,
                user_id=owner_id,
                source_id=int(source_id),
                api_key=api_key,
                supply_id=sid,
                mode="picking_list",
            )
            pdf_bytes = wb_detail.render_picking_list_pdf(
                payload,
                repo=repository,
                user_id=owner_id,
                variant=pick_variant,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        safe_name = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in sid)[:80]
        suffix = "Extended" if pick_variant == "extended" else "Summary"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": (
                    f'inline; filename="PickingList_{suffix}_{safe_name}.pdf"'
                )
            },
        )

    @app.get("/api/wb-fbs/supplies/{supply_id}/picking-list")
    def wb_fbs_supply_picking_list(
        request: Request,
        supply_id: str,
        source_id: int,
        variant: str = "summary",
    ) -> Response:
        """A4 picking list as HTML for browser print (CSS layout stays intact).

        ``variant``: ``summary`` (default compact sheet) or ``extended`` (detailed).
        """
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid or not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и supply_id")
        pick_variant = str(variant or "summary").strip().lower()
        if pick_variant not in {"summary", "extended"}:
            pick_variant = "summary"
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            payload = wb_detail.build_article_groups_for_print(
                repository,
                user_id=owner_id,
                source_id=int(source_id),
                api_key=api_key,
                supply_id=sid,
                mode="picking_list",
            )
            html_doc = wb_detail.render_picking_list_html(
                payload, for_pdf=False, variant=pick_variant
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return Response(
            content=html_doc,
            media_type="text/html; charset=utf-8",
            headers={
                "Cache-Control": "no-store, no-cache, must-revalidate",
                "Pragma": "no-cache",
                "X-Feedpilot-Build": "picking-20260807c",
            },
        )

    @app.get("/api/wb-fbs/supplies/{supply_id}/sticker-groups")
    def wb_fbs_supply_sticker_groups(
        request: Request,
        supply_id: str,
        source_id: int,
    ) -> dict[str, object]:
        """Grouped products for «Печать стикеров по категориям» modal."""
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid or not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и supply_id")
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            return wb_detail.list_sticker_print_groups(
                repository,
                user_id=owner_id,
                source_id=int(source_id),
                api_key=api_key,
                supply_id=sid,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.get("/api/wb-fbs/supplies/{supply_id}/stickers-print")
    def wb_fbs_supply_stickers_print(
        request: Request,
        supply_id: str,
        source_id: int,
        order_ids: str = "",
    ) -> Response:
        """58×40 thermal stickers HTML: article separators + WB stickers.

        Optional ``order_ids`` (comma-separated) prints only selected orders.
        """
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid or not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и supply_id")
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        selected_ids: list[int] = []
        for part in str(order_ids or "").replace(";", ",").split(","):
            part = part.strip()
            if not part:
                continue
            try:
                selected_ids.append(int(part))
            except (TypeError, ValueError):
                continue
        try:
            payload = wb_detail.build_article_groups_for_print(
                repository,
                user_id=owner_id,
                source_id=int(source_id),
                api_key=api_key,
                supply_id=sid,
                mode="stickers",
                order_ids_filter=selected_ids or None,
            )
            if selected_ids and not (payload.get("groups") or []):
                raise ValueError("Нет стикеров для выбранных товаров")
            html_doc = wb_detail.render_stickers_print_html(payload)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return Response(
            content=html_doc,
            media_type="text/html; charset=utf-8",
            headers={
                "Cache-Control": "no-store, no-cache, must-revalidate",
                "Pragma": "no-cache",
                "X-Feedpilot-Build": "picking-20260809a",
            },
        )

    @app.get("/api/wb-fbs/supplies/{supply_id}/kiz")
    def wb_fbs_supply_kiz_list(
        request: Request,
        supply_id: str,
        source_id: int,
    ) -> dict[str, object]:
        """Orders in supply that need КИЗ (sgtin), with sticker numbers and codes."""
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid or not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и supply_id")
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            return wb_detail.build_kiz_marking_payload(
                repository,
                user_id=owner_id,
                source_id=int(source_id),
                api_key=api_key,
                supply_id=sid,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.get("/api/wb-fbs/supplies/{supply_id}/kiz/status")
    def wb_fbs_supply_kiz_status(
        request: Request,
        supply_id: str,
        source_id: int,
    ) -> dict[str, object]:
        """Live КИЗ check (metaDetails) without opening the marking modal."""
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid or not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и supply_id")
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            return wb_detail.check_supply_kiz_status(
                repository,
                user_id=owner_id,
                source_id=int(source_id),
                api_key=api_key,
                supply_id=sid,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.get("/api/wb-fbs/supplies/{supply_id}/cancelled")
    def wb_fbs_supply_cancelled_orders(
        request: Request,
        supply_id: str,
        source_id: int,
    ) -> dict[str, object]:
        """Live check for cancelled orders still present in the supply."""
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid or not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и supply_id")
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            return wb_detail.list_supply_cancelled_orders(
                repository,
                user_id=owner_id,
                source_id=int(source_id),
                api_key=api_key,
                supply_id=sid,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.put("/api/wb-fbs/supplies/{supply_id}/kiz")
    async def wb_fbs_supply_kiz_save(
        request: Request,
        supply_id: str,
        source_id: int,
    ) -> dict[str, object]:
        """Save КИЗ codes to WB API and mirror into local wb_fbs_orders."""
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid or not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и supply_id")
        try:
            body = await request.json()
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Некорректный JSON") from exc
        items = body.get("items") if isinstance(body, dict) else None
        if not isinstance(items, list):
            raise HTTPException(status_code=400, detail="Укажите items[]")
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            only_local = bool(items) and all(
                isinstance(it, dict) and bool(it.get("local_only")) for it in items
            )
            if only_local:
                # Silent autosave after scan: scope from local DB only.
                # Never call get_supply_detail here — that hits Marketplace + stickers
                # and would lag / rate-limit operators during rapid scanning.
                allowed = set(
                    wb_detail._local_order_ids_for_supply(
                        repository,
                        user_id=owner_id,
                        source_id=int(source_id),
                        supply_id=sid,
                    )
                )
            else:
                # Full Save: scope to this supply's kiz-required orders.
                detail = wb_detail.get_supply_detail(
                    repository,
                    user_id=owner_id,
                    source_id=int(source_id),
                    api_key=api_key,
                    supply_id=sid,
                )
                allowed = {
                    int(o["order_id"])
                    for o in (detail.get("orders") or [])
                    if o.get("kiz_required") and o.get("order_id") is not None
                }
            result = wb_detail.save_kiz_marking(
                api_key=api_key,
                items=items,
                allowed_order_ids=allowed,
                repo=repository,
                user_id=owner_id,
                source_id=int(source_id),
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        # Local-only autosave does not change WB-facing badges enough to bust cache
        # on every scan (multi-op / supply-detail thrash). Full Save still invalidates.
        if not only_local:
            wb_detail.invalidate_supply_detail_cache(
                user_id=owner_id, source_id=int(source_id), supply_id=sid
            )
        return result

    @app.get("/api/wb-fbs/supplies/{supply_id}/pick-verify")
    def wb_fbs_supply_pick_verify_list(
        request: Request,
        supply_id: str,
        source_id: int,
    ) -> dict[str, object]:
        """Orders without КИЗ: local EAN-13 pick-check payload (owner only)."""
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        if not _is_wb_fbs_tenant_owner(user):
            raise HTTPException(
                status_code=403,
                detail="Проверка ШК доступна только главному пользователю",
            )
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid or not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и supply_id")
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            return wb_detail.build_pick_verify_payload(
                repository,
                user_id=owner_id,
                source_id=int(source_id),
                api_key=api_key,
                supply_id=sid,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.put("/api/wb-fbs/supplies/{supply_id}/pick-verify")
    async def wb_fbs_supply_pick_verify_save(
        request: Request,
        supply_id: str,
        source_id: int,
    ) -> dict[str, object]:
        """Save local ШК pick-check (no Wildberries API). Owner only."""
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        if not _is_wb_fbs_tenant_owner(user):
            raise HTTPException(
                status_code=403,
                detail="Проверка ШК доступна только главному пользователю",
            )
        owner_id = _supply_owner_id(user)
        sid = str(supply_id or "").strip()
        if not sid or not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и supply_id")
        try:
            body = await request.json()
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Некорректный JSON") from exc
        items = body.get("items") if isinstance(body, dict) else None
        if not isinstance(items, list):
            raise HTTPException(status_code=400, detail="Укажите items[]")
        try:
            # Local-only endpoint — never call get_supply_detail here.
            # That hits Marketplace + stickers and would lag / rate-limit on every
            # silent autosave scan. Scope from local supply order ids.
            allowed = set(
                wb_detail._local_order_ids_for_supply(
                    repository,
                    user_id=owner_id,
                    source_id=int(source_id),
                    supply_id=sid,
                )
            )
            return wb_detail.save_pick_verify(
                repo=repository,
                user_id=owner_id,
                source_id=int(source_id),
                items=items,
                allowed_order_ids=allowed,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except RuntimeError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.get("/api/wb-fbs/orders/{order_id}/sticker-print")
    def wb_fbs_order_sticker_print(
        request: Request,
        order_id: int,
        source_id: int,
    ) -> Response:
        """Single order sticker 58×40 HTML for print."""
        from . import wb_fbs_detail as wb_detail

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        if not source_id or not order_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и order_id")
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        client = wb_fbs_mod.WbFbsClient(api_key)
        try:
            stickers = client.get_order_stickers(
                [int(order_id)], sticker_type="png", width=58, height=40
            )
        except Exception as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        file_b64 = ""
        for s in stickers:
            if isinstance(s, dict) and str(s.get("file") or "").strip():
                file_b64 = str(s.get("file") or "").strip()
                break
        try:
            html_doc = wb_detail.render_single_sticker_html(
                order_id=int(order_id), file_b64=file_b64
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return Response(content=html_doc, media_type="text/html; charset=utf-8")

    @app.get("/api/wb-fbs/sync/status")
    def get_wb_fbs_sync_status(request: Request) -> dict[str, object]:
        _require_user(request)
        return wb_fbs_mod.get_sync_state()

    @app.post("/api/wb-fbs/sync/stop")
    def stop_wb_fbs_sync(request: Request) -> dict[str, object]:
        _require_user(request)
        if wb_fbs_mod.request_sync_stop():
            return {"ok": True, "message": "Остановка синхронизации ВБ ФБС…"}
        return {"ok": False, "message": "Синхронизация не запущена"}

    @app.get("/api/wb-fbs/auto-sync-settings")
    def get_wb_fbs_auto_sync_settings(request: Request) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        if not _is_wb_fbs_tenant_owner(user):
            raise HTTPException(
                status_code=403,
                detail="Настройки автоматики доступны только главному пользователю",
            )
        owner_id = _supply_owner_id(user)
        try:
            settings = repository.get_wb_fbs_auto_sync_settings(user_id=owner_id)
        except RuntimeError as exc:
            raise HTTPException(status_code=404, detail=str(exc)) from exc
        settings["can_edit"] = True
        return settings

    @app.put("/api/wb-fbs/auto-sync-settings")
    def update_wb_fbs_auto_sync_settings(
        request: Request, payload: WbFbsAutoSyncSettingsRequest
    ) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        if not _is_wb_fbs_tenant_owner(user):
            raise HTTPException(
                status_code=403,
                detail="Настройки автоматики доступны только главному пользователю",
            )
        owner_id = _supply_owner_id(user)
        try:
            interval_minutes = payload.validated_interval_minutes()
            collect_interval_minutes = payload.validated_collect_interval_minutes()
            updated = repository.save_wb_fbs_auto_sync_settings(
                user_id=owner_id,
                enabled=bool(payload.enabled),
                interval_minutes=interval_minutes,
                lookback_days=int(payload.lookback_days),
                active_from=str(payload.active_from or "12:00"),
                active_to=str(payload.active_to or "06:00"),
                collect_mgt_enabled=bool(payload.collect_mgt_enabled),
                collect_mgt_interval_minutes=collect_interval_minutes,
                collect_mgt_active_from=str(payload.collect_mgt_active_from or "12:00"),
                collect_mgt_active_to=str(payload.collect_mgt_active_to or "06:00"),
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        if not updated:
            raise HTTPException(status_code=404, detail="Пользователь не найден")
        settings = repository.get_wb_fbs_auto_sync_settings(user_id=owner_id)
        settings["can_edit"] = True
        return {"ok": True, "settings": settings}

    # ── WB FBS → Честный знак: вывод / возврат КИЗ (new block) ─────────────

    @app.get("/api/supply-chz-settings")
    def get_supply_chz_settings(request: Request) -> dict[str, object]:
        from . import wb_kiz_circulation as kiz_circ

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        return kiz_circ.get_chz_settings(repository, user_id=_supply_owner_id(user))

    @app.put("/api/supply-chz-settings")
    def put_supply_chz_settings(
        request: Request, payload: UpsertSupplyChzSettingsRequest
    ) -> dict[str, object]:
        from . import wb_kiz_circulation as kiz_circ

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        if not _is_wb_fbs_tenant_owner(user):
            raise HTTPException(
                status_code=403,
                detail="Настройки ЧЗ доступны только главному пользователю",
            )
        try:
            return kiz_circ.upsert_chz_settings(
                repository,
                user_id=_supply_owner_id(user),
                is_enabled=bool(payload.is_enabled),
                participant_inn=str(payload.participant_inn or ""),
                product_group=str(payload.product_group or ""),
                api_base=payload.api_base,
                kpp=payload.kpp,
                fias_id=payload.fias_id,
                return_type=payload.return_type,
                cert_thumbprint=payload.cert_thumbprint,
                wb_analytics_api_key=payload.wb_analytics_api_key,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.post("/api/supply-chz-settings/auth-challenge")
    def supply_chz_auth_challenge(request: Request) -> dict[str, object]:
        """Return True API auth challenge (uuid + data) for browser УКЭП signing."""
        from . import wb_kiz_circulation as kiz_circ

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        settings = kiz_circ.get_chz_settings(repository, user_id=_supply_owner_id(user))
        client = kiz_circ.chz_client_from_settings(settings)
        try:
            challenge = client.auth_key()
        except kiz_circ.ChzTrueApiError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return {
            "ok": True,
            "uuid": challenge["uuid"],
            "data": challenge["data"],
            "inn": settings.get("participant_inn") or "",
            "cert_thumbprint": settings.get("cert_thumbprint") or "",
            "api_base": settings.get("api_base") or "prod",
        }

    @app.post("/api/supply-chz-settings/auth-complete")
    def supply_chz_auth_complete(
        request: Request, payload: WbKizChzAuthRequest
    ) -> dict[str, object]:
        from . import wb_kiz_circulation as kiz_circ

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        settings = kiz_circ.get_chz_settings(repository, user_id=_supply_owner_id(user))
        client = kiz_circ.chz_client_from_settings(settings)
        inn = str(payload.inn or settings.get("participant_inn") or "").strip()
        try:
            token = client.simple_sign_in(
                uuid=str(payload.uuid or ""),
                signature_b64=str(payload.signature_base64 or ""),
                inn=inn,
            )
        except kiz_circ.ChzTrueApiError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return {"ok": True, "token": token}

    @app.get("/api/wb-fbs/kiz-circulation")
    def wb_fbs_kiz_circulation_overview(
        request: Request, source_id: int
    ) -> dict[str, object]:
        from . import wb_kiz_circulation as kiz_circ

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        _require_wb_fbs_kiz_owner(user)
        owner_id = _supply_owner_id(user)
        if not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id")
        _ = _wb_fbs_source_key(owner_id, int(source_id))
        return kiz_circ.get_overview(
            repository, user_id=owner_id, source_id=int(source_id)
        )

    @app.get("/api/wb-fbs/kiz-circulation/events")
    def wb_fbs_kiz_circulation_events(
        request: Request,
        source_id: int,
        status: str = "",
        operation_type: int = 0,
        limit: int = 200,
    ) -> dict[str, object]:
        from . import wb_kiz_circulation as kiz_circ

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        _require_wb_fbs_kiz_owner(user)
        owner_id = _supply_owner_id(user)
        if not source_id:
            raise HTTPException(status_code=400, detail="Укажите source_id")
        # List must stay fast: no archive hydrate, no FBO purge (purge runs in sync).
        # Heal stuck submitted statuses; join/refresh happens inside list_events.
        api_key = _wb_fbs_source_key(owner_id, int(source_id))
        try:
            healed = kiz_circ.heal_submitted_terminal_statuses(
                repository, user_id=owner_id, source_id=int(source_id)
            )
        except Exception:
            healed = {"healed": 0}
        try:
            items = kiz_circ.list_events(
                repository,
                user_id=owner_id,
                source_id=int(source_id),
                status=str(status or ""),
                operation_type=int(operation_type) if operation_type else None,
                limit=int(limit or 200),
                api_key=api_key,
                hydrate_orders=False,
                refresh_statuses=True,
            )
        except Exception as exc:
            raise HTTPException(
                status_code=400, detail=f"Ошибка списка КИЗ: {exc}"
            ) from exc
        return {
            "items": items,
            "total": len(items),
            "healed": int((healed or {}).get("healed") or 0),
            "not_fbs_skipped": 0,
            "not_fbs_purged": 0,
            "not_sold_skipped": 0,
            "not_return_skipped": 0,
            "fbs_requeued": 0,
        }

    @app.post("/api/wb-fbs/kiz-circulation/chz/reconcile")
    def wb_fbs_kiz_circulation_chz_reconcile(
        request: Request, payload: WbKizChzReconcileRequest
    ) -> dict[str, object]:
        """Poll CHZ for submitted docs and refresh local statuses."""
        from . import wb_kiz_circulation as kiz_circ

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        _require_wb_fbs_kiz_owner(user)
        owner_id = _supply_owner_id(user)
        sid = int(payload.source_id or 0)
        if sid <= 0:
            raise HTTPException(status_code=400, detail="Укажите source_id")
        _ = _wb_fbs_source_key(owner_id, sid)
        token = str(payload.token or "").strip()
        # Always heal local inconsistencies even without a CHZ token.
        healed = kiz_circ.heal_submitted_terminal_statuses(
            repository, user_id=owner_id, source_id=sid
        )
        if not token:
            return {
                "ok": True,
                "healed": healed,
                "docs_checked": 0,
                "accepted": 0,
                "failed": 0,
                "token_required": True,
            }
        settings = kiz_circ.get_chz_settings(repository, user_id=owner_id)
        client = kiz_circ.chz_client_from_settings(settings)
        client.set_token(token)
        try:
            recon = kiz_circ.reconcile_submitted_with_chz(
                repository, client, user_id=owner_id, source_id=sid
            )
        except Exception as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return {"ok": True, "healed": healed, **recon}

    @app.post("/api/wb-fbs/kiz-circulation/chz/cis-status")
    def wb_fbs_kiz_circulation_chz_cis_status(
        request: Request, payload: WbKizChzCisStatusRequest
    ) -> dict[str, object]:
        """Refresh True API CIS card status (в обороте / выведен) for table rows."""
        from . import wb_kiz_circulation as kiz_circ

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        _require_wb_fbs_kiz_owner(user)
        owner_id = _supply_owner_id(user)
        sid = int(payload.source_id or 0)
        if sid <= 0:
            raise HTTPException(status_code=400, detail="Укажите source_id")
        _ = _wb_fbs_source_key(owner_id, sid)
        token = str(payload.token or "").strip()
        if not token:
            raise HTTPException(
                status_code=400,
                detail="Нужен токен ЧЗ (подпишите УКЭП)",
            )
        settings = kiz_circ.get_chz_settings(repository, user_id=owner_id)
        client = kiz_circ.chz_client_from_settings(settings)
        client.set_token(token)
        keys = [
            str(k).strip()
            for k in (payload.event_keys or [])
            if str(k or "").strip()
        ]
        try:
            out = kiz_circ.refresh_cis_statuses(
                repository,
                client,
                user_id=owner_id,
                source_id=sid,
                event_keys=keys or None,
                product_group=str(settings.get("product_group") or ""),
                limit=max(len(keys), 1) if keys else 2000,
            )
        except Exception as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return {"ok": True, **out}

    @app.post("/api/wb-fbs/kiz-circulation/sync")
    def wb_fbs_kiz_circulation_sync(
        request: Request, payload: WbKizCirculationSyncRequest
    ) -> dict[str, object]:
        from . import wb_kiz_circulation as kiz_circ

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        _require_wb_fbs_kiz_owner(user)
        owner_id = _supply_owner_id(user)
        sid = int(payload.source_id or 0)
        if sid <= 0:
            raise HTTPException(status_code=400, detail="Укажите source_id")
        # Validate FBS source exists. Analytics enriches fiscal; Marketplace alone
        # is enough for Ежедневный вывод (sold / отказ / дефект with КИЗ).
        marketplace_key = _wb_fbs_source_key(owner_id, sid)
        api_key = kiz_circ.get_wb_analytics_api_key(repository, user_id=owner_id)
        if not api_key and not marketplace_key:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Нужен токен WB Marketplace FBS и/или «Аналитика» "
                    "(Поставки → источник / Настройки → ЧЗ)"
                ),
            )
        try:
            started = kiz_circ.create_excise_sync_run(
                repository,
                user_id=owner_id,
                source_id=sid,
                date_from=str(payload.date_from or ""),
                date_to=str(payload.date_to or ""),
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

        run_id = int(started.get("run_id") or 0)

        def _worker() -> None:
            try:
                result = kiz_circ.sync_excise_report(
                    repository,
                    user_id=owner_id,
                    source_id=sid,
                    api_key=api_key,
                    date_from=str(payload.date_from or ""),
                    date_to=str(payload.date_to or ""),
                    run_id=run_id,
                    marketplace_api_key=marketplace_key,
                )
                if result.get("cancelled"):
                    _log.info("kiz circulation sync cancelled run_id=%s", run_id)
            except Exception as exc:
                if isinstance(exc, kiz_circ.SyncCancelled):
                    _log.info("kiz circulation sync cancelled run_id=%s", run_id)
                    return
                _log.exception("kiz circulation async sync failed: %s", exc)
                # sync_excise_report already finishes the run with the full log;
                # do not overwrite progress with the short create-time log.
                try:
                    existing = kiz_circ.get_run(
                        repository, user_id=owner_id, run_id=run_id
                    )
                    st = str((existing or {}).get("status") or "")
                    if st in {"ok", "error", "cancelled"}:
                        return
                except Exception:
                    _log.exception("kiz circulation read run after failure failed")
                try:
                    kiz_circ._finish_run(
                        repository,
                        run_id=run_id,
                        status="error",
                        log=[str(started.get("log") or ""), f"Ошибка: {exc}"],
                        error_text=str(exc),
                    )
                except Exception:
                    _log.exception("kiz circulation finish error run failed")

        threading.Thread(
            target=_worker,
            name=f"kiz-excise-sync-{owner_id}-{sid}-{run_id}",
            daemon=True,
        ).start()
        return started

    @app.post("/api/wb-fbs/kiz-circulation/sync/cancel")
    def wb_fbs_kiz_circulation_sync_cancel(
        request: Request, payload: WbKizCirculationSyncCancelRequest
    ) -> dict[str, object]:
        """Stop an in-flight «Ежедневный вывод» for this source (or explicit run_id)."""
        from . import wb_kiz_circulation as kiz_circ

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        _require_wb_fbs_kiz_owner(user)
        owner_id = _supply_owner_id(user)
        rid = int(payload.run_id or 0)
        sid = int(payload.source_id or 0)
        if rid <= 0:
            if sid <= 0:
                raise HTTPException(
                    status_code=400, detail="Укажите source_id или run_id"
                )
            active = kiz_circ.find_active_excise_sync_run(
                repository, user_id=owner_id, source_id=sid
            )
            if not active:
                return {
                    "ok": True,
                    "already_finished": True,
                    "message": "Активной выгрузки нет",
                }
            rid = int(active.get("id") or 0)
        try:
            return kiz_circ.cancel_excise_sync_run(
                repository, user_id=owner_id, run_id=rid
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    @app.get("/api/wb-fbs/kiz-circulation/runs/{run_id}")
    def wb_fbs_kiz_circulation_run(
        request: Request, run_id: int
    ) -> dict[str, object]:
        from . import wb_kiz_circulation as kiz_circ

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        _require_wb_fbs_kiz_owner(user)
        run = kiz_circ.get_run(
            repository, user_id=_supply_owner_id(user), run_id=int(run_id)
        )
        if not run:
            raise HTTPException(status_code=404, detail="Прогон не найден")
        return run

    @app.post("/api/wb-fbs/kiz-circulation/chz/prepare")
    def wb_fbs_kiz_circulation_chz_prepare(
        request: Request,
        source_id: int | None = None,
        payload: WbKizChzPrepareRequest | None = None,
    ) -> dict[str, object]:
        from . import wb_kiz_circulation as kiz_circ

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        _require_wb_fbs_kiz_owner(user)
        owner_id = _supply_owner_id(user)
        body = payload or WbKizChzPrepareRequest()
        sid = int(source_id or body.source_id or 0)
        if sid <= 0:
            raise HTTPException(status_code=400, detail="Укажите source_id")
        # Marketplace key required to verify wbStatus=sold before CHZ withdraw.
        api_key = _wb_fbs_source_key(owner_id, sid)
        keys = [str(k).strip() for k in (body.event_keys or []) if str(k or "").strip()]
        _log.info(
            "CHZ prepare start user=%s source_id=%s keys=%s",
            owner_id,
            sid,
            len(keys),
        )
        try:
            result = kiz_circ.prepare_chz_batches(
                repository,
                user_id=owner_id,
                source_id=sid,
                event_keys=keys or None,
                api_key=api_key,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        except Exception as exc:
            _log.exception("CHZ prepare failed user=%s source_id=%s", owner_id, sid)
            raise HTTPException(
                status_code=500,
                detail=f"Ошибка prepare: {exc}",
            ) from exc
        docs_n = len(result.get("documents") or []) if isinstance(result, dict) else 0
        _log.info(
            "CHZ prepare done user=%s source_id=%s documents=%s",
            owner_id,
            sid,
            docs_n,
        )
        return result

    @app.post("/api/wb-fbs/kiz-circulation/chz/submit")
    def wb_fbs_kiz_circulation_chz_submit(
        request: Request, payload: WbKizChzSubmitRequest
    ) -> dict[str, object]:
        from . import wb_kiz_circulation as kiz_circ

        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        _require_wb_fbs_kiz_owner(user)
        owner_id = _supply_owner_id(user)
        sid = int(payload.source_id or 0)
        if sid <= 0:
            raise HTTPException(status_code=400, detail="Укажите source_id")
        token = str(payload.token or "").strip()
        if not token:
            raise HTTPException(status_code=400, detail="Нет токена ЧЗ")
        settings = kiz_circ.get_chz_settings(repository, user_id=owner_id)
        client = kiz_circ.chz_client_from_settings(settings)
        client.set_token(token)
        results: list[dict[str, object]] = []
        log_lines: list[str] = []
        # Reconcile in-flight submitted docs before accepting new ones.
        try:
            recon = kiz_circ.reconcile_submitted_with_chz(
                repository, client, user_id=owner_id, source_id=sid
            )
            if recon.get("docs_checked"):
                log_lines.append(
                    "сверка submitted: "
                    f"док={recon.get('docs_checked')}, "
                    f"принято={recon.get('accepted')}, "
                    f"отклонено={recon.get('failed')}"
                )
        except Exception as exc:
            log_lines.append(f"сверка submitted пропущена: {exc}")

        for doc in list(payload.documents or []):
            keys = [str(k) for k in (doc.event_keys or []) if str(k).strip()]
            title = f"{doc.doc_type} · {len(keys)} КИЗ"
            payload_b64 = str(getattr(doc, "product_document_b64", "") or "").strip()
            try:
                doc_id = client.create_document(
                    doc_type=str(doc.doc_type or ""),
                    product_group=str(
                        doc.product_group or settings.get("product_group") or ""
                    ),
                    product_document=dict(doc.product_document or {}),
                    product_document_b64=payload_b64,
                    signature_b64=str(doc.signature_base64 or ""),
                )
            except Exception as exc:
                # Only mark error when CHZ create itself failed (safe to retry).
                kiz_circ.mark_events_error(
                    repository,
                    user_id=owner_id,
                    source_id=sid,
                    event_keys=keys,
                    error_text=str(exc),
                )
                results.append(
                    {
                        "ok": False,
                        "doc_type": doc.doc_type,
                        "error": str(exc),
                        "event_count": len(keys),
                    }
                )
                log_lines.append(f"{title} → ошибка: {exc}")
                continue

            # Document already exists in CHZ — never flip back to error on local faults.
            chz_status = "submitted"
            local_status = "submitted"
            chz_err = ""
            try:
                kiz_circ.mark_events_submitted(
                    repository,
                    user_id=owner_id,
                    source_id=sid,
                    event_keys=keys,
                    chz_doc_id=doc_id,
                    doc_type=str(doc.doc_type or ""),
                    run_id=payload.run_id,
                )
                try:
                    import time as _time

                    info: dict = {}
                    for _attempt in range(4):
                        info = client.document_info(doc_id)
                        chz_err = kiz_circ.extract_chz_doc_errors(info)
                        chz_status = (
                            kiz_circ.extract_chz_doc_status(info) or ""
                        )
                        if not chz_status and chz_err:
                            chz_status = "CHECKED_NOT_OK"
                        elif not chz_status:
                            chz_status = "submitted"
                        if kiz_circ.classify_chz_doc_status(chz_status) != "submitted":
                            break
                        _time.sleep(1.2)
                    if not chz_err:
                        chz_err = kiz_circ.extract_chz_doc_errors(info)
                    local_status = kiz_circ.apply_chz_doc_status(
                        repository,
                        user_id=owner_id,
                        source_id=sid,
                        event_keys=keys,
                        chz_doc_id=doc_id,
                        chz_status=chz_status,
                        error_text=chz_err,
                    )
                    if local_status == "error":
                        _log.warning(
                            "CHZ submit doc %s → %s: %s",
                            doc_id,
                            chz_status,
                            (chz_err or chz_status)[:500],
                        )
                except Exception as poll_exc:
                    _log.warning(
                        "CHZ submit doc %s status poll failed: %s",
                        doc_id,
                        poll_exc,
                    )
            except Exception as exc:
                log_lines.append(
                    f"{title} → создан в ЧЗ {doc_id}, локальный статус: {exc}"
                )
            ok_doc = local_status != "error"
            results.append(
                {
                    "ok": ok_doc,
                    "doc_type": doc.doc_type,
                    "chz_doc_id": doc_id,
                    "chz_status": chz_status,
                    "local_status": local_status,
                    "event_count": len(keys),
                    "error": chz_err if not ok_doc else "",
                }
            )
            suffix = f" · {chz_err}" if (not ok_doc and chz_err) else ""
            log_lines.append(
                f"{title} → {doc_id} (ЧЗ: {chz_status}, локально: {local_status}){suffix}"
            )
        ok_n = sum(1 for r in results if r.get("ok"))
        return {
            "ok": ok_n == len(results) and bool(results),
            "submitted": ok_n,
            "failed": len(results) - ok_n,
            "results": results,
            "log": "\n".join(log_lines),
        }

    @app.post("/api/wb-fbs/sync")
    def sync_wb_fbs(request: Request, source_id: int | None = None) -> dict[str, object]:
        """Sync all visible FBS sources (name contains ФБС/FBS). source_id is ignored."""
        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        wb_fbs_mod.ensure_wb_fbs_tables(repository)
        # source_id kept for API compatibility but sync always covers the full FBS set.
        _ = source_id
        jobs = wb_fbs_mod.list_fbs_sync_jobs(repository, user_id=owner_id)
        role = str(user.get("role") or ROLE_USER)
        if role not in ROLE_CAN_ACCESS_SETTINGS:
            perms = repository.get_manager_supply_permissions(manager_user_id=int(user["id"]))
            allowed = {
                str(sid)
                for sid, sv in (perms.get("sources") or {}).items()
                if isinstance(sv, dict) and sv.get("wb_fbs")
            }
            jobs = [j for j in jobs if str(j.get("source_id")) in allowed]
        if not jobs:
            # Distinguish "no FBS sources" vs "sources without keys".
            named = [
                s
                for s in repository.list_supply_sources(user_id=owner_id)
                if (s.get("marketplace") or "wb").lower() == "wb"
                and s.get("is_enabled")
                and wb_fbs_mod.is_fbs_source_name(s.get("name"))
            ]
            if role not in ROLE_CAN_ACCESS_SETTINGS:
                perms = repository.get_manager_supply_permissions(manager_user_id=int(user["id"]))
                allowed = {
                    str(sid)
                    for sid, sv in (perms.get("sources") or {}).items()
                    if isinstance(sv, dict) and sv.get("wb_fbs")
                }
                named = [s for s in named if str(s.get("id")) in allowed]
            if not named:
                return {
                    "ok": False,
                    "message": "Нет источников с «ФБС» в названии. Добавьте источник в Поставки → Настройки → Источники.",
                }
            return {"ok": False, "message": wb_fbs_mod.SCOPE_ERROR_MESSAGE}
        ok, message = wb_fbs_mod.start_sync_thread(
            repo=repository,
            user_id=owner_id,
            sources=jobs,
        )
        return {
            "ok": ok,
            "message": message,
            "source_ids": [int(j["source_id"]) for j in jobs],
            "sources_count": len(jobs),
        }

    @app.get("/api/wb-fbs/collect-mgt/preview")
    def wb_fbs_collect_mgt_preview(request: Request, source_id: int) -> dict[str, object]:
        """Plan «Собрать все МГТ-заказы» for the current FBS source (New tab)."""
        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        sid = int(source_id or 0)
        if not sid:
            raise HTTPException(status_code=400, detail="Укажите source_id")
        _wb_fbs_source_key(owner_id, sid)  # validates FBS source exists
        return wb_fbs_mod.preview_collect_mgt(
            repository, user_id=owner_id, source_id=sid
        )

    @app.post("/api/wb-fbs/collect-mgt")
    async def wb_fbs_collect_mgt(request: Request) -> dict[str, object]:
        """Create/choose supplies and add all New-tab MGT orders for one FBS source."""
        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        try:
            payload = await request.json()
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Некорректный JSON") from exc
        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail="Некорректное тело запроса")
        sid = int(payload.get("source_id") or 0)
        if not sid:
            raise HTTPException(status_code=400, detail="Укажите source_id")
        api_key = _wb_fbs_source_key(owner_id, sid)
        decisions_raw = payload.get("decisions") or []
        if not isinstance(decisions_raw, list):
            raise HTTPException(status_code=400, detail="decisions должен быть списком")
        decisions: list[dict[str, object]] = [
            d for d in decisions_raw if isinstance(d, dict)
        ]
        result = wb_fbs_mod.execute_collect_mgt(
            repository,
            user_id=owner_id,
            source_id=sid,
            api_key=api_key,
            decisions=decisions,
        )
        # Manual collect is also a "last MGT collect" event (auto path updates separately).
        try:
            repository.mark_wb_fbs_collect_mgt_at(user_id=owner_id)
        except Exception:
            pass
        return result

    def _wb_fbs_parse_order_ids(payload: dict[str, object]) -> list[int]:
        raw = payload.get("order_ids") or []
        if not isinstance(raw, list):
            raise HTTPException(status_code=400, detail="order_ids должен быть списком")
        out: list[int] = []
        for x in raw:
            try:
                out.append(int(x))
            except (TypeError, ValueError):
                continue
        if not out:
            raise HTTPException(status_code=400, detail="Укажите order_ids")
        if len(out) > 5000:
            raise HTTPException(status_code=400, detail="Слишком много заказов (макс. 5000)")
        return out

    @app.post("/api/wb-fbs/selection/preview")
    async def wb_fbs_selection_preview(request: Request) -> dict[str, object]:
        """Validate selected New-tab orders; return name suggestion + compatible supplies."""
        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        try:
            payload = await request.json()
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Некорректный JSON") from exc
        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail="Некорректное тело запроса")
        sid = int(payload.get("source_id") or 0)
        if not sid:
            raise HTTPException(status_code=400, detail="Укажите source_id")
        _wb_fbs_source_key(owner_id, sid)
        order_ids = _wb_fbs_parse_order_ids(payload)
        return wb_fbs_mod.preview_selection_supply(
            repository, user_id=owner_id, source_id=sid, order_ids=order_ids
        )

    @app.post("/api/wb-fbs/selection/create-supply")
    async def wb_fbs_selection_create_supply(request: Request) -> dict[str, object]:
        """Create supply and add selected New-tab orders (WB POST /api/v3/supplies + PATCH orders)."""
        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        try:
            payload = await request.json()
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Некорректный JSON") from exc
        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail="Некорректное тело запроса")
        sid = int(payload.get("source_id") or 0)
        if not sid:
            raise HTTPException(status_code=400, detail="Укажите source_id")
        api_key = _wb_fbs_source_key(owner_id, sid)
        order_ids = _wb_fbs_parse_order_ids(payload)
        name = str(payload.get("name") or "").strip()
        return wb_fbs_mod.create_supply_from_selection(
            repository,
            user_id=owner_id,
            source_id=sid,
            api_key=api_key,
            order_ids=order_ids,
            name=name,
        )

    @app.post("/api/wb-fbs/selection/add-to-supply")
    async def wb_fbs_selection_add_to_supply(request: Request) -> dict[str, object]:
        """Add selected New-tab orders to an existing open supply."""
        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        try:
            payload = await request.json()
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Некорректный JSON") from exc
        if not isinstance(payload, dict):
            raise HTTPException(status_code=400, detail="Некорректное тело запроса")
        sid = int(payload.get("source_id") or 0)
        if not sid:
            raise HTTPException(status_code=400, detail="Укажите source_id")
        api_key = _wb_fbs_source_key(owner_id, sid)
        order_ids = _wb_fbs_parse_order_ids(payload)
        supply_id = str(payload.get("supply_id") or "").strip()
        if not supply_id:
            raise HTTPException(status_code=400, detail="Укажите supply_id")
        return wb_fbs_mod.add_selection_to_supply(
            repository,
            user_id=owner_id,
            source_id=sid,
            api_key=api_key,
            order_ids=order_ids,
            supply_id=supply_id,
        )

    @app.delete("/api/wb-fbs/orders")
    def clear_wb_fbs_orders(request: Request, source_id: int) -> dict[str, object]:
        """Owner/admin only. UI button removed; kept for emergency admin use."""
        user = _require_user(request)
        if str(user.get("role") or "") not in ROLE_CAN_ACCESS_SETTINGS:
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        sid = int(source_id)
        src = repository.get_supply_source_with_key(user_id=owner_id, source_id=sid)
        if not src or not wb_fbs_mod.is_fbs_source_name(src.get("name")):
            raise HTTPException(status_code=404, detail="Источник ФБС не найден")
        deleted = wb_fbs_mod.clear_source_data(repository, user_id=owner_id, source_id=sid)
        return {"ok": True, **deleted}

    @app.post("/api/wb-fbs/stickers/orders")
    async def wb_fbs_order_stickers(request: Request) -> Response:
        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        payload = await request.json()
        source_id = int(payload.get("source_id") or 0)
        order_ids = [int(x) for x in (payload.get("order_ids") or []) if str(x).strip().isdigit() or isinstance(x, int)]
        sticker_type = str(payload.get("type") or "png").strip().lower()
        if sticker_type not in {"png", "svg", "zplv", "zplh"}:
            sticker_type = "png"
        if not source_id or not order_ids:
            raise HTTPException(status_code=400, detail="Укажите source_id и order_ids")
        src_full = repository.get_supply_source_with_key(user_id=owner_id, source_id=source_id)
        if not src_full or not src_full.get("api_key"):
            raise HTTPException(status_code=400, detail="Источник не найден")
        client = wb_fbs_mod.WbFbsClient(str(src_full["api_key"]))
        try:
            stickers = client.get_order_stickers(order_ids[:100], sticker_type=sticker_type)
        except Exception as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return JSONResponse({"stickers": stickers})

    @app.get("/api/wb-fbs/stickers/supply/{supply_id}")
    def wb_fbs_supply_sticker(
        request: Request,
        supply_id: str,
        source_id: int,
        type: str = "png",
        disposition: str = "inline",
    ) -> Response:
        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        sticker_type = type if type in {"png", "svg", "zplv", "zplh"} else "png"
        disp = "attachment" if str(disposition or "").strip().lower() == "attachment" else "inline"
        src_full = repository.get_supply_source_with_key(user_id=owner_id, source_id=source_id)
        if not src_full or not src_full.get("api_key"):
            raise HTTPException(status_code=400, detail="Источник не найден")
        client = wb_fbs_mod.WbFbsClient(str(src_full["api_key"]))
        try:
            raw = client.get_supply_barcode(supply_id, sticker_type=sticker_type)
        except Exception as exc:
            text = str(exc)
            lower = text.lower()
            # WB: barcode only after PATCH .../deliver («передана в доставку»).
            if (
                "409" in lower
                or "deliver" in lower
                or "доставк" in lower
                or "not transfer" in lower
                or "не передан" in lower
            ):
                raise HTTPException(
                    status_code=409,
                    detail=(
                        "QR-код поставки доступен только после передачи "
                        "поставки в доставку"
                    ),
                ) from exc
            raise HTTPException(status_code=400, detail=text) from exc
        if sticker_type == "png":
            media = "image/png"
        elif sticker_type == "svg":
            media = "image/svg+xml"
        else:
            media = "application/octet-stream"
        return Response(
            content=raw,
            media_type=media,
            headers={"Content-Disposition": f'{disp}; filename="supply_{supply_id}.{sticker_type}"'},
        )

    @app.post("/api/wb-fbs/stickers/boxes")
    async def wb_fbs_box_stickers(request: Request) -> Response:
        user = _require_user(request)
        if not _can_view_wb_fbs(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        payload = await request.json()
        source_id = int(payload.get("source_id") or 0)
        supply_id = str(payload.get("supply_id") or "").strip()
        box_ids = [str(x) for x in (payload.get("box_ids") or []) if str(x).strip()]
        sticker_type = str(payload.get("type") or "png").strip().lower()
        if sticker_type not in {"png", "svg", "zplv", "zplh"}:
            sticker_type = "png"
        if not source_id or not supply_id:
            raise HTTPException(status_code=400, detail="Укажите source_id и supply_id")
        src_full = repository.get_supply_source_with_key(user_id=owner_id, source_id=source_id)
        if not src_full or not src_full.get("api_key"):
            raise HTTPException(status_code=400, detail="Источник не найден")
        client = wb_fbs_mod.WbFbsClient(str(src_full["api_key"]))
        # If box_ids omitted — load from local cache / API
        if not box_ids:
            supplies = wb_fbs_mod.list_supplies(
                repository, user_id=owner_id, source_id=source_id, only_open=False
            )
            for s in supplies:
                if str(s.get("supply_id")) == supply_id:
                    for b in s.get("boxes") or []:
                        bid = str((b.get("id") if isinstance(b, dict) else b) or "").strip()
                        if bid:
                            box_ids.append(bid)
                    break
            if not box_ids:
                try:
                    boxes = client.get_supply_boxes(supply_id)
                    for b in boxes:
                        bid = str((b.get("id") if isinstance(b, dict) else b) or "").strip()
                        if bid:
                            box_ids.append(bid)
                except Exception:
                    pass
        if not box_ids:
            raise HTTPException(status_code=404, detail="Короба не найдены")
        try:
            stickers = wb_fbs_mod.fetch_trbx_stickers(
                api_key=str(src_full["api_key"]),
                supply_id=supply_id,
                box_ids=box_ids,
                sticker_type=sticker_type,
            )
        except Exception as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
        return JSONResponse({"stickers": stickers, "box_ids": box_ids})

    # ── OZON Supplies endpoints (isolated from WB) ──────────────────────────

    _ozon_sync_state: dict[str, object] = {
        "in_progress": False,
        "synced": 0,
        "total": 0,
        "message": "",
        "errors": [],
        "processed_sources": [],
        "failed_sources": [],
        "cancel_requested": False,
    }
    _ozon_sync_lock = threading.Lock()

    @app.get("/api/ozon-supplies")
    def list_ozon_supplies(request: Request) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        items = repository.list_ozon_supply_items(user_id=owner_id)
        return {"items": items, "total": len(items)}

    @app.get("/api/ozon-supplies/{supply_order_id}/goods")
    def get_ozon_supply_goods(request: Request, supply_order_id: int) -> list[dict[str, object]]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        # Build name map from our product catalog — OZON: by SKU first
        name_map = repository.get_product_name_by_ozon_sku(user_id=owner_id)
        name_map_art = repository.get_product_name_by_article(user_id=owner_id)
        # Editable in Ozon LK — do not trust local composition cache for these states.
        _ozon_composition_refresh_states = frozenset({"DATA_FILLING", "READY_TO_SUPPLY"})

        def _apply_product_names(rows: list[dict[str, object]]) -> list[dict[str, object]]:
            for g in rows:
                sku_key = str(g.get("sku") or "").strip()
                offer_id = str(g.get("offer_id") or "").strip()
                g["product_name"] = (
                    name_map.get(sku_key) or name_map_art.get(offer_id) or offer_id or g.get("name") or ""
                )
            return rows

        item_row = repository.get_ozon_supply_item_row(user_id=owner_id, supply_order_id=supply_order_id)
        if not item_row:
            _log.warning("ozon goods: item_row not found for supply_order_id=%d user=%d", supply_order_id, owner_id)
            return []
        state = str(item_row.get("state") or "")
        cached = repository.get_ozon_supply_goods(user_id=owner_id, supply_order_id=supply_order_id)
        # Frozen supplies: serve DB cache. Editable: refresh from Ozon so LK edits show up.
        if cached and state not in _ozon_composition_refresh_states:
            return _apply_product_names(cached)

        # Lazy-load / refresh from OZON API
        import urllib.request as _ul, json as _jj, ssl as _sl
        src = repository.get_supply_source_with_key(user_id=owner_id, source_id=int(item_row["source_id"]))
        if not src or not src.get("api_key"):
            _log.warning("ozon goods: no api_key for source_id=%s", item_row.get("source_id"))
            return _apply_product_names(cached) if cached else []
        client_id = str(src.get("client_id") or "")
        api_key = str(src["api_key"])
        ctx = _sl.create_default_context()
        ozon_headers = {"Client-Id": client_id, "Api-Key": api_key,
                        "Content-Type": "application/json", "User-Agent": "Mozilla/5.0"}

        # Step 1: get bundle_ids from raw_json (fast path)
        raw = _jj.loads(item_row.get("raw_json") or "{}")
        supplies_raw = raw.get("supplies") or []
        bundle_ids = [s["bundle_id"] for s in supplies_raw if s.get("bundle_id")]

        # Step 2: if no bundle_ids in raw_json, re-fetch from v3 API (fallback)
        if not bundle_ids:
            _log.info("ozon goods: no bundle_ids in raw_json for supply_order_id=%d, fetching fresh", supply_order_id)
            try:
                req_det = _ul.Request("https://api-seller.ozon.ru/v3/supply-order/get",
                    data=_jj.dumps({"order_ids": [str(supply_order_id)]}).encode(),
                    method="POST", headers=ozon_headers)
                with _ul.urlopen(req_det, context=ctx, timeout=15) as r_det:
                    det_resp = _jj.loads(r_det.read())
                for order in (det_resp.get("orders") or []):
                    for s in (order.get("supplies") or []):
                        if s.get("bundle_id"):
                            bundle_ids.append(s["bundle_id"])
            except Exception as ex:
                _log.warning("ozon goods re-fetch order details: %s", ex)

        if not bundle_ids:
            _log.warning("ozon goods: still no bundle_ids for supply_order_id=%d", supply_order_id)
            return _apply_product_names(cached) if cached else []

        goods = []
        try:
            body = _jj.dumps({"bundle_ids": bundle_ids, "limit": 100, "last_id": ""}).encode()
            req = _ul.Request("https://api-seller.ozon.ru/v1/supply-order/bundle", data=body, method="POST",
                headers=ozon_headers)
            with _ul.urlopen(req, context=ctx, timeout=15) as r:
                resp = _jj.loads(r.read())
            goods = resp.get("items") or []
            if item_row.get("id"):
                repository.upsert_ozon_supply_goods(supply_item_id=int(item_row["id"]), goods=goods)
                total_qty = sum(int(g.get("quantity") or 0) for g in goods)
                repository.update_ozon_supply_total_quantity(
                    supply_order_id=supply_order_id, total_quantity=total_qty)
        except Exception as ex:
            _log.warning("ozon goods bundle call sid=%d: %s", supply_order_id, ex)
            if cached:
                return _apply_product_names(cached)
        result = repository.get_ozon_supply_goods(user_id=owner_id, supply_order_id=supply_order_id) or [
            {"sku": g.get("sku"), "name": g.get("name"), "quantity": g.get("quantity"),
             "barcode": g.get("barcode"), "offer_id": g.get("offer_id")} for g in goods
        ]
        return _apply_product_names(result)

    def _ozon_cargoes_api_payload(cached_raw: object) -> dict[str, object]:
        """Shape cargoes-info response from cached JSON (legacy list or v2)."""
        from .ozon_etrn import parse_cargoes_cache

        parsed = parse_cargoes_cache(cached_raw)
        return {
            "groups": parsed["groups"],
            "transport_cargoes": parsed["transport_cargoes"],
        }

    @app.get("/api/ozon-supplies/{supply_order_id}/cargoes-info")
    def get_ozon_supply_cargoes(request: Request, supply_order_id: int) -> dict[str, object]:
        """Fetch cargo places from Ozon (boxes + transport pallets) and cache."""
        import urllib.request as _ul, json as _jj, ssl as _sl
        from .ozon_etrn import build_ozon_cargoes_cache

        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        item_row = repository.get_ozon_supply_item_row(user_id=owner_id, supply_order_id=supply_order_id)
        if not item_row:
            return {"ok": False, "groups": [], "transport_cargoes": []}
        cached_raw = None
        cached = str(item_row.get("cargoes_json") or "")
        if cached and cached not in ("[]", "null"):
            try:
                cached_raw = _jj.loads(cached)
            except Exception:
                cached_raw = None
        src = repository.get_supply_source_with_key(user_id=owner_id, source_id=int(item_row["source_id"]))
        if not src or not src.get("api_key"):
            if cached_raw is not None:
                return {"ok": True, "cached": True, **_ozon_cargoes_api_payload(cached_raw)}
            return {"ok": False, "groups": [], "transport_cargoes": [], "error": "no_key"}
        client_id = str(src.get("client_id") or "")
        api_key = str(src["api_key"])
        ctx = _sl.create_default_context()
        ozon_headers = {"Client-Id": client_id, "Api-Key": api_key,
                        "Content-Type": "application/json", "User-Agent": "Mozilla/5.0"}
        try:
            # /v1/cargoes/* uses supply_id (from supplies[].supply_id in order),
            # NOT order_id. Extract from raw_json stored during sync.
            actual_supply_id = supply_order_id  # fallback
            raw_json_str = str(item_row.get("raw_json") or "")
            if raw_json_str:
                try:
                    raw = _jj.loads(raw_json_str)
                    supplies_raw = raw.get("supplies") or []
                    if supplies_raw:
                        sid = int(supplies_raw[0].get("supply_id") or 0)
                        if sid > 0:
                            actual_supply_id = sid
                except Exception:
                    pass
            body = _jj.dumps({"supply_ids": [actual_supply_id]}).encode()
            req = _ul.Request("https://api-seller.ozon.ru/v1/cargoes/get",
                data=body, method="POST", headers=ozon_headers)
            with _ul.urlopen(req, context=ctx, timeout=10) as r:
                data = _jj.loads(r.read())
            cargoes = []
            for s in (data.get("supply") or []):
                cargoes.extend(s.get("cargoes") or [])
            supplies_cargoes: list = []
            try:
                t_req = _ul.Request(
                    "https://api-seller.ozon.ru/v1/cargoes/supplies/get",
                    data=body, method="POST", headers=ozon_headers,
                )
                with _ul.urlopen(t_req, context=ctx, timeout=10) as tr:
                    t_data = _jj.loads(tr.read())
                supplies_cargoes = list(t_data.get("supplies_cargoes") or [])
            except Exception as tex:
                _log.debug("ozon transport cargoes sid=%d: %s", supply_order_id, tex)
            cache_obj = build_ozon_cargoes_cache(
                flat_cargoes=cargoes,
                supplies_cargoes=supplies_cargoes,
            )
            cargoes_json_str = _jj.dumps(cache_obj, ensure_ascii=False)
            repository.update_ozon_supply_cargoes(
                supply_order_id=supply_order_id, cargoes_json=cargoes_json_str)
            return {"ok": True, **_ozon_cargoes_api_payload(cache_obj)}
        except Exception as ex:
            code = getattr(ex, "code", None)
            if code == 403:
                payload = _ozon_cargoes_api_payload(cached_raw) if cached_raw is not None else {
                    "groups": [], "transport_cargoes": [],
                }
                return {"ok": False, "error": "no_role", "cached": bool(cached_raw), **payload}
            _log.warning("ozon cargoes fetch sid=%d: %s", supply_order_id, ex)
            if cached_raw is not None:
                return {
                    "ok": True,
                    "cached": True,
                    "error": str(ex)[:100],
                    **_ozon_cargoes_api_payload(cached_raw),
                }
            return {"ok": False, "groups": [], "transport_cargoes": [], "error": str(ex)[:100]}

    @app.get("/api/ozon-supplies/{supply_order_id}/vehicle")
    def get_ozon_supply_vehicle(request: Request, supply_order_id: int) -> dict[str, object]:
        """Fetch vehicle/driver info from OZON /v1/supply-order/details and cache it."""
        import urllib.request as _ul, json as _jj, ssl as _sl
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        item_row = repository.get_ozon_supply_item_row(user_id=owner_id, supply_order_id=supply_order_id)
        if not item_row:
            return {"ok": False, "vehicle": None}
        # Always try fresh OZON API; fall back to cache only on failure
        cached_vehicle = None
        cached = str(item_row.get("vehicle_json") or "")
        if cached and cached not in ("{}", "null"):
            try:
                cached_vehicle = _jj.loads(cached)
            except Exception:
                pass
        src = repository.get_supply_source_with_key(user_id=owner_id, source_id=int(item_row["source_id"]))
        if not src or not src.get("api_key"):
            if cached_vehicle is not None:
                return {"ok": True, "vehicle": cached_vehicle, "cached": True}
            return {"ok": False, "vehicle": None, "error": "no_key"}
        client_id = str(src.get("client_id") or "")
        api_key = str(src["api_key"])
        ctx = _sl.create_default_context()
        ozon_headers = {"Client-Id": client_id, "Api-Key": api_key,
                        "Content-Type": "application/json", "User-Agent": "Mozilla/5.0"}
        try:
            body = _jj.dumps({"order_id": supply_order_id}).encode()
            req = _ul.Request("https://api-seller.ozon.ru/v1/supply-order/details",
                data=body, method="POST", headers=ozon_headers)
            with _ul.urlopen(req, context=ctx, timeout=10) as r:
                data = _jj.loads(r.read())
            vehicle_obj = data.get("vehicle") or {}
            vehicle_val = vehicle_obj.get("value") or {}
            vehicle_json_str = _jj.dumps(vehicle_val, ensure_ascii=False)
            repository.update_ozon_supply_vehicle(
                supply_order_id=supply_order_id, vehicle_json=vehicle_json_str)
            return {"ok": True, "vehicle": vehicle_val}
        except Exception as ex:
            code = getattr(ex, "code", None)
            if code == 403:
                return {"ok": False, "vehicle": cached_vehicle, "error": "no_role",
                        "cached": bool(cached_vehicle)}
            _log.warning("ozon vehicle fetch sid=%d: %s", supply_order_id, ex)
            if cached_vehicle is not None:
                return {"ok": True, "vehicle": cached_vehicle, "cached": True, "error": str(ex)[:100]}
            return {"ok": False, "vehicle": None, "error": str(ex)[:100]}

    def _ozon_get_doc_data(owner_id: int, supply_order_id: int) -> dict:
        """Collect all data needed for OZON document generation."""
        import urllib.request as _ul, json as _jj, ssl as _sl
        item_row = repository.get_ozon_supply_item_row(user_id=owner_id, supply_order_id=supply_order_id)
        if not item_row:
            return {}
        src = repository.get_supply_source_with_key(user_id=owner_id, source_id=int(item_row["source_id"]))
        if not src or not src.get("api_key"):
            return {"item": item_row}
        api_key = str(src["api_key"])
        client_id = str(src.get("client_id") or "")
        ctx = _sl.create_default_context()
        hdrs = {"Client-Id": client_id, "Api-Key": api_key,
                "Content-Type": "application/json", "User-Agent": "Mozilla/5.0"}

        # Driver from supply-order/details
        driver_name, driver_docs, vehicle_num = "", "", ""
        cached_v = str(item_row.get("vehicle_json") or "")
        if cached_v and cached_v != "{}":
            try:
                v = _jj.loads(cached_v)
                driver_name = str(v.get("driver_name") or "")
                driver_docs = str(v.get("driver_phone") or "")
                vehicle_num = f"{v.get('vehicle_model','') or ''} {v.get('vehicle_number','') or ''}".strip()
            except Exception:
                pass
        if not driver_name:
            try:
                body = _jj.dumps({"order_id": supply_order_id}).encode()
                req = _ul.Request("https://api-seller.ozon.ru/v1/supply-order/details",
                    data=body, method="POST", headers=hdrs)
                with _ul.urlopen(req, context=ctx, timeout=10) as r:
                    det = _jj.loads(r.read())
                v = (det.get("vehicle") or {}).get("value") or {}
                driver_name = str(v.get("driver_name") or "")
                driver_docs = str(v.get("driver_phone") or "")
                vehicle_num = f"{v.get('vehicle_model','') or ''} {v.get('vehicle_number','') or ''}".strip()
                if v:
                    repository.update_ozon_supply_vehicle(
                        supply_order_id=supply_order_id,
                        vehicle_json=_jj.dumps(v, ensure_ascii=False))
            except Exception:
                pass

        # Goods from bundle (cached or fresh)
        goods = repository.get_ozon_supply_goods(user_id=owner_id, supply_order_id=supply_order_id)
        if not goods:
            raw_json_str = str(item_row.get("raw_json") or "")
            bundle_ids = []
            if raw_json_str:
                try:
                    raw = _jj.loads(raw_json_str)
                    for s in (raw.get("supplies") or []):
                        bid = str(s.get("bundle_id") or "")
                        if bid:
                            bundle_ids.append(bid)
                            break
                except Exception:
                    pass
            if bundle_ids:
                try:
                    body2 = _jj.dumps({"bundle_ids": bundle_ids, "limit": 100, "last_id": ""}).encode()
                    req2 = _ul.Request("https://api-seller.ozon.ru/v1/supply-order/bundle",
                        data=body2, method="POST", headers=hdrs)
                    with _ul.urlopen(req2, context=ctx, timeout=15) as r2:
                        bdata = _jj.loads(r2.read())
                    goods_raw = bdata.get("items") or []
                    if goods_raw and item_row.get("id"):
                        repository.upsert_ozon_supply_goods(
                            supply_item_id=int(item_row["id"]), goods=goods_raw)
                    goods = repository.get_ozon_supply_goods(
                        user_id=owner_id, supply_order_id=supply_order_id) or [
                        {"sku": g.get("sku"), "name": g.get("name"),
                         "quantity": g.get("quantity"), "offer_id": g.get("offer_id")} for g in goods_raw]
                except Exception:
                    pass

        # Prices by product_id (get product_ids first then prices)
        price_map: dict[str, float] = {}
        if goods:
            offer_ids = list({str(g.get("offer_id") or "") for g in goods if g.get("offer_id")})
            if offer_ids:
                try:
                    body3 = _jj.dumps({"filter": {"offer_id": offer_ids, "visibility": "ALL"},
                                       "last_id": "", "limit": len(offer_ids) + 10}).encode()
                    req3 = _ul.Request("https://api-seller.ozon.ru/v3/product/list",
                        data=body3, method="POST", headers=hdrs)
                    with _ul.urlopen(req3, context=ctx, timeout=10) as r3:
                        pl = _jj.loads(r3.read())
                    pid_map = {i["offer_id"]: i["product_id"] for i in (pl.get("result") or {}).get("items", [])}
                    if pid_map:
                        body4 = _jj.dumps({"filter": {"product_id": list(pid_map.values()),
                                                       "visibility": "ALL"}, "last_id": "", "limit": 200}).encode()
                        req4 = _ul.Request("https://api-seller.ozon.ru/v5/product/info/prices",
                            data=body4, method="POST", headers=hdrs)
                        with _ul.urlopen(req4, context=ctx, timeout=10) as r4:
                            price_data = _jj.loads(r4.read())
                        for pi in (price_data.get("items") or []):
                            oid = str(pi.get("offer_id") or "")
                            price_val = float(pi.get("price", {}).get("price") or 0)
                            if oid and price_val:
                                price_map[oid] = price_val
                except Exception:
                    pass

        # Legal entity — same matching as eTrN/Заявка XML (prefer entry with phone).
        from .ozon_etrn import _find_legal_entity as _find_le
        entities = repository.list_supply_legal_entities(user_id=owner_id)
        le = _find_le(entities, str(item_row.get("supplier_name") or ""))

        # Product name map — OZON uses SKU for lookup
        name_map = repository.get_product_name_by_ozon_sku(user_id=owner_id)

        return {
            "item": item_row,
            "owner_id": owner_id,
            "driver_name": driver_name,
            "driver_docs": driver_docs,
            "vehicle_num": vehicle_num,
            "goods": goods,
            "price_map": price_map,
            "le": le,
            "name_map": name_map,
        }

    def _build_ozon_poa_html(data: dict, include_signature: bool = True) -> str:
        """Build PoA M-2 HTML for OZON supply (identical structure to WB PoA)."""
        import html as _hm
        from datetime import datetime as _dtt
        e = _hm.escape
        item = data["item"]
        driver_name = str(data.get("driver_name") or "")
        driver_docs  = str(data.get("driver_docs") or "")
        le = data.get("le") or {}
        goods = data.get("goods") or []
        name_map = data.get("name_map") or {}
        now = _dtt.now()
        date_display = now.strftime("%d.%m.%Y")
        supply_num = str(item.get("supply_order_number") or "")
        wh = str(item.get("warehouse_name") or "")
        org_full = str(le.get("full_name") or le.get("short_name") or "")
        org_req  = str(le.get("requisites") or "")
        org_line = ", ".join(filter(None, [org_full, org_req]))
        supplier_short = str(le.get("short_name") or "")
        signatories = str(le.get("signatories") or supplier_short or "")
        UL = "_" * 30

        goods_rows_html = ""
        for i, g in enumerate(goods):
            sku_key = str(g.get("sku") or "")
            offer_id = str(g.get("offer_id") or "")
            name = name_map.get(sku_key) or name_map.get(offer_id) or offer_id or str(g.get("name") or "Товар")
            qty  = g.get("quantity") or "—"
            goods_rows_html += (f"<tr>"
                                f"<td style='border:1px solid #000;padding:0 2pt;text-align:center;white-space:nowrap;line-height:1.1'>{i+1}</td>"
                                f"<td class='mat-name' style='border:1px solid #000;padding:0 2pt;text-align:left;white-space:normal;line-height:1.1'>{e(name)}</td>"
                                f"<td style='border:1px solid #000;padding:0 2pt;text-align:center;white-space:nowrap;line-height:1.1'>шт.</td>"
                                f"<td style='border:1px solid #000;padding:0 2pt;text-align:center;white-space:nowrap;line-height:1.1'>{qty}</td>"
                                f"</tr>")

        owner_id = data.get("owner_id") or 0
        sig_block = "&nbsp;" * 20
        if include_signature:
            sig_data = None
            try:
                sig_data = repository.get_legal_entity_signature(
                    user_id=owner_id, entity_id=int(le.get("id") or 0))
            except Exception:
                pass
            if sig_data:
                sig_block = f"<img src='data:image/png;base64,{sig_data}' style='max-height:25mm;max-width:60mm;object-fit:contain;vertical-align:middle' />"

        docs_line = f"<p>{e(driver_docs)}</p>" if driver_docs else ""

        return f"""<!DOCTYPE html>
<html xmlns:o="urn:schemas-microsoft-com:office:office"
      xmlns:w="urn:schemas-microsoft-com:office:word"
      xmlns="http://www.w3.org/TR/REC-html40">
<head><meta charset="utf-8">
<!--[if gte mso 9]><xml><w:WordDocument><w:View>Print</w:View></w:WordDocument></xml><![endif]-->
<style>
  @page {{ size: 210mm 297mm; margin: 15mm 10mm 15mm 25mm; }}
  @page Section1 {{
    size: 210.0mm 297.0mm;
    margin: 15.0mm 10.0mm 15.0mm 25.0mm;
    mso-header-margin: 0mm;
    mso-footer-margin: 0mm;
    mso-paper-source: 0;
  }}
  div.Section1 {{ page: Section1; }}
  body {{ font-family: "Times New Roman", serif; font-size: 9pt; line-height: 1.05; }}
  .small {{ font-size: 8pt; text-align: center; }}
  .underline {{ text-decoration: underline; }}
  .center {{ text-align: center; }}
  .right {{ text-align: right; }}
  .bold {{ font-weight: bold; }}
  table.outer {{ width: 100%; border-collapse: collapse; margin-bottom: 6pt; }}
  table.codes {{ border-collapse: collapse; margin-left: auto; font-size: 9pt; border: 1px solid #000; }}
  table.codes td {{ border: 1px solid #000; padding: 2pt 6pt; }}
  table.mat {{ width: 100%; border-collapse: collapse; margin-top: 4pt; font-size: 9pt; border: 1px solid #000; }}
  table.mat td, table.mat th {{ border: 1px solid #000; padding: 0pt 2pt; text-align: center; white-space: nowrap; line-height: 1.1; }}
  table.mat td.mat-name, table.mat th.mat-name {{ text-align: left; white-space: normal; }}
  .dotline {{ display: inline-block; border-bottom: 1px solid #000; min-width: 120pt; }}
  p {{ margin: 0; padding: 0; }}
  p {{ margin: 3pt 0; }}
</style>
</head>
<body><div class="Section1">

<table class="outer">
  <tr>
    <td style="width:55%;vertical-align:top">
      Организация <span class="underline">{e(org_full)}</span>
    </td>
    <td style="width:45%;vertical-align:top;text-align:right;font-size:8pt">
      Типовая межотраслевая форма № М-2<br>
      Утверждена постановлением Госстата России от 30.10.97 № 71а<br><br>
      <table class="codes" border="1" cellspacing="0">
        <tr><td colspan="2" class="bold center">Коды</td></tr>
        <tr><td>Форма по ОКУД</td><td>0315001</td></tr>
        <tr><td>по ОКПО</td><td>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</td></tr>
      </table>
    </td>
  </tr>
</table>

<p style="text-align:center;font-size:12pt;font-weight:bold;margin:4pt 0 2pt"><b>Доверенность № {e(supply_num)}</b></p>

<p>Дата выдачи <span class="underline bold">{date_display}</span></p>
<p>Доверенность действительна 14 дней с даты подписания.</p>
<p><span class="underline">{e(org_line)}</span></p>
<p class="small">(наименование потребителя и его адрес)</p>
<p><span class="underline">{e(org_line)}</span></p>
<p class="small">(наименование плательщика и его адрес)</p>

<p>
  Доверенность выдана &nbsp;&nbsp;
  <span class="underline" style="min-width:60pt;display:inline-block">водителю</span>
  &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;
  <span class="underline">{e(driver_name)}</span>
</p>
<p class="small" style="padding-left:108pt">(должность) &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; (фамилия, имя, отчество)</p>

<p>
  На отправку груза от &nbsp;&nbsp;
  <span class="underline">&nbsp;&nbsp;&nbsp;&nbsp;{e(supplier_short)}&nbsp;&nbsp;&nbsp;&nbsp;</span>
</p>
<p class="small" style="text-align:center">наименование поставщика</p>

<p>
  материальных ценностей. Основание: №<span class="underline bold">{e(supply_num)}</span>
  &nbsp; от &nbsp;
  <span class="underline bold">{date_display}</span>
</p>
<p class="small">наименование, номер и дата документа</p>

<p style="margin-top:3pt">Перечень материальных ценностей, подлежащих доставке</p>
<table class="mat" border="1" cellspacing="0">
  <colgroup><col style="width:5%"><col style="width:75%"><col style="width:10%"><col style="width:10%"></colgroup>
  <tr>
    <th style="border:1px solid #000;padding:0 2pt;white-space:nowrap;line-height:1.1">№</th>
    <th style="border:1px solid #000;padding:0 2pt;text-align:left;white-space:normal;line-height:1.1">Материальные ценности</th>
    <th style="border:1px solid #000;padding:0 2pt;white-space:nowrap;line-height:1.1">Ед. изм.</th>
    <th style="border:1px solid #000;padding:0 2pt;white-space:nowrap;line-height:1.1">Кол-во</th>
  </tr>
  {goods_rows_html}
</table>

<p style="margin-top:6pt">
  Подпись лица, получившего доверенность удостоверяем.
  &nbsp;&nbsp;&nbsp;&nbsp;
  <span class="dotline">&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</span>
  &nbsp;&nbsp;
  ({e(driver_name)})
</p>

<table style="width:100%;margin-top:6pt;border-collapse:collapse">
  <tr>
    <td style="width:25%;vertical-align:bottom">Руководитель<br><span style="font-size:8pt">М.П.</span></td>
    <td style="width:30%;vertical-align:bottom;text-align:center">
      <span class="dotline">&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</span><br>
      <span class="small">подпись</span>
    </td>
    <td style="width:45%;vertical-align:bottom;text-align:center">
      {sig_block}<br>({e(signatories)})<br>
      <span class="small">расшифровка подписи</span>
    </td>
  </tr>
</table>

<table style="width:100%;margin-top:4pt;border-collapse:collapse">
  <tr>
    <td style="width:25%;vertical-align:bottom">Главный бухгалтер</td>
    <td style="width:30%;vertical-align:bottom;text-align:center">
      <span class="dotline">&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</span><br>
      <span class="small">подпись</span>
    </td>
    <td style="width:45%;vertical-align:bottom;text-align:center">
      ({e(signatories)})<br>
      <span class="small">расшифровка подписи</span>
    </td>
  </tr>
</table>

</div></body></html>"""

    @app.post("/api/ozon-supplies/combined-poa.doc")
    def get_ozon_combined_poa(request: Request, body: OzonCombinedDocsRequest) -> "Response":
        """Generate combined PoA for multiple OZON supplies (same LE, same driver)."""
        import html as _hm
        from fastapi.responses import Response
        from urllib.parse import quote as _qp
        from datetime import datetime as _dtt

        user = _require_user(request)
        if not _can_view_supplies(user): raise HTTPException(status_code=403)
        owner_id = _supply_owner_id(user)
        e = _hm.escape
        supply_ids = body.supply_ids
        if not supply_ids: raise HTTPException(status_code=400, detail="supply_ids required")

        now = _dtt.now()
        date_display = now.strftime("%d.%m.%Y")
        seq = repository.next_ttn_number()
        doc_num = f"{now.strftime('%d%m%Y')}_{seq}"

        # Collect and merge goods across all supplies
        all_goods: dict[str, dict] = {}
        le = {}
        driver_name = ""
        supplier_short = ""
        name_map = repository.get_product_name_by_article(user_id=owner_id)

        for sid in supply_ids:
            try:
                data = _ozon_get_doc_data(owner_id, sid)
            except Exception:
                data = {}
            if not data: continue
            if not le:
                le = data.get("le") or {}
                supplier_short = str(le.get("short_name") or "")
            if not driver_name:
                driver_name = data.get("driver_name") or ""
            for g in (data.get("goods") or []):
                oid = str(g.get("offer_id") or "")
                qty = int(g.get("quantity") or 0)
                sku_k = str(g.get("sku") or ""); nm = name_map.get(sku_k) or name_map.get(oid) or oid or str(g.get("name") or "Товар")
                if oid in all_goods:
                    all_goods[oid]["quantity"] += qty
                else:
                    all_goods[oid] = {"offer_id": oid, "name": nm, "quantity": qty}

        goods = list(all_goods.values())
        supply_nums = ", ".join(str(sid) for sid in supply_ids)

        data_combined = {
            "item": {"supply_order_number": doc_num, "warehouse_name": supplier_short},
            "owner_id": owner_id, "driver_name": driver_name, "driver_docs": "",
            "le": le, "goods": goods, "price_map": {}, "name_map": name_map,
        }
        html_content = "\uFEFF" + _build_ozon_poa_html(data_combined, include_signature=False)
        fname = f"Доверенность суммарная {doc_num}, {supplier_short}.doc"
        return Response(content=html_content.encode("utf-8"), media_type="application/msword",
                        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{_qp(fname)}"})

    @app.post("/api/ozon-supplies/combined-poa.pdf")
    def get_ozon_combined_poa_pdf(request: Request, body: OzonCombinedDocsRequest) -> "Response":
        """Print-ready PDF of combined PoA via LibreOffice."""
        import subprocess as _sp, tempfile as _tf, pathlib as _pl, os as _os
        from fastapi.responses import Response
        from urllib.parse import quote as _qp
        from datetime import datetime as _dtt

        user = _require_user(request)
        if not _can_view_supplies(user): raise HTTPException(status_code=403)
        owner_id = _supply_owner_id(user)
        supply_ids = body.supply_ids
        if not supply_ids: raise HTTPException(status_code=400, detail="supply_ids required")

        now = _dtt.now()
        seq = repository.next_ttn_number()
        doc_num = f"{now.strftime('%d%m%Y')}_{seq}"

        all_goods: dict[str, dict] = {}
        le = {}
        driver_name = ""
        supplier_short = ""
        name_map = repository.get_product_name_by_ozon_sku(user_id=owner_id)
        name_map_art = repository.get_product_name_by_article(user_id=owner_id)

        for sid in supply_ids:
            try: data = _ozon_get_doc_data(owner_id, sid)
            except Exception: data = {}
            if not data: continue
            if not le:
                le = data.get("le") or {}
                supplier_short = str(le.get("short_name") or "")
            if not driver_name: driver_name = data.get("driver_name") or ""
            for g in (data.get("goods") or []):
                oid = str(g.get("offer_id") or "")
                sku_k = str(g.get("sku") or "")
                qty = int(g.get("quantity") or 0)
                nm = name_map.get(sku_k) or name_map_art.get(oid) or oid or str(g.get("name") or "Товар")
                if oid in all_goods: all_goods[oid]["quantity"] += qty
                else: all_goods[oid] = {"offer_id": oid, "sku": sku_k, "name": nm, "quantity": qty}

        data_combined = {
            "item": {"supply_order_number": doc_num, "warehouse_name": supplier_short},
            "owner_id": owner_id, "driver_name": driver_name, "driver_docs": "",
            "le": le, "goods": list(all_goods.values()), "price_map": {}, "name_map": name_map,
        }
        html_content = _build_ozon_poa_html(data_combined, include_signature=True)
        tmp_dir = _tf.mkdtemp()
        html_path = _pl.Path(tmp_dir) / "combined_poa.html"
        pdf_path  = _pl.Path(tmp_dir) / "combined_poa.pdf"
        html_path.write_text(html_content, encoding="utf-8")
        env = dict(_os.environ)
        env.update({"HOME": "/tmp", "XDG_CACHE_HOME": "/tmp/.cache",
                    "XDG_CONFIG_HOME": "/tmp/.config", "DCONF_PROFILE": "empty"})
        _sp.run(["soffice", "--headless", "--convert-to", "pdf",
                 "--outdir", str(tmp_dir), str(html_path)],
                capture_output=True, env=env, timeout=60)
        if not pdf_path.exists():
            raise HTTPException(status_code=500, detail="LibreOffice не смог сгенерировать PDF")
        fname = f"Доверенность суммарная {doc_num}, {supplier_short}.pdf"
        return Response(content=pdf_path.read_bytes(), media_type="application/pdf",
                        headers={"Content-Disposition": f"inline; filename*=UTF-8''{_qp(fname)}"})

    @app.post("/api/ozon-supplies/combined-ttn.pdf")
    def get_ozon_combined_ttn_pdf(request: Request, body: OzonCombinedDocsRequest) -> "Response":
        """Print-ready PDF of combined TTN via LibreOffice."""
        import subprocess as _sp, tempfile as _tf, pathlib as _pl, os as _os
        import zipfile as _zf, io as _io
        from fastapi.responses import Response
        from urllib.parse import quote as _qp

        user = _require_user(request)
        if not _can_view_supplies(user): raise HTTPException(status_code=403)
        owner_id = _supply_owner_id(user)

        # Reuse combined-ttn.docx logic to build the DOCX bytes
        class _FakeBody:
            supply_ids = body.supply_ids
        # Call the existing endpoint function directly
        docx_resp = get_ozon_combined_ttn(request, body)
        if not hasattr(docx_resp, "body"):
            raise HTTPException(status_code=500, detail="Не удалось собрать DOCX")

        tmp_dir  = _tf.mkdtemp()
        docx_path = _pl.Path(tmp_dir) / "combined_ttn.docx"
        pdf_path  = _pl.Path(tmp_dir) / "combined_ttn.pdf"
        docx_path.write_bytes(docx_resp.body)
        env = dict(_os.environ)
        env.update({"HOME": "/tmp", "XDG_CACHE_HOME": "/tmp/.cache",
                    "XDG_CONFIG_HOME": "/tmp/.config", "DCONF_PROFILE": "empty"})
        _sp.run(["soffice", "--headless", "--convert-to", "pdf",
                 "--outdir", str(tmp_dir), str(docx_path)],
                capture_output=True, env=env, timeout=60)
        if not pdf_path.exists():
            raise HTTPException(status_code=500, detail="LibreOffice не смог сгенерировать PDF")
        fname = "ТТН суммарная.pdf"
        return Response(content=pdf_path.read_bytes(), media_type="application/pdf",
                        headers={"Content-Disposition": f"inline; filename*=UTF-8''{_qp(fname)}"})

    @app.post("/api/ozon-supplies/combined-ttn.docx")
    def get_ozon_combined_ttn(request: Request, body: OzonCombinedDocsRequest) -> "Response":
        """Generate combined TTN DOCX for multiple OZON supplies."""
        import zipfile as _zf, io as _io, re as _re, html as _html_esc
        from fastapi.responses import Response
        from urllib.parse import quote as _qp
        from datetime import datetime as _dtt

        user = _require_user(request)
        if not _can_view_supplies(user): raise HTTPException(status_code=403)
        owner_id = _supply_owner_id(user)
        supply_ids = body.supply_ids
        if not supply_ids: raise HTTPException(status_code=400, detail="supply_ids required")

        now = _dtt.now()
        seq = repository.next_ttn_number()
        doc_num = f"{now.strftime('%d%m%Y')}_{seq}"
        supply_date_disp = now.strftime("%d.%m.%Y")

        # Collect data
        all_goods: dict[str, dict] = {}
        price_map: dict[str, float] = {}
        le = {}
        supplier_short = ""
        driver_name = ""
        name_map = repository.get_product_name_by_article(user_id=owner_id)
        wh_name = ""
        transit_wh_name = ""

        for sid in supply_ids:
            try:
                data = _ozon_get_doc_data(owner_id, sid)
            except Exception:
                data = {}
            if not data: continue
            if not le:
                le = data.get("le") or {}
                supplier_short = str(le.get("short_name") or "")
                _item0 = data.get("item") or {}
                wh_name = str(_item0.get("warehouse_name") or "")
                transit_wh_name = str(_item0.get("transit_warehouse_name") or "")
            if not driver_name:
                driver_name = data.get("driver_name") or ""
            price_map.update(data.get("price_map") or {})
            for g in (data.get("goods") or []):
                oid = str(g.get("offer_id") or "")
                qty = int(g.get("quantity") or 0)
                sku_k = str(g.get("sku") or ""); nm = name_map.get(sku_k) or name_map.get(oid) or oid or str(g.get("name") or "Товар")
                if oid in all_goods:
                    all_goods[oid]["quantity"] += qty
                else:
                    all_goods[oid] = {"offer_id": oid, "name": nm, "quantity": qty}

        goods = list(all_goods.values())
        org_full = str(le.get("full_name") or supplier_short)
        org_req = str(le.get("requisites") or "")
        org_line = ", ".join(filter(None, [org_full, org_req]))
        signatories = str(le.get("signatories") or supplier_short or "—")
        VAT_RATE = 0.22

        pickup_wh = (transit_wh_name or wh_name).strip()
        warehouses = repository.list_supply_warehouses(user_id=owner_id)
        wh_addr = next(
            (repository.warehouse_address_line(w) for w in warehouses
             if str(w.get("warehouse_name") or "").strip() == pickup_wh),
            "",
        )
        recipient_line = "ООО «РВБ»" + (f", {wh_addr}" if wh_addr else "")

        def _rubles_in_words(n: int) -> str:
            ones_m=["","один","два","три","четыре","пять","шесть","семь","восемь","девять"]
            ones_f=["","одна","две","три","четыре","пять","шесть","семь","восемь","девять"]
            teens=["десять","одиннадцать","двенадцать","тринадцать","четырнадцать","пятнадцать","шестнадцать","семнадцать","восемнадцать","девятнадцать"]
            tens=["","","двадцать","тридцать","сорок","пятьдесят","шестьдесят","семьдесят","восемьдесят","девяносто"]
            hunds=["","сто","двести","триста","четыреста","пятьсот","шестьсот","семьсот","восемьсот","девятьсот"]
            def chunk(x,fem):
                r,w=x%100,[]
                h=x//100
                if h: w.append(hunds[h])
                if 10<=r<=19: w.append(teens[r-10])
                else:
                    if r//10: w.append(tens[r//10])
                    d=r%10
                    if d: w.append((ones_f if fem else ones_m)[d])
                return w
            if n==0: return "ноль рублей 00 копеек"
            w=[]
            bn,mn,th,ru=n//1000000000,(n//1000000)%1000,(n//1000)%1000,n%1000
            def suf(x,forms): return forms[1 if x%10==1 and x%100!=11 else 2 if x%10 in(2,3,4) and x%100 not in range(12,15) else 3]
            if bn: w.extend(chunk(bn,False)); w.append(suf(bn,["","миллиард","миллиарда","миллиардов"]))
            if mn: w.extend(chunk(mn,False)); w.append(suf(mn,["","миллион","миллиона","миллионов"]))
            if th: w.extend(chunk(th,True)); w.append(suf(th,["","тысяча","тысячи","тысяч"]))
            if ru: w.extend(chunk(ru,False))
            w.append(suf(ru,["рублей","рубль","рубля","рублей"]))
            w.append("00 копеек")
            return " ".join(w)

        rows_data = []
        total_qty = total_excl = total_vat = total_incl = 0
        for i, g in enumerate(goods):
            oid = str(g.get("offer_id") or "")
            sku_k = str(g.get("sku") or ""); nm = name_map.get(sku_k) or name_map.get(oid) or oid or str(g.get("name") or "Товар")
            qty = int(g.get("quantity") or 0)
            price = float(price_map.get(oid) or 0)
            excl = round(price / (1 + VAT_RATE), 2) if price else 0.0
            vat_amt = round(price * VAT_RATE / (1 + VAT_RATE), 2) if price else 0.0
            sum_excl = round(excl * qty, 2); sum_vat = round(vat_amt * qty, 2); sum_incl = round(price * qty, 2)
            total_qty += qty; total_excl += sum_excl; total_vat += sum_vat; total_incl += sum_incl
            rows_data.append({"num": str(i+1), "name": nm, "qty": qty,
                "price_excl": f"{excl:.2f}" if price else "—", "amt_excl": f"{sum_excl:.2f}" if price else "—",
                "vat_amt": f"{sum_vat:.2f}" if price else "—", "amt_incl": f"{sum_incl:.2f}" if price else "—"})

        t_excl = f"{total_excl:.2f}"; t_vat = f"{total_vat:.2f}"; t_incl = f"{total_incl:.2f}"
        amt_words = _rubles_in_words(round(total_incl)) if total_incl else "Ноль рублей 00 копеек"

        tpl_path = STATIC_DIR / "torg12_tpl.docx"
        with open(tpl_path, "rb") as f: tpl_bytes = f.read()
        with _zf.ZipFile(_io.BytesIO(tpl_bytes)) as zin:
            all_files = {n: zin.read(n) for n in zin.namelist()}
        doc_xml = all_files["word/document.xml"].decode("utf-8")

        row_rx = _re.compile(r'(<w:tr[\s>](?:(?!</w:tr>).)*?\{\{GOODS_NAME\}\}.*?</w:tr>)', _re.DOTALL)
        m = row_rx.search(doc_xml)
        if m and rows_data:
            row_tpl = m.group(1); multi = ""
            for rd in rows_data:
                r = row_tpl
                for ph, val in [("{{ROW_NUM}}", rd["num"]), ("{{GOODS_NAME}}", _html_esc.escape(rd["name"])),
                                 ("{{PRICE}}", rd["price_excl"]), ("{{ROW_AMOUNT_EXCL}}", rd["amt_excl"]),
                                 ("{{ROW_VAT_SUM}}", rd["vat_amt"]), ("{{ROW_AMOUNT_INCL}}", rd["amt_incl"]),
                                 ("{{ROW_QTY}}", str(rd["qty"]))]:
                    r = r.replace(ph, val)
                multi += r
            doc_xml = doc_xml.replace(row_tpl, multi, 1)

        # Recipient for combined OZON TTN: first supply's transit/dest warehouse address
        first_item = {}
        for sid in supply_ids:
            try:
                _fd = _ozon_get_doc_data(owner_id, sid)
            except Exception:
                _fd = {}
            if _fd:
                first_item = _fd.get("item") or {}
                break
        dest_wh = str(first_item.get("warehouse_name") or wh_name or "").strip()
        transit_wh = str(first_item.get("transit_warehouse_name") or "").strip()
        pickup_wh = transit_wh or dest_wh
        warehouses = repository.list_supply_warehouses(user_id=owner_id)
        wh_addr = next(
            (repository.warehouse_address_line(w) for w in warehouses
             if str(w.get("warehouse_name") or "").strip() == pickup_wh),
            "",
        )
        recipient_line = "ООО «Интернет решения»" + (f", {wh_addr}" if wh_addr else "")

        mon_names = ['января','февраля','марта','апреля','мая','июня','июля','августа','сентября','октября','ноября','декабря']
        for ph, val in [
            ("{{TTN_NUMBER}}", doc_num), ("{{ORG_FULL}}", org_line), ("{{SUPPLIER}}", org_line),
            ("{{PAYER}}", org_line), ("{{RECIPIENT}}", recipient_line),
            ("{{ORDER_DATE}}", doc_num), ("{{DOC_NUM_VAL}}", doc_num),
            ("{{DOC_DATE_VAL}}", supply_date_disp),
            ("{{GOODS_NAME}}", rows_data[0]["name"] if rows_data else "Товар"),
            ("{{ROW_NUM}}", "1"), ("{{PRICE}}", rows_data[0]["price_excl"] if rows_data else "—"),
            ("{{ROW_AMOUNT_EXCL}}", rows_data[0]["amt_excl"] if rows_data else "—"),
            ("{{ROW_VAT_SUM}}", rows_data[0]["vat_amt"] if rows_data else "—"),
            ("{{ROW_AMOUNT_INCL}}", rows_data[0]["amt_incl"] if rows_data else "—"),
            ("{{QTY}}", str(total_qty)), ("{{QTY_SHT}}", f"{total_qty} шт"),
            ("{{TOTAL_EXCL}}", t_excl), ("{{TOTAL_VAT}}", t_vat), ("{{TOTAL_INCL}}", t_incl),
            ("{{AMOUNT}}", t_excl), ("{{VAT_SUM}}", t_vat), ("{{AMOUNT_WITH_VAT}}", t_incl),
            ("{{AMOUNT_WORDS}}", amt_words), ("{{PAGES_COUNT}}", "1"),
            ("{{ITEMS_COUNT}}", str(len(rows_data) or 1)),
            ("{{TOTAL_RUB}}", str(int(total_incl)) if total_incl else "0"),
            ("{{TOTAL_KOP}}", str(round((total_incl % 1)*100)).zfill(2) if total_incl else "00"),
            ("{{SUPPLY_ID}}", doc_num),
            ("{{DOC_DATE_FULL}}", f"«{now.strftime('%d')}» {mon_names[now.month-1]} {now.year}"),
            ("{{ISSUED_BY}}", supplier_short or "—"), ("{{SIGNATORIES}}", signatories),
            ("{{PROD_HEAD}}", next(
                (str(p.get("head_name") or "").strip() or "—"
                 for p in repository.list_supply_productions(user_id=owner_id)
                 if str(p.get("name") or "").strip() == str(first_item.get("production") or "").strip()),
                "—",
            )),
            ("{{SIGN_SUPPLIER}}", supplier_short), ("{{SIGN_DRIVER}}", driver_name),
        ]:
            doc_xml = doc_xml.replace(ph, val)
        doc_xml = doc_xml.replace("{{ROW_QTY}}", str(total_qty))
        all_files["word/document.xml"] = doc_xml.encode("utf-8")

        buf = _io.BytesIO()
        with _zf.ZipFile(buf, "w", _zf.ZIP_DEFLATED) as zout:
            for n, fdata in all_files.items(): zout.writestr(n, fdata)
        fname = f"ТТН суммарная {doc_num}, {supplier_short}.docx"
        return Response(content=buf.getvalue(),
                        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{_qp(fname)}"})

    @app.get("/api/ozon-supplies/{supply_order_id}/poa.doc")
    def get_ozon_poa_doc(request: Request, supply_order_id: int) -> "Response":
        from fastapi.responses import Response
        from urllib.parse import quote as _qp
        user = _require_user(request)
        if not _can_view_supplies(user): raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        data = _ozon_get_doc_data(owner_id, supply_order_id)
        if not data: raise HTTPException(status_code=404, detail="Поставка не найдена")
        html_content = "\uFEFF" + _build_ozon_poa_html(data, include_signature=False)
        supply_num = str(data["item"].get("supply_order_number") or supply_order_id)
        fname = f"Доверенность_OZON_{supply_num}.doc"
        return Response(content=html_content.encode("utf-8"), media_type="application/msword",
                        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{_qp(fname)}"})

    @app.get("/api/ozon-supplies/{supply_order_id}/poa.pdf")
    def get_ozon_poa_pdf(request: Request, supply_order_id: int) -> "Response":
        """Generate OZON Power of Attorney HTML → PDF via LibreOffice."""
        import subprocess as _sp, tempfile as _tf, pathlib as _pl, os as _os
        from fastapi.responses import Response
        from urllib.parse import quote as _qp

        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        data = _ozon_get_doc_data(owner_id, supply_order_id)
        if not data:
            raise HTTPException(status_code=404, detail="Поставка не найдена")

        supply_num = str(data["item"].get("supply_order_number") or supply_order_id)
        html_content = _build_ozon_poa_html(data, include_signature=True)

        tmp_dir = _tf.mkdtemp()
        html_path = _pl.Path(tmp_dir) / "ozon_poa.html"
        pdf_path = _pl.Path(tmp_dir) / "ozon_poa.pdf"
        html_path.write_text(html_content, encoding="utf-8")
        env = dict(_os.environ)
        env.update({"HOME": "/tmp", "XDG_CACHE_HOME": "/tmp/.cache",
                    "XDG_CONFIG_HOME": "/tmp/.config", "DCONF_PROFILE": "empty"})
        _sp.run(["soffice", "--headless", "--convert-to", "pdf",
                 "--outdir", str(tmp_dir), str(html_path)], capture_output=True, env=env, timeout=60)
        if not pdf_path.exists():
            raise HTTPException(status_code=500, detail="LibreOffice не смог сгенерировать PDF")
        pdf_bytes = pdf_path.read_bytes()
        fname = f"Доверенность_OZON_{supply_num}.pdf"
        return Response(content=pdf_bytes, media_type="application/pdf",
                        headers={"Content-Disposition": f"inline; filename*=UTF-8''{_qp(fname)}"})

    def _build_ozon_ttn_docx(request: Request, supply_order_id: int, owner_id: int):
        """Build OZON TTN DOCX bytes. Returns (bytes, supply_num) or None."""
        import zipfile as _zf, io as _io, re as _re, html as _html_esc
        from datetime import datetime as _dtt

        data = _ozon_get_doc_data(owner_id, supply_order_id)
        if not data:
            return None

        item = data["item"]
        le = data["le"]
        goods = data["goods"]
        price_map = data["price_map"]
        name_map = data["name_map"]
        driver_name = data.get("driver_name") or "—"

        def _rubles_in_words(n: int) -> str:
            ones_m = ["","один","два","три","четыре","пять","шесть","семь","восемь","девять"]
            ones_f = ["","одна","две","три","четыре","пять","шесть","семь","восемь","девять"]
            teens  = ["десять","одиннадцать","двенадцать","тринадцать","четырнадцать",
                      "пятнадцать","шестнадцать","семнадцать","восемнадцать","девятнадцать"]
            tens   = ["","","двадцать","тридцать","сорок","пятьдесят","шестьдесят","семьдесят","восемьдесят","девяносто"]
            hunds  = ["","сто","двести","триста","четыреста","пятьсот","шестьсот","семьсот","восемьсот","девятьсот"]
            def chunk(x, fem):
                r,w = x%100,[]
                h = x//100
                if h: w.append(hunds[h])
                if 10<=r<=19: w.append(teens[r-10])
                else:
                    if r//10: w.append(tens[r//10])
                    d = r%10
                    if d: w.append((ones_f if fem else ones_m)[d])
                return w
            if n==0: return "ноль рублей 00 копеек"
            w=[]
            bn,mn,th,ru = n//1000000000,(n//1000000)%1000,(n//1000)%1000,n%1000
            def suf(x,forms): return forms[1 if x%10==1 and x%100!=11 else 2 if x%10 in(2,3,4) and x%100 not in range(12,15) else 3]
            if bn: w.extend(chunk(bn,False)); w.append(suf(bn,["","миллиард","миллиарда","миллиардов"]))
            if mn: w.extend(chunk(mn,False)); w.append(suf(mn,["","миллион","миллиона","миллионов"]))
            if th: w.extend(chunk(th,True));  w.append(suf(th,["","тысяча","тысячи","тысяч"]))
            if ru: w.extend(chunk(ru,False))
            w.append(suf(ru,["рублей","рубль","рубля","рублей"]))
            w.append("00 копеек")
            return " ".join(w)

        now = _dtt.now()
        supply_num = str(item.get("supply_order_number") or supply_order_id)
        supply_date_disp = now.strftime("%d.%m.%Y")
        org_full = str(le.get("full_name") or le.get("short_name") or "")
        org_req = str(le.get("requisites") or "")
        org_line = ", ".join(filter(None, [org_full, org_req]))
        supplier_short = str(le.get("short_name") or "")
        signatories = str(le.get("signatories") or supplier_short or "—")
        VAT_RATE = 0.22

        rows_data = []
        total_qty = 0
        total_excl = 0.0
        total_vat = 0.0
        total_incl = 0.0
        for i, g in enumerate(goods):
            offer_id = str(g.get("offer_id") or "")
            sku_key = str(g.get("sku") or "")
            name = name_map.get(sku_key) or name_map.get(offer_id) or offer_id or str(g.get("name") or "Товар")
            qty = int(g.get("quantity") or 0)
            price = float(price_map.get(offer_id) or 0)
            excl = round(price / (1 + VAT_RATE), 2) if price else 0.0
            vat_amt = round(price * VAT_RATE / (1 + VAT_RATE), 2) if price else 0.0
            sum_excl = round(excl * qty, 2)
            sum_vat = round(vat_amt * qty, 2)
            sum_incl = round(price * qty, 2)
            total_qty += qty
            total_excl += sum_excl
            total_vat += sum_vat
            total_incl += sum_incl
            rows_data.append({
                "num": str(i + 1),
                "name": name,
                "qty": qty,
                "price_excl": f"{excl:.2f}" if price else "—",
                "amt_excl": f"{sum_excl:.2f}" if price else "—",
                "vat_amt": f"{sum_vat:.2f}" if price else "—",
                "amt_incl": f"{sum_incl:.2f}" if price else "—",
            })

        t_excl = f"{total_excl:.2f}"
        t_vat  = f"{total_vat:.2f}"
        t_incl = f"{total_incl:.2f}"
        amt_words = _rubles_in_words(round(total_incl)) if total_incl else "Ноль рублей 00 копеек"

        tpl_path = STATIC_DIR / "torg12_tpl.docx"
        with open(tpl_path, "rb") as f:
            tpl_bytes = f.read()
        with _zf.ZipFile(_io.BytesIO(tpl_bytes)) as zin:
            all_files = {name: zin.read(name) for name in zin.namelist()}
        doc_xml = all_files["word/document.xml"].decode("utf-8")

        row_rx = _re.compile(r'(<w:tr[\s>](?:(?!</w:tr>).)*?\{\{GOODS_NAME\}\}.*?</w:tr>)', _re.DOTALL)
        m = row_rx.search(doc_xml)
        if m and rows_data:
            row_tpl = m.group(1)
            multi = ""
            for rd in rows_data:
                r = row_tpl
                r = r.replace("{{ROW_NUM}}",         rd["num"])
                r = r.replace("{{GOODS_NAME}}",       _html_esc.escape(rd["name"]))
                r = r.replace("{{PRICE}}",            _html_esc.escape(rd["price_excl"]))
                r = r.replace("{{ROW_AMOUNT_EXCL}}",  _html_esc.escape(rd["amt_excl"]))
                r = r.replace("{{ROW_VAT_SUM}}",      _html_esc.escape(rd["vat_amt"]))
                r = r.replace("{{ROW_AMOUNT_INCL}}",  _html_esc.escape(rd["amt_incl"]))
                r = r.replace("{{ROW_QTY}}",          str(rd["qty"]))
                multi += r
            doc_xml = doc_xml.replace(row_tpl, multi, 1)

        # Recipient: ООО «Интернет решения» + address of initial (transit) warehouse, else destination
        dest_wh = str(item.get("warehouse_name") or "").strip()
        transit_wh = str(item.get("transit_warehouse_name") or "").strip()
        pickup_wh = transit_wh or dest_wh
        warehouses = repository.list_supply_warehouses(user_id=owner_id)
        wh_addr = next(
            (repository.warehouse_address_line(w) for w in warehouses
             if str(w.get("warehouse_name") or "").strip() == pickup_wh),
            "",
        )
        recipient_line = "ООО «Интернет решения»" + (f", {wh_addr}" if wh_addr else "")

        mon_names = ['января','февраля','марта','апреля','мая','июня','июля','августа','сентября','октября','ноября','декабря']
        for ph, val in [
            ("{{TTN_NUMBER}}",      supply_num),
            ("{{ORG_FULL}}",        org_line),
            ("{{SUPPLIER}}",        org_line),
            ("{{PAYER}}",           org_line),
            ("{{RECIPIENT}}",       recipient_line),
            ("{{ORDER_DATE}}",      supply_num),
            ("{{DOC_NUM_VAL}}",     supply_num),
            ("{{DOC_DATE_VAL}}",    supply_date_disp),
            ("{{GOODS_NAME}}",      rows_data[0]["name"] if rows_data else "Товар"),
            ("{{ROW_NUM}}",         "1"),
            ("{{PRICE}}",           rows_data[0]["price_excl"] if rows_data else "—"),
            ("{{ROW_AMOUNT_EXCL}}", rows_data[0]["amt_excl"] if rows_data else "—"),
            ("{{ROW_VAT_SUM}}",     rows_data[0]["vat_amt"] if rows_data else "—"),
            ("{{ROW_AMOUNT_INCL}}", rows_data[0]["amt_incl"] if rows_data else "—"),
            ("{{QTY}}",             str(total_qty)),
            ("{{QTY_SHT}}",         f"{total_qty} шт"),
            ("{{TOTAL_EXCL}}",      t_excl),
            ("{{TOTAL_VAT}}",       t_vat),
            ("{{TOTAL_INCL}}",      t_incl),
            ("{{AMOUNT}}",          t_excl),
            ("{{VAT_SUM}}",         t_vat),
            ("{{AMOUNT_WITH_VAT}}", t_incl),
            ("{{AMOUNT_WORDS}}",    amt_words),
            ("{{PAGES_COUNT}}",     "1"),
            ("{{ITEMS_COUNT}}",     str(len(rows_data) or 1)),
            ("{{TOTAL_RUB}}",       str(int(total_incl)) if total_incl else "0"),
            ("{{TOTAL_KOP}}",       str(round((total_incl % 1) * 100)).zfill(2) if total_incl else "00"),
            ("{{SUPPLY_ID}}",       supply_num),
            ("{{DOC_DATE_FULL}}",   f"«{now.strftime('%d')}» {mon_names[now.month-1]} {now.year}"),
            ("{{ISSUED_BY}}",       supplier_short or "—"),
            ("{{SIGNATORIES}}",     signatories),
            ("{{PROD_HEAD}}",       next(
                (str(p.get("head_name") or "").strip() or "—"
                 for p in repository.list_supply_productions(user_id=owner_id)
                 if str(p.get("name") or "").strip() == str(item.get("production") or "").strip()),
                "—",
            )),
            ("{{SIGN_SUPPLIER}}",   supplier_short),
            ("{{SIGN_DRIVER}}",     driver_name),
        ]:
            doc_xml = doc_xml.replace(ph, val)
        doc_xml = doc_xml.replace("{{ROW_QTY}}", str(total_qty))

        all_files["word/document.xml"] = doc_xml.encode("utf-8")

        buf = _io.BytesIO()
        with _zf.ZipFile(buf, "w", _zf.ZIP_DEFLATED) as zout:
            for name, fdata in all_files.items():
                zout.writestr(name, fdata)
        wh_name = str(item.get("warehouse_name") or "")
        return buf.getvalue(), supply_num, supplier_short, supply_date_disp, wh_name

    @app.get("/api/ozon-supplies/{supply_order_id}/ttn.docx")
    def get_ozon_ttn_docx(request: Request, supply_order_id: int) -> "Response":
        from fastapi.responses import Response
        from urllib.parse import quote as _qp
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        result = _build_ozon_ttn_docx(request, supply_order_id, _supply_owner_id(user))
        if result is None:
            raise HTTPException(status_code=404, detail="Поставка не найдена")
        docx_bytes, supply_num, supplier_short, supply_date_disp, wh_name = result
        fname = f"ТТН №{supply_num}, {supplier_short} от {supply_date_disp}, {wh_name}.docx"
        return Response(content=docx_bytes,
                        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{_qp(fname)}"})

    def _ozon_prepare_doc_xml_context(request: Request, supply_order_id: int) -> tuple[dict, dict, object]:
        """Shared context for Ozon eTrN / ЭЗЗ XML downloads."""
        from . import ozon_etrn as _ozon_etrn

        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        data = _ozon_get_doc_data(owner_id, supply_order_id)
        if not data or not data.get("item"):
            raise HTTPException(status_code=404, detail="Поставка не найдена")
        item = data["item"]
        cargoes_json = item.get("cargoes_json")
        try:
            import json as _jj
            cargo_resp = get_ozon_supply_cargoes(request, supply_order_id) or {}
            cache_obj = {
                "version": 2,
                "groups": list(cargo_resp.get("groups") or []),
                "transport_cargoes": list(cargo_resp.get("transport_cargoes") or []),
            }
            if cache_obj["groups"] or cache_obj["transport_cargoes"]:
                cargoes_json = _jj.dumps(cache_obj, ensure_ascii=False)
        except Exception as ex:
            _log.debug("ozon xml cargoes refresh sid=%s: %s", supply_order_id, ex)
        ctx = _ozon_etrn.collect_ozon_etrn_context(
            repository=repository,
            owner_id=owner_id,
            item=item,
            driver_name=str(data.get("driver_name") or ""),
            driver_phone=str(data.get("driver_docs") or ""),
            vehicle_line=str(data.get("vehicle_num") or ""),
        )
        return data, ctx, cargoes_json

    @app.get("/api/ozon-supplies/{supply_order_id}/zakaz.xml")
    def get_ozon_zakaz_xml(request: Request, supply_order_id: int) -> "Response":
        """Download ЭЗЗ (заказ-заявка) title-1 XML draft for Kontur.Logistics."""
        from fastapi.responses import Response
        from urllib.parse import quote as _qp
        from . import ozon_zakaz as _ozon_zakaz

        data, ctx, cargoes_json = _ozon_prepare_doc_xml_context(request, supply_order_id)
        item = data["item"]
        try:
            le_ctx = dict(ctx.get("le") or data.get("le") or {})
            # Explicit shipper phone (same idea as carrier_fields.carrier_phone).
            shipper_phone = str(ctx.get("shipper_phone") or le_ctx.get("phone") or "").strip()
            if not shipper_phone:
                for ent in ctx.get("legal_entities") or []:
                    if str(ent.get("phone") or "").strip():
                        shipper_phone = str(ent.get("phone") or "").strip()
                        le_ctx["phone"] = shipper_phone
                        break
            xml_bytes = _ozon_zakaz.build_ozon_zakaz_xml(
                item=item,
                le=le_ctx,
                driver_name=str(ctx.get("driver_name") or ""),
                driver_phone=str(ctx.get("driver_phone") or ""),
                driver_documents=str(ctx.get("driver_documents") or ""),
                driver_fields=ctx.get("driver_fields") or None,
                vehicle_line=str(ctx.get("vehicle_line") or ""),
                vehicle_json=item.get("vehicle_json"),
                vehicle_fields=ctx.get("vehicle_fields") or None,
                cargoes_json=cargoes_json,
                load_address=str(ctx.get("load_address") or ""),
                load_addr_fields=ctx.get("load_addr_fields") or None,
                delivery_address=str(ctx.get("delivery_address") or ""),
                delivery_addr_fields=ctx.get("delivery_addr_fields") or None,
                carrier_text=str(ctx.get("carrier_text") or ""),
                carrier_fields=ctx.get("carrier_fields") or None,
                loader_name=str(ctx.get("loader_name") or ""),
                shipper_phone=shipper_phone,
                legal_entities=list(ctx.get("legal_entities") or []),
            )
        except Exception as exc:
            _log.exception("ozon zakaz xml failed for %s", supply_order_id)
            raise HTTPException(status_code=500, detail=f"Не удалось сформировать Заявку: {exc}") from exc
        supply_num = str(item.get("supply_order_number") or supply_order_id)
        supplier_short = str((ctx.get("le") or data.get("le") or {}).get("short_name") or "")
        fname = f"Заявка №{supply_num}{', ' + supplier_short if supplier_short else ''}.xml"
        return Response(
            content=xml_bytes,
            media_type="application/xml; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{_qp(fname)}"},
        )

    @app.get("/api/ozon-supplies/{supply_order_id}/etrn.xml")
    def get_ozon_etrn_xml(request: Request, supply_order_id: int) -> "Response":
        """Download eTrN title-1 XML draft for Kontur.Logistics upload."""
        from fastapi.responses import Response
        from urllib.parse import quote as _qp
        from . import ozon_etrn as _ozon_etrn

        data, ctx, cargoes_json = _ozon_prepare_doc_xml_context(request, supply_order_id)
        item = data["item"]
        try:
            xml_bytes = _ozon_etrn.build_ozon_etrn_xml(
                item=item,
                le=ctx.get("le") or data.get("le") or {},
                driver_name=str(ctx.get("driver_name") or ""),
                driver_phone=str(ctx.get("driver_phone") or ""),
                driver_documents=str(ctx.get("driver_documents") or ""),
                driver_fields=ctx.get("driver_fields") or None,
                vehicle_line=str(ctx.get("vehicle_line") or ""),
                vehicle_json=item.get("vehicle_json"),
                vehicle_fields=ctx.get("vehicle_fields") or None,
                cargoes_json=cargoes_json,
                load_address=str(ctx.get("load_address") or ""),
                load_addr_fields=ctx.get("load_addr_fields") or None,
                delivery_address=str(ctx.get("delivery_address") or ""),
                delivery_addr_fields=ctx.get("delivery_addr_fields") or None,
                carrier_text=str(ctx.get("carrier_text") or ""),
                carrier_fields=ctx.get("carrier_fields") or None,
                loader_name=str(ctx.get("loader_name") or ""),
            )
        except Exception as exc:
            _log.exception("ozon etrn xml failed for %s", supply_order_id)
            raise HTTPException(status_code=500, detail=f"Не удалось сформировать эТрН: {exc}") from exc
        supply_num = str(item.get("supply_order_number") or supply_order_id)
        supplier_short = str((ctx.get("le") or data.get("le") or {}).get("short_name") or "")
        fname = f"эТрН №{supply_num}{', ' + supplier_short if supplier_short else ''}.xml"
        return Response(
            content=xml_bytes,
            media_type="application/xml; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{_qp(fname)}"},
        )

    def _ozon_build_edo_xml(request: Request, supply_order_id: int, doc_type: str) -> tuple[bytes, str, dict]:
        """Build Заявка/эТрН XML for CryptoPro signing + Contour send."""
        import base64 as _b64
        from . import ozon_etrn as _ozon_etrn
        from . import ozon_zakaz as _ozon_zakaz

        doc_type = str(doc_type or "").strip().lower()
        if doc_type not in ("zakaz", "etrn"):
            raise HTTPException(status_code=400, detail="doc_type: zakaz или etrn")
        data, ctx, cargoes_json = _ozon_prepare_doc_xml_context(request, supply_order_id)
        item = data["item"]
        le_ctx = dict(ctx.get("le") or data.get("le") or {})
        common = dict(
            item=item,
            le=le_ctx,
            driver_name=str(ctx.get("driver_name") or ""),
            driver_phone=str(ctx.get("driver_phone") or ""),
            driver_documents=str(ctx.get("driver_documents") or ""),
            driver_fields=ctx.get("driver_fields") or None,
            vehicle_line=str(ctx.get("vehicle_line") or ""),
            vehicle_json=item.get("vehicle_json"),
            vehicle_fields=ctx.get("vehicle_fields") or None,
            cargoes_json=cargoes_json,
            load_address=str(ctx.get("load_address") or ""),
            load_addr_fields=ctx.get("load_addr_fields") or None,
            delivery_address=str(ctx.get("delivery_address") or ""),
            delivery_addr_fields=ctx.get("delivery_addr_fields") or None,
            carrier_text=str(ctx.get("carrier_text") or ""),
            carrier_fields=ctx.get("carrier_fields") or None,
            loader_name=str(ctx.get("loader_name") or ""),
        )
        supply_num = str(item.get("supply_order_number") or supply_order_id)
        if doc_type == "zakaz":
            shipper_phone = str(ctx.get("shipper_phone") or le_ctx.get("phone") or "").strip()
            xml_bytes = _ozon_zakaz.build_ozon_zakaz_xml(
                **common,
                shipper_phone=shipper_phone,
                legal_entities=list(ctx.get("legal_entities") or []),
            )
            # ИдФайл for Contour filename — extract from XML root attr when possible.
            fname = f"ON_ZAKZVGO_{supply_num}.xml"
            try:
                import xml.etree.ElementTree as ET
                root = ET.fromstring(xml_bytes)
                if root.attrib.get("ИдФайл"):
                    fname = f"{root.attrib['ИдФайл']}.xml"
            except Exception:
                pass
        else:
            xml_bytes = _ozon_etrn.build_ozon_etrn_xml(**common)
            fname = f"ON_TRNACLGROT_{supply_num}.xml"
            try:
                import xml.etree.ElementTree as ET
                root = ET.fromstring(xml_bytes)
                if root.attrib.get("ИдФайл"):
                    fname = f"{root.attrib['ИдФайл']}.xml"
            except Exception:
                pass
        meta = {
            "doc_type": doc_type,
            "supply_order_id": supply_order_id,
            "filename": fname,
            "xml_base64": _b64.b64encode(xml_bytes).decode("ascii"),
            "xml_size": len(xml_bytes),
        }
        return xml_bytes, fname, meta

    @app.get("/api/supply-edo-settings")
    def get_supply_edo_settings(request: Request) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        return repository.get_supply_edo_settings(user_id=_supply_owner_id(user))

    @app.put("/api/supply-edo-settings")
    def put_supply_edo_settings(request: Request, payload: UpsertSupplyEdoSettingsRequest) -> dict[str, object]:
        user = _require_user(request)
        if str(user.get("role") or "") not in ROLE_CAN_ACCESS_SETTINGS and not bool(
            user.get("can_supply_settings") or user.get("can_view_settings")
        ):
            raise HTTPException(status_code=403, detail="Нет доступа к настройкам ЭДО")
        return repository.upsert_supply_edo_settings(
            user_id=_supply_owner_id(user),
            api_url=payload.api_url,
            api_key=payload.api_key,
            diadoc_url=payload.diadoc_url,
            diadoc_client_id=payload.diadoc_client_id,
            diadoc_login=payload.diadoc_login,
            diadoc_password=payload.diadoc_password,
            diadoc_from_box_id=payload.diadoc_from_box_id,
            diadoc_to_box_id=payload.diadoc_to_box_id,
            cert_thumbprint=payload.cert_thumbprint,
            is_enabled=payload.is_enabled,
        )

    @app.post("/api/supply-edo-settings/test")
    def test_supply_edo_settings(request: Request) -> dict[str, object]:
        """Проверка ключа Contour.Логистика (+ опционально Diadoc auth)."""
        from .kontur_logistics import KonturLogisticsClient
        from .kontur_diadoc import KonturDiadocClient

        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        settings = repository.get_supply_edo_settings(user_id=_supply_owner_id(user), include_secrets=True)
        out: dict[str, object] = {"logistics": None, "diadoc": None}
        if settings.get("api_key"):
            client = KonturLogisticsClient(api_url=str(settings.get("api_url") or ""), api_key=str(settings["api_key"]))
            res = client.ping()
            out["logistics"] = {
                "ok": res.ok,
                "status_code": res.status_code,
                "error": res.error,
                "org": res.data if res.ok else None,
            }
        else:
            out["logistics"] = {"ok": False, "error": "API-ключ Логистики не задан"}
        if settings.get("diadoc_client_id") and settings.get("diadoc_login") and settings.get("diadoc_password"):
            dclient = KonturDiadocClient(
                api_url=str(settings.get("diadoc_url") or ""),
                client_id=str(settings.get("diadoc_client_id") or ""),
                login=str(settings.get("diadoc_login") or ""),
                password=str(settings.get("diadoc_password") or ""),
            )
            ares = dclient.authenticate()
            out["diadoc"] = {"ok": ares.ok, "status_code": ares.status_code, "error": ares.error}
        else:
            out["diadoc"] = {"ok": False, "error": "Diadoc не настроен (нужен для Заявки)"}
        return out

    @app.get("/api/ozon-supplies/{supply_order_id}/edo/prepare")
    def prepare_ozon_edo_xml(request: Request, supply_order_id: int, doc_type: str = "etrn") -> dict[str, object]:
        """XML для подписи КриптоПро перед отправкой в ЭДО."""
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        _xml, _fname, meta = _ozon_build_edo_xml(request, supply_order_id, doc_type)
        settings = repository.get_supply_edo_settings(user_id=_supply_owner_id(user))
        meta["cert_thumbprint"] = str(settings.get("cert_thumbprint") or "")
        diadoc_ready = bool(
            settings.get("diadoc_client_id")
            and settings.get("diadoc_login")
            and settings.get("has_diadoc_password")
            and settings.get("diadoc_from_box_id")
            and settings.get("diadoc_to_box_id")
        )
        # эТрН → Logistics API key; Заявка → Diadoc creds (either channel is enough to enable UI).
        meta["edo_enabled"] = bool(
            settings.get("is_enabled") and (settings.get("has_api_key") or diadoc_ready)
        )
        meta["logistics_ready"] = bool(settings.get("has_api_key"))
        meta["diadoc_ready"] = diadoc_ready
        return meta

    @app.post("/api/ozon-supplies/{supply_order_id}/edo/send")
    def send_ozon_edo_document(
        request: Request, supply_order_id: int, payload: OzonEdoSendRequest
    ) -> dict[str, object]:
        """Подписанный XML → Contour.Логистика (эТрН) или Diadoc (Заявка)."""
        import base64 as _b64
        import json as _jj
        from .kontur_logistics import KonturLogisticsClient, status_label
        from .kontur_diadoc import KonturDiadocClient

        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        doc_type = str(payload.doc_type or "").strip().lower()
        if doc_type not in ("zakaz", "etrn"):
            raise HTTPException(status_code=400, detail="doc_type: zakaz или etrn")
        sig_b64 = str(payload.signature_base64 or "").strip()
        if not sig_b64:
            raise HTTPException(status_code=400, detail="Нужна подпись CryptoPro (signature_base64)")
        try:
            signature_bytes = _b64.b64decode(sig_b64)
        except Exception as exc:
            raise HTTPException(status_code=400, detail="Некорректный signature_base64") from exc
        if payload.xml_base64:
            try:
                xml_bytes = _b64.b64decode(payload.xml_base64)
                fname = f"{'ON_ZAKZVGO' if doc_type == 'zakaz' else 'ON_TRNACLGROT'}_{supply_order_id}.xml"
                try:
                    import xml.etree.ElementTree as ET
                    root = ET.fromstring(xml_bytes)
                    if root.attrib.get("ИдФайл"):
                        fname = f"{root.attrib['ИдФайл']}.xml"
                except Exception:
                    pass
            except Exception as exc:
                raise HTTPException(status_code=400, detail="Некорректный xml_base64") from exc
        else:
            xml_bytes, fname, _meta = _ozon_build_edo_xml(request, supply_order_id, doc_type)

        settings = repository.get_supply_edo_settings(user_id=owner_id, include_secrets=True)
        if not settings.get("is_enabled"):
            raise HTTPException(status_code=400, detail="ЭДО выключен в настройках")

        sig_name = fname + ".sig"
        if doc_type == "etrn":
            if not settings.get("api_key"):
                raise HTTPException(status_code=400, detail="Задайте API-ключ Contour.Логистика в Настройки → ЭДО")
            client = KonturLogisticsClient(
                api_url=str(settings.get("api_url") or ""),
                api_key=str(settings["api_key"]),
            )
            res = client.send_waybill(
                xml_bytes=xml_bytes,
                xml_filename=fname,
                signature_bytes=signature_bytes,
                signature_filename=sig_name,
            )
            if not res.ok:
                repository.upsert_ozon_edo_document(
                    user_id=owner_id,
                    supply_order_id=supply_order_id,
                    doc_type=doc_type,
                    channel="logistics",
                    status="error",
                    status_label="Ошибка отправки",
                    last_error=res.error or f"HTTP {res.status_code}",
                    raw_json=res.raw[:8000],
                )
                raise HTTPException(status_code=502, detail=res.error or "Ошибка Contour.Логистика")
            tid = str(res.data.get("transportationId") or res.data.get("transportation_id") or "").strip()
            st = repository.upsert_ozon_edo_document(
                user_id=owner_id,
                supply_order_id=supply_order_id,
                doc_type=doc_type,
                channel="logistics",
                transportation_id=tid,
                status="sent",
                status_label=status_label("NewTransportation"),
                last_error="",
                raw_json=_jj.dumps(res.data, ensure_ascii=False)[:8000],
                mark_sent=True,
            )
            return {"ok": True, "doc_type": doc_type, "channel": "logistics", "document": st}

        # zakaz → Diadoc LogisticsOrderRequest
        if not (
            settings.get("diadoc_client_id")
            and settings.get("diadoc_login")
            and settings.get("diadoc_password")
            and settings.get("diadoc_from_box_id")
            and settings.get("diadoc_to_box_id")
        ):
            raise HTTPException(
                status_code=400,
                detail="Для Заявки заполните Diadoc в Настройки → ЭДО (Client ID, логин, пароль, From/To BoxId)",
            )
        dclient = KonturDiadocClient(
            api_url=str(settings.get("diadoc_url") or ""),
            client_id=str(settings.get("diadoc_client_id") or ""),
            login=str(settings.get("diadoc_login") or ""),
            password=str(settings.get("diadoc_password") or ""),
        )
        res = dclient.send_order_request(
            from_box_id=str(settings.get("diadoc_from_box_id") or ""),
            to_box_id=str(settings.get("diadoc_to_box_id") or ""),
            xml_bytes=xml_bytes,
            signature_bytes=signature_bytes,
        )
        if not res.ok:
            repository.upsert_ozon_edo_document(
                user_id=owner_id,
                supply_order_id=supply_order_id,
                doc_type=doc_type,
                channel="diadoc",
                status="error",
                status_label="Ошибка отправки",
                last_error=res.error or f"HTTP {res.status_code}",
                raw_json=res.raw[:8000],
            )
            raise HTTPException(status_code=502, detail=res.error or "Ошибка Diadoc")
        ids = KonturDiadocClient.parse_post_message_ids(res.data)
        # kl-id may appear later in OuterDocflow; store message/entity for status poll
        st = repository.upsert_ozon_edo_document(
            user_id=owner_id,
            supply_order_id=supply_order_id,
            doc_type=doc_type,
            channel="diadoc",
            message_id=ids.get("message_id") or "",
            entity_id=ids.get("entity_id") or "",
            status="sent",
            status_label="Отправлено в Diadoc (ожидание статусов ГИС ЭПД)",
            last_error="",
            raw_json=_jj.dumps(res.data, ensure_ascii=False)[:8000] if isinstance(res.data, (dict, list)) else (res.raw[:8000]),
            mark_sent=True,
        )
        return {"ok": True, "doc_type": doc_type, "channel": "diadoc", "document": st}

    @app.get("/api/ozon-supplies/{supply_order_id}/edo/status")
    def get_ozon_edo_status(request: Request, supply_order_id: int) -> dict[str, object]:
        """Проверка стадии документов ЭДО по поставке."""
        import json as _jj
        from .kontur_logistics import KonturLogisticsClient, status_label
        from .kontur_diadoc import KonturDiadocClient

        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        docs = repository.list_ozon_edo_documents(user_id=owner_id, supply_order_id=supply_order_id)
        settings = repository.get_supply_edo_settings(user_id=owner_id, include_secrets=True)
        refreshed: list[dict] = []
        for doc in docs:
            dtype = str(doc.get("doc_type") or "")
            channel = str(doc.get("channel") or "")
            tid = str(doc.get("transportation_id") or "").strip()
            if channel == "logistics" and tid and settings.get("api_key"):
                client = KonturLogisticsClient(
                    api_url=str(settings.get("api_url") or ""),
                    api_key=str(settings["api_key"]),
                )
                res = client.get_transportation(tid)
                if res.ok:
                    parsed = KonturLogisticsClient.parse_transportation_status(res.data if isinstance(res.data, dict) else {})
                    doc = repository.upsert_ozon_edo_document(
                        user_id=owner_id,
                        supply_order_id=supply_order_id,
                        doc_type=dtype,
                        channel="logistics",
                        transportation_id=parsed.get("transportation_id") or tid,
                        status=str(parsed.get("status") or ""),
                        status_label=str(parsed.get("status_label") or status_label(parsed.get("status"))),
                        mintrans_id=str(parsed.get("mintrans_id") or ""),
                        mintrans_status=str(parsed.get("mintrans_status") or ""),
                        last_error=str(parsed.get("mintrans_errors") or ""),
                        raw_json=_jj.dumps(res.data, ensure_ascii=False)[:8000],
                    )
                else:
                    doc = repository.upsert_ozon_edo_document(
                        user_id=owner_id,
                        supply_order_id=supply_order_id,
                        doc_type=dtype,
                        channel="logistics",
                        transportation_id=tid,
                        status=str(doc.get("status") or "unknown"),
                        status_label=str(doc.get("status_label") or ""),
                        last_error=res.error or f"HTTP {res.status_code}",
                    )
            elif channel == "diadoc" and doc.get("message_id") and settings.get("diadoc_client_id"):
                # Pull document meta for stage hints (LastOuterDocflows / DocflowStatus)
                dclient = KonturDiadocClient(
                    api_url=str(settings.get("diadoc_url") or ""),
                    client_id=str(settings.get("diadoc_client_id") or ""),
                    login=str(settings.get("diadoc_login") or ""),
                    password=str(settings.get("diadoc_password") or ""),
                )
                box_id = str(settings.get("diadoc_from_box_id") or "")
                entity_id = str(doc.get("entity_id") or "")
                if box_id and entity_id:
                    res = dclient.get_document(
                        box_id=box_id,
                        message_id=str(doc.get("message_id") or ""),
                        entity_id=entity_id,
                    )
                    if res.ok and isinstance(res.data, dict):
                        parsed = KonturDiadocClient.parse_document_status(res.data)
                        tid = str(doc.get("transportation_id") or parsed.get("kl_id") or "")
                        doc = repository.upsert_ozon_edo_document(
                            user_id=owner_id,
                            supply_order_id=supply_order_id,
                            doc_type=dtype,
                            channel="diadoc",
                            transportation_id=tid,
                            message_id=str(doc.get("message_id") or ""),
                            entity_id=entity_id,
                            status=str(parsed.get("status") or "sent"),
                            status_label=str(parsed.get("status_label") or ""),
                            mintrans_id=str(parsed.get("mintrans_id") or ""),
                            mintrans_status=str(parsed.get("mintrans_status") or ""),
                            last_error="",
                            raw_json=_jj.dumps(res.data, ensure_ascii=False)[:8000],
                        )
            refreshed.append(doc)
        diadoc_ready = bool(
            settings.get("diadoc_client_id")
            and settings.get("diadoc_login")
            and settings.get("has_diadoc_password")
            and settings.get("diadoc_from_box_id")
            and settings.get("diadoc_to_box_id")
        )
        return {
            "supply_order_id": supply_order_id,
            "documents": refreshed or docs,
            "edo_configured": bool(settings.get("has_api_key") or diadoc_ready),
        }

    @app.get("/api/ozon-supplies/{supply_order_id}/ttn.pdf")
    def get_ozon_ttn_pdf_ep(request: Request, supply_order_id: int) -> "Response":
        import subprocess as _sp, tempfile as _tf, pathlib as _pl, os as _os
        import io as _io, zipfile as _zf
        from fastapi.responses import Response
        from urllib.parse import quote as _qp
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        result = _build_ozon_ttn_docx(request, supply_order_id, _supply_owner_id(user))
        if result is None:
            raise HTTPException(status_code=404, detail="Поставка не найдена")
        docx_bytes, supply_num, supplier_short, supply_date_disp, wh_name = result
        tmp_dir   = _tf.mkdtemp()
        docx_path = _pl.Path(tmp_dir) / f"ozon_ttn_{supply_order_id}.docx"
        pdf_path  = _pl.Path(tmp_dir) / f"ozon_ttn_{supply_order_id}.pdf"
        docx_path.write_bytes(docx_bytes)
        env = dict(_os.environ)
        env.update({"HOME": "/tmp", "XDG_CACHE_HOME": "/tmp/.cache",
                    "XDG_CONFIG_HOME": "/tmp/.config", "DCONF_PROFILE": "empty"})
        _sp.run(["soffice", "--headless", "--convert-to", "pdf", "--outdir", str(tmp_dir), str(docx_path)],
                capture_output=True, env=env, timeout=60)
        if not pdf_path.exists():
            raise HTTPException(status_code=500, detail="LibreOffice не смог сгенерировать PDF")
        fname = f"ТТН №{supply_num}, {supplier_short} от {supply_date_disp}, {wh_name}.pdf"
        return Response(content=pdf_path.read_bytes(), media_type="application/pdf",
                        headers={"Content-Disposition": f"inline; filename*=UTF-8''{_qp(fname)}"})

    @app.patch("/api/ozon-supplies/{supply_order_id}/manual-fields")
    def update_ozon_manual_fields(request: Request, supply_order_id: int, payload: dict) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        repository.update_ozon_supply_manual_fields(
            user_id=_supply_owner_id(user), supply_order_id=supply_order_id,
            pallets_count=str(payload.get("pallets_count") or ""),
            driver_name=str(payload.get("driver_name") or ""),
            notes=str(payload.get("notes") or ""),
            production=str(payload.get("production") or ""),
        )
        return {"ok": True}

    @app.get("/api/ozon-supplies/sync/status")
    def get_ozon_sync_status(request: Request) -> dict[str, object]:
        _require_user(request)
        with _ozon_sync_lock:
            return dict(_ozon_sync_state)

    @app.post("/api/ozon-supplies/sync/stop")
    def stop_ozon_sync(request: Request) -> dict[str, object]:
        _require_user(request)
        with _ozon_sync_lock:
            if _ozon_sync_state.get("in_progress"):
                _ozon_sync_state["cancel_requested"] = True
                return {"ok": True, "message": "Остановка синхронизации OZON…"}
        return {"ok": False, "message": "Синхронизация не запущена"}

    @app.post("/api/ozon-supplies/sync")
    def sync_ozon_supplies(request: Request) -> dict[str, object]:
        import threading as _thr, urllib.request as _ul, json as _jj, ssl as _sl
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()  # run migrations including supplier_name column
        with _ozon_sync_lock:
            if _ozon_sync_state.get("in_progress"):
                return {"ok": False, "message": "Синхронизация уже запущена"}

        # Get OZON sources only
        sources = [s for s in repository.list_supply_sources(user_id=owner_id)
                   if (s.get("marketplace") or "wb").lower() == "ozon" and s.get("is_enabled")]
        if not sources:
            return {"ok": False, "message": "Нет активных источников OZON"}

        # v3 API states (confirmed by live testing — v2 states return 404)
        # Keep list for sync/purge must include all non-cancelled progressive states,
        # otherwise local rows in e.g. ACCEPTANCE_AT_STORAGE_WAREHOUSE get wiped.
        ACTIVE_STATES = [
            "DATA_FILLING",
            "READY_TO_SUPPLY",
            "ACCEPTED_AT_SUPPLY_WAREHOUSE",
            "IN_TRANSIT",
            "ACCEPTANCE_AT_STORAGE_WAREHOUSE",
            "REPORTS_CONFIRMATION_AWAITING",
            "REPORT_REJECTED",
            "COMPLETED",
        ]
        # Composition can still change in Ozon LK while supply is being filled / ready.
        # Later states are frozen — skip bundle re-fetch there for sync speed.
        COMPOSITION_REFRESH_STATES = frozenset({"DATA_FILLING", "READY_TO_SUPPLY"})
        import threading as _thr
        from datetime import datetime as _odt, timezone as _otz, timedelta as _otd
        _ozon_now = _odt.now(_otz.utc)
        _ozon_date_from = (_ozon_now - _otd(days=90)).strftime("%Y-%m-%d")
        _ozon_today = _ozon_now.strftime("%Y-%m-%d")

        def _run_ozon_sync():
            with _ozon_sync_lock:
                _ozon_sync_state.update({
                    "in_progress": True,
                    "synced": 0,
                    "total": 0,
                    "errors": [],
                    "processed_sources": [],
                    "failed_sources": [],
                    "message": "Запуск…",
                    "cancel_requested": False,
                })
            total_synced = 0
            errors: list[str] = []
            processed_sources: list[dict[str, object]] = []
            failed_sources: list[dict[str, object]] = []
            try:
                for src in sources:
                    src_name = str(src.get("name") or "?")
                    src_synced_before = total_synced
                    # list_supply_sources strips the real key — fetch it properly
                    src_full = repository.get_supply_source_with_key(
                        user_id=owner_id, source_id=int(src["id"])
                    )
                    if not src_full:
                        continue
                    api_key = str(src_full.get("api_key") or "")
                    client_id = str(src_full.get("client_id") or "")
                    if not api_key or not client_id:
                        continue
                    ctx = _sl.create_default_context()
                    headers = {"Client-Id": client_id, "Api-Key": api_key,
                               "Content-Type": "application/json", "User-Agent": "Mozilla/5.0"}

                    # Get default legal entity for OZON (prefer ООО, fallback to first)
                    _legal_entities = repository.list_supply_legal_entities(user_id=owner_id)
                    _ozon_le = (
                        next((e for e in _legal_entities if "ООО" in str(e.get("short_name") or "")), None)
                        or (_legal_entities[0] if _legal_entities else None)
                    )
                    _default_supplier = str(_ozon_le.get("short_name") or "") if _ozon_le else None

                    # Step 0: build cluster cache (macrolocal_cluster_id → name)
                    cluster_cache: dict[int, str] = {}
                    try:
                        req_cl = _ul.Request(
                            "https://api-seller.ozon.ru/v2/cluster/list",
                            data=_jj.dumps({"cluster_ids": []}).encode(), method="POST", headers=headers,
                        )
                        with _ul.urlopen(req_cl, context=ctx, timeout=15) as r_cl:
                            cl_data = _jj.loads(r_cl.read())
                        for item in (cl_data.get("result") or []):
                            cid = int(item.get("macrolocal_cluster_id") or 0)
                            cname = str((item.get("data") or {}).get("macrolocal_cluster", {}).get("name") or "")
                            if cid and cname:
                                cluster_cache[cid] = cname
                    except Exception as ex:
                        _log.warning("ozon cluster cache: %s", ex)

                    # Pre-load existing quantities — skip bundle API for frozen supplies (speed).
                    # Editable states (DATA_FILLING / READY_TO_SUPPLY) always refresh composition:
                    # seller can change goods in Ozon LK after the first sync.
                    # Vehicle/cargoes are always refreshed below.
                    existing_items = repository.list_ozon_supply_items(
                        user_id=owner_id, source_id=int(src["id"])
                    )
                    existing_qty: dict[int, int] = {
                        int(item.get("supply_order_id") or 0): int(item.get("total_quantity") or 0)
                        for item in existing_items
                        if item.get("supply_order_id")
                    }

                    # Step 1: paginate supply order IDs via /v3/supply-order/list
                    # Pagination is cursor-based: last_id string returned in response
                    with _ozon_sync_lock:
                        _ozon_sync_state["message"] = f"«{src['name']}»: получение списка поставок…"
                    all_order_ids: list[str] = []
                    last_id_cursor = ""
                    list_ok = True
                    while True:
                        req_body = {"filter": {"states": ACTIVE_STATES}, "limit": 100, "sort_by": 1}
                        if last_id_cursor:
                            req_body["last_id"] = last_id_cursor
                        try:
                            req = _ul.Request(
                                "https://api-seller.ozon.ru/v3/supply-order/list",
                                data=_jj.dumps(req_body).encode(), method="POST", headers=headers,
                            )
                            with _ul.urlopen(req, context=ctx, timeout=20) as r:
                                resp = _jj.loads(r.read())
                            page_ids = [str(x) for x in (resp.get("order_ids") or [])]
                            if not page_ids:
                                break
                            all_order_ids.extend(page_ids)
                            last_id_cursor = str(resp.get("last_id") or "")
                            if not last_id_cursor or len(page_ids) < 100:
                                break
                        except Exception as ex:
                            list_ok = False
                            code = getattr(ex, "code", None)
                            _log.error("ozon sync list error src=%s: %s", src_name, ex, exc_info=True)
                            if code == 403:
                                # Typical for returns-only / limited API keys — skip source, keep syncing others.
                                err_text = "нет прав на поставки (403 Forbidden)"
                            else:
                                err_text = f"list: {ex}"
                            errors.append(f"«{src_name}»: {err_text}")
                            failed_sources.append({"name": src_name, "error": err_text})
                            break

                    # List failed (e.g. 403) — do not treat as processed / synced.
                    if not list_ok:
                        continue

                    # Purge local rows for cancelled/deleted Ozon orders.
                    # Sync list excludes CANCELLED; orders deleted in Ozon LK disappear from
                    # the active list but previously stayed forever as DATA_FILLING locally.
                    keep_ids = [int(x) for x in all_order_ids if str(x).isdigit() and int(x) > 0]
                    try:
                        deleted_count = repository.delete_ozon_supply_items_not_in(
                            source_id=int(src["id"]),
                            keep_order_ids=keep_ids,
                            delete_all_if_empty=True,
                        )
                        if deleted_count:
                            _log.info(
                                "ozon sync: removed %d cancelled/missing supplies for source %s",
                                deleted_count,
                                src.get("id"),
                            )
                    except Exception as del_ex:
                        _log.error("ozon sync purge error src=%s: %s", src_name, del_ex, exc_info=True)
                        errors.append(f"«{src_name}»: purge: {del_ex}")

                    if not all_order_ids:
                        try:
                            repository.mark_supply_source_synced(source_id=int(src["id"]))
                        except Exception:
                            pass
                        processed_sources.append({"name": src_name, "synced": 0})
                        continue

                    with _ozon_sync_lock:
                        _ozon_sync_state["total"] = len(all_order_ids)
                        _ozon_sync_state["message"] = f"«{src['name']}»: загрузка {len(all_order_ids)} поставок…"

                    # Step 2: get details in batches of 50 via /v3/supply-order/get
                    for i in range(0, len(all_order_ids), 50):
                        batch = all_order_ids[i:i+50]
                        order_ids_in_batch: list[int] = []
                        qty_batch: list = []  # (order_id, bundle_id, item_id)
                        order_supply_ids: dict[int, int] = {}  # order_id → OZON supply_id for cargoes
                        try:
                            body2 = _jj.dumps({"order_ids": batch}).encode()
                            req2 = _ul.Request(
                                "https://api-seller.ozon.ru/v3/supply-order/get",
                                data=body2, method="POST", headers=headers,
                            )
                            with _ul.urlopen(req2, context=ctx, timeout=30) as r2:
                                det_resp = _jj.loads(r2.read())
                            for order in (det_resp.get("orders") or []):
                                oid = int(order.get("order_id") or 0)
                                if not oid: continue
                                # Filter: created in last 30 days OR delivery scheduled in the future
                                created_date = str(order.get("created_date") or "")
                                ts_outer_f = order.get("timeslot") or {}
                                ts_inner_f = ts_outer_f.get("timeslot") or {}
                                timeslot_date = str(ts_inner_f.get("from") or "")[:10]
                                is_recent = created_date[:10] >= _ozon_date_from
                                is_future_delivery = bool(timeslot_date) and timeslot_date >= _ozon_today
                                if not is_recent and not is_future_delivery:
                                    continue
                                # supply_date: order.timeslot.timeslot.from
                                supply_date = None
                                ts_outer = order.get("timeslot") or {}
                                ts_inner = ts_outer.get("timeslot") or {}
                                if ts_inner.get("from"):
                                    supply_date = str(ts_inner["from"])
                                # warehouse resolution: crossdock → storage_warehouse is final dest
                                supplies_list = order.get("supplies") or []
                                is_crossdock = any(s.get("is_crossdock") for s in supplies_list)
                                dropoff_obj = order.get("drop_off_warehouse") or {}
                                dropoff_id = int(dropoff_obj.get("warehouse_id") or 0) or None
                                dropoff_name = str(dropoff_obj.get("name") or "") or None
                                if is_crossdock and supplies_list:
                                    supply0 = supplies_list[0]
                                    storage_obj = supply0.get("storage_warehouse") or {}
                                    final_name = str(storage_obj.get("name") or "") or None
                                    # If storage_warehouse is null, resolve via macrolocal_cluster_id
                                    if not final_name:
                                        cid = int(supply0.get("macrolocal_cluster_id") or 0)
                                        if cid and cid in cluster_cache:
                                            final_name = cluster_cache[cid]
                                    wh_id = dropoff_id
                                    wh_name = final_name or dropoff_name
                                    transit_wh_name = dropoff_name
                                else:
                                    wh_id = dropoff_id
                                    wh_name = dropoff_name
                                    transit_wh_name = None
                                data = {
                                    "supply_order_id": oid,
                                    "supply_order_number": str(order.get("order_number") or ""),
                                    "state": str(order.get("state") or ""),
                                    "creation_date": created_date or None,
                                    "supply_date": supply_date,
                                    "dropoff_warehouse_id": wh_id,
                                    "warehouse_name": wh_name,
                                    "transit_warehouse_name": transit_wh_name,
                                    "is_crossdock": is_crossdock,
                                    "total_quantity": 0,
                                    "creation_flow": None,
                                    "supplier_name": _default_supplier,
                                    "supplies": supplies_list,
                                }
                                item_id = repository.upsert_ozon_supply_item(source_id=int(src["id"]), data=data)
                                order_ids_in_batch.append(oid)
                                for s in supplies_list:
                                    sid = int(s.get("supply_id") or 0)
                                    if sid > 0:
                                        order_supply_ids[oid] = sid
                                        break
                                # Bundle: first load, or editable states where LK composition may change
                                state = str(order.get("state") or "")
                                need_bundle = (
                                    existing_qty.get(oid, 0) == 0
                                    or state in COMPOSITION_REFRESH_STATES
                                )
                                if need_bundle:
                                    for s in supplies_list:
                                        bid = str(s.get("bundle_id") or "")
                                        if bid:
                                            qty_batch.append((oid, bid, item_id))
                                            break  # one bundle per order
                                total_synced += 1
                        except Exception as ex:
                            _log.error("ozon supply get batch: %s", ex, exc_info=True)
                            errors.append(str(ex))

                        # Load quantities for this batch via bundle API (1 call per order)
                        import time as _t
                        if qty_batch:
                            with _ozon_sync_lock:
                                _ozon_sync_state["message"] = f"«{src['name']}»: количество товаров ({len(qty_batch)})…"
                        for (order_id, bundle_id, item_id) in qty_batch:
                            if _ozon_sync_state.get("cancel_requested"):
                                break
                            for attempt in range(3):
                                try:
                                    bdy = _jj.dumps({"bundle_ids": [bundle_id], "limit": 100, "last_id": ""}).encode()
                                    bq_req = _ul.Request(
                                        "https://api-seller.ozon.ru/v1/supply-order/bundle",
                                        data=bdy, method="POST", headers=headers,
                                    )
                                    with _ul.urlopen(bq_req, context=ctx, timeout=15) as bq_r:
                                        bq_resp = _jj.loads(bq_r.read())
                                    goods = bq_resp.get("items") or []
                                    total_qty = sum(int(g.get("quantity") or 0) for g in goods)
                                    # Always replace local composition after a successful bundle fetch
                                    # (including empty / reduced qty — otherwise LK edits stay stale).
                                    if item_id:
                                        repository.upsert_ozon_supply_goods(
                                            supply_item_id=item_id, goods=goods
                                        )
                                    repository.update_ozon_supply_total_quantity(
                                        supply_order_id=order_id, total_quantity=total_qty
                                    )
                                    break  # success
                                except Exception as bex:
                                    code = getattr(bex, "code", None)
                                    if code == 429:
                                        wait = (attempt + 1) * 2  # 2s, 4s, 6s
                                        _log.info("ozon bundle 429 order_id=%d, retry in %ds", order_id, wait)
                                        _t.sleep(wait)
                                    else:
                                        _log.warning("ozon bundle qty order_id=%d: %s", order_id, bex)
                                        break
                            _t.sleep(0.4)  # 400ms between calls ≈ 2.5 req/s

                        with _ozon_sync_lock:
                            _ozon_sync_state["synced"] = total_synced
                            if _ozon_sync_state.get("cancel_requested"):
                                break

                        # Always refresh vehicle/driver + cargo places
                        if order_ids_in_batch:
                            with _ozon_sync_lock:
                                _ozon_sync_state["message"] = (
                                    f"«{src['name']}»: водители и места 0/{len(order_ids_in_batch)}…"
                                )
                        for vi, order_id in enumerate(order_ids_in_batch):
                            if _ozon_sync_state.get("cancel_requested"):
                                break
                            if vi % 5 == 0:
                                with _ozon_sync_lock:
                                    _ozon_sync_state["message"] = (
                                        f"«{src['name']}»: водители и места {vi}/{len(order_ids_in_batch)}…"
                                    )
                            try:
                                vbody = _jj.dumps({"order_id": order_id}).encode()
                                v_req = _ul.Request(
                                    "https://api-seller.ozon.ru/v1/supply-order/details",
                                    data=vbody, method="POST", headers=headers,
                                )
                                with _ul.urlopen(v_req, context=ctx, timeout=10) as vr:
                                    vdet = _jj.loads(vr.read())
                                vval = (vdet.get("vehicle") or {}).get("value") or {}
                                repository.update_ozon_supply_vehicle(
                                    supply_order_id=order_id,
                                    vehicle_json=_jj.dumps(vval, ensure_ascii=False))
                                actual_supply_id = order_supply_ids.get(order_id) or 0
                                if not actual_supply_id:
                                    for s in (vdet.get("supplies") or []):
                                        sid = int(s.get("supply_id") or 0)
                                        if sid > 0:
                                            actual_supply_id = sid
                                            break
                                if actual_supply_id:
                                    from .ozon_etrn import build_ozon_cargoes_cache as _build_cargo_cache
                                    cbody = _jj.dumps({"supply_ids": [actual_supply_id]}).encode()
                                    c_req = _ul.Request(
                                        "https://api-seller.ozon.ru/v1/cargoes/get",
                                        data=cbody, method="POST", headers=headers,
                                    )
                                    with _ul.urlopen(c_req, context=ctx, timeout=10) as cr:
                                        cdet = _jj.loads(cr.read())
                                    cargoes = []
                                    for s in (cdet.get("supply") or []):
                                        cargoes.extend(s.get("cargoes") or [])
                                    supplies_cargoes: list = []
                                    try:
                                        t_req = _ul.Request(
                                            "https://api-seller.ozon.ru/v1/cargoes/supplies/get",
                                            data=cbody, method="POST", headers=headers,
                                        )
                                        with _ul.urlopen(t_req, context=ctx, timeout=10) as tr:
                                            tdet = _jj.loads(tr.read())
                                        supplies_cargoes = list(tdet.get("supplies_cargoes") or [])
                                    except Exception:
                                        supplies_cargoes = []
                                    cache_obj = _build_cargo_cache(
                                        flat_cargoes=cargoes,
                                        supplies_cargoes=supplies_cargoes,
                                    )
                                    repository.update_ozon_supply_cargoes(
                                        supply_order_id=order_id,
                                        cargoes_json=_jj.dumps(cache_obj, ensure_ascii=False))
                            except Exception as vex:
                                code = getattr(vex, "code", None)
                                if code == 403:
                                    break  # No role for this key — skip all vehicle fetches
                                elif code == 429:
                                    _t.sleep(2)
                                else:
                                    _log.debug("ozon vehicle/cargoes order_id=%d: %s", order_id, vex)
                            _t.sleep(0.25)

                    # Mark source as synced (updates last_synced_at timestamp)
                    try:
                        repository.mark_supply_source_synced(source_id=int(src["id"]))
                    except Exception:
                        pass
                    processed_sources.append({
                        "name": src_name,
                        "synced": max(0, total_synced - src_synced_before),
                    })
                    with _ozon_sync_lock:
                        if _ozon_sync_state.get("cancel_requested"):
                            break
            finally:
                cancelled = bool(_ozon_sync_state.get("cancel_requested"))
                with _ozon_sync_lock:
                    if cancelled:
                        msg = f"Остановлено. Загружено {total_synced} поставок OZON."
                    else:
                        # Keep success line clean — errors go to failed_sources for the UI.
                        msg = f"Готово. Загружено {total_synced} поставок OZON."
                    _ozon_sync_state.update({
                        "in_progress": False,
                        "synced": total_synced,
                        "errors": errors,
                        "processed_sources": processed_sources,
                        "failed_sources": failed_sources,
                        "cancel_requested": False,
                        "message": msg,
                    })

        _thr.Thread(target=_run_ozon_sync, daemon=True).start()
        return {"ok": True, "message": "Синхронизация OZON запущена"}

    @app.delete("/api/ozon-supplies")
    def clear_ozon_supplies(request: Request) -> dict[str, object]:
        user = _require_user(request)
        if str(user.get("role") or "") not in ROLE_CAN_ACCESS_SETTINGS:
            raise HTTPException(status_code=403, detail="Нет доступа")
        deleted = repository.clear_ozon_supply_items(user_id=_supply_owner_id(user))
        return {"ok": True, "deleted": deleted}

    # ── OZON return giveout barcode (Получение возвратов) ───────────────────

    def _ozon_giveout_post(
        *,
        client_id: str,
        api_key: str,
        path: str,
        payload: dict | None = None,
        timeout: int = 30,
    ) -> dict:
        """POST to Ozon return/giveout API. Returns parsed JSON dict."""
        import json as _jj
        import ssl as _sl
        import urllib.error as _ue
        import urllib.request as _ul

        body = _jj.dumps(payload if payload is not None else {}).encode("utf-8")
        headers = {
            "Client-Id": client_id,
            "Api-Key": api_key,
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0",
        }
        req = _ul.Request(
            f"https://api-seller.ozon.ru{path}",
            data=body,
            method="POST",
            headers=headers,
        )
        ctx = _sl.create_default_context()
        try:
            with _ul.urlopen(req, context=ctx, timeout=timeout) as resp:
                raw = resp.read()
        except _ue.HTTPError as exc:
            err_body = ""
            try:
                err_body = exc.read().decode("utf-8", errors="replace")[:500]
            except Exception:
                pass
            raise HTTPException(
                status_code=400,
                detail=f"Ozon API {path}: HTTP {exc.code}" + (f" — {err_body}" if err_body else ""),
            ) from exc
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Ozon API {path}: {exc}") from exc
        if not raw:
            return {}
        try:
            data = _jj.loads(raw.decode("utf-8"))
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Ozon API {path}: некорректный JSON") from exc
        return data if isinstance(data, dict) else {}

    def _ozon_decode_file_content(
        value: object,
        *,
        expect: str = "png",
    ) -> bytes | None:
        """Decode Ozon base64 payload and validate magic bytes."""
        import base64 as _b64

        if value is None:
            return None
        if isinstance(value, (bytes, bytearray)):
            raw = bytes(value)
        else:
            text = str(value).strip()
            if not text:
                return None
            if text.startswith("data:") and "," in text:
                text = text.split(",", 1)[1]
            try:
                raw = _b64.b64decode(text, validate=False)
            except Exception:
                try:
                    raw = text.encode("latin-1")
                except Exception:
                    return None
        if not raw:
            return None
        # Reject false-positive base64 that is not a real image/pdf.
        if expect == "pdf":
            if not raw.startswith(b"%PDF"):
                return None
        else:
            if not raw.startswith(b"\x89PNG\r\n\x1a\n"):
                return None
        return raw

    def _ozon_extract_giveout_bytes(data: dict, *, expect: str = "png") -> bytes | None:
        """Extract PNG/PDF bytes from giveout API response.

        Live Seller API returns ``{"png": "<base64>"}`` / ``{"pdf": "<base64>"}``.
        Older docs also mention ``file_content`` — support both.
        """
        if not isinstance(data, dict):
            return None
        if expect == "pdf":
            keys = ("pdf", "file_content", "barcode_pdf")
        else:
            keys = ("png", "file_content", "barcode_png")
        for key in keys:
            raw = _ozon_decode_file_content(data.get(key), expect=expect)
            if raw:
                return raw
        return None

    def _ozon_find_giveout_source(owner_id: int) -> dict:
        """Pick first enabled Ozon supply source where giveout is enabled."""
        sources = [
            s
            for s in repository.list_supply_sources(user_id=owner_id)
            if (s.get("marketplace") or "wb").lower() == "ozon" and s.get("is_enabled")
        ]
        if not sources:
            raise HTTPException(status_code=400, detail="Нет активных источников OZON")
        last_err = "Нет источника OZON с доступом к получению возвратов"
        for src in sources:
            src_full = repository.get_supply_source_with_key(
                user_id=owner_id, source_id=int(src["id"])
            )
            if not src_full:
                continue
            api_key = str(src_full.get("api_key") or "").strip()
            client_id = str(src_full.get("client_id") or "").strip()
            if not api_key or not client_id:
                continue
            try:
                enabled_data = _ozon_giveout_post(
                    client_id=client_id,
                    api_key=api_key,
                    path="/v1/return/giveout/is-enabled",
                )
            except HTTPException as exc:
                last_err = str(exc.detail)
                continue
            if not bool(enabled_data.get("enabled")):
                last_err = (
                    f"У источника «{src_full.get('name') or src['id']}» "
                    "получение возвратов по штрихкоду недоступно"
                )
                continue
            return {
                "id": int(src_full["id"]),
                "name": str(src_full.get("name") or ""),
                "client_id": client_id,
                "api_key": api_key,
            }
        raise HTTPException(status_code=400, detail=last_err)

    def _ozon_fetch_giveout_list(client_id: str, api_key: str) -> list[dict[str, object]]:
        """Fetch returns waiting at drop-off points for the barcode modal table.

        Portal «Список возвратов» comes from ``/v1/returns/company/fbs/info``
        (drop-off points with returns_count), not from empty ``giveout/list``.
        """
        rows: list[dict[str, object]] = []
        last_id = 0
        for _ in range(20):
            pagination: dict[str, object] = {"limit": 100}
            if last_id:
                pagination["last_id"] = last_id
            data = _ozon_giveout_post(
                client_id=client_id,
                api_key=api_key,
                path="/v1/returns/company/fbs/info",
                payload={"filter": {}, "pagination": pagination},
            )
            points = data.get("drop_off_points") or []
            if not isinstance(points, list) or not points:
                break
            page_last_id = last_id
            last_item = points[-1]
            if isinstance(last_item, dict):
                try:
                    page_last_id = int(last_item.get("id") or last_id)
                except Exception:
                    page_last_id = last_id
            for p in points:
                if not isinstance(p, dict):
                    continue
                try:
                    qty_n = int(p.get("returns_count") or 0)
                except Exception:
                    qty_n = 0
                try:
                    pid = int(p.get("id") or p.get("place_id") or 0)
                except Exception:
                    pid = 0
                rows.append(
                    {
                        "giveout_id": pid,
                        "warehouse_name": str(p.get("name") or "").strip(),
                        "warehouse_address": str(p.get("address") or "").strip(),
                        "quantity": qty_n,
                    }
                )
            if page_last_id == last_id:
                break
            last_id = page_last_id
            if not bool(data.get("has_next")) or len(points) < 100:
                break
        return rows

    def _ozon_giveout_moscow_tz():
        from zoneinfo import ZoneInfo

        try:
            return ZoneInfo("Europe/Moscow")
        except Exception:
            from datetime import timezone as _tz

            return _tz(timedelta(hours=3))

    def _ozon_parse_iso_dt(value: object):
        from datetime import datetime as _dt

        raw = str(value or "").strip()
        if not raw:
            return None
        try:
            dt = _dt.fromisoformat(raw.replace("Z", "+00:00"))
        except Exception:
            return None
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=UTC)
        return dt

    def _ozon_giveout_valid_until_label(valid_until_iso: str = "") -> str:
        """Format giveout validity like the modal: '4 августа, 15:38' (Moscow)."""
        from datetime import datetime as _dt

        months = (
            "",
            "января",
            "февраля",
            "марта",
            "апреля",
            "мая",
            "июня",
            "июля",
            "августа",
            "сентября",
            "октября",
            "ноября",
            "декабря",
        )
        tz = _ozon_giveout_moscow_tz()
        dt = _ozon_parse_iso_dt(valid_until_iso)
        if dt is None:
            dt = _dt.now(tz) + timedelta(hours=24)
        else:
            dt = dt.astimezone(tz)
        return f"{dt.day} {months[dt.month]}, {dt.strftime('%H:%M')}"

    @app.post("/api/ozon-returns/giveout")
    async def ozon_returns_giveout_refresh(request: Request) -> dict[str, object]:
        """Return giveout barcode modal payload; reset only when needed or forced."""
        import base64 as _b64
        from datetime import datetime as _dt

        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()

        force = False
        try:
            payload = await request.json()
            if isinstance(payload, dict):
                force = bool(payload.get("force"))
        except Exception:
            force = False

        src = _ozon_find_giveout_source(owner_id)
        client_id = src["client_id"]
        api_key = src["api_key"]
        source_id = int(src["id"])

        now_utc = _dt.now(UTC)
        stored_reset_raw = repository.get_ozon_giveout_barcode_reset_at(
            user_id=owner_id, source_id=source_id
        )
        stored_reset_at = _ozon_parse_iso_dt(stored_reset_raw)

        # Validity window: 24h from last reset. Auto-reset when unknown or ≤6h left.
        should_reset = bool(force)
        if not should_reset:
            if stored_reset_at is None:
                should_reset = True
            else:
                remaining = (stored_reset_at + timedelta(hours=24)) - now_utc
                if remaining <= timedelta(hours=6):
                    should_reset = True

        png_bytes: bytes | None = None
        reset_at = stored_reset_at or now_utc
        did_reset = False
        if should_reset:
            reset_data = _ozon_giveout_post(
                client_id=client_id,
                api_key=api_key,
                path="/v1/return/giveout/barcode-reset",
            )
            png_bytes = _ozon_extract_giveout_bytes(reset_data, expect="png")
            reset_at = now_utc
            did_reset = True
            repository.set_ozon_giveout_barcode_reset_at(
                user_id=owner_id,
                source_id=source_id,
                reset_at=reset_at.isoformat(),
            )

        if not png_bytes:
            png_data = _ozon_giveout_post(
                client_id=client_id,
                api_key=api_key,
                path="/v1/return/giveout/get-png",
            )
            png_bytes = _ozon_extract_giveout_bytes(png_data, expect="png")
        if not png_bytes:
            raise HTTPException(status_code=400, detail="Не удалось получить изображение штрихкода")

        barcode = ""
        try:
            barcode_data = _ozon_giveout_post(
                client_id=client_id,
                api_key=api_key,
                path="/v1/return/giveout/barcode",
            )
            barcode = str(barcode_data.get("barcode") or "").strip()
        except Exception as ex:
            _log.warning("ozon giveout barcode text failed source=%s: %s", source_id, ex)
        try:
            giveouts = _ozon_fetch_giveout_list(client_id, api_key)
        except Exception as ex:
            _log.warning("ozon giveout list failed source=%s: %s", source_id, ex)
            giveouts = []

        valid_until_dt = reset_at + timedelta(hours=24)
        valid_until_label = _ozon_giveout_valid_until_label(valid_until_dt.isoformat())

        return {
            "ok": True,
            "source_id": source_id,
            "source_name": src["name"],
            "barcode": barcode,
            "barcode_png_base64": _b64.b64encode(png_bytes).decode("ascii"),
            "valid_until": valid_until_dt.isoformat(),
            "valid_until_label": valid_until_label,
            "reset": did_reset,
            "force": force,
            "giveouts": giveouts,
        }

    def _ozon_build_giveout_print_html(
        *,
        png_bytes: bytes,
        barcode: str = "",
        valid_until_label: str = "",
    ) -> str:
        """A4 print page: compact barcode (~2x bar height) + pickup instructions."""
        import base64 as _b64
        import html as _html
        import io as _io

        from PIL import Image as _PilImage

        # Validate image; print size is fixed in mm so PNG pixel size cannot blow up the page.
        _PilImage.open(_io.BytesIO(png_bytes)).convert("RGB")
        b64 = _b64.b64encode(png_bytes).decode("ascii")
        code = _html.escape(str(barcode or "").strip())
        valid = _html.escape(str(valid_until_label or "").strip())
        valid_html = (
            f'<div class="valid">Действует до {valid}</div>' if valid else ""
        )
        code_html = f'<div class="code">{code}</div>' if code else ""
        return f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8" />
  <title>Штрихкод получения возвратов Ozon</title>
  <style>
    @page {{ size: A4 portrait; margin: 14mm 16mm; }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: Arial, Helvetica, sans-serif;
      color: #0f172a;
      font-size: 14px;
      line-height: 1.45;
      background: #fff;
    }}
    .toolbar {{ margin: 0 0 12px; }}
    .toolbar button {{
      min-height: 36px; padding: 8px 12px; font-size: 14px;
      border: 1px solid #cbd5e1; border-radius: 8px; background: #fff; cursor: pointer;
    }}
    .sheet {{ max-width: 180mm; margin: 0 auto; }}
    .barcode-wrap {{
      display: flex;
      flex-direction: column;
      align-items: center;
      margin: 0 0 20px;
    }}
    .barcode-wrap img {{
      display: block;
      width: 110mm;
      max-width: 100%;
      /* ~2x typical Ozon bar height, compact sticker — not full page */
      height: 28mm;
      object-fit: fill;
      image-rendering: pixelated;
      background: #fff;
    }}
    .code {{
      margin-top: 10px;
      font-size: 16px;
      font-weight: 700;
      letter-spacing: 0.02em;
      font-variant-numeric: tabular-nums;
    }}
    .valid {{
      margin-top: 4px;
      font-size: 13px;
      color: #475569;
    }}
    .instructions {{
      margin: 0;
      padding: 0 0 0 20px;
    }}
    .instructions > li {{
      margin: 0 0 10px;
    }}
    .instructions ul {{
      margin: 6px 0 0;
      padding: 0 0 0 18px;
    }}
    .instructions ul li {{
      margin: 0 0 6px;
    }}
    .footer {{
      margin-top: 12px;
      color: #334155;
    }}
    @media print {{
      .no-print {{ display: none !important; }}
      body {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    }}
  </style>
</head>
<body>
  <div class="toolbar no-print"><button type="button" onclick="window.print()">Печать</button></div>
  <div class="sheet">
    <div class="barcode-wrap">
      <img src="data:image/png;base64,{b64}" alt="Штрихкод возвратов Ozon" />
      {code_html}
      {valid_html}
    </div>
    <ol class="instructions">
      <li>Проверьте срок действия штрихкода. Если он истёк или вы не успеете приехать вовремя, распечатайте новый.</li>
      <li>Покажите штрихкод сотруднику пункта выдачи, сортировочного центра или курьеру Ozon.</li>
      <li>Проверьте товары:
        <ul>
          <li>В пункте выдачи или в сортировочном центре — сравните выданные товары со списком в отчёте по возвратам. Если выдали не все возвраты, попросите у сотрудника акт об отказе в выдаче.</li>
          <li>У курьера — сравните выданные товары со списком в отчёте по возвратам. Если выдали не все возвраты, убедитесь, что курьер изменил список в приложении.</li>
        </ul>
      </li>
      <li>Снова покажите штрихкод сотруднику пункта выдачи заказов или курьеру Ozon. В сортировочном центре достаточно показать штрихкод только один раз.</li>
    </ol>
    <div class="footer">Все документы отправим на электронную почту продавца.</div>
  </div>
  <script>
    window.addEventListener("load", function () {{
      setTimeout(function () {{ window.print(); }}, 300);
    }});
  </script>
</body>
</html>"""

    @app.get("/api/ozon-returns/giveout/pdf")
    def ozon_returns_giveout_pdf(
        request: Request,
        source_id: int = 0,
        valid_until: str = "",
        valid_until_label: str = "",
    ) -> Response:
        """Print page: giveout barcode (~2x taller) + Ozon pickup instructions."""
        from fastapi.responses import HTMLResponse

        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        if source_id:
            src_full = repository.get_supply_source_with_key(user_id=owner_id, source_id=source_id)
            if not src_full or (src_full.get("marketplace") or "").lower() != "ozon":
                raise HTTPException(status_code=400, detail="Источник OZON не найден")
            api_key = str(src_full.get("api_key") or "").strip()
            client_id = str(src_full.get("client_id") or "").strip()
            if not api_key or not client_id:
                raise HTTPException(status_code=400, detail="Нет ключей источника OZON")
            src = {"id": source_id, "client_id": client_id, "api_key": api_key}
        else:
            src = _ozon_find_giveout_source(owner_id)

        png_data = _ozon_giveout_post(
            client_id=src["client_id"],
            api_key=src["api_key"],
            path="/v1/return/giveout/get-png",
        )
        png_bytes = _ozon_extract_giveout_bytes(png_data, expect="png")
        if not png_bytes:
            raise HTTPException(status_code=400, detail="Не удалось получить изображение штрихкода")

        barcode = ""
        try:
            barcode_data = _ozon_giveout_post(
                client_id=src["client_id"],
                api_key=src["api_key"],
                path="/v1/return/giveout/barcode",
            )
            barcode = str(barcode_data.get("barcode") or "").strip()
        except Exception as ex:
            _log.warning("ozon giveout print barcode text failed source=%s: %s", src["id"], ex)

        label = str(valid_until_label or "").strip()
        if not label:
            label = _ozon_giveout_valid_until_label(valid_until)

        try:
            html_doc = _ozon_build_giveout_print_html(
                png_bytes=png_bytes,
                barcode=barcode,
                valid_until_label=label,
            )
        except Exception as ex:
            _log.warning("ozon giveout print html failed: %s", ex)
            raise HTTPException(status_code=400, detail="Не удалось сформировать страницу штрихкода") from ex

        return HTMLResponse(content=html_doc)

    # ── End OZON Supplies ────────────────────────────────────────────────────

    @app.get("/api/supplies/combined-poa.pdf")
    def get_combined_poa_pdf(request: Request, ids: str = ""):
        """Generate combined PoA PDF for multiple supplies via LibreOffice."""
        import subprocess as _sp, tempfile as _tf, pathlib as _pl, os as _os
        import html as _hm, urllib.request as _ul, json as _jm, ssl as _sl
        from fastapi.responses import Response
        from datetime import datetime as _dtt

        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        supply_ids = [int(x) for x in ids.split(",") if x.strip().isdigit()]
        if not supply_ids:
            raise HTTPException(status_code=400, detail="Укажите ids поставок")

        # Collect all goods from all selected supplies (with lazy WB API fetch if not cached)
        entities = repository.list_supply_legal_entities(user_id=owner_id)
        drivers = repository.list_supply_drivers(user_id=owner_id)
        name_map = repository.get_product_name_by_article(user_id=owner_id)

        def _wb_get_goods(supply_id: int, api_key: str) -> list:
            ctx = _sl.create_default_context()
            req = _ul.Request(
                f"https://supplies-api.wildberries.ru/api/v1/supplies/{supply_id}/goods",
                method="GET", headers={"Authorization": api_key, "User-Agent": "Mozilla/5.0"}
            )
            try:
                with _ul.urlopen(req, context=ctx, timeout=15) as r:
                    return _jm.loads(r.read()) or []
            except Exception:
                return []

        all_goods = []
        ref_item = None
        for sid in supply_ids:
            item_row = repository.get_supply_item_row(user_id=owner_id, supply_id=sid)
            if item_row and ref_item is None:
                ref_item = dict(item_row)
            goods = repository.get_supply_goods(user_id=owner_id, supply_id=sid)
            # Lazy load from WB API if not cached
            if not goods and item_row:
                try:
                    src = repository.get_supply_source_with_key(
                        user_id=owner_id, source_id=int(item_row["source_id"])
                    )
                    if src and src.get("api_key"):
                        wb_goods = _wb_get_goods(sid, str(src["api_key"]))
                        if wb_goods and isinstance(wb_goods, list):
                            repository.upsert_supply_goods(
                                supply_item_id=int(item_row["id"]), goods=wb_goods
                            )
                            goods = repository.get_supply_goods(user_id=owner_id, supply_id=sid)
                except Exception as ex:
                    _log.warning("combined-poa lazy goods sid=%d: %s", sid, ex)
            for g in goods:
                vc = str(g.get("vendor_code") or "")
                g["product_name"] = name_map.get(vc) or vc or ""
            all_goods.extend(goods)

        if not ref_item:
            raise HTTPException(status_code=404, detail="Поставки не найдены")

        supplier_short = str(ref_item.get("supplier_name") or "").strip()
        le = (
            next((e for e in entities if str(e.get("short_name") or "").strip() == supplier_short), None)
            or next((e for e in entities if str(e.get("short_name") or "").strip().lower() == supplier_short.lower()), None)
            or next((e for e in entities if str(e.get("full_name") or "").strip() == supplier_short), None)
            or {}
        )
        org_full = le.get("full_name") or supplier_short
        org_line = ", ".join(filter(None, [org_full, le.get("requisites") or ""]))
        signatories = le.get("signatories") or supplier_short
        driver_name = str(ref_item.get("driver_name") or "")
        driver_obj = next((d for d in drivers if d.get("full_name") == driver_name), {})
        driver_docs = repository.driver_documents_line(driver_obj)

        now = _dtt.now()
        date_display = now.strftime("%d.%m.%Y")
        doc_num = f"{now.strftime('%d%m%Y')}_1"

        e = _hm.escape
        UL = "_" * 30
        sig_name = e(signatories) if signatories and signatories != "—" else ""
        _td = 'style="border:1px solid black;padding:2pt 4pt;font-size:9pt"'
        _tbl = 'border="1" cellspacing="0" width="100%" style="border-collapse:collapse;margin:0;table-layout:fixed;font-size:9pt"'

        goods_rows = "".join(
            f'<tr><td {_td} align="center">{i+1}</td><td {_td}>{e(g.get("product_name") or "Товар")}</td>'
            f'<td {_td} align="center">шт.</td><td {_td} align="center">{g.get("quantity") or "—"}</td></tr>'
            for i, g in enumerate(all_goods)
        )

        html_content = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<style>
@page{{size:210mm 297mm;margin:15mm 10mm 15mm 25mm}}
body{{font-family:"Times New Roman",serif;font-size:11pt;line-height:1.3}}
p{{margin:2pt 0}}tr{{page-break-inside:avoid}}
</style></head><body>
<table width="100%" cellspacing="0" cellpadding="0"><tr>
  <td width="55%" valign="top" style="font-size:11pt"><b>Организация:</b> {e(org_full)}</td>
  <td width="45%" valign="top" align="right" style="font-size:8pt">
    Типовая межотраслевая форма № М-2<br>Утверждена постановлением Госстата России от 30.10.97 № 71а
    <table border="1" cellspacing="0" cellpadding="1" align="right" style="font-size:7pt;margin-top:2pt">
      <tr><td colspan="2" align="center"><b>Коды</b></td></tr>
      <tr><td style="padding:1pt 3pt">Форма по ОКУД</td><td style="padding:1pt 3pt">0315001</td></tr>
      <tr><td style="padding:1pt 3pt">по ОКПО</td><td style="padding:1pt 3pt">&nbsp;&nbsp;&nbsp;&nbsp;</td></tr>
    </table>
  </td>
</tr></table>
<p align="center" style="font-size:14pt;margin:10pt 0 4pt"><b>Доверенность № {e(doc_num)}</b></p>
<p>Дата выдачи <b><u>{e(date_display)}</u></b></p>
<p>Доверенность действительна 14 дней с даты подписания.</p>
<p style="margin-top:4pt">{e(org_line)}</p><p style="font-size:8pt;text-align:center">наименование потребителя и его адрес</p>
<p style="margin-top:4pt">{e(org_line)}</p><p style="font-size:8pt;text-align:center">наименование плательщика и его адрес</p>
<p style="margin-top:6pt">Доверенность выдана &nbsp;<u>водителю</u>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; <u>{e(driver_name)}</u></p>
<p style="font-size:8pt">должность &nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp; фамилия, имя, отчество</p>
{f"<p>{e(driver_docs)}</p>" if driver_docs else ""}
<p style="margin-top:4pt">На отправку груза от &nbsp;<u>&nbsp;{e(supplier_short)}&nbsp;</u></p>
<p style="font-size:8pt;text-align:center">наименование поставщика</p>
<p style="margin-top:4pt">материальных ценностей по суммарным транспортным накладным поставок: {", ".join(str(x) for x in supply_ids)}</p>
<p style="margin-top:6pt">Перечень материальных ценностей, подлежащих доставке</p>
<table border="1" cellspacing="0" width="100%" style="border-collapse:collapse;margin:0;table-layout:fixed;font-size:9pt">
  <colgroup><col width="15%"><col width="45%"><col width="20%"><col width="20%"></colgroup>
  <tr style="page-break-inside:avoid">
    <th style="border:1px solid black;padding:2pt 4pt;font-size:8pt;font-weight:bold" align="center">Номер по порядку</th>
    <th style="border:1px solid black;padding:2pt 4pt;font-size:8pt;font-weight:bold" align="center">Материальные ценности</th>
    <th style="border:1px solid black;padding:2pt 4pt;font-size:8pt;font-weight:bold" align="center">Единица измерения</th>
    <th style="border:1px solid black;padding:2pt 4pt;font-size:8pt;font-weight:bold" align="center">Количество</th>
  </tr>
  {goods_rows}
</table>
<p style="margin-top:8pt">Подпись лица, получившего доверенность удостоверяем. &nbsp;&nbsp;&nbsp;&nbsp; {UL} &nbsp;&nbsp; ({e(driver_name)})</p>
<table width="100%" cellspacing="0" cellpadding="2" style="margin-top:6pt">
  <tr>
    <td width="25%" valign="bottom">Руководитель<br><small>М.П.</small></td>
    <td width="30%" valign="bottom" align="center">{UL}<br><small>подпись</small></td>
    <td width="45%" valign="bottom" align="center">{sig_name}<br><small>расшифровка подписи</small></td>
  </tr>
</table>
<table width="100%" cellspacing="0" cellpadding="2" style="margin-top:6pt">
  <tr>
    <td width="25%" valign="bottom">Главный бухгалтер</td>
    <td width="30%" valign="bottom" align="center">{UL}<br><small>подпись</small></td>
    <td width="45%" valign="bottom" align="center">{sig_name}<br><small>расшифровка подписи</small></td>
  </tr>
</table>
</body></html>"""

        tmp_dir = _tf.mkdtemp()
        html_path = _pl.Path(tmp_dir) / "combined_poa.html"
        pdf_path  = _pl.Path(tmp_dir) / "combined_poa.pdf"
        html_path.write_text(html_content, encoding="utf-8")
        lo_env = dict(_os.environ)
        for k,v in [("HOME",tmp_dir),("XDG_CACHE_HOME",tmp_dir),("XDG_CONFIG_HOME",tmp_dir),
                    ("XDG_RUNTIME_DIR",tmp_dir),("DCONF_PROFILE","/dev/null")]:
            lo_env[k]=v
        lo_ok=False
        for binary in ("/usr/bin/soffice","/usr/lib/libreoffice/program/soffice","soffice","libreoffice"):
            try:
                r=_sp.run([binary,"--headless","--norestore",f"-env:UserInstallation=file://{tmp_dir}/lo_profile","--convert-to","pdf","--outdir",tmp_dir,str(html_path)],capture_output=True,timeout=60,env=lo_env)
                if r.returncode==0 and pdf_path.exists(): lo_ok=True; break
            except FileNotFoundError: continue
            except _sp.TimeoutExpired: raise HTTPException(status_code=504,detail="Таймаут")
        if not lo_ok: raise HTTPException(status_code=500,detail="Ошибка конвертации PDF")
        return Response(content=pdf_path.read_bytes(),media_type="application/pdf",headers={"Content-Disposition":'inline; filename="combined_poa.pdf"'})

    @app.get("/api/supplies/combined-ttn.pdf")
    def get_combined_ttn_pdf(request: Request, ids: str = ""):
        """Generate combined TTN PDF for multiple supplies via LibreOffice (uses torg12_tpl.docx)."""
        import subprocess as _sp, tempfile as _tf, zipfile as _zf, io as _io
        import pathlib as _pl, os as _os, re as _re
        import html as _hm, urllib.request as _ul, json as _jm, ssl as _sl
        from fastapi.responses import Response
        from datetime import datetime as _dtt

        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        supply_ids = [int(x) for x in ids.split(",") if x.strip().isdigit()]
        if not supply_ids:
            raise HTTPException(status_code=400, detail="Укажите ids поставок")

        # Reuse the same logic as single TTN but aggregate all goods (with lazy WB API fetch)
        entities = repository.list_supply_legal_entities(user_id=owner_id)
        name_map = repository.get_product_name_by_article(user_id=owner_id)
        all_goods = []
        ref_item = None
        for sid in supply_ids:
            item_row = repository.get_supply_item_row(user_id=owner_id, supply_id=sid)
            if item_row and ref_item is None:
                ref_item = dict(item_row)
            goods = repository.get_supply_goods(user_id=owner_id, supply_id=sid)
            if not goods and item_row:
                try:
                    src = repository.get_supply_source_with_key(user_id=owner_id, source_id=int(item_row["source_id"]))
                    if src and src.get("api_key"):
                        ctx2 = _sl.create_default_context()
                        req2 = _ul.Request(f"https://supplies-api.wildberries.ru/api/v1/supplies/{sid}/goods",
                            method="GET", headers={"Authorization":str(src["api_key"]),"User-Agent":"Mozilla/5.0"})
                        with _ul.urlopen(req2, context=ctx2, timeout=15) as rr:
                            wb_g = _jm.loads(rr.read()) or []
                        if wb_g and isinstance(wb_g, list):
                            repository.upsert_supply_goods(supply_item_id=int(item_row["id"]), goods=wb_g)
                            goods = repository.get_supply_goods(user_id=owner_id, supply_id=sid)
                except Exception as ex:
                    _log.warning("combined-ttn lazy goods sid=%d: %s", sid, ex)
            for g in goods:
                vc = str(g.get("vendor_code") or "")
                g["product_name"] = name_map.get(vc) or vc or ""
            all_goods.extend(goods)
        if not ref_item:
            raise HTTPException(status_code=404, detail="Поставки не найдены")

        supplier_short = str(ref_item.get("supplier_name") or "").strip()
        le = (
            next((e for e in entities if str(e.get("short_name") or "").strip() == supplier_short), None)
            or next((e for e in entities if str(e.get("short_name") or "").strip().lower() == supplier_short.lower()), None)
            or next((e for e in entities if str(e.get("full_name") or "").strip() == supplier_short), None)
            or {}
        )
        org_line = ", ".join(filter(None, [le.get("full_name") or supplier_short, le.get("requisites") or ""]))
        driver_name = str(ref_item.get("driver_name") or "")

        now = _dtt.now()
        doc_num = f"{now.strftime('%d%m%Y')}_1"
        date_disp = now.strftime("%d.%m.%Y")

        # Fetch prices
        nm_prices: dict[int,float] = {}
        try:
            src = repository.get_supply_source_with_key(user_id=owner_id, source_id=int(ref_item["source_id"]))
            if src and src.get("api_key"):
                ctx = _sl.create_default_context()
                offset = 0
                while True:
                    url2 = f"https://discounts-prices-api.wildberries.ru/api/v2/list/goods/filter?limit=1000&offset={offset}"
                    req2 = _ul.Request(url2,method="GET",headers={"Authorization":str(src["api_key"]),"User-Agent":"Mozilla/5.0"})
                    with _ul.urlopen(req2,context=ctx,timeout=15) as resp:
                        pg=_jm.loads(resp.read()).get("data",{}).get("listGoods",[])
                    for g in pg:
                        nm=int(g.get("nmID") or 0); sz=g.get("sizes") or []
                        dp=float(sz[0].get("discountedPrice",0)) if sz else 0.0
                        if nm and dp>0: nm_prices[nm]=dp
                    offset+=len(pg)
                    if len(pg)<1000: break
        except Exception as ex:
            _log.warning("combined-ttn prices: %s", ex)

        VAT=0.22
        def fmt2(x): return f"{x:,.2f}".replace(",", " ").replace(".", ",")

        rows_data=[]
        total_excl=total_vat=total_incl=0.0
        qty_total=sum(int(g.get("quantity") or 0) for g in all_goods)
        for i,g in enumerate(all_goods):
            qty=int(g.get("quantity") or 0); nm=int(g.get("nm_id") or 0)
            pi=nm_prices.get(nm); pe=pi/(1+VAT) if pi else None
            ae=pe*qty if pe is not None else None; va=ae*VAT if ae is not None else None; ai=ae+va if ae is not None else None
            if ae is not None: total_excl+=ae; total_vat+=va; total_incl+=ai
            rows_data.append({"num":i+1,"name":g.get("product_name") or g.get("vendor_code") or "Товар","qty":qty,
                "price_excl":fmt2(pe) if pe is not None else "—","amt_excl":fmt2(ae) if ae is not None else "—",
                "vat_amt":fmt2(va) if va is not None else "—","amt_incl":fmt2(ai) if ai is not None else "—"})
        t_excl=fmt2(total_excl) if total_excl else "—"
        t_vat=fmt2(total_vat) if total_vat else "—"
        t_incl=fmt2(total_incl) if total_incl else "—"
        amt_words=""

        def _rubles_words(n):
            if n<=0: return "ноль рублей 00 копеек"
            ones_f=["","одна","две","три","четыре","пять","шесть","семь","восемь","девять"]
            tens=["","","двадцать","тридцать","сорок","пятьдесят","шестьдесят","семьдесят","восемьдесят","девяносто"]
            hunds=["","сто","двести","триста","четыреста","пятьсот","шестьсот","семьсот","восемьсот","девятьсот"]
            teens=["десять","одиннадцать","двенадцать","тринадцать","четырнадцать","пятнадцать","шестнадцать","семнадцать","восемнадцать","девятнадцать"]
            def chunk(x,fem=False):
                of=["","один","два","три","четыре","пять","шесть","семь","восемь","девять"]
                r,w=x%100,[]
                if x//100: w.append(hunds[x//100])
                if 10<=r<=19: w.append(teens[r-10])
                else:
                    if r//10: w.append(tens[r//10])
                    if r%10: w.append((ones_f if fem else of)[r%10])
                return w
            w=[]; th=(n//1000)%1000; ru=n%1000
            if th:
                w.extend(chunk(th,True))
                w.append(["тысяч","тысяча","тысячи","тысяч"][1 if th%10==1 and th%100!=11 else 2 if th%10 in(2,3,4) and th%100 not in range(12,15) else 3])
            if ru: w.extend(chunk(ru,False))
            rw=["рублей","рубль","рубля","рублей"][1 if ru%10==1 and n%100!=11 else 2 if ru%10 in(2,3,4) and n%100 not in range(12,15) else 3]
            return " ".join(w+[rw,"00 копеек"])
        if total_incl: amt_words=_rubles_words(round(total_incl))

        tpl_path=STATIC_DIR/"torg12_tpl.docx"
        with open(tpl_path,"rb") as f: tpl_bytes=f.read()
        with _zf.ZipFile(_io.BytesIO(tpl_bytes)) as zin:
            all_files={name:zin.read(name) for name in zin.namelist()}
        doc_xml=all_files["word/document.xml"].decode("utf-8")
        row_rx=_re.compile(r'(<w:tr[\s>](?:(?!</w:tr>).)*?\{\{GOODS_NAME\}\}.*?</w:tr>)',_re.DOTALL)
        m=row_rx.search(doc_xml)
        import html as _htmlm
        if m and rows_data:
            rt=m.group(1); mul=""
            for rd in rows_data:
                r=rt.replace("{{ROW_NUM}}",str(rd["num"])).replace("{{GOODS_NAME}}",_htmlm.escape(rd["name"])).replace("{{PRICE}}",_htmlm.escape(rd["price_excl"])).replace("{{ROW_AMOUNT_EXCL}}",_htmlm.escape(rd["amt_excl"])).replace("{{ROW_VAT_SUM}}",_htmlm.escape(rd["vat_amt"])).replace("{{ROW_AMOUNT_INCL}}",_htmlm.escape(rd["amt_incl"])).replace("{{ROW_QTY}}",str(rd["qty"]))
                mul+=r
            doc_xml=doc_xml.replace(rt,mul,1)
        supply_ids_str = ", ".join(str(x) for x in supply_ids)
        # Recipient: ООО «РВБ» + transit (initial) warehouse address from first supply
        dest_wh = str((ref_item or {}).get("warehouse_name") or "").strip()
        transit_wh = str((ref_item or {}).get("transit_warehouse_name") or "").strip()
        pickup_wh = transit_wh or dest_wh
        warehouses = repository.list_supply_warehouses(user_id=owner_id)
        wh_addr = next(
            (repository.warehouse_address_line(w) for w in warehouses
             if str(w.get("warehouse_name") or "").strip() == pickup_wh),
            "",
        )
        recipient_line = "ООО «РВБ»" + (f", {wh_addr}" if wh_addr else "")
        for ph,val in [("{{TTN_NUMBER}}",doc_num),("{{ORG_FULL}}",org_line),("{{SUPPLIER}}",org_line),("{{PAYER}}",org_line),("{{RECIPIENT}}",recipient_line),("{{ORDER_DATE}}",supply_ids_str),("{{DOC_NUM_VAL}}",doc_num),("{{DOC_DATE_VAL}}",date_disp),("{{GOODS_NAME}}",rows_data[0]["name"] if rows_data else "Товар"),("{{ROW_NUM}}","1"),("{{PRICE}}",rows_data[0]["price_excl"] if rows_data else "—"),("{{ROW_AMOUNT_EXCL}}",rows_data[0]["amt_excl"] if rows_data else "—"),("{{ROW_VAT_SUM}}",rows_data[0]["vat_amt"] if rows_data else "—"),("{{ROW_AMOUNT_INCL}}",rows_data[0]["amt_incl"] if rows_data else "—"),("{{QTY}}",str(qty_total)),("{{QTY_SHT}}",f"{qty_total} шт"),("{{TOTAL_EXCL}}",t_excl),("{{TOTAL_VAT}}",t_vat),("{{TOTAL_INCL}}",t_incl),("{{AMOUNT}}",t_excl),("{{VAT_SUM}}",t_vat),("{{AMOUNT_WITH_VAT}}",t_incl),("{{TOTAL_RUB}}",str(int(total_incl)) if total_incl else "0"),("{{TOTAL_KOP}}","00"),("{{PAGES_COUNT}}","1"),("{{ITEMS_COUNT}}",str(len(rows_data))),("{{SUPPLY_ID}}",doc_num),("{{DOC_DATE_FULL}}",f"«{now.strftime('%d')}» {['января','февраля','марта','апреля','мая','июня','июля','августа','сентября','октября','ноября','декабря'][now.month-1]} {now.year}"),("{{ISSUED_BY}}",supplier_short or "—"),("{{SIGNATORIES}}",le.get("signatories") or supplier_short or "—"),("{{PROD_HEAD}}",le.get("signatories") or supplier_short or "—"),("{{SIGN_SUPPLIER}}",supplier_short),("{{SIGN_DRIVER}}",driver_name),("{{AMOUNT_WORDS}}",amt_words)]:
            doc_xml=doc_xml.replace(ph,val)
        doc_xml=doc_xml.replace("{{ROW_QTY}}",str(qty_total))
        all_files["word/document.xml"]=doc_xml.encode("utf-8")
        tmp_dir=_tf.mkdtemp(); docx_path=_pl.Path(tmp_dir)/f"combined_ttn.docx"; pdf_path=_pl.Path(tmp_dir)/f"combined_ttn.pdf"
        buf=_io.BytesIO()
        with _zf.ZipFile(buf,"w",_zf.ZIP_DEFLATED) as zout:
            for name,data in all_files.items(): zout.writestr(name,data)
        docx_path.write_bytes(buf.getvalue())
        lo_env=dict(_os.environ)
        for k,v in [("HOME",tmp_dir),("XDG_CACHE_HOME",tmp_dir),("XDG_CONFIG_HOME",tmp_dir),("XDG_RUNTIME_DIR",tmp_dir),("DCONF_PROFILE","/dev/null")]: lo_env[k]=v
        lo_ok=False
        for binary in ("/usr/bin/soffice","/usr/lib/libreoffice/program/soffice","soffice","libreoffice"):
            try:
                r=_sp.run([binary,"--headless","--norestore",f"-env:UserInstallation=file://{tmp_dir}/lo_profile","--convert-to","pdf","--outdir",tmp_dir,str(docx_path)],capture_output=True,timeout=60,env=lo_env)
                if r.returncode==0 and pdf_path.exists(): lo_ok=True; break
            except FileNotFoundError: continue
            except _sp.TimeoutExpired: raise HTTPException(status_code=504,detail="Таймаут")
        if not lo_ok: raise HTTPException(status_code=500,detail="Ошибка конвертации PDF")
        return Response(content=pdf_path.read_bytes(),media_type="application/pdf",headers={"Content-Disposition":'inline; filename="combined_ttn.pdf"'})

    @app.get("/api/supplies/{supply_id}/ttn-error-test")
    def ttn_error_test(request: Request, supply_id: int) -> dict[str, object]:
        """Debug: run ttn.pdf and return error detail instead of 500."""
        import subprocess as _sp, tempfile as _tf, zipfile as _zf, io as _io
        import re as _re, pathlib as _pl, traceback as _tb
        from fastapi.responses import FileResponse
        try:
            return get_ttn_pdf(request, supply_id)
        except HTTPException as e:
            return {"ok": False, "status": e.status_code, "detail": e.detail}
        except Exception as ex:
            return {"ok": False, "error": str(ex), "trace": _tb.format_exc()[-2000:]}

    @app.delete("/api/supplies")
    def clear_supplies(request: Request) -> dict[str, object]:
        user = _require_user(request)
        if str(user.get("role") or "") not in ROLE_CAN_ACCESS_SETTINGS:
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        deleted = repository.clear_supply_items(user_id=owner_id)
        return {"ok": True, "deleted": deleted}

    @app.get("/api/supply-drivers")
    def list_supply_drivers(request: Request) -> list[dict[str, object]]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        return repository.list_supply_drivers(user_id=owner_id)

    @app.post("/api/supply-drivers")
    def create_supply_driver(request: Request, payload: CreateSupplyDriverRequest) -> dict[str, object]:
        user = _require_user(request)
        # Drivers: accessible to owners AND managers with can_supplies
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        fio = repository._normalize_driver_fio_fields(
            last_name=payload.last_name,
            first_name=payload.first_name,
            middle_name=payload.middle_name,
            full_name=payload.full_name,
        )
        name = fio["full_name"]
        if not name:
            raise HTTPException(status_code=400, detail="Укажите фамилию или ФИО")
        # Always save under owner's user_id so drivers are shared across team
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        if repository.driver_exists(user_id=owner_id, full_name=name):
            raise HTTPException(status_code=409, detail=f"Водитель «{name}» уже существует")
        return repository.create_supply_driver(
            user_id=owner_id,
            full_name=name,
            last_name=fio["last_name"],
            first_name=fio["first_name"],
            middle_name=fio["middle_name"],
            phone=payload.phone,
            documents=payload.documents,
            in_person=payload.in_person,
            vehicles=payload.vehicles,
            carrier=payload.carrier,
            carrier_name=payload.carrier_name,
            carrier_inn=payload.carrier_inn,
            carrier_kpp=payload.carrier_kpp,
            carrier_phone=payload.carrier_phone,
            carrier_fns_id=payload.carrier_fns_id,
            carrier_addr_index=payload.carrier_addr_index,
            carrier_addr_region_code=payload.carrier_addr_region_code,
            carrier_addr_district=payload.carrier_addr_district,
            carrier_addr_city=payload.carrier_addr_city,
            carrier_addr_settlement=payload.carrier_addr_settlement,
            carrier_addr_street=payload.carrier_addr_street,
            carrier_addr_house=payload.carrier_addr_house,
            carrier_addr_corpus=payload.carrier_addr_corpus,
            carrier_addr_flat=payload.carrier_addr_flat,
            carrier_addr_fias=payload.carrier_addr_fias,
            doc_vu_series=payload.doc_vu_series,
            doc_vu_number=payload.doc_vu_number,
            doc_vu_issuer=payload.doc_vu_issuer,
            doc_vu_date=payload.doc_vu_date,
            doc_inn_fl=payload.doc_inn_fl,
        )

    @app.patch("/api/supply-drivers/{driver_id}")
    def update_supply_driver_endpoint(request: Request, driver_id: int, payload: UpdateSupplyDriverRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        fio = repository._normalize_driver_fio_fields(
            last_name=payload.last_name,
            first_name=payload.first_name,
            middle_name=payload.middle_name,
            full_name=payload.full_name,
        )
        name = fio["full_name"]
        if not name:
            raise HTTPException(status_code=400, detail="Укажите фамилию или ФИО")
        ok = repository.update_supply_driver(
            user_id=_supply_owner_id(user),
            driver_id=driver_id,
            full_name=name,
            last_name=fio["last_name"],
            first_name=fio["first_name"],
            middle_name=fio["middle_name"],
            phone=payload.phone,
            documents=payload.documents,
            in_person=payload.in_person,
            vehicles=payload.vehicles,
            carrier=payload.carrier,
            carrier_name=payload.carrier_name,
            carrier_inn=payload.carrier_inn,
            carrier_kpp=payload.carrier_kpp,
            carrier_phone=payload.carrier_phone,
            carrier_fns_id=payload.carrier_fns_id,
            carrier_addr_index=payload.carrier_addr_index,
            carrier_addr_region_code=payload.carrier_addr_region_code,
            carrier_addr_district=payload.carrier_addr_district,
            carrier_addr_city=payload.carrier_addr_city,
            carrier_addr_settlement=payload.carrier_addr_settlement,
            carrier_addr_street=payload.carrier_addr_street,
            carrier_addr_house=payload.carrier_addr_house,
            carrier_addr_corpus=payload.carrier_addr_corpus,
            carrier_addr_flat=payload.carrier_addr_flat,
            carrier_addr_fias=payload.carrier_addr_fias,
            doc_vu_series=payload.doc_vu_series,
            doc_vu_number=payload.doc_vu_number,
            doc_vu_issuer=payload.doc_vu_issuer,
            doc_vu_date=payload.doc_vu_date,
            doc_inn_fl=payload.doc_inn_fl,
        )
        if not ok:
            raise HTTPException(status_code=404, detail="Водитель не найден")
        return {"ok": True}

    @app.delete("/api/supply-drivers/{driver_id}")
    def delete_supply_driver(request: Request, driver_id: int) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        ok = repository.delete_supply_driver(user_id=owner_id, driver_id=driver_id)
        if not ok:
            raise HTTPException(status_code=404, detail="Водитель не найден")
        return {"ok": True}

    @app.get("/api/supply-warehouses")
    def list_supply_warehouses(request: Request) -> list[dict[str, object]]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        return repository.list_supply_warehouses(user_id=owner_id)

    @app.post("/api/supply-warehouses")
    def create_supply_warehouse(request: Request, payload: CreateSupplyWarehouseRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        name = payload.warehouse_name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Название склада не может быть пустым")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        existing = repository.list_supply_warehouses(user_id=owner_id)
        if any(str(w.get("warehouse_name") or "").strip().lower() == name.lower() for w in existing):
            raise HTTPException(status_code=400, detail=f"Склад «{name}» уже существует")
        try:
            return repository.create_supply_warehouse(
                user_id=owner_id,
                warehouse_name=name,
                address=payload.address.strip(),
                addr_index=payload.addr_index,
                addr_region_code=payload.addr_region_code,
                addr_district=payload.addr_district,
                addr_city=payload.addr_city,
                addr_settlement=payload.addr_settlement,
                addr_street=payload.addr_street,
                addr_house=payload.addr_house,
                addr_corpus=payload.addr_corpus,
                addr_flat=payload.addr_flat,
            )
        except Exception as ex:
            msg = str(ex).lower()
            if "unique" in msg or "duplicate" in msg:
                raise HTTPException(status_code=400, detail=f"Склад «{name}» уже существует") from ex
            raise

    @app.patch("/api/supply-warehouses/{warehouse_id}")
    def update_supply_warehouse_endpoint(request: Request, warehouse_id: int, payload: UpdateSupplyWarehouseRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        name = payload.warehouse_name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Название не может быть пустым")
        ok = repository.update_supply_warehouse(
            user_id=_supply_owner_id(user),
            warehouse_id=warehouse_id,
            warehouse_name=name,
            address=payload.address.strip(),
            addr_index=payload.addr_index,
            addr_region_code=payload.addr_region_code,
            addr_district=payload.addr_district,
            addr_city=payload.addr_city,
            addr_settlement=payload.addr_settlement,
            addr_street=payload.addr_street,
            addr_house=payload.addr_house,
            addr_corpus=payload.addr_corpus,
            addr_flat=payload.addr_flat,
        )
        if not ok:
            raise HTTPException(status_code=404, detail="Склад не найден")
        return {"ok": True}

    @app.delete("/api/supply-warehouses/{warehouse_id}")
    def delete_supply_warehouse(request: Request, warehouse_id: int) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        ok = repository.delete_supply_warehouse(user_id=_supply_owner_id(user), warehouse_id=warehouse_id)
        if not ok:
            raise HTTPException(status_code=404, detail="Склад не найден")
        return {"ok": True}

    @app.get("/api/supply-legal-entities")
    def list_supply_legal_entities(request: Request) -> list[dict[str, object]]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        return repository.list_supply_legal_entities(user_id=owner_id)

    @app.post("/api/supply-legal-entities")
    def create_supply_legal_entity(request: Request, payload: CreateSupplyLegalEntityRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        name = payload.short_name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Короткое наименование не может быть пустым")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        return repository.create_supply_legal_entity(
            user_id=owner_id,
            short_name=name,
            full_name=payload.full_name.strip(),
            requisites=payload.requisites,
            signatories=payload.signatories,
            in_person=payload.in_person,
            basis=payload.basis,
            address=payload.address,
            phone=payload.phone,
            addr_index=payload.addr_index,
            addr_region_code=payload.addr_region_code,
            addr_district=payload.addr_district,
            addr_city=payload.addr_city,
            addr_settlement=payload.addr_settlement,
            addr_street=payload.addr_street,
            addr_house=payload.addr_house,
            addr_corpus=payload.addr_corpus,
            addr_flat=payload.addr_flat,
            addr_fias=payload.addr_fias,
            signature_image=payload.signature_image,
        )

    @app.patch("/api/supply-legal-entities/{entity_id}")
    def update_supply_legal_entity_endpoint(request: Request, entity_id: int, payload: UpdateSupplyLegalEntityRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        name = payload.short_name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Короткое наименование не может быть пустым")
        repository._ensure_supply_tables()
        ok = repository.update_supply_legal_entity(
            user_id=_supply_owner_id(user),
            entity_id=entity_id,
            short_name=name,
            full_name=payload.full_name.strip(),
            requisites=payload.requisites,
            signatories=payload.signatories,
            in_person=payload.in_person,
            basis=payload.basis,
            address=payload.address,
            phone=payload.phone,
            addr_index=payload.addr_index,
            addr_region_code=payload.addr_region_code,
            addr_district=payload.addr_district,
            addr_city=payload.addr_city,
            addr_settlement=payload.addr_settlement,
            addr_street=payload.addr_street,
            addr_house=payload.addr_house,
            addr_corpus=payload.addr_corpus,
            addr_flat=payload.addr_flat,
            addr_fias=payload.addr_fias,
            signature_image=payload.signature_image,
            clear_signature=payload.clear_signature,
        )
        if not ok:
            raise HTTPException(status_code=404, detail="Юридическое лицо не найдено")
        return {"ok": True}

    @app.get("/api/supply-legal-entities/{entity_id}/signature")
    def get_legal_entity_signature(request: Request, entity_id: int) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        img = repository.get_legal_entity_signature(user_id=_supply_owner_id(user), entity_id=entity_id)
        return {"signature_image": img}

    @app.delete("/api/supply-legal-entities/{entity_id}")
    def delete_supply_legal_entity(request: Request, entity_id: int) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        ok = repository.delete_supply_legal_entity(user_id=_supply_owner_id(user), entity_id=entity_id)
        if not ok:
            raise HTTPException(status_code=404, detail="Юридическое лицо не найдено")
        return {"ok": True}

    @app.get("/api/supply-productions")
    def list_supply_productions(request: Request) -> list[dict[str, object]]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        repository._ensure_supply_tables()
        return repository.list_supply_productions(user_id=_supply_owner_id(user))

    @app.post("/api/supply-productions")
    def create_supply_production(request: Request, payload: CreateSupplyProductionRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        if not payload.name.strip():
            raise HTTPException(status_code=400, detail="Название не может быть пустым")
        repository._ensure_supply_tables()
        return repository.create_supply_production(
            user_id=_supply_owner_id(user),
            name=payload.name,
            head_name=payload.head_name,
            address=payload.address,
            load_contact=payload.load_contact,
            addr_index=payload.addr_index,
            addr_region_code=payload.addr_region_code,
            addr_district=payload.addr_district,
            addr_city=payload.addr_city,
            addr_settlement=payload.addr_settlement,
            addr_street=payload.addr_street,
            addr_house=payload.addr_house,
            addr_corpus=payload.addr_corpus,
            addr_flat=payload.addr_flat,
        )

    @app.patch("/api/supply-productions/{production_id}")
    def update_supply_production(request: Request, production_id: int, payload: UpdateSupplyProductionRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        if not payload.name.strip():
            raise HTTPException(status_code=400, detail="Название не может быть пустым")
        repository._ensure_supply_tables()
        ok = repository.update_supply_production(
            user_id=_supply_owner_id(user),
            production_id=production_id,
            name=payload.name,
            head_name=payload.head_name,
            address=payload.address,
            load_contact=payload.load_contact,
            addr_index=payload.addr_index,
            addr_region_code=payload.addr_region_code,
            addr_district=payload.addr_district,
            addr_city=payload.addr_city,
            addr_settlement=payload.addr_settlement,
            addr_street=payload.addr_street,
            addr_house=payload.addr_house,
            addr_corpus=payload.addr_corpus,
            addr_flat=payload.addr_flat,
        )
        if not ok:
            raise HTTPException(status_code=404, detail="Производство не найдено")
        return {"ok": True}

    @app.delete("/api/supply-productions/{production_id}")
    def delete_supply_production(request: Request, production_id: int) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        ok = repository.delete_supply_production(
            user_id=_supply_owner_id(user), production_id=production_id
        )
        if not ok:
            raise HTTPException(status_code=404, detail="Производство не найдено")
        return {"ok": True}

    # ── Stock ledger (Поставки → Остатки) ────────────────────────────────────

    def _stock_productions_for_user(user: dict[str, object]) -> list[dict[str, object]]:
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()
        repository.ensure_supply_balances_tables()
        productions = repository.list_supply_productions(user_id=owner_id)
        allowed = _allowed_stock_production_ids(user)
        if allowed is not None:
            allowed_set = {int(x) for x in allowed}
            productions = [
                p for p in productions if int(p.get("id") or 0) in allowed_set
            ]
        return [
            {"id": int(p.get("id") or 0), "name": str(p.get("name") or "")}
            for p in productions
            if int(p.get("id") or 0) > 0
        ]

    def _default_stock_production_id(user: dict[str, object]) -> int | None:
        prods = _stock_productions_for_user(user)
        if not prods:
            return None
        return int(prods[0]["id"])

    def _parse_stock_date(raw: str, *, today: str) -> str:
        s = str(raw or "").strip()
        if not s:
            return today
        try:
            from datetime import date as _date

            return _date.fromisoformat(s).isoformat()
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Некорректная дата") from exc

    def _normalize_stock_line_items(
        raw_items: list[dict[str, object]],
    ) -> list[dict[str, object]]:
        out: list[dict[str, object]] = []
        for item in raw_items or []:
            item_type = str(item.get("item_type") or "").strip().lower()
            if item_type not in {"material", "product"}:
                continue
            try:
                item_id = int(item.get("item_id") or 0)
                qty = float(item.get("qty") if "qty" in item else item.get("quantity"))
            except (TypeError, ValueError):
                continue
            if item_id <= 0:
                continue
            row: dict[str, object] = {
                "item_type": item_type,
                "item_id": item_id,
                "qty": qty,
            }
            cmt = str(item.get("comment") or "").strip()
            if cmt:
                row["comment"] = cmt[:500]
            out.append(row)
        return out

    @app.get("/api/supply-balances/meta")
    def supply_balances_meta(request: Request) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supply_stock(user):
            raise HTTPException(status_code=403, detail="Нет доступа к остаткам")
        productions = _stock_productions_for_user(user)
        return {
            "today": _moscow_today(),
            "productions": productions,
            "production_id": int(productions[0]["id"]) if productions else None,
            # Matrix cells are read-only; mutations go through receipt/adjustment.
            "can_edit": False,
            "ledger": True,
        }

    @app.get("/api/supply-balances")
    def get_supply_balances(
        request: Request,
        production_id: int = 0,
        as_of: str = "",
        history: int = 0,
    ) -> dict[str, object]:
        """Read-only ledger snapshot. ``as_of`` = Moscow date (default today).

        By default returns a single date column (``as_of``). Pass ``history=1``
        to include all movement dates ≤ ``as_of`` as extra columns.
        """
        user = _require_user(request)
        if not _can_view_supply_stock(user):
            raise HTTPException(status_code=403, detail="Нет доступа к остаткам")
        owner_id = _supply_owner_id(user)
        productions = _stock_productions_for_user(user)
        today = _moscow_today()
        include_history = int(history or 0) != 0
        if not productions:
            return {
                "today": today,
                "as_of": today,
                "production_id": None,
                "dates": [today],
                "history": include_history,
                "rows": [],
                "productions": [],
                "can_edit": False,
                "ledger": True,
            }
        prod_ids = [int(p["id"]) for p in productions]
        pid = int(production_id or 0)
        if pid <= 0 or pid not in prod_ids:
            pid = prod_ids[0]
        as_of_raw = str(as_of or "").strip()
        if as_of_raw:
            as_of_date = _parse_stock_date(as_of_raw, today=today)
        else:
            as_of_date = today
        # Idempotent import of legacy editable snapshots into the ledger.
        try:
            repository.migrate_legacy_supply_balances_to_movements(
                user_id=owner_id,
                production_id=pid,
                created_by=int(user.get("id") or 0) or None,
            )
        except Exception:
            pass
        if include_history:
            move_dates = repository.list_supply_stock_movement_dates(
                user_id=owner_id, production_id=pid, as_of=as_of_date
            )
            dates = [d for d in move_dates if d <= as_of_date]
            if as_of_date not in dates:
                dates = dates + [as_of_date]
        else:
            # Default UX: one column for the current slice (today or as_of).
            dates = [as_of_date]
        # Precompute cumulative balances for each column date.
        bal_by_date: dict[str, dict[tuple[str, int], float]] = {
            d: repository.sum_supply_stock_balances(
                user_id=owner_id, production_id=pid, as_of=d
            )
            for d in dates
        }
        vis_rows = repository.list_supply_balance_visibility(user_id=owner_id)
        vis_map = {
            (str(v.get("item_type") or ""), int(v.get("item_id") or 0)): bool(
                v.get("visible", True)
            )
            for v in vis_rows
        }
        sort_map = {
            (str(v.get("item_type") or ""), int(v.get("item_id") or 0)): int(
                v.get("sort_order") or 0
            )
            for v in vis_rows
        }
        min_map: dict[tuple[str, int], float | None] = {
            (str(v.get("item_type") or ""), int(v.get("item_id") or 0)): (
                repository._parse_supply_balance_min_qty(v.get("min_qty"))
            )
            for v in vis_rows
        }

        def _stock_balance_num(raw: object) -> float:
            if raw is None or raw == "":
                return 0.0
            try:
                val = float(raw)  # type: ignore[arg-type]
            except (TypeError, ValueError):
                return 0.0
            if val != val:  # NaN
                return 0.0
            return val

        def _row_min_fields(
            item_type: str, item_id: int, balance: object
        ) -> dict[str, object]:
            min_qty = min_map.get((item_type, item_id))
            below = (
                min_qty is not None and _stock_balance_num(balance) < float(min_qty)
            )
            return {"min_qty": min_qty, "below_min": below}

        materials = repository.list_feedback_materials(user_id=owner_id)
        products = repository.list_product_photos(user_id=owner_id)
        fbs_barcodes = repository.get_wb_fbs_barcodes_by_product_id(user_id=owner_id)
        material_rows: list[dict[str, object]] = []
        product_rows: list[dict[str, object]] = []
        for m in materials:
            mid = int(m.get("id") or 0)
            if mid <= 0:
                continue
            if not vis_map.get(("material", mid), True):
                continue
            values = {
                d: bal_by_date[d].get(("material", mid))
                for d in dates
            }
            balance = bal_by_date[as_of_date].get(("material", mid))
            row: dict[str, object] = {
                "item_type": "material",
                "item_id": mid,
                "name": str(m.get("name") or ""),
                "unit": str(m.get("unit") or "шт"),
                "values": values,
                "balance": balance,
                "sort_order": sort_map.get(("material", mid), 10**9),
            }
            row.update(_row_min_fields("material", mid, balance))
            material_rows.append(row)
        for p in products:
            pid_item = int(p.get("id") or 0)
            if pid_item <= 0:
                continue
            if not vis_map.get(("product", pid_item), True):
                continue
            values = {
                d: bal_by_date[d].get(("product", pid_item))
                for d in dates
            }
            article = str(p.get("supplier_article") or "").strip()
            wb_nmid = str(p.get("wb_nmid") or "").strip()
            ozon_sku = str(p.get("ozon_sku") or "").strip()
            # ШК as in поставки: from FBS order skus; ozon_sku as extra if present.
            # Do not put seller article into barcodes — it is shown as «Арт.» separately.
            barcodes: list[str] = []
            for b in fbs_barcodes.get(pid_item) or []:
                text = str(b or "").strip()
                if text and text != article and text not in barcodes:
                    barcodes.append(text)
            if ozon_sku and ozon_sku != article and ozon_sku not in barcodes:
                barcodes.append(ozon_sku)
            balance = bal_by_date[as_of_date].get(("product", pid_item))
            row = {
                "item_type": "product",
                "item_id": pid_item,
                "name": str(p.get("name") or ""),
                "unit": "шт",
                "supplier_article": article,
                "wb_nmid": wb_nmid,
                "ozon_sku": ozon_sku,
                "photo_url": (
                    f"/api/products/photo/{pid_item}"
                    if p.get("photo_path")
                    else None
                ),
                "barcodes": barcodes,
                "values": values,
                "balance": balance,
                "sort_order": sort_map.get(("product", pid_item), 10**9),
            }
            row.update(_row_min_fields("product", pid_item, balance))
            product_rows.append(row)
        material_rows.sort(
            key=lambda r: repository.supply_balance_item_sort_key(
                item_type="material",
                item_id=int(r["item_id"]),
                name=str(r.get("name") or ""),
                sort_map=sort_map,
            )
        )
        product_rows.sort(
            key=lambda r: repository.supply_balance_item_sort_key(
                item_type="product",
                item_id=int(r["item_id"]),
                name=str(r.get("name") or ""),
                sort_map=sort_map,
            )
        )
        rows = material_rows + product_rows
        return {
            "today": today,
            "as_of": as_of_date,
            "production_id": pid,
            "dates": dates,
            "history": include_history,
            "rows": rows,
            "productions": productions,
            "can_edit": False,
            "ledger": True,
        }

    @app.post("/api/supply-balances/receipt")
    def post_supply_stock_receipt(
        request: Request, payload: SupplyStockReceiptRequest
    ) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supply_stock(user):
            raise HTTPException(status_code=403, detail="Нет доступа к остаткам")
        owner_id = _supply_owner_id(user)
        pid = _default_stock_production_id(user)
        if not pid:
            raise HTTPException(
                status_code=400,
                detail="Добавьте производство в Поставки → Настройки → Производства",
            )
        today = _moscow_today()
        date_s = _parse_stock_date(payload.date, today=today)
        lines = _normalize_stock_line_items(list(payload.items or []))
        lines = [x for x in lines if float(x["qty"]) > 0]
        if not lines:
            raise HTTPException(status_code=400, detail="Добавьте позиции с количеством")
        import uuid as _uuid

        items = [
            {
                **line,
                "source_id": f"receipt:{date_s}:{line['item_type']}:{line['item_id']}:{_uuid.uuid4().hex[:10]}",
            }
            for line in lines
        ]
        saved = repository.add_supply_stock_movements(
            user_id=owner_id,
            production_id=pid,
            movement_date=date_s,
            kind="receipt",
            source_type="manual_receipt",
            items=items,
            comment=str(payload.comment or "").strip(),
            created_by=int(user.get("id") or 0) or None,
        )
        return {"ok": True, "saved": saved, "date": date_s, "production_id": pid}

    @app.post("/api/supply-balances/adjustment")
    def post_supply_stock_adjustment(
        request: Request, payload: SupplyStockAdjustmentRequest
    ) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supply_stock(user):
            raise HTTPException(status_code=403, detail="Нет доступа к остаткам")
        owner_id = _supply_owner_id(user)
        pid = _default_stock_production_id(user)
        if not pid:
            raise HTTPException(
                status_code=400,
                detail="Добавьте производство в Поставки → Настройки → Производства",
            )
        mode = str(payload.mode or "adjustment").strip().lower()
        if mode not in {"opening", "adjustment"}:
            raise HTTPException(status_code=400, detail="Некорректный тип корректировки")
        qty_mode = str(payload.quantity_mode or "absolute").strip().lower()
        if qty_mode not in {"absolute", "delta"}:
            qty_mode = "absolute"
        today = _moscow_today()
        date_s = _parse_stock_date(payload.date, today=today)
        lines = _normalize_stock_line_items(list(payload.items or []))
        if not lines:
            raise HTTPException(status_code=400, detail="Добавьте позиции")
        current = repository.sum_supply_stock_balances(
            user_id=owner_id, production_id=pid, as_of=date_s
        )
        import uuid as _uuid

        items: list[dict[str, object]] = []
        for line in lines:
            key = (str(line["item_type"]), int(line["item_id"]))
            raw_qty = float(line["qty"])
            if qty_mode == "absolute":
                delta = raw_qty - float(current.get(key) or 0)
            else:
                delta = raw_qty
            if delta == 0:
                continue
            row_out: dict[str, object] = {
                "item_type": key[0],
                "item_id": key[1],
                "qty": delta,
                "source_id": (
                    f"{mode}:{date_s}:{key[0]}:{key[1]}:{_uuid.uuid4().hex[:10]}"
                ),
            }
            if line.get("comment"):
                row_out["comment"] = str(line.get("comment") or "")
            items.append(row_out)
        if not items:
            # Still freeze current FBS deliveries: user confirmed the on-hand figure.
            settled = 0
            try:
                settled = repository.settle_open_wb_fbs_orders_for_stock(
                    user_id=owner_id,
                    production_id=pid,
                    reason=mode,
                )
            except Exception:
                settled = 0
            return {
                "ok": True,
                "saved": 0,
                "date": date_s,
                "production_id": pid,
                "message": "Изменений нет",
                "fbs_settled": settled,
            }
        saved = repository.add_supply_stock_movements(
            user_id=owner_id,
            production_id=pid,
            movement_date=date_s,
            kind=mode,
            source_type=f"manual_{mode}",
            items=items,
            comment=str(payload.comment or "").strip(),
            created_by=int(user.get("id") or 0) or None,
        )
        # Physical count after correction already includes open deliveries —
        # freeze those FBS order ids so sync will not deduct them again.
        settled = 0
        try:
            settled = repository.settle_open_wb_fbs_orders_for_stock(
                user_id=owner_id,
                production_id=pid,
                reason=mode,
            )
        except Exception:
            settled = 0
        return {
            "ok": True,
            "saved": saved,
            "date": date_s,
            "production_id": pid,
            "fbs_settled": settled,
        }

    @app.put("/api/supply-balances")
    def save_supply_balances(
        request: Request, payload: SupplyBalanceSaveRequest
    ) -> dict[str, object]:
        """Deprecated: matrix cell edit. Prefer receipt/adjustment ledger APIs."""
        user = _require_user(request)
        if not _can_view_supply_stock(user):
            raise HTTPException(status_code=403, detail="Нет доступа к остаткам")
        raise HTTPException(
            status_code=410,
            detail="Редактирование ячеек отключено. Используйте «Добавить на склад» или «Корректировка».",
        )

    @app.get("/api/supply-balances/visibility")
    def get_supply_balance_visibility(request: Request) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supply_stock(user):
            raise HTTPException(status_code=403, detail="Нет доступа к остаткам")
        owner_id = _supply_owner_id(user)
        repository.ensure_supply_balances_tables()
        vis_rows = repository.list_supply_balance_visibility(user_id=owner_id)
        vis_map = {
            (str(v.get("item_type") or ""), int(v.get("item_id") or 0)): bool(
                v.get("visible", True)
            )
            for v in vis_rows
        }
        sort_map = {
            (str(v.get("item_type") or ""), int(v.get("item_id") or 0)): int(
                v.get("sort_order") or 0
            )
            for v in vis_rows
        }
        min_map = {
            (str(v.get("item_type") or ""), int(v.get("item_id") or 0)): (
                repository._parse_supply_balance_min_qty(v.get("min_qty"))
            )
            for v in vis_rows
        }
        materials = repository.list_feedback_materials(user_id=owner_id)
        products = repository.list_product_photos(user_id=owner_id)
        items: list[dict[str, object]] = []
        material_items: list[dict[str, object]] = []
        product_items: list[dict[str, object]] = []
        for m in materials:
            mid = int(m.get("id") or 0)
            if mid <= 0:
                continue
            material_items.append(
                {
                    "item_type": "material",
                    "item_id": mid,
                    "name": str(m.get("name") or ""),
                    "unit": str(m.get("unit") or "шт"),
                    "visible": vis_map.get(("material", mid), True),
                    "sort_order": sort_map.get(("material", mid), 10**9),
                    "min_qty": min_map.get(("material", mid)),
                }
            )
        for p in products:
            pid_item = int(p.get("id") or 0)
            if pid_item <= 0:
                continue
            article = str(p.get("supplier_article") or "").strip()
            wb_nmid = str(p.get("wb_nmid") or "").strip()
            ozon_sku = str(p.get("ozon_sku") or "").strip()
            product_items.append(
                {
                    "item_type": "product",
                    "item_id": pid_item,
                    "name": str(p.get("name") or ""),
                    "unit": "шт",
                    "supplier_article": article,
                    "wb_nmid": wb_nmid,
                    "ozon_sku": ozon_sku,
                    "photo_url": (
                        f"/api/products/photo/{pid_item}"
                        if p.get("photo_path")
                        else None
                    ),
                    "barcodes": [x for x in (article, ozon_sku) if x],
                    "visible": vis_map.get(("product", pid_item), True),
                    "sort_order": sort_map.get(("product", pid_item), 10**9),
                    "min_qty": min_map.get(("product", pid_item)),
                }
            )
        material_items.sort(
            key=lambda r: repository.supply_balance_item_sort_key(
                item_type="material",
                item_id=int(r["item_id"]),
                name=str(r.get("name") or ""),
                sort_map=sort_map,
            )
        )
        product_items.sort(
            key=lambda r: repository.supply_balance_item_sort_key(
                item_type="product",
                item_id=int(r["item_id"]),
                name=str(r.get("name") or ""),
                sort_map=sort_map,
            )
        )
        # Normalize sequential sort_order for the editor (materials then products).
        for idx, row in enumerate(material_items):
            row["sort_order"] = idx
        for idx, row in enumerate(product_items):
            row["sort_order"] = idx
        items = material_items + product_items
        return {"items": items}

    @app.put("/api/supply-balances/visibility")
    def save_supply_balance_visibility(
        request: Request, payload: SupplyBalanceVisibilityRequest
    ) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supply_stock(user):
            raise HTTPException(status_code=403, detail="Нет доступа к остаткам")
        # Catalog visibility is account-level; managers with stock access may toggle.
        owner_id = _supply_owner_id(user)
        saved = repository.set_supply_balance_visibility(
            user_id=owner_id, items=list(payload.items or [])
        )
        return {"ok": True, "saved": saved}

    @app.get("/api/supply-balances/movements")
    def get_supply_balance_movements(
        request: Request,
        production_id: int = 0,
        item_type: str = "",
        item_id: int = 0,
        limit: int = 100,
    ) -> dict[str, object]:
        """Read-only ledger journal for one Остатки row."""
        user = _require_user(request)
        if not _can_view_supply_stock(user):
            raise HTTPException(status_code=403, detail="Нет доступа к остаткам")
        owner_id = _supply_owner_id(user)
        itype = str(item_type or "").strip().lower()
        if itype not in {"material", "product"}:
            raise HTTPException(status_code=400, detail="Некорректный тип позиции")
        try:
            iid = int(item_id or 0)
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail="Некорректный id позиции") from exc
        if iid <= 0:
            raise HTTPException(status_code=400, detail="Некорректный id позиции")
        prods = _stock_productions_for_user(user)
        if not prods:
            raise HTTPException(
                status_code=400,
                detail="Добавьте производство в Поставки → Настройки → Производства",
            )
        prod_ids = {int(p["id"]) for p in prods}
        pid = int(production_id or 0)
        if pid <= 0 or pid not in prod_ids:
            pid = int(prods[0]["id"])
        try:
            lim = int(limit or 100)
        except (TypeError, ValueError):
            lim = 100
        lim = max(1, min(lim, 200))
        name = ""
        unit = "шт"
        found = False
        if itype == "material":
            for m in repository.list_feedback_materials(user_id=owner_id):
                if int(m.get("id") or 0) == iid:
                    name = str(m.get("name") or "").strip() or f"Материал #{iid}"
                    unit = str(m.get("unit") or "шт")
                    found = True
                    break
        else:
            for p in repository.list_product_photos(user_id=owner_id):
                if int(p.get("id") or 0) == iid:
                    name = str(p.get("name") or "").strip() or f"Товар #{iid}"
                    found = True
                    break
        if not found:
            raise HTTPException(status_code=404, detail="Позиция не найдена")
        bal_map = repository.sum_supply_stock_balances(
            user_id=owner_id, production_id=pid, as_of=_moscow_today()
        )
        balance = bal_map.get((itype, iid))
        rows = repository.list_supply_stock_movements_for_item(
            user_id=owner_id,
            production_id=pid,
            item_type=itype,
            item_id=iid,
            limit=lim,
        )
        kind_labels = {
            "opening": "Начальный остаток",
            "receipt": "Приход",
            "fbs_ship": "Списание FBS",
            "adjustment": "Корректировка",
            "fbs_reverse": "Возврат FBS",
        }
        creator_ids = {
            int(r["created_by"])
            for r in rows
            if r.get("created_by") not in (None, "")
        }
        creator_names: dict[int, str] = {}
        for uid in creator_ids:
            u = repository.get_user_by_id(uid)
            if not u:
                continue
            label = str(u.get("full_name") or "").strip() or str(u.get("email") or "").strip()
            if label:
                creator_names[uid] = label

        def _fbs_order_ref(source_type: str, source_id: str) -> str:
            st = str(source_type or "")
            sid = str(source_id or "").strip()
            if not sid:
                return ""
            if st not in {"wb_fbs_order", "wb_fbs_order_reverse"} and not sid[:1].isdigit():
                return ""
            # source_id forms: "123", "123:s:1", "123:r:1"
            head = sid.split(":", 1)[0].strip()
            if head.isdigit():
                return f"Заказ #{head}"
            return ""

        items: list[dict[str, object]] = []
        for r in rows:
            kind = str(r.get("kind") or "")
            created_by = r.get("created_by")
            try:
                created_by_i = int(created_by) if created_by not in (None, "") else None
            except (TypeError, ValueError):
                created_by_i = None
            source_type = str(r.get("source_type") or "")
            source_id = str(r.get("source_id") or "")
            order_ref = _fbs_order_ref(source_type, source_id)
            comment = str(r.get("comment") or "").strip()
            if order_ref and order_ref not in comment:
                comment = f"{comment} · {order_ref}".strip(" ·") if comment else order_ref
            items.append(
                {
                    "id": int(r.get("id") or 0),
                    "movement_date": str(r.get("movement_date") or ""),
                    "kind": kind,
                    "kind_label": kind_labels.get(kind, kind or "Движение"),
                    "qty": r.get("qty"),
                    "comment": comment,
                    "source_type": source_type,
                    "source_id": source_id,
                    "created_at": str(r.get("created_at") or ""),
                    "created_by": created_by_i,
                    "created_by_name": (
                        creator_names.get(created_by_i)
                        if created_by_i is not None
                        else ("Система" if kind.startswith("fbs_") else "")
                    ),
                }
            )
        return {
            "production_id": pid,
            "item_type": itype,
            "item_id": iid,
            "name": name,
            "unit": unit,
            "balance": balance,
            "items": items,
        }

    @app.get("/api/supply-contractors")
    def list_supply_contractors(request: Request) -> list[dict[str, object]]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        repository._ensure_supply_tables()
        return repository.list_supply_contractors(user_id=_supply_owner_id(user))

    @app.post("/api/supply-contractors")
    def create_supply_contractor(request: Request, payload: CreateSupplyContractorRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        if not payload.name.strip():
            raise HTTPException(status_code=400, detail="Название не может быть пустым")
        return repository.create_supply_contractor(
            user_id=_supply_owner_id(user), name=payload.name, requisites=payload.requisites
        )

    @app.patch("/api/supply-contractors/{contractor_id}")
    def update_supply_contractor_ep(request: Request, contractor_id: int, payload: UpdateSupplyContractorRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        if not payload.name.strip():
            raise HTTPException(status_code=400, detail="Название не может быть пустым")
        ok = repository.update_supply_contractor(
            user_id=_supply_owner_id(user), contractor_id=contractor_id,
            name=payload.name, requisites=payload.requisites
        )
        if not ok:
            raise HTTPException(status_code=404, detail="Контрагент не найден")
        return {"ok": True}

    @app.delete("/api/supply-contractors/{contractor_id}")
    def delete_supply_contractor_ep(request: Request, contractor_id: int) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        ok = repository.delete_supply_contractor(
            user_id=_supply_owner_id(user), contractor_id=contractor_id
        )
        if not ok:
            raise HTTPException(status_code=404, detail="Контрагент не найден")
        return {"ok": True}

    # ── Supply PoA Records ────────────────────────────────────────────────────

    # ── Certificates ──────────────────────────────────────────────────────────

    @app.get("/api/certificates")
    def list_certificates(request: Request) -> list[dict[str, object]]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        repository._ensure_supply_tables()
        return repository.list_certificates(user_id=_supply_owner_id(user))

    @app.post("/api/certificates")
    def create_certificate(request: Request, body: CertificateCreateRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        repository._ensure_supply_tables()
        cert_id = repository.create_certificate(
            user_id=_supply_owner_id(user),
            legal_entity_short=body.legal_entity_short,
            category=body.category,
            number=body.number,
            expiry_date=body.expiry_date,
            verification_url=body.verification_url,
            image_data=body.image_data,
            doc_type=body.doc_type,
        )
        return {"ok": True, "id": cert_id}

    @app.put("/api/certificates/{cert_id}")
    def update_certificate(request: Request, cert_id: int, body: CertificateCreateRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        ok = repository.update_certificate(
            user_id=_supply_owner_id(user), cert_id=cert_id,
            legal_entity_short=body.legal_entity_short, category=body.category,
            number=body.number, expiry_date=body.expiry_date,
            verification_url=body.verification_url, image_data=body.image_data,
            doc_type=body.doc_type,
        )
        return {"ok": ok}

    @app.delete("/api/certificates/{cert_id}")
    def delete_certificate(request: Request, cert_id: int) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        # Only managers (non-feedback_manager roles) can delete
        if str(user.get("role") or "") in TENANT_MANAGER_ROLES:
            raise HTTPException(status_code=403, detail="Только менеджер может удалять сертификаты")
        ok = repository.delete_certificate(user_id=_supply_owner_id(user), cert_id=cert_id)
        return {"ok": ok}

    @app.get("/api/supply-poa-records")
    def list_poa_records(request: Request) -> list[dict[str, object]]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        repository._ensure_supply_tables()
        records = repository.list_supply_poa_records(user_id=_supply_owner_id(user))
        # Strip signature_image from list to keep response small
        for r in records:
            r.pop("le_signature_image", None)
        return records

    @app.post("/api/supply-poa-records")
    def create_poa_record(request: Request, payload: CreatePoARecordRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        from datetime import datetime as _dtt
        poa_date = _dtt.now().strftime("%d.%m.%Y")
        record = repository.create_supply_poa_record(
            user_id=_supply_owner_id(user),
            legal_entity_id=payload.legal_entity_id,
            contractor_id=payload.contractor_id,
            driver_id=payload.driver_id,
            poa_date=poa_date,
            driver_manual_name=payload.driver_manual_name,
            driver_manual_docs=payload.driver_manual_docs,
        )
        record.pop("le_signature_image", None)
        return record

    @app.patch("/api/supply-poa-records/{record_id}")
    def update_poa_record(request: Request, record_id: int, payload: UpdatePoARecordRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        ok = repository.update_supply_poa_record(
            user_id=_supply_owner_id(user), record_id=record_id,
            legal_entity_id=payload.legal_entity_id, contractor_id=payload.contractor_id,
            driver_id=payload.driver_id, driver_manual_name=payload.driver_manual_name,
            driver_manual_docs=payload.driver_manual_docs,
        )
        if not ok:
            raise HTTPException(status_code=404, detail="Доверенность не найдена")
        return {"ok": True}

    @app.delete("/api/supply-poa-records/{record_id}")
    def delete_poa_record(request: Request, record_id: int) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        ok = repository.delete_supply_poa_record(user_id=_supply_owner_id(user), record_id=record_id)
        if not ok:
            raise HTTPException(status_code=404, detail="Доверенность не найдена")
        return {"ok": True}

    def _build_poa_html(record: dict, include_signature: bool = True) -> str:
        """Build the PoA HTML document from a record dict (with joined data)."""
        import html as _hm
        e = _hm.escape
        le_full   = e(str(record.get("le_full") or ""))
        le_in_p   = e(str(record.get("le_in_person") or ""))
        le_basis  = e(str(record.get("le_basis") or ""))
        le_sig    = e(str(record.get("le_signatories") or ""))
        sig_img   = record.get("le_signature_image") or "" if include_signature else ""
        # Use manual driver if driver_id is 0/null
        driver_id = int(record.get("driver_id") or 0)
        if driver_id > 0:
            d_in_person = str(record.get("d_in_person") or "")
            d_full = str(record.get("d_full") or "")
            d_docs = str(record.get("d_docs") or "")
            # Use "В лице" if filled, otherwise fallback to ФИО + документы
            driver_str = e(d_in_person) if d_in_person else (f"{e(d_full)}, {e(d_docs)}".strip(", ") if d_docs else e(d_full))
        else:
            d_full = str(record.get("driver_manual_name") or "")
            d_docs = str(record.get("driver_manual_docs") or "")
            driver_str = f"{e(d_full)}, {e(d_docs)}".strip(", ") if d_docs else e(d_full)
        c_name    = e(str(record.get("c_name") or ""))
        c_req     = e(str(record.get("c_req") or ""))
        poa_date  = e(str(record.get("poa_date") or ""))
        contractor_str = f"{c_name} {c_req}".strip() if c_req else c_name
        sig_html = f'<img src="{sig_img}" style="max-height:25mm;max-width:60mm;object-fit:contain;vertical-align:middle" />' if sig_img else "&nbsp;" * 20

        return f"""<!DOCTYPE html>
<html xmlns:o="urn:schemas-microsoft-com:office:office" xmlns:w="urn:schemas-microsoft-com:office:word" xmlns="http://www.w3.org/TR/REC-html40">
<head><meta charset="utf-8">
<!--[if gte mso 9]><xml>
<w:WordDocument>
  <w:View>Print</w:View>
  <w:Zoom>100</w:Zoom>
  <w:DoNotOptimizeForBrowser/>
</w:WordDocument>
</xml><![endif]-->
<style>
  @page {{ size: 210mm 297mm; margin: 20mm 20mm 20mm 30mm; }}
  @page Section1 {{
    size: 210.0mm 297.0mm;
    margin: 20.0mm 20.0mm 20.0mm 30.0mm;
    mso-header-margin: 0mm;
    mso-footer-margin: 0mm;
    mso-paper-source: 0;
  }}
  div.Section1 {{ page: Section1; }}
  body {{ font-family: "Times New Roman", serif; font-size: 12pt; line-height: 1.5; color: #000; }}
  .center {{ text-align: center; }}
  .right {{ text-align: right; }}
  .small {{ font-size: 9pt; }}
  .bold {{ font-weight: bold; }}
  .sig-line {{ border-bottom: 1px solid #000; display: inline-block; width: 200pt; vertical-align: bottom; }}
</style></head>
<body><div class="Section1">
<p style="text-align:center;font-weight:bold;font-size:14pt;margin:0 0 4pt">ДОВЕРЕННОСТЬ №Б/Н</p>
<table width="100%" cellspacing="0" cellpadding="0" style="margin:0 0 16pt">
  <tr>
    <td style="text-align:left;font-size:12pt">г. Иваново</td>
    <td style="text-align:right;font-size:12pt">{poa_date}</td>
  </tr>
</table>
<p style="text-align:justify;margin-bottom:8pt">
  {le_full} в лице генерального директора {le_in_p}, действующей на основании {le_basis},
</p>
<p style="text-align:center;font-size:9pt;margin-bottom:16pt">(Доверитель)</p>
<p style="text-align:justify;margin-bottom:16pt">
  настоящей доверенностью уполномачивает гражданина РФ: <b>{driver_str}</b> получать от <b>{contractor_str}</b> товарно-материальные ценности (ткань), с правом подписания товарной накладной, УПД, акта выполненных работ и отчета агента.
</p>
<p style="margin-bottom:32pt"><b>Срок действия доверенности – 1 (один) год.</b></p>
<p style="margin-bottom:32pt">
  Образец подписи доверенного лица &nbsp;&nbsp;&nbsp;&nbsp; <span style="border-bottom:1px solid #000;display:inline-block;width:240pt;vertical-align:bottom">&nbsp;</span> &nbsp;&nbsp;&nbsp;&nbsp; удостоверяю
</p>
<table style="width:100%;border-collapse:collapse;margin-top:20pt">
  <tr>
    <td style="width:40%;vertical-align:bottom;font-size:12pt">Генеральный директор</td>
    <td style="width:35%;vertical-align:bottom;text-align:center">{sig_html}</td>
    <td style="width:25%;vertical-align:bottom;text-align:center">
      {le_sig}<br><span class="small">(расшифровка подписи)</span>
    </td>
  </tr>
</table>
</div></body></html>"""

    @app.get("/api/supply-poa-records/{record_id}/html")
    def get_poa_html(request: Request, record_id: int):
        """Return the PoA as HTML for browser print."""
        from fastapi.responses import HTMLResponse
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        records = repository.list_supply_poa_records(user_id=_supply_owner_id(user))
        record = next((r for r in records if r["id"] == record_id), None)
        if not record:
            raise HTTPException(status_code=404, detail="Не найдено")
        return HTMLResponse(content=_build_poa_html(record, include_signature=False))

    @app.get("/api/supply-poa-records/{record_id}/pdf")
    def get_poa_pdf(request: Request, record_id: int):
        """Generate PoA PDF via LibreOffice."""
        import subprocess as _sp, tempfile as _tf, pathlib as _pl, os as _os
        from fastapi.responses import Response
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        records = repository.list_supply_poa_records(user_id=_supply_owner_id(user))
        record = next((r for r in records if r["id"] == record_id), None)
        if not record:
            raise HTTPException(status_code=404, detail="Не найдено")

        html_content = _build_poa_html(record)
        tmp_dir = _tf.mkdtemp()
        html_path = _pl.Path(tmp_dir) / "poa.html"
        pdf_path  = _pl.Path(tmp_dir) / "poa.pdf"
        html_path.write_text(html_content, encoding="utf-8")
        lo_env = dict(_os.environ)
        for k, v in [("HOME",tmp_dir),("XDG_CACHE_HOME",tmp_dir),("XDG_CONFIG_HOME",tmp_dir),
                     ("XDG_RUNTIME_DIR",tmp_dir),("DCONF_PROFILE","/dev/null")]:
            lo_env[k] = v
        lo_ok = False
        for binary in ("/usr/bin/soffice","/usr/lib/libreoffice/program/soffice","soffice","libreoffice"):
            try:
                r = _sp.run([binary,"--headless","--norestore",f"-env:UserInstallation=file://{tmp_dir}/lo_profile","--convert-to","pdf","--outdir",tmp_dir,str(html_path)],capture_output=True,timeout=60,env=lo_env)
                if r.returncode == 0 and pdf_path.exists(): lo_ok = True; break
            except FileNotFoundError: continue
            except _sp.TimeoutExpired: raise HTTPException(status_code=504,detail="Таймаут")
        if not lo_ok: raise HTTPException(status_code=500,detail="Ошибка генерации PDF")
        import re as _re
        def _poa_fn(rec, ext):
            le = _re.sub(r'[/\\?%*:|"<>]', '', str(rec.get("le_short") or ""))
            cn = _re.sub(r'[/\\?%*:|"<>]', '', str(rec.get("c_name") or ""))
            dr_id = int(rec.get("driver_id") or 0)
            dr = str(rec.get("d_full") if dr_id > 0 else rec.get("driver_manual_name") or "")
            dr = _re.sub(r'[/\\?%*:|"<>]', '', dr)
            name = f"{le}_{cn}_{dr}.{ext}".strip("_")
            return name or f"POA_{record_id}.{ext}"
        fname = _poa_fn(record, "pdf")
        from urllib.parse import quote as _qp
        return Response(content=pdf_path.read_bytes(),media_type="application/pdf",headers={"Content-Disposition":f"attachment; filename*=UTF-8''{_qp(fname)}"})

    @app.get("/api/supply-poa-records/{record_id}/doc")
    def get_poa_doc(request: Request, record_id: int):
        """Return the PoA as .doc (HTML Word format) for download."""
        from fastapi.responses import Response
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        records = repository.list_supply_poa_records(user_id=_supply_owner_id(user))
        record = next((r for r in records if r["id"] == record_id), None)
        if not record:
            raise HTTPException(status_code=404, detail="Не найдено")
        import re as _re2
        def _poa_fn2(rec, ext):
            le = _re2.sub(r'[/\\?%*:|"<>]', '', str(rec.get("le_short") or ""))
            cn = _re2.sub(r'[/\\?%*:|"<>]', '', str(rec.get("c_name") or ""))
            dr_id = int(rec.get("driver_id") or 0)
            dr = str(rec.get("d_full") if dr_id > 0 else rec.get("driver_manual_name") or "")
            dr = _re2.sub(r'[/\\?%*:|"<>]', '', dr)
            name = f"{le}_{cn}_{dr}.{ext}".strip("_")
            return name or f"POA_{record_id}.{ext}"
        html_content = "\uFEFF" + _build_poa_html(record, include_signature=False)
        fname_doc = _poa_fn2(record, "doc")
        from urllib.parse import quote as _qp2
        return Response(content=html_content.encode("utf-8"),media_type="application/msword",headers={"Content-Disposition":f"attachment; filename*=UTF-8''{_qp2(fname_doc)}"})

    @app.patch("/api/supplies/{supply_id}/manual-fields")
    def update_supply_manual_fields(
        request: Request,
        supply_id: int,
        payload: SupplyManualFieldsRequest,
    ) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        ok = repository.update_supply_manual_fields(
            user_id=owner_id,
            supply_id=supply_id,
            pass_number=payload.pass_number,
            pallets_count=payload.pallets_count,
            driver_name=payload.driver_name,
            notes=payload.notes,
            production=payload.production,
            drivers_json=payload.drivers_json,
        )
        if not ok:
            raise HTTPException(status_code=404, detail="Поставка не найдена")
        return {"ok": True}

    @app.get("/api/supplies/sync/status")
    def get_supply_sync_status(request: Request) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        with supply_sync_lock:
            return dict(supply_sync_state)

    @app.post("/api/supplies/sync")
    def sync_supplies(request: Request, payload: SyncSuppliesRequest) -> dict[str, object]:
        user = _require_user(request)
        if not _can_view_supplies(user):
            raise HTTPException(status_code=403, detail="Нет доступа")
        owner_id = _supply_owner_id(user)
        repository._ensure_supply_tables()

        with supply_sync_lock:
            if supply_sync_state.get("in_progress"):
                # Reset if stuck (thread died without cleanup)
                started = str(supply_sync_state.get("started_at") or "")
                if started:
                    from datetime import datetime as _dt2, timezone as _tz2
                    try:
                        age = (_dt2.now(_tz2.utc) - _dt2.fromisoformat(started.replace("Z", "+00:00"))).seconds
                        if age < 600:  # allow reset after 10 min
                            return {"ok": False, "message": "Синхронизация уже запущена"}
                    except Exception:
                        pass
                supply_sync_state["in_progress"] = False

        sources = repository.list_supply_sources(user_id=owner_id)
        if payload.source_id:
            sources = [s for s in sources if s["id"] == payload.source_id]
        # WB sync: only process WB sources (marketplace == 'wb' or null/empty)
        active_sources = [s for s in sources if s.get("is_enabled")
                          and (s.get("marketplace") or "wb").lower() == "wb"]
        if not active_sources:
            return {"ok": True, "synced": 0, "message": "Нет активных источников"}

        import urllib.request as _urllib
        import json as _json_mod
        import ssl as _ssl
        import time as _time
        from datetime import datetime as _dt, timezone as _tz, timedelta as _td

        def _wb_request(method: str, url: str, api_key: str, body: dict | None = None):
            data = _json_mod.dumps(body).encode() if body else None
            headers = {
                "Authorization": api_key,
                "Content-Type": "application/json",
                "User-Agent": "FeedPilot/1.0",
            }
            req = _urllib.Request(url, data=data, headers=headers, method=method)
            ctx = _ssl.create_default_context()
            # Retry up to 3 times with backoff: handles 429, 503, network errors
            for attempt in range(3):
                try:
                    with _urllib.urlopen(req, timeout=30, context=ctx) as r:
                        return r.status, _json_mod.loads(r.read() or b"{}")
                except Exception as e:
                    code = getattr(e, "code", None)
                    if code in (429, 503):
                        # Rate limited or service unavailable — wait and retry
                        wait = (attempt + 1) * 2  # 2s, 4s, 6s
                        _log.warning("WB supplies API %d, retry %d in %ds", code, attempt + 1, wait)
                        _time.sleep(wait)
                        continue
                    return (int(code) if code else 0), {}
            return 0, {}

        def _wb_post(url: str, api_key: str, body: dict):
            s, d = _wb_request("POST", url, api_key, body)
            return s, d if isinstance(d, list) else []

        def _run_sync():
            _now = _dt.now(_tz.utc)
            date_from = (_now - _td(days=30)).strftime("%Y-%m-%d")
            date_to = (_now + _td(days=1)).strftime("%Y-%m-%d")
            active_statuses = {1, 2, 3, 4, 5, 6}  # 3=Отгрузка разрешена, 6=Отгружено на воротах
            total_synced = 0
            errors: list[str] = []

            with supply_sync_lock:
                supply_sync_state.update({
                    "in_progress": True, "page": 0, "synced": 0, "total": 0,
                    "errors": [], "message": "Запуск…",
                    "started_at": _dt.now(_tz.utc).isoformat(), "finished_at": None,
                })

            try:
                for src in active_sources:
                    src_full = repository.get_supply_source_with_key(
                        user_id=owner_id, source_id=int(src["id"])
                    )
                    if not src_full:
                        continue
                    api_key = str(src_full.get("api_key") or "")
                    if not api_key:
                        errors.append(f"Источник «{src['name']}»: нет API-ключа")
                        continue
                    source_id = int(src["id"])
                    synced_this_source = 0
                    with supply_sync_lock:
                        supply_sync_state["message"] = f"«{src['name']}»: загрузка списка…"
                        supply_sync_state["page"] = 1
                    try:
                        # WB API ignores dateFrom/dateTo — always returns ALL supplies.
                        # Fetch once (API also ignores pagination) and filter client-side.
                        http_status, items = _wb_post(
                            "https://supplies-api.wildberries.ru/api/v1/supplies",
                            api_key,
                            {"dateFrom": "2020-01-01", "dateTo": "2099-12-31",
                             "status": "ALL", "page": 1, "pageSize": 1000},
                        )
                        if http_status == 401:
                            errors.append(f"«{src['name']}»: неверный API-ключ")
                        elif isinstance(items, list):
                            # Collect ALL supply IDs returned by WB for this source
                            # (used later to purge cancelled/removed supplies from DB)
                            all_wb_ids = [int(x.get("supplyID") or 0) for x in items if x.get("supplyID")]

                            # Client-side filter: supplyDate >= last 30 days AND active status
                            items = [
                                x for x in items
                                if (x.get("supplyDate") or "")[:10] >= date_from
                                and int(x.get("statusID") or 0) in active_statuses
                            ]
                            with supply_sync_lock:
                                supply_sync_state["total"] = (supply_sync_state.get("total") or 0) + len(items)
                                supply_sync_state["message"] = (
                                    f"«{src['name']}»: найдено {len(items)} поставок, загрузка деталей…"
                                )
                            item_errors = 0
                            for item in items:
                                supply_wb_id = int(item.get("supplyID") or 0)
                                if not supply_wb_id:
                                    continue
                                status_id = int(item.get("statusID") or 0)
                                if status_id not in active_statuses:
                                    continue
                                try:
                                    # For active (1,2,4): always fetch details
                                    # For accepted (5): fetch only if not already cached
                                    need_details = status_id in {1, 2, 3, 4}  # 5,6 — fetch only if missing
                                    if not need_details:
                                        existing = repository.get_supply_item_row(
                                            user_id=owner_id, supply_id=supply_wb_id
                                        )
                                        need_details = not (existing and existing.get("warehouse_name"))
                                    if need_details:
                                        det_status, det_data = _wb_request(
                                            "GET",
                                            f"https://supplies-api.wildberries.ru/api/v1/supplies/{supply_wb_id}",
                                            api_key,
                                        )
                                        if det_status == 200 and isinstance(det_data, dict):
                                            item.update({k: v for k, v in det_data.items() if v is not None})
                                        _time.sleep(2.0)  # WB Supplies API: interval ≥ 2s per docs
                                    item["supplyID"] = supply_wb_id
                                    repository.upsert_supply_item(source_id=source_id, data=item)
                                    synced_this_source += 1
                                    with supply_sync_lock:
                                        supply_sync_state["synced"] = total_synced + synced_this_source
                                except Exception as item_exc:
                                    item_errors += 1
                                    err_msg = f"{type(item_exc).__name__}: {item_exc}"
                                    _log.error("supply upsert error supply_id=%s: %s", supply_wb_id, err_msg, exc_info=True)
                                    if item_errors == 1:
                                        # Show first error in status
                                        with supply_sync_lock:
                                            supply_sync_state["message"] = f"Ошибка поставки {supply_wb_id}: {err_msg}"
                                        errors.append(f"Поставка {supply_wb_id}: {err_msg}")

                            # Remove cancelled/removed supplies from DB:
                            # WB returned all_wb_ids — anything in DB but NOT in that list
                            # was cancelled or deleted on WB side.
                            if all_wb_ids:
                                deleted_count = repository.delete_supply_items_not_in(
                                    source_id=source_id, keep_supply_ids=all_wb_ids
                                )
                                if deleted_count:
                                    _log.info("supply sync: removed %d cancelled/missing supplies for source %d", deleted_count, source_id)

                            repository.mark_supply_source_synced(source_id=source_id)
                            total_synced += synced_this_source
                            # Restore manually-entered fields (pass, pallets, driver)
                            # that survived a previous clear_supply_items call.
                            try:
                                repository.restore_supply_manual_fields(user_id=owner_id)
                            except Exception:
                                pass
                    except Exception as exc:
                        _log.error("supply sync source %d: %s", source_id, exc, exc_info=True)
                        err_msg = f"{type(exc).__name__}: {exc}"
                        errors.append(f"«{src['name']}»: {err_msg}")
                        with supply_sync_lock:
                            supply_sync_state["message"] = f"Ошибка: {err_msg}"
            finally:
                with supply_sync_lock:
                    supply_sync_state.update({
                        "in_progress": False,
                        "synced": total_synced,
                        "errors": errors,
                        "message": f"Готово. Загружено {total_synced} поставок." + (
                            f" Ошибки: {'; '.join(errors)}" if errors else ""
                        ),
                        "finished_at": _dt.now(_tz.utc).isoformat(),
                    })

        t = threading.Thread(target=_run_sync, daemon=True)
        t.start()
        return {"ok": True, "started": True, "message": "Синхронизация запущена"}

    # ── End supply module endpoints ───────────────────────────────────────────

    @app.on_event("shutdown")
    def stop_auto_sync_worker() -> None:
        auto_sync_stop_event.set()
        stock_scheduler.stop()
        wb_fbs_scheduler.stop()
        worker = auto_sync_worker.get("thread")
        if isinstance(worker, threading.Thread) and worker.is_alive():
            worker.join(timeout=1.5)

    @app.post("/api/admin/reviews-clear")
    def admin_clear_reviews(request: Request, payload: ClearReviewsRequest) -> dict[str, object]:
        actor = _require_admin(request)
        if payload.user_id is None:
            target_user_id = _tenant_owner_id(actor) if not _is_super_admin(actor) else int(actor["id"])
        else:
            target_user_id = int(payload.user_id)
            _target_user_for_admin_scope(actor=actor, target_user_id=target_user_id)
        deleted = repository.clear_reviews(user_id=target_user_id)
        return {"ok": True, "deleted": deleted, "user_id": target_user_id}

    @app.post("/api/admin/conversations-clear")
    def admin_clear_conversations_v2(request: Request, payload: ClearConversationsRequest) -> dict[str, object]:
        actor = _require_admin(request)
        if payload.user_id is None:
            target_user_id = _tenant_owner_id(actor) if not _is_super_admin(actor) else int(actor["id"])
        else:
            target_user_id = int(payload.user_id)
            _target_user_for_admin_scope(actor=actor, target_user_id=target_user_id)
        deleted = repository.clear_conversations(user_id=target_user_id, kind=payload.kind, source=payload.source)
        return {"ok": True, "deleted": deleted, "user_id": target_user_id}

    @app.exception_handler(HTTPException)
    def http_exception_handler(_: Request, exc: HTTPException) -> JSONResponse:
        return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})

    return app


app = create_app()

def _render_template(name: str, context: dict[str, str] | None = None) -> str:
    template_path = TEMPLATES_DIR / name
    html = template_path.read_text(encoding="utf-8")
    for key, value in (context or {}).items():
        html = html.replace(f"{{{{{key}}}}}", value)
    return html


def build_landing_html() -> str:
    return _render_template("landing.html")


def build_login_html(error: str | None = None) -> str:
    error_html = f"<p class='error'>{escape(error)}</p>" if error else ""
    return _render_template("login.html", {"ERROR_HTML": error_html})


def build_register_html(error: str | None = None) -> str:
    error_html = f"<p class='error'>{escape(error)}</p>" if error else ""
    return _render_template("register.html", {"ERROR_HTML": error_html})


def build_wb_fbs_tsd_html(user: dict[str, object], repository=None) -> str:
    """Standalone ТСД page HTML (warehouse handheld). Isolated from app.html/app.js."""
    safe_email = escape(str(user["email"]))
    role = str(user.get("role") or ROLE_USER)
    user_id = int(user.get("id") or 0)
    owner_user_id = int(user.get("owner_user_id") or user_id or 0)
    is_tenant_owner = (
        role in ROLE_CAN_ACCESS_SETTINGS
        and user_id > 0
        and owner_user_id == user_id
    )
    can_view = role in ROLE_CAN_ACCESS_SETTINGS
    if not can_view and bool(user.get("can_supplies")) and repository is not None:
        perms = repository.get_manager_supply_permissions(manager_user_id=user_id)
        sources = perms.get("sources") or {}
        can_view = any(
            bool(v.get("wb_fbs_tsd"))
            for v in sources.values()
            if isinstance(v, dict)
        )
    return _render_template(
        "wb_fbs_tsd.html",
        {
            "SAFE_EMAIL": safe_email,
            "CAN_VIEW_WB_FBS_TSD": "true" if can_view else "false",
            "IS_TENANT_OWNER": "true" if is_tenant_owner else "false",
        },
    )


def build_app_html(user: dict[str, object], repository=None) -> str:
    safe_email = escape(str(user["email"]))
    role = str(user.get("role") or ROLE_USER)
    is_super_admin = bool(user.get("is_super_admin"))
    user_id = int(user.get("id") or 0)
    owner_user_id = int(user.get("owner_user_id") or user_id or 0)
    is_tenant_owner = (
        role in ROLE_CAN_ACCESS_SETTINGS
        and user_id > 0
        and owner_user_id == user_id
    )
    role_labels = {
        ROLE_ADMIN: "администратор",
        ROLE_USER: "пользователь",
        ROLE_FEEDBACK_MANAGER: "менеджер обратной связи",
        "production_manager": "начальник производства",
        "manager": "менеджер",
    }
    safe_role = escape(role_labels.get(role, role))
    can_view_analytics = role in ROLE_CAN_ACCESS_ANALYTICS
    can_view_settings = role in ROLE_CAN_ACCESS_SETTINGS
    can_view_supplies = (
        role in ROLE_CAN_ACCESS_SETTINGS
        or bool(user.get("can_supplies"))
    )
    # Granular supply permissions for managers
    _supply_perms: dict = {}
    if can_view_supplies and role not in ROLE_CAN_ACCESS_SETTINGS and repository is not None:
        _supply_perms = repository.get_manager_supply_permissions(manager_user_id=user_id)
    _sp_sources = _supply_perms.get("sources") or {}
    can_view_wb_supplies = (
        role in ROLE_CAN_ACCESS_SETTINGS
        or any(v.get("wb") for v in _sp_sources.values() if isinstance(v, dict))
    )
    can_view_wb_fbs_supplies = (
        role in ROLE_CAN_ACCESS_SETTINGS
        or any(v.get("wb_fbs") for v in _sp_sources.values() if isinstance(v, dict))
    )
    can_view_wb_fbs_tsd = (
        role in ROLE_CAN_ACCESS_SETTINGS
        or any(v.get("wb_fbs_tsd") for v in _sp_sources.values() if isinstance(v, dict))
    )
    can_view_ozon_supplies = (
        role in ROLE_CAN_ACCESS_SETTINGS
        or any(v.get("ozon") for v in _sp_sources.values() if isinstance(v, dict))
    )
    can_view_supply_poa = (
        role in ROLE_CAN_ACCESS_SETTINGS
        or bool(_supply_perms.get("can_supply_poa"))
    )
    can_view_supply_settings = (
        role in ROLE_CAN_ACCESS_SETTINGS
        or bool(_supply_perms.get("can_supply_settings"))
    )
    can_view_supply_certs = (
        role in ROLE_CAN_ACCESS_SETTINGS
        or bool(_supply_perms.get("can_supply_certs"))
    )
    can_supply_planning = is_tenant_owner or bool(user.get("can_supply_planning"))
    can_supply_stock = is_tenant_owner or bool(user.get("can_supply_stock"))
    import json as _json_stock
    if is_tenant_owner:
        _stock_prods_js = "null"
    else:
        try:
            _stock_list = _json_stock.loads(str(user.get("stock_productions") or "[]"))
            _stock_prods_js = _json_stock.dumps(
                _stock_list if isinstance(_stock_list, list) else []
            )
        except Exception:
            _stock_prods_js = "[]"
    can_view_any_supply = (
        can_view_wb_supplies
        or can_view_wb_fbs_supplies
        or can_view_wb_fbs_tsd
        or can_view_ozon_supplies
        or can_view_supply_poa
        or can_view_supply_settings
        or can_view_supply_certs
        or can_supply_planning
        or can_supply_stock
    )
    chats_sync_on = sync_chats_enabled()
    if role in ROLE_CAN_ACCESS_SETTINGS:
        can_view_feedback = True
        can_view_reviews = True
        can_view_questions = True
        can_view_chats = True
    elif repository is not None:
        _perms = repository.list_manager_permissions(manager_user_id=user_id)
        can_view_reviews = any(bool(p.get("can_reviews")) for p in _perms)
        can_view_questions = any(bool(p.get("can_questions")) for p in _perms)
        can_view_chats = any(bool(p.get("can_chats")) for p in _perms)
        can_view_feedback = can_view_reviews or can_view_questions or can_view_chats
    else:
        can_view_feedback = True
        can_view_reviews = True
        can_view_questions = True
        can_view_chats = True
    # Hide chats UI while chat marketplace sync is globally disabled.
    if not chats_sync_on:
        can_view_chats = False
        can_view_feedback = can_view_reviews or can_view_questions or can_view_chats
    can_view_salary = is_tenant_owner or bool(user.get("can_salary"))
    can_view_salary_settings = is_tenant_owner or bool(user.get("can_salary_settings"))
    can_view_salary_report = is_tenant_owner or bool(user.get("can_salary_report"))
    can_view_salary_zp_export = is_tenant_owner or bool(user.get("can_salary_zp_export"))
    import json as _json_salary
    _raw_prods = user.get("salary_productions")
    if is_tenant_owner:
        _salary_prods_js = "null"
    else:
        try:
            _prods_list = _json_salary.loads(str(_raw_prods or "[]"))
            _salary_prods_js = _json_salary.dumps(_prods_list if isinstance(_prods_list, list) else [])
        except Exception:
            _salary_prods_js = "[]"
    admin_link = '<a class="navbtn nav-admin" href="/admin"><span class="nav-item-icon">○</span> Админ-панель</a>' if role == ROLE_ADMIN else ""
    nav_team = (
        '<a id="nav-team" class="nav-item nav-item-bottom" href="#" onclick="showSection(\'team\')"><span class="nav-item-icon">◫</span> Команда</a>'
        if is_tenant_owner
        else ""
    )
    nav_analytics = (
        '<a id="nav-analytics" class="nav-item" href="#" onclick="showSection(\'analytics\')"><span class="nav-item-icon">∑</span> Аналитика</a>'
        if can_view_analytics
        else ""
    )
    nav_settings = (
        '<a id="nav-settings" class="navbtn" href="#" onclick="showSection(\'settings\')">Настройки</a>'
        if can_view_settings
        else ""
    )
    _planning_link = ('<a id="nav-supply-planning" class="nav-item" href="#" onclick="showSection(\'supply-planning\')"><span class="nav-item-icon">▤</span> Планирование</a>'
                      if can_supply_planning and can_view_supplies else "")
    _stock_link = ('<a id="nav-supplies-balances" class="nav-item" href="#" onclick="showSection(\'supplies-balances\')"><span class="nav-item-icon">▤</span> Остатки</a>'
                   if can_supply_stock else "")
    _wb_link = ('<a id="nav-supplies-wb" class="nav-item" href="#" onclick="showSection(\'supplies-wb\')"><span class="nav-item-icon">▦</span> ВБ</a>'
                if can_view_wb_supplies else "")
    _wb_fbs_link = ('<a id="nav-supplies-wb-fbs" class="nav-item" href="#" onclick="showSection(\'supplies-wb-fbs\')"><span class="nav-item-icon">▣</span> ВБ ФБС</a>'
                    if can_view_wb_fbs_supplies else "")
    # Managers with only ТСД (no desktop ВБ ФБС) get a direct link — avoid broken FBS section.
    _tsd_link = (
        '<a id="nav-wb-fbs-tsd" class="nav-item" href="/wb-fbs/tsd"><span class="nav-item-icon">▣</span> ТСД</a>'
        if (can_view_wb_fbs_tsd and not can_view_wb_fbs_supplies)
        else ""
    )
    _ozon_link = ('<a id="nav-supplies-ozon" class="nav-item" href="#" onclick="showSection(\'supplies-ozon\')"><span class="nav-item-icon">◉</span> ОЗОН</a>'
                  if can_view_ozon_supplies else "")
    _poa_link = ('<a id="nav-supplies-poa" class="nav-item" href="#" onclick="showSection(\'supplies-poa\')"><span class="nav-item-icon">☐</span> Доверенности</a>'
                 if can_view_supply_poa else "")
    _certs_link = ('<a id="nav-supplies-certificates" class="nav-item" href="#" onclick="showSection(\'supplies-certificates\')"><span class="nav-item-icon">✦</span> Сертификаты</a>'
                   if can_view_supply_certs else "")
    nav_supplies_wb = (
        (_wb_link + _wb_fbs_link + _tsd_link + _ozon_link + _stock_link + _poa_link + _certs_link)
        if can_view_supplies or can_supply_stock or can_view_wb_fbs_tsd
        else ""
    )
    nav_supplies_settings = (
        '<a id="nav-supplies-settings" class="nav-item" href="#" onclick="showSection(\'supplies-settings\')"><span class="nav-item-icon">≡</span> Настройки</a>'
        if (can_view_settings or can_view_supply_settings) else ""
    )
    return _render_template(
        "app.html",
        {
            "SAFE_EMAIL": safe_email,
            "SAFE_ROLE": safe_role,
            "ADMIN_LINK": admin_link,
            "NAV_TEAM": nav_team,
            "NAV_ANALYTICS": nav_analytics,
            "NAV_SETTINGS": nav_settings,
            "NAV_SETTINGS_SUB": (
                '<a id="nav-settings" class="nav-item" href="#" onclick="showSection(\'settings\')"><span class="nav-item-icon">≡</span> Настройки</a>'
                if can_view_settings else ""
            ),
            "NAV_SUPPLIES_WB": nav_supplies_wb,
            "NAV_SUPPLIES_PLANNING": _planning_link,
            "NAV_SUPPLIES_SETTINGS": nav_supplies_settings,
            "CAN_VIEW_ANY_SUPPLY": "true" if can_view_any_supply else "false",
            "CAN_VIEW_ANALYTICS": "true" if can_view_analytics else "false",
            "CAN_VIEW_SETTINGS": "true" if can_view_settings else "false",
            "CAN_VIEW_SUPPLIES": "true" if can_view_supplies else "false",
            "CAN_VIEW_WB_SUPPLIES": "true" if can_view_wb_supplies else "false",
            "CAN_VIEW_WB_FBS_SUPPLIES": "true" if can_view_wb_fbs_supplies else "false",
            "CAN_VIEW_WB_FBS_TSD": "true" if can_view_wb_fbs_tsd else "false",
            "CAN_VIEW_OZON_SUPPLIES": "true" if can_view_ozon_supplies else "false",
            "CAN_VIEW_FEEDBACK": "true" if can_view_feedback else "false",
            "CAN_VIEW_REVIEWS": "true" if can_view_reviews else "false",
            "CAN_VIEW_QUESTIONS": "true" if can_view_questions else "false",
            "CAN_VIEW_CHATS": "true" if can_view_chats else "false",
            "SYNC_CHATS_ENABLED": "true" if chats_sync_on else "false",
            "CAN_SUPPLY_PLANNING": "true" if can_supply_planning else "false",
            "CAN_SUPPLY_STOCK": "true" if can_supply_stock else "false",
            "CAN_STOCK_PRODUCTIONS": _stock_prods_js,
            "CAN_VIEW_SALARY": "true" if can_view_salary else "false",
            "CAN_VIEW_SALARY_SETTINGS": "true" if can_view_salary_settings else "false",
            "CAN_VIEW_SALARY_REPORT": "true" if can_view_salary_report else "false",
            "CAN_VIEW_SALARY_ZP_EXPORT": "true" if can_view_salary_zp_export else "false",
            "CAN_SALARY_PRODUCTIONS": _salary_prods_js,
            "HIDE_FEEDBACK_SECTION": "" if (can_view_feedback or can_view_settings or can_view_analytics) else "style=\"display:none\"",
            "IS_ADMIN": "true" if role == ROLE_ADMIN else "false",
            "IS_SUPER_ADMIN": "true" if is_super_admin else "false",
            "IS_TENANT_OWNER": "true" if is_tenant_owner else "false",
        },
    )


def build_admin_html(user: dict[str, object]) -> str:
    safe_email = escape(str(user["email"]))
    return _render_template("admin.html", {"SAFE_EMAIL": safe_email})
